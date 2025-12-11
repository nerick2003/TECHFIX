from __future__ import annotations

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog, colorchooser
import tkinter.font as tkfont
import sqlite3
from pathlib import Path
from datetime import datetime, timedelta, date
import calendar
from typing import Optional, Sequence, List, Dict
import json
import sys
import subprocess
import platform
import os
import logging
import smtplib
import tempfile
from email.message import EmailMessage
try:
    import winreg
except Exception:
    winreg = None

# Support running as a module (package) or as a script
try:
    if __package__:
        from . import db  # type: ignore
        from .accounting import AccountingEngine, JournalLine  # type: ignore
    else:
        raise ImportError
except Exception:
    import os, sys
    sys.path.append(os.path.dirname(os.path.dirname(__file__)))
    from techfix import db  # type: ignore
    from techfix.accounting import AccountingEngine, JournalLine  # type: ignore


logger = logging.getLogger(__name__)

THEMES = {
    "Light": {
        "app_bg": "#f5f7fb",
        "surface_bg": "#ffffff",
        "accent_color": "#2563eb",
        "accent_hover": "#1d4ed8",
        "accent_disabled": "#93c5fd",
        "text_primary": "#1f2937",
        "text_secondary": "#4b5563",
        "subtitle_fg": "#dbeafe",
        "entry_border": "#d8dee9",
        "entry_disabled_bg": "#e5e7eb",
        "table_stripe": "#eef2ff",
        "tree_heading_bg": "#e2e8f0",
        "tree_selected_bg": "#dbeafe",
        "tab_selected_bg": "#e0ecff",
        "tab_active_bg": "#e8f0ff",
    },
    "Dark": {
        "app_bg": "#111827",
        "surface_bg": "#1f2937",
        "accent_color": "#2563eb",
        "accent_hover": "#1d4ed8",
        "accent_disabled": "#1e3a8a",
        "text_primary": "#f9fafb",
        "text_secondary": "#9ca3af",
        "subtitle_fg": "#bfdbfe",
        "entry_border": "#374151",
        "entry_disabled_bg": "#4b5563",
        "table_stripe": "#1b2534",
        "tree_heading_bg": "#374151",
        "tree_selected_bg": "#1e3a8a",
        "tab_selected_bg": "#1e293b",
        "tab_active_bg": "#24324a",
    },
}

# Predefined color theme combinations
COLOR_THEMES = {
    "Green + White": {
        "app_bg": "#f0fdf4",
        "surface_bg": "#ffffff",
        "accent_color": "#10b981",
        "accent_hover": "#059669",
        "accent_disabled": "#86efac",
        "text_primary": "#064e3b",
        "text_secondary": "#047857",
        "subtitle_fg": "#d1fae5",
        "entry_border": "#a7f3d0",
        "entry_disabled_bg": "#d1fae5",
        "table_stripe": "#ecfdf5",
        "tree_heading_bg": "#d1fae5",
        "tree_selected_bg": "#a7f3d0",
        "tab_selected_bg": "#d1fae5",
        "tab_active_bg": "#ecfdf5",
    },
    "Purple + Dark": {
        "app_bg": "#1e1b2e",
        "surface_bg": "#2d1b3d",
        "accent_color": "#a855f7",
        "accent_hover": "#9333ea",
        "accent_disabled": "#6b21a8",
        "text_primary": "#f3e8ff",
        "text_secondary": "#c084fc",
        "subtitle_fg": "#d8b4fe",
        "entry_border": "#581c87",
        "entry_disabled_bg": "#3b0764",
        "table_stripe": "#1a0a2e",
        "tree_heading_bg": "#3b0764",
        "tree_selected_bg": "#6b21a8",
        "tab_selected_bg": "#581c87",
        "tab_active_bg": "#6b21a8",
    },
    "Red + Black": {
        "app_bg": "#1a0a0a",
        "surface_bg": "#2d1b1b",
        "accent_color": "#ef4444",
        "accent_hover": "#dc2626",
        "accent_disabled": "#991b1b",
        "text_primary": "#fee2e2",
        "text_secondary": "#fca5a5",
        "subtitle_fg": "#fecaca",
        "entry_border": "#7f1d1d",
        "entry_disabled_bg": "#450a0a",
        "table_stripe": "#0f0505",
        "tree_heading_bg": "#450a0a",
        "tree_selected_bg": "#991b1b",
        "tab_selected_bg": "#7f1d1d",
        "tab_active_bg": "#991b1b",
    },
    "Orange + Cream": {
        "app_bg": "#fff7ed",
        "surface_bg": "#ffffff",
        "accent_color": "#f97316",
        "accent_hover": "#ea580c",
        "accent_disabled": "#fb923c",
        "text_primary": "#7c2d12",
        "text_secondary": "#9a3412",
        "subtitle_fg": "#fed7aa",
        "entry_border": "#fdba74",
        "entry_disabled_bg": "#ffedd5",
        "table_stripe": "#fff7ed",
        "tree_heading_bg": "#ffedd5",
        "tree_selected_bg": "#fed7aa",
        "tab_selected_bg": "#ffedd5",
        "tab_active_bg": "#fff7ed",
    },
    "Teal + Gray": {
        "app_bg": "#0f172a",
        "surface_bg": "#1e293b",
        "accent_color": "#14b8a6",
        "accent_hover": "#0d9488",
        "accent_disabled": "#0f766e",
        "text_primary": "#ccfbf1",
        "text_secondary": "#5eead4",
        "subtitle_fg": "#99f6e4",
        "entry_border": "#134e4a",
        "entry_disabled_bg": "#0f766e",
        "table_stripe": "#0a1628",
        "tree_heading_bg": "#134e4a",
        "tree_selected_bg": "#0f766e",
        "tab_selected_bg": "#134e4a",
        "tab_active_bg": "#0f766e",
    },
    "Pink + White": {
        "app_bg": "#fdf2f8",
        "surface_bg": "#ffffff",
        "accent_color": "#ec4899",
        "accent_hover": "#db2777",
        "accent_disabled": "#f9a8d4",
        "text_primary": "#831843",
        "text_secondary": "#be185d",
        "subtitle_fg": "#fbcfe8",
        "entry_border": "#f9a8d4",
        "entry_disabled_bg": "#fce7f3",
        "table_stripe": "#fdf2f8",
        "tree_heading_bg": "#fce7f3",
        "tree_selected_bg": "#fbcfe8",
        "tab_selected_bg": "#fce7f3",
        "tab_active_bg": "#fdf2f8",
    },
    "Amber + Dark": {
        "app_bg": "#1c1917",
        "surface_bg": "#292524",
        "accent_color": "#f59e0b",
        "accent_hover": "#d97706",
        "accent_disabled": "#92400e",
        "text_primary": "#fef3c7",
        "text_secondary": "#fde68a",
        "subtitle_fg": "#fcd34d",
        "entry_border": "#78350f",
        "entry_disabled_bg": "#451a03",
        "table_stripe": "#0f0a05",
        "tree_heading_bg": "#451a03",
        "tree_selected_bg": "#92400e",
        "tab_selected_bg": "#78350f",
        "tab_active_bg": "#92400e",
    },
    "Cyan + Dark": {
        "app_bg": "#0c1419",
        "surface_bg": "#1a252e",
        "accent_color": "#06b6d4",
        "accent_hover": "#0891b2",
        "accent_disabled": "#155e75",
        "text_primary": "#cffafe",
        "text_secondary": "#67e8f9",
        "subtitle_fg": "#a5f3fc",
        "entry_border": "#164e63",
        "entry_disabled_bg": "#0e7490",
        "table_stripe": "#0a1217",
        "tree_heading_bg": "#164e63",
        "tree_selected_bg": "#155e75",
        "tab_selected_bg": "#164e63",
        "tab_active_bg": "#155e75",
    },
    "Indigo + Light": {
        "app_bg": "#eef2ff",
        "surface_bg": "#ffffff",
        "accent_color": "#6366f1",
        "accent_hover": "#4f46e5",
        "accent_disabled": "#c7d2fe",
        "text_primary": "#1e1b4b",
        "text_secondary": "#4338ca",
        "subtitle_fg": "#e0e7ff",
        "entry_border": "#c7d2fe",
        "entry_disabled_bg": "#eef2ff",
        "table_stripe": "#f5f7fb",
        "tree_heading_bg": "#e0e7ff",
        "tree_selected_bg": "#c7d2fe",
        "tab_selected_bg": "#e0e7ff",
        "tab_active_bg": "#eef2ff",
    },
    "Emerald + White": {
        "app_bg": "#ecfdf5",
        "surface_bg": "#ffffff",
        "accent_color": "#10b981",
        "accent_hover": "#059669",
        "accent_disabled": "#86efac",
        "text_primary": "#064e3b",
        "text_secondary": "#047857",
        "subtitle_fg": "#d1fae5",
        "entry_border": "#a7f3d0",
        "entry_disabled_bg": "#d1fae5",
        "table_stripe": "#f0fdf4",
        "tree_heading_bg": "#d1fae5",
        "tree_selected_bg": "#a7f3d0",
        "tab_selected_bg": "#d1fae5",
        "tab_active_bg": "#ecfdf5",
    },
    "Rose + Light": {
        "app_bg": "#fff1f2",
        "surface_bg": "#ffffff",
        "accent_color": "#f43f5e",
        "accent_hover": "#e11d48",
        "accent_disabled": "#fda4af",
        "text_primary": "#881337",
        "text_secondary": "#be123c",
        "subtitle_fg": "#ffe4e6",
        "entry_border": "#fda4af",
        "entry_disabled_bg": "#ffe4e6",
        "table_stripe": "#fff1f2",
        "tree_heading_bg": "#ffe4e6",
        "tree_selected_bg": "#fda4af",
        "tab_selected_bg": "#ffe4e6",
        "tab_active_bg": "#fff1f2",
    },
    "Violet + Dark": {
        "app_bg": "#1e1b2e",
        "surface_bg": "#2d1b3d",
        "accent_color": "#8b5cf6",
        "accent_hover": "#7c3aed",
        "accent_disabled": "#6b21a8",
        "text_primary": "#f3e8ff",
        "text_secondary": "#c084fc",
        "subtitle_fg": "#ddd6fe",
        "entry_border": "#581c87",
        "entry_disabled_bg": "#3b0764",
        "table_stripe": "#1a0a2e",
        "tree_heading_bg": "#3b0764",
        "tree_selected_bg": "#6b21a8",
        "tab_selected_bg": "#581c87",
        "tab_active_bg": "#6b21a8",
    },
    "Sky + Blue": {
        "app_bg": "#f0f9ff",
        "surface_bg": "#ffffff",
        "accent_color": "#0ea5e9",
        "accent_hover": "#0284c7",
        "accent_disabled": "#7dd3fc",
        "text_primary": "#0c4a6e",
        "text_secondary": "#0369a1",
        "subtitle_fg": "#bae6fd",
        "entry_border": "#7dd3fc",
        "entry_disabled_bg": "#e0f2fe",
        "table_stripe": "#f0f9ff",
        "tree_heading_bg": "#e0f2fe",
        "tree_selected_bg": "#bae6fd",
        "tab_selected_bg": "#e0f2fe",
        "tab_active_bg": "#f0f9ff",
    },
    "Slate + Gray": {
        "app_bg": "#0f172a",
        "surface_bg": "#1e293b",
        "accent_color": "#64748b",
        "accent_hover": "#475569",
        "accent_disabled": "#334155",
        "text_primary": "#f1f5f9",
        "text_secondary": "#cbd5e1",
        "subtitle_fg": "#94a3b8",
        "entry_border": "#334155",
        "entry_disabled_bg": "#1e293b",
        "table_stripe": "#0a0f1a",
        "tree_heading_bg": "#334155",
        "tree_selected_bg": "#475569",
        "tab_selected_bg": "#334155",
        "tab_active_bg": "#475569",
    },
    "Lime + Dark": {
        "app_bg": "#1a1f0f",
        "surface_bg": "#2d3417",
        "accent_color": "#84cc16",
        "accent_hover": "#65a30d",
        "accent_disabled": "#365314",
        "text_primary": "#f7fee7",
        "text_secondary": "#d9f99d",
        "subtitle_fg": "#bef264",
        "entry_border": "#365314",
        "entry_disabled_bg": "#1a2e05",
        "table_stripe": "#0f1408",
        "tree_heading_bg": "#365314",
        "tree_selected_bg": "#3f6212",
        "tab_selected_bg": "#365314",
        "tab_active_bg": "#3f6212",
    },
    "Fuchsia + Dark": {
        "app_bg": "#1a0a1a",
        "surface_bg": "#2d1b2d",
        "accent_color": "#d946ef",
        "accent_hover": "#c026d3",
        "accent_disabled": "#86198f",
        "text_primary": "#fdf4ff",
        "text_secondary": "#f0abfc",
        "subtitle_fg": "#f5d0fe",
        "entry_border": "#86198f",
        "entry_disabled_bg": "#701a75",
        "table_stripe": "#0f050f",
        "tree_heading_bg": "#701a75",
        "tree_selected_bg": "#86198f",
        "tab_selected_bg": "#701a75",
        "tab_active_bg": "#86198f",
    },
}

# Core typography for the entire app
FONT_BASE = "{Segoe UI} 10"
FONT_BOLD = "{Segoe UI Semibold} 11"
FONT_TAB = "{Segoe UI Semibold} 10"
FONT_MONO = "{Cascadia Mono} 11"

# Icon system - Unicode/emoji icons for tabs and actions
ICONS = {
    "transactions": "🧾",
    "journal": "📓",
    "ledger": "📚",
    "trial_balance": "🧮",
    "adjustments": "⚙️",
    "financial_statements": "💰",
    "closing": "🔒",
    "post_closing": "📈",
    "export": "⬇️",
    "audit": "🧪",
    "help": "❓",
    "save": "💾",
    "delete": "🗑️",
    "add": "➕",
    "edit": "✏️",
    "search": "🔍",
    "filter": "🔽",
    "refresh": "🔄",
    "settings": "⚙️",
    "close": "✕",
    "check": "✓",
    "warning": "⚠️",
    "info": "ℹ️",
    "error": "❌",
    "success": "✅",
    "loading": "⏳",
    "chart": "📈",
    "document": "📄",
    "calendar": "📅",
}

# Status indicator colors
STATUS_COLORS = {
    "success": "#10b981",
    "warning": "#f59e0b",
    "error": "#ef4444",
    "info": "#3b82f6",
    "pending": "#6b7280",
}


class TechFixApp(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title("TechFix Solutions - Accounting System")
        
        # Set minimum window size
        self.minsize(1024, 720)
        
        # Make the root window expandable
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        # Centralize unexpected Tk callback errors into the Python logger so that
        # hard‑to‑reproduce UI issues are easier to diagnose.
        try:
            self.report_callback_exception = self._report_callback_exception  # type: ignore[attr-defined]
        except Exception:
            # If Tk refuses the override for any reason, continue without crashing.
            logger.debug("Unable to hook Tk report_callback_exception", exc_info=True)
        # Start in a large centered window (easier for modern UI)
        self.update_idletasks()
        width, height = 1200, 800
        screen_w = self.winfo_screenwidth()
        screen_h = self.winfo_screenheight()
        x = max(0, (screen_w - width) // 2)
        y = max(0, (screen_h - height) // 2)
        self.geometry(f"{width}x{height}+{x}+{y}")
        # Load settings first (including theme preference)
        saved_theme = None
        try:
            self._load_window_settings()
            # Check if a theme was loaded from settings
            if hasattr(self, 'theme_name') and self.theme_name in THEMES:
                saved_theme = self.theme_name
        except Exception:
            pass

        # Apply saved theme if available, otherwise use system theme
        try:
            theme_to_apply = saved_theme if saved_theme else self._get_system_theme()
            self._apply_theme(theme_to_apply, initial=True)
        except Exception:
            pass

        # Bind F11 to toggle fullscreen (still available)
        self.bind('<F11>', lambda e: self.attributes('-fullscreen', not self.attributes('-fullscreen')))
        self.bind('<Escape>', lambda e: self.attributes('-fullscreen', False))
        self.protocol('WM_DELETE_WINDOW', self._on_close)
        
        # Handle window resize
        self.bind('<Configure>', self._on_window_resize)

        db.init_db(reset=False)
        conn = db.get_connection()
        db.seed_chart_of_accounts(conn)
        # Ensure default user exists for authentication
        db.ensure_default_role_and_user(conn=conn)
        conn.close()

        # Initialize authentication (but don't show login yet)
        try:
            from . import auth
            self.auth_module = auth
            self.current_user = None
            self.session_token = None
        except Exception as e:
            logger.warning(f"Authentication not available: {e}")
            self.auth_module = None
            self.current_user = None
            self.session_token = None

        self.engine = AccountingEngine()
        self.periods: List = []
        self.current_period_id: Optional[int] = self.engine.current_period_id
        self.cycle_status_rows: List = []
        self._scroll_canvases: list[tk.Canvas] = []
        self._nav_buttons: list = []
        self.period_form_cache: Dict[int, Dict[str, str]] = {}
        self._last_period_id: Optional[int] = self.current_period_id
        self._accounts_prefilled: bool = False
        self._accounts_modified_manually: bool = False
        self._rules_map: dict = {}

        self.style = ttk.Style(self)
        # Initialize theme and palette before building UI
        self.theme_name = "Light"
        self.palette = THEMES[self.theme_name].copy()  # Use copy to avoid modifying original
        self._configure_style()
        
        # Show login dialog before building UI (window is hidden)
        if self.auth_module:
            # Hide the main window - login dialog will use it as parent
            self.withdraw()
            try:
                from . import login_dialog
                # Pass self as parent so login dialog uses Toplevel instead of new root
                login_result = login_dialog.show_login_dialog(self.auth_module, self.palette, parent=self)
                if not login_result:
                    # Login cancelled
                    self.destroy()
                    return
                # Set user info from login
                self.current_user = login_result.get("user")
                self.session_token = login_result.get("session_token")
            except Exception as e:
                logger.error(f"Login dialog error: {e}", exc_info=True)
                self.destroy()
                return
            # Show main window after successful login
            self.deiconify()
        
        self._build_ui()
        
        # Apply theme to all widgets after UI is built
        
        self._update_theme_widgets()
        self._load_periods()
        self._update_theme_widgets()
        self._load_all_views()
        
        # Initialize undo/redo button states
        self._update_undo_redo_states()
        
        try:
            self._load_inference_rules()
        except Exception:
            pass

        try:
            self._start_theme_monitor()
        except Exception:
            pass

    def destroy(self) -> None:
        try:
            self.engine.close()
        finally:
            super().destroy()

    def _apply_theme(self, name: str, *, initial: bool = False) -> None:
        if name not in THEMES or name == self.theme_name:
            return

        if initial:
            self.theme_name = name
            self.palette = THEMES[name].copy()  # Use copy to avoid modifying original
            # Load custom colors if available
            custom_colors = self._load_custom_colors()
            if custom_colors:
                self.palette.update(custom_colors)
            self._configure_style()
            self._update_theme_widgets()
            try:
                self.set_status(f"Theme: {name}")
                self._save_window_settings()
            except Exception:
                pass
            return

        try:
            if getattr(self, '_theme_animating', False):
                return
            self._animate_theme_switch(name)
            try:
                self._apply_resize_layout()
                self._capture_ui_snapshot(label=f"theme:{name}")
            except Exception:
                pass
        except Exception:
            # Fallback to immediate apply if animation fails
            self.theme_name = name
            self.palette = THEMES[name].copy()  # Use copy to avoid modifying original
            # Load custom colors if available
            custom_colors = self._load_custom_colors()
            if custom_colors:
                self.palette.update(custom_colors)
            self._configure_style()
            self._update_theme_widgets()
            try:
                self.set_status(f"Theme: {name}")
                self._save_window_settings()
            except Exception:
                pass

    def _configure_style(self) -> None:
        colors = self.palette

        self.configure(bg=colors["app_bg"])
        self.option_add("*Font", FONT_BASE)
        self.option_add("*TCombobox*Listbox.font", FONT_BASE)
        # Ensure dropdown popups (Combobox listbox) and menus are readable in both themes
        try:
            self.option_add("*TCombobox*Listbox.background", colors.get("surface_bg", "#ffffff"))
            self.option_add("*TCombobox*Listbox.foreground", colors.get("text_primary", "#000000"))
            self.option_add("*TCombobox*Listbox.selectBackground", colors.get("accent_color", "#2563eb"))
            self.option_add("*TCombobox*Listbox.selectForeground", "#ffffff")
            self.option_add("*Menu.background", colors.get("surface_bg", "#ffffff"))
            self.option_add("*Menu.foreground", colors.get("text_primary", "#000000"))
            self.option_add("*Menu.activeBackground", colors.get("accent_color", "#2563eb"))
            self.option_add("*Menu.activeForeground", "#ffffff")
            self.option_add("*Menu.relief", "flat")
        except Exception:
            pass

        try:
            self.style.theme_use("clam")
        except tk.TclError:
            pass

        self.style.configure("TFrame", background=colors["app_bg"])
        self.style.configure("Techfix.App.TFrame", background=colors["app_bg"])
        self.style.configure("Techfix.Surface.TFrame", background=colors["surface_bg"])
        self.style.configure("TLabel", background=colors["surface_bg"], foreground=colors["text_primary"], font=FONT_BASE)

        self.style.configure(
            "Techfix.Headline.TLabel",
            background=colors["accent_color"],
            foreground="#ffffff",
            font="{Segoe UI Semibold} 14",
        )
        self.style.configure(
            "Techfix.Subtitle.TLabel",
            background=colors["accent_color"],
            foreground=colors["subtitle_fg"],
            font=FONT_BASE,
        )
        try:
            tkfont.Font(name="TechfixBrandFont", family="Arial Black", size=17)
        except Exception:
            try:
                tkfont.Font(name="TechfixBrandFont", family="Arial", size=17, weight="bold")
            except Exception:
                try:
                    tkfont.Font(name="TechfixBrandFont", family="Segoe UI", size=17, weight="bold")
                except Exception:
                    pass
        self.style.configure(
            "Techfix.SidebarBrand.TLabel",
            background=colors["surface_bg"],
            foreground=colors["accent_color"],
            font="TechfixBrandFont",
        )
        self.style.configure(
            "Techfix.Brand.TLabel",
            background=colors["app_bg"],
            foreground=colors["accent_color"],
            font="{Segoe UI Semibold} 16",
        )

        self.style.configure(
            "Techfix.TLabelframe",
            background=colors["surface_bg"],
            foreground=colors["text_secondary"],
            borderwidth=0,
            padding=12,
        )
        self.style.configure(
            "Techfix.TLabelframe.Label",
            background=colors["surface_bg"],
            foreground=colors["accent_color"],
            font=FONT_BOLD,
        )

        # Monkeypatch ttk.Label to inherit visible bg/fg from parent when not explicitly provided.
        # This reduces text 'halo' artifacts on colored backgrounds by ensuring labels draw with
        # a matching background color and a contrasting foreground.
        try:
            _orig_ttk_label_init = ttk.Label.__init__

            def _contrast_color(hexcolor: str) -> str:
                try:
                    h = hexcolor.lstrip('#')
                    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
                    # relative luminance
                    lum = (0.2126 * (r/255.0) + 0.7152 * (g/255.0) + 0.0722 * (b/255.0))
                    return '#ffffff' if lum < 0.55 else colors.get('text_primary', '#000000')
                except Exception:
                    return colors.get('text_primary', '#000000')

            def _patched_label_init(self, master=None, *args, **kwargs):
                try:
                    # If background not given, attempt to inherit parent's bg
                    if ('background' not in kwargs) and ('bg' not in kwargs):
                        if master is not None:
                            try:
                                parent_bg = master.cget('bg')
                                if parent_bg:
                                    kwargs['background'] = parent_bg
                            except Exception:
                                pass

                    # If foreground not given, pick a contrasting color based on background
                    if 'foreground' not in kwargs and 'fg' not in kwargs:
                        bg = kwargs.get('background') or (master.cget('bg') if master is not None else None)
                        if isinstance(bg, str) and bg.startswith('#'):
                            kwargs['foreground'] = _contrast_color(bg)
                        else:
                            kwargs['foreground'] = colors.get('text_primary', '#000000')
                except Exception:
                    pass
                return _orig_ttk_label_init(self, master, *args, **kwargs)

            ttk.Label.__init__ = _patched_label_init
        except Exception:
            pass

        self.style.configure(
            "Techfix.TNotebook",
            background=colors["app_bg"],
            borderwidth=0,
            padding=6,
        )
        self.style.configure(
            "Techfix.AppBar.TLabel",
            background=colors["app_bg"],
            foreground=colors["text_secondary"],
            font=FONT_BASE,
        )
        self.style.configure(
            "Techfix.TNotebook.Tab",
            background=colors["surface_bg"],
            foreground=colors["text_primary"],
            padding=(18, 10),
            font=FONT_TAB,
        )
        self.style.map(
            "Techfix.TNotebook.Tab",
            background=[("selected", colors["tab_selected_bg"]), ("active", colors["tab_active_bg"])],
            foreground=[("selected", colors["text_primary"])],
        )

        self.style.configure(
            "Techfix.TButton",
            background=colors["accent_color"],
            foreground="#ffffff",
            padding=(16, 8),
            borderwidth=0,
            # remove focus thickness/color to avoid drawing focus outlines that look like shadows
        )
        self.style.map(
            "Techfix.TButton",
            background=[("active", colors["accent_hover"]), ("disabled", colors["accent_disabled"])],
            foreground=[("disabled", "#ffffff")],
        )
        
        # Enhanced button style with rounded corners and shadow effect
        self.style.configure(
            "Techfix.Enhanced.TButton",
            background=colors["accent_color"],
            foreground="#ffffff",
            padding=(18, 10),
            borderwidth=0,
            relief=tk.RAISED,
        )
        self.style.map(
            "Techfix.Enhanced.TButton",
            background=[("active", colors["accent_hover"]), ("disabled", colors["accent_disabled"])],
            foreground=[("disabled", "#ffffff")],
        )

        # Navigation / Sidebar button style
        self.style.configure(
            "Techfix.Nav.TButton",
            background=colors["surface_bg"],
            foreground=colors["text_primary"],
            padding=(12, 10),
            anchor="w",
            font=FONT_BASE,
            relief="flat",
        )
        self.style.map(
            "Techfix.Nav.TButton",
            background=[("active", colors.get("tab_active_bg")), ("selected", colors.get("tab_selected_bg"))],
            foreground=[("active", colors.get("text_primary"))],
        )
        # Selected variant: accent background with white foreground
        self.style.configure(
            "Techfix.Nav.Selected.TButton",
            background=colors["accent_color"],
            foreground="#ffffff",
            padding=(12, 10),
            anchor="w",
            font=FONT_BASE,
            relief="flat",
        )
        self.style.map(
            "Techfix.Nav.Selected.TButton",
            background=[("active", colors.get("accent_hover"))],
            foreground=[("active", "#ffffff")],
        )

        # Theme toggle button styles
        self.style.configure(
            "Techfix.Theme.TButton",
            background=colors["surface_bg"],
            foreground=colors["text_primary"],
            padding=(10, 6),
            borderwidth=1,
            relief="solid",
        )
        self.style.map(
            "Techfix.Theme.TButton",
            background=[("active", colors.get("tab_active_bg"))],
            foreground=[("active", colors.get("text_primary"))],
        )
        self.style.configure(
            "Techfix.Theme.Selected.TButton",
            background=colors["accent_color"],
            foreground="#ffffff",
            padding=(10, 6),
            borderwidth=1,
            relief="solid",
        )
        self.style.map(
            "Techfix.Theme.Selected.TButton",
            background=[("active", colors.get("accent_hover"))],
            foreground=[("active", "#ffffff")],
        )

        self.style.configure(
            "Techfix.TCheckbutton",
            background=colors["surface_bg"],
            foreground=colors["text_primary"],
            padding=4,
        )
        self.style.configure(
            "Techfix.TEntry",
            fieldbackground=colors["surface_bg"],
            background=colors["surface_bg"],
            foreground=colors["text_primary"],
            borderwidth=0,
            padding=6,
        )
        self.style.map(
            "Techfix.TEntry",
            fieldbackground=[("disabled", colors["entry_disabled_bg"])],
        )

        self.style.configure(
            "Techfix.TCombobox",
            fieldbackground=colors["surface_bg"],
            background=colors["surface_bg"],
            foreground=colors["text_primary"],
            borderwidth=0,
            padding=6,
        )

        self.style.configure(
            "Techfix.Treeview",
            background=colors["surface_bg"],
            fieldbackground=colors["surface_bg"],
            foreground=colors["text_primary"],
            font=FONT_BASE,
            rowheight=26,
            borderwidth=0,
        )
        self.style.configure(
            "Techfix.Treeview.Heading",
            background=colors["tree_heading_bg"],
            foreground=colors["text_primary"],
            font=FONT_BOLD,
        )
        # Ensure header hover/active states stay consistent in both light and dark themes
        self.style.map(
            "Techfix.Treeview.Heading",
            background=[
                ("active", colors["tree_heading_bg"]),
                ("pressed", colors["tree_heading_bg"]),
            ],
            foreground=[
                ("active", colors["text_primary"]),
                ("pressed", colors["text_primary"]),
            ],
        )
        self.style.map(
            "Techfix.Treeview",
            background=[("selected", colors["tree_selected_bg"])],
            foreground=[("selected", colors["text_primary"])],
        )
        self.style.layout("Techfix.Treeview", self.style.layout("Treeview"))

        self.style.configure(
            "Techfix.StatusBar.TLabel",
            background=colors.get("app_bg", "#ffffff"),
            foreground=colors.get("text_secondary", "#4b5563"),
            font=FONT_BASE,
        )

        self.style.configure(
            "Techfix.Danger.TButton",
            background="#dc2626",
            foreground="#ffffff",
            padding=(16, 8),
            borderwidth=0,
        )
        self.style.map(
            "Techfix.Danger.TButton",
            background=[("active", "#b91c1c"), ("disabled", "#ef4444")],
            foreground=[("disabled", "#ffffff")],
        )

    def _update_theme_widgets(self) -> None:
        """Apply current palette colors to widgets that need manual updates."""
        try:
            colors = self.palette

            # Update root/background colors
            self.configure(bg=colors.get("app_bg", "#ffffff"))

            # Update any stored scroll canvases
            if hasattr(self, '_scroll_canvases'):
                for c in getattr(self, '_scroll_canvases'):
                    try:
                        c.configure(bg=colors.get('surface_bg', '#ffffff'))
                    except Exception:
                        pass

            # Update any Text widgets created for financial statements and help text
            for attr in ('income_text', 'balance_sheet_text', 'cash_flow_text', 'close_log', 'txn_memo', 'help_text'):
                if hasattr(self, attr):
                    w = getattr(self, attr)
                    try:
                        # Special handling for help_text to preserve flat appearance
                        is_help_text = (attr == 'help_text')
                        w.configure(
                            bg=colors.get('surface_bg', '#ffffff'),
                            fg=colors.get('text_primary', '#000000'),
                            insertbackground=colors.get('accent_color', '#2563eb') if is_help_text else colors.get('text_primary', '#000000'),
                            selectbackground=colors.get('accent_color', '#2563eb'),
                            selectforeground='#ffffff'
                        )
                        if not is_help_text:
                            try:
                                w.configure(
                                    bd=1,
                                    relief=tk.SOLID,
                                    highlightthickness=1,
                                    highlightbackground=colors.get('entry_border', '#d8dee9'),
                                    highlightcolor=colors.get('accent_color', '#2563eb'),
                                )
                            except Exception:
                                pass
                        else:
                            # Preserve flat appearance for help_text
                            try:
                                w.configure(
                                    bd=0,
                                    relief=tk.FLAT,
                                    highlightthickness=0,
                                )
                            except Exception:
                                pass
                        # Update commonly used tags so previously-inserted tagged text remains visible after theme change
                        try:
                            w.tag_configure('header', foreground=colors.get('accent_color', '#2563eb'))
                            w.tag_configure('subheader', foreground=colors.get('text_secondary', '#4b5563'))
                            w.tag_configure('section', foreground=colors.get('text_primary', '#1f2937'))
                            # Ensure totals in text widgets are bold and clearly visible
                            w.tag_configure('total', foreground=colors.get('text_primary', '#1f2937'), font=("Segoe UI", 10, "bold"))
                            w.tag_configure('net', foreground=colors.get('accent_color', '#2563eb'))
                            w.tag_configure('warning', foreground='red')
                        except Exception:
                            pass
                    except Exception:
                        pass

            # Update treeview heading/background via styles (ensure maps are applied)
            try:
                self.style.configure('Techfix.Treeview', background=colors.get('surface_bg'))
                self.style.configure('Techfix.Treeview.Heading', background=colors.get('tree_heading_bg'))
            except Exception:
                pass

            # Refresh sidebar brand canvas to match theme
            try:
                if hasattr(self, 'sidebar_brand_canvas'):
                    self.sidebar_brand_canvas.configure(bg=colors.get('surface_bg', '#ffffff'))
                    self._draw_sidebar_brand()
            except Exception:
                pass

            # If financial statements already have content, re-generate them
            # so tag colors and formatting are applied for the new theme.
            try:
                if hasattr(self, 'income_text') and self.income_text.get("1.0", tk.END).strip():
                    try:
                        # Re-generate financial statement text without updating cycle status
                        self._regenerate_financial_statements(self.fs_date_to.get().strip() if hasattr(self, 'fs_date_to') else None)
                    except Exception:
                        pass
            except Exception:
                pass

            # Ensure existing Treeview widgets have readable row colors (tag and retag rows)
            try:
                for name in dir(self):
                    try:
                        obj = getattr(self, name)
                    except Exception:
                        continue
                    if isinstance(obj, ttk.Treeview):
                        try:
                            # Configure a generic 'row' tag with proper colors
                            obj.tag_configure(
                                'row',
                                background=colors.get('surface_bg', '#ffffff'),
                                foreground=colors.get('text_primary', '#000000'),
                                font=FONT_BASE,
                            )
                            # Ensure totals rows reflect current theme and stand out in bold
                            obj.tag_configure(
                                'totals',
                                background=colors.get('tab_selected_bg', '#e0ecff'),
                                foreground=colors.get('text_primary', '#000000'),
                                font=FONT_BOLD,
                            )
                            # Retag existing rows so the tag applies immediately
                            for iid in obj.get_children():
                                tags = tuple(obj.item(iid, 'tags') or ())
                                if 'row' not in tags:
                                    obj.item(iid, tags=(*tags, 'row'))
                        except Exception:
                            pass
            except Exception:
                pass

            # Update header frame and labels to match theme
            try:
                if hasattr(self, 'header_frame'):
                    header = self.header_frame
                    header.configure(bg=colors.get('accent_color', '#2563eb'))
                    
                    # Update header left and right frames
                    if hasattr(self, 'header_left_frame'):
                        self.header_left_frame.configure(bg=colors.get('accent_color', '#2563eb'))
                        for widget in self.header_left_frame.winfo_children():
                            if isinstance(widget, tk.Label):
                                widget.configure(bg=colors.get('accent_color', '#2563eb'))
                                if widget.cget('text') == "Integrated accounting workspace":
                                    widget.configure(fg=colors.get('subtitle_fg', '#dbeafe'))
                    
                    if hasattr(self, 'header_right_frame'):
                        self.header_right_frame.configure(bg=colors.get('accent_color', '#2563eb'))
                        # Update all frames and buttons in header_right
                        for widget in self.header_right_frame.winfo_children():
                            if isinstance(widget, tk.Frame):
                                widget.configure(bg=colors.get('accent_color', '#2563eb'))
                                # Update search frame and its children
                                for child in widget.winfo_children():
                                    if isinstance(child, tk.Frame):
                                        child.configure(bg=colors.get('accent_color', '#2563eb'))
                                    elif isinstance(child, tk.Entry):
                                        # Update search entry highlight color
                                        child.configure(highlightcolor=colors.get('accent_hover', colors.get('accent_color', '#2563eb')))
                                    elif isinstance(child, tk.Button):
                                        # Update search button colors
                                        child.configure(fg=colors.get('accent_color', '#2563eb'))
                            elif isinstance(widget, tk.Button):
                                # Update button foreground color to match theme
                                widget.configure(fg=colors.get('accent_color', '#2563eb'))
                    
                    # Update header buttons (user menu and notifications)
                    if hasattr(self, 'user_menu_btn'):
                        self.user_menu_btn.configure(fg=colors.get('accent_color', '#2563eb'))
                    if hasattr(self, 'notif_btn'):
                        self.notif_btn.configure(fg=colors.get('accent_color', '#2563eb'))
                    
                    # Update gradient canvas if it exists
                    for widget in header.winfo_children():
                        if isinstance(widget, tk.Canvas):
                            try:
                                # Recreate gradient with new colors
                                header.update_idletasks()
                                w = header.winfo_width() or 1200
                                h = header.winfo_height() or 60
                                widget.destroy()
                                gradient_canvas = self._create_gradient_canvas(
                                    header, w, h,
                                    colors.get('accent_color', '#2563eb'),
                                    colors.get('accent_hover', colors.get('accent_color', '#2563eb')),
                                    direction='horizontal'
                                )
                                gradient_canvas.place(x=0, y=0, relwidth=1, relheight=1)
                                gradient_canvas.lower()
                            except Exception:
                                pass
                        elif isinstance(widget, tk.Label):
                            widget.configure(bg=colors.get('accent_color', '#2563eb'))
                            # Update subtitle foreground color
                            if widget.cget('text') == "Integrated accounting workspace":
                                widget.configure(fg=colors.get('subtitle_fg', '#dbeafe'))
            except Exception:
                pass

            # Update sidebar indicators and button styles to match theme
            try:
                if hasattr(self, '_nav_buttons'):
                    cur = getattr(self, '_current_nav_index', None)
                    for ind, btn, nav_idx in self._nav_buttons:
                        try:
                            ind.configure(bg=(colors.get('accent_color') if cur == nav_idx else colors.get('surface_bg')),
                                          highlightthickness=0, bd=0, relief=tk.FLAT)
                        except Exception:
                            pass
                        try:
                            btn.configure(style=('Techfix.Nav.Selected.TButton' if cur == nav_idx else 'Techfix.Nav.TButton'))
                        except Exception:
                            pass
            except Exception:
                pass

            try:
                self._style_all_combobox_popdowns()
            except Exception:
                pass
            try:
                self._style_menus()
            except Exception:
                pass

        except Exception:
            # Don't let theme update errors break app initialization
            pass

    def _draw_sidebar_brand(self) -> None:
        try:
            if not hasattr(self, 'sidebar_brand_canvas'):
                return
            c = self.sidebar_brand_canvas
            c.delete('all')
            # Intentionally leave the brand area blank (no 'Techfix' text)
        except Exception:
            pass

    def _animate_theme_switch(self, name: str, *, duration_ms: int = 300) -> None:
        try:
            self._theme_animating = True
            self.update_idletasks()
            w = max(1, self.winfo_width() or 1200)
            h = max(1, self.winfo_height() or 800)

            new_palette = THEMES[name]
            overlay = tk.Canvas(self, highlightthickness=0, bd=0, bg=self.palette.get('app_bg', '#ffffff'))
            overlay.place(x=0, y=0, width=w, height=h)
            try:
                overlay.lift()
            except Exception:
                pass

            steps = max(10, int(duration_ms / 15))
            dy = h / float(steps)
            if name == "Dark":
                rect = overlay.create_rectangle(0, 0, w, 1, fill=new_palette.get('surface_bg', '#ffffff'), outline="")
                def coords_for_step(i):
                    return (0, 0, w, int(min(h, i * dy)))
            else:
                rect = overlay.create_rectangle(0, h-1, w, h, fill=new_palette.get('surface_bg', '#ffffff'), outline="")
                def coords_for_step(i):
                    y1 = int(max(0, h - i * dy))
                    return (0, y1, w, h)
            i = 0

            def sweep():
                nonlocal i
                i += 1
                overlay.coords(rect, *coords_for_step(i))
                if i < steps:
                    self.after(15, sweep)
                else:
                    try:
                        self.theme_name = name
                        self.palette = new_palette
                        self._configure_style()
                        self._update_theme_widgets()
                        self.set_status(f"Theme: {name}")
                        self._save_window_settings()
                    except Exception:
                        pass
                    try:
                        overlay.place_forget()
                        overlay.destroy()
                    except Exception:
                        pass
                    self._theme_animating = False

            sweep()
        except Exception:
            self._theme_animating = False

    def _get_system_theme(self) -> str:
        try:
            os_name = platform.system()
            if os_name == "Windows" and winreg:
                try:
                    key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"Software\Microsoft\Windows\CurrentVersion\Themes\Personalize")
                    val, _ = winreg.QueryValueEx(key, "AppsUseLightTheme")
                    return "Light" if int(val) != 0 else "Dark"
                except Exception:
                    return "Light"
            if os_name == "Darwin":
                try:
                    out = subprocess.check_output(["defaults", "read", "-g", "AppleInterfaceStyle"], stderr=subprocess.STDOUT)
                    s = out.decode().strip()
                    return "Dark" if s.lower().startswith("dark") else "Light"
                except Exception:
                    return "Light"
            return "Light"
        except Exception:
            return "Light"

    def _start_theme_monitor(self) -> None:
        try:
            self._last_system_theme = self._get_system_theme()
            def _poll():
                try:
                    cur = self._get_system_theme()
                    if cur != getattr(self, "_last_system_theme", cur):
                        self._last_system_theme = cur
                        try:
                            self._apply_theme(cur)
                        except Exception:
                            pass
                except Exception:
                    pass
                try:
                    self.after(3000, _poll)
                except Exception:
                    pass
            _poll()
        except Exception:
            pass

    def _open_color_picker(self) -> None:
        """Open a color picker dialog to customize GUI colors."""
        try:
            # Create a dialog window
            dialog = tk.Toplevel(self)
            dialog.title("Customize Colors")
            dialog.transient(self)
            dialog.grab_set()
            
            # Make dialog modal and center it
            dialog.update_idletasks()
            width, height = 600, 500
            x = (dialog.winfo_screenwidth() // 2) - (width // 2)
            y = (dialog.winfo_screenheight() // 2) - (height // 2)
            dialog.geometry(f"{width}x{height}+{x}+{y}")
            dialog.resizable(False, False)
            
            # Apply current theme to dialog
            dialog.configure(bg=self.palette.get("surface_bg", "#ffffff"))
            
            # Create main frame
            main_frame = ttk.Frame(dialog, style="Techfix.Surface.TFrame", padding=20)
            main_frame.pack(fill=tk.BOTH, expand=True)
            
            # Title
            title_label = ttk.Label(
                main_frame,
                text="Customize GUI Colors",
                font=FONT_BOLD,
                style="Techfix.Brand.TLabel"
            )
            title_label.pack(pady=(0, 20))
            
            # Description
            desc_label = ttk.Label(
                main_frame,
                text="Choose colors for different GUI elements. Changes apply immediately.",
                style="TLabel"
            )
            desc_label.pack(pady=(0, 20))
            
            # Color selection frame with scrollable area
            canvas = tk.Canvas(main_frame, bg=self.palette.get("surface_bg", "#ffffff"), highlightthickness=0)
            scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
            scrollable_frame = ttk.Frame(canvas, style="Techfix.Surface.TFrame")
            
            scrollable_frame.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
            )
            
            canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)
            
            # Color definitions with user-friendly names
            color_defs = [
                ("app_bg", "Application Background", "Main window background color"),
                ("surface_bg", "Surface Background", "Panel and card background color"),
                ("accent_color", "Accent Color", "Primary buttons and highlights"),
                ("accent_hover", "Accent Hover", "Button hover state"),
                ("text_primary", "Primary Text", "Main text color"),
                ("text_secondary", "Secondary Text", "Secondary text color"),
                ("entry_border", "Input Border", "Input field border color"),
                ("table_stripe", "Table Stripe", "Alternating table row color"),
                ("tree_heading_bg", "Tree Heading", "Table header background"),
                ("tree_selected_bg", "Selected Item", "Selected item background"),
            ]
            
            # Store color variables
            color_vars = {}
            color_buttons = {}
            
            # Create color picker rows
            for i, (key, label, desc) in enumerate(color_defs):
                row_frame = ttk.Frame(scrollable_frame, style="Techfix.Surface.TFrame")
                row_frame.pack(fill=tk.X, pady=8, padx=10)
                
                # Label and description
                label_frame = ttk.Frame(row_frame, style="Techfix.Surface.TFrame")
                label_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
                
                ttk.Label(
                    label_frame,
                    text=label,
                    font=FONT_BASE,
                    style="TLabel"
                ).pack(anchor=tk.W)
                
                ttk.Label(
                    label_frame,
                    text=desc,
                    font="{Segoe UI} 9",
                    foreground=self.palette.get("text_secondary", "#666666"),
                    style="TLabel"
                ).pack(anchor=tk.W)
                
                # Color button
                current_color = self.palette.get(key, "#000000")
                color_vars[key] = tk.StringVar(value=current_color)
                
                def make_color_button(k=key, v=color_vars[key]):
                    btn = tk.Button(
                        row_frame,
                        text="Choose Color",
                        bg=v.get(),
                        fg="#ffffff" if self._is_dark_color(v.get()) else "#000000",
                        command=lambda k=k, v=v: self._pick_color(k, v, btn),
                        relief=tk.RAISED,
                        bd=2,
                        padx=15,
                        pady=5,
                        font=FONT_BASE
                    )
                    btn.pack(side=tk.RIGHT)
                    color_buttons[k] = btn
                    return btn
                
                make_color_button()
            
            # Pack canvas and scrollbar
            canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            # Button frame
            button_frame = ttk.Frame(main_frame, style="Techfix.Surface.TFrame")
            button_frame.pack(fill=tk.X, pady=(20, 0))
            
            # Reset to default button
            def reset_to_default():
                try:
                    # Reset to current theme's default colors
                    base_theme = self.theme_name
                    if base_theme in THEMES:
                        self.palette = THEMES[base_theme].copy()
                        # Clear custom colors from settings
                        self._save_custom_colors(None)
                        self._apply_theme(base_theme, initial=True)
                        dialog.destroy()
                        messagebox.showinfo("Colors Reset", f"Colors reset to {base_theme} theme defaults.")
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to reset colors: {e}")
            
            ttk.Button(
                button_frame,
                text="Reset to Default",
                command=reset_to_default,
                style="Techfix.TButton"
            ).pack(side=tk.LEFT, padx=(0, 10))
            
            # Apply button
            def apply_colors():
                try:
                    # Update palette with selected colors
                    for key, var in color_vars.items():
                        self.palette[key] = var.get()
                    
                    # Save custom colors
                    custom_colors = {key: var.get() for key, var in color_vars.items()}
                    self._save_custom_colors(custom_colors)
                    
                    # Apply the updated theme
                    self._configure_style()
                    self._update_theme_widgets()
                    
                    dialog.destroy()
                    messagebox.showinfo("Colors Applied", "Custom colors have been applied!")
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to apply colors: {e}")
            
            ttk.Button(
                button_frame,
                text="Apply Colors",
                command=apply_colors,
                style="Techfix.TButton"
            ).pack(side=tk.LEFT, padx=(0, 10))
            
            # Cancel button
            ttk.Button(
                button_frame,
                text="Cancel",
                command=dialog.destroy,
                style="Techfix.TButton"
            ).pack(side=tk.RIGHT)
            
            # Update canvas scroll region
            def update_scroll_region():
                canvas.update_idletasks()
                canvas.configure(scrollregion=canvas.bbox("all"))
            
            scrollable_frame.after(100, update_scroll_region)
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open color picker: {e}")
    
    def _pick_color(self, key: str, var: tk.StringVar, button: tk.Button) -> None:
        """Open color chooser for a specific color key."""
        try:
            color = colorchooser.askcolor(
                title=f"Choose Color for {key}",
                color=var.get(),
                parent=self
            )
            if color[1]:  # User selected a color
                var.set(color[1])
                button.configure(bg=color[1])
                button.configure(fg="#ffffff" if self._is_dark_color(color[1]) else "#000000")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to pick color: {e}")
    
    def _is_dark_color(self, hex_color: str) -> bool:
        """Check if a hex color is dark (for text contrast)."""
        try:
            h = hex_color.lstrip('#')
            r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
            # Calculate relative luminance
            lum = (0.2126 * (r/255.0) + 0.7152 * (g/255.0) + 0.0722 * (b/255.0))
            return lum < 0.5
        except Exception:
            return False
    
    def _save_custom_colors(self, custom_colors: Optional[Dict[str, str]]) -> None:
        """Save custom colors to settings file."""
        try:
            settings_path = db.DB_DIR / "settings.json"
            if settings_path.exists():
                data = json.loads(settings_path.read_text(encoding="utf-8"))
            else:
                data = {}
            
            if custom_colors:
                data["custom_colors"] = custom_colors
                data["custom_colors_theme"] = self.theme_name  # Remember which theme was customized
            else:
                data.pop("custom_colors", None)
                data.pop("custom_colors_theme", None)
            
            settings_path.parent.mkdir(parents=True, exist_ok=True)
            settings_path.write_text(json.dumps(data, indent=2), encoding="utf-8")
        except Exception as e:
            logger.error(f"Failed to save custom colors: {e}")
    
    def _load_custom_colors(self) -> Optional[Dict[str, str]]:
        """Load custom colors from settings file."""
        try:
            # Ensure theme_name is set
            if not hasattr(self, 'theme_name') or not self.theme_name:
                return None
                
            settings_path = db.DB_DIR / "settings.json"
            if settings_path.exists():
                data = json.loads(settings_path.read_text(encoding="utf-8"))
                custom_colors = data.get("custom_colors")
                custom_theme = data.get("custom_colors_theme")
                # Only apply custom colors if they match the current theme
                if custom_colors and custom_theme == self.theme_name:
                    return custom_colors
            return None
        except Exception as e:
            logger.error(f"Failed to load custom colors: {e}")
            return None

    def _open_color_theme_picker(self) -> None:
        """Open a color theme picker dialog with predefined color combinations."""
        try:
            # Create a dialog window
            dialog = tk.Toplevel(self)
            dialog.title("Choose Color Theme")
            dialog.transient(self)
            dialog.grab_set()
            
            # Make dialog modal and center it
            dialog.update_idletasks()
            width, height = 700, 650
            x = (dialog.winfo_screenwidth() // 2) - (width // 2)
            y = (dialog.winfo_screenheight() // 2) - (height // 2)
            dialog.geometry(f"{width}x{height}+{x}+{y}")
            dialog.resizable(True, True)  # Allow resizing so buttons are always accessible
            
            # Apply current theme to dialog
            dialog.configure(bg=self.palette.get("surface_bg", "#ffffff"))
            
            # Create main frame
            main_frame = ttk.Frame(dialog, style="Techfix.Surface.TFrame", padding=20)
            main_frame.pack(fill=tk.BOTH, expand=True)
            
            # Title
            title_label = ttk.Label(
                main_frame,
                text="Choose a Color Theme",
                font=FONT_BOLD,
                style="Techfix.Brand.TLabel"
            )
            title_label.pack(pady=(0, 10))
            
            # Description
            desc_label = ttk.Label(
                main_frame,
                text="Select a color combination to apply to your GUI",
                style="TLabel"
            )
            desc_label.pack(pady=(0, 20))
            
            # Create a container for the scrollable area and buttons
            content_container = ttk.Frame(main_frame, style="Techfix.Surface.TFrame")
            content_container.pack(fill=tk.BOTH, expand=True)
            
            # Create scrollable frame for color theme cards
            canvas = tk.Canvas(content_container, bg=self.palette.get("surface_bg", "#ffffff"), highlightthickness=0)
            scrollbar = ttk.Scrollbar(content_container, orient="vertical", command=canvas.yview)
            scrollable_frame = ttk.Frame(canvas, style="Techfix.Surface.TFrame")
            
            scrollable_frame.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
            )
            
            canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)
            
            # Create color theme cards
            selected_theme = [None]  # Use list to allow modification in nested function
            
            def create_theme_card(theme_name: str, theme_colors: Dict[str, str], parent_frame: tk.Widget):
                """Create a visual card for a color theme."""
                card_frame = tk.Frame(
                    parent_frame,
                    bg=self.palette.get("surface_bg", "#ffffff"),
                    relief=tk.RAISED,
                    bd=2,
                    cursor="hand2"
                )
                card_frame.pack(fill=tk.X, pady=8, padx=10)
                
                # Card header with theme name
                header = tk.Frame(card_frame, bg=theme_colors["accent_color"], height=40)
                header.pack(fill=tk.X)
                tk.Label(
                    header,
                    text=theme_name,
                    bg=theme_colors["accent_color"],
                    fg="#ffffff",
                    font=FONT_BOLD
                ).pack(pady=10)
                
                # Color preview area
                preview_frame = tk.Frame(card_frame, bg=theme_colors["app_bg"])
                preview_frame.pack(fill=tk.X, padx=2, pady=2)
                
                # Show color swatches in a grid
                swatch_frame = tk.Frame(preview_frame, bg=theme_colors["app_bg"])
                swatch_frame.pack(fill=tk.X, padx=10, pady=10)
                
                # Main colors to display
                colors_to_show = [
                    ("Background", theme_colors["app_bg"]),
                    ("Surface", theme_colors["surface_bg"]),
                    ("Accent", theme_colors["accent_color"]),
                    ("Text", theme_colors["text_primary"]),
                ]
                
                # Create a grid of color swatches
                for i, (label, color) in enumerate(colors_to_show):
                    row = i // 2
                    col = i % 2
                    
                    # Container for each swatch
                    swatch_container = tk.Frame(swatch_frame, bg=theme_colors["app_bg"])
                    swatch_container.grid(row=row, column=col, padx=5, pady=5, sticky="w")
                    
                    # Color swatch
                    swatch = tk.Frame(swatch_container, bg=color, width=50, height=25, relief=tk.SUNKEN, bd=1)
                    swatch.pack(side=tk.LEFT, padx=(0, 5))
                    
                    # Label
                    tk.Label(
                        swatch_container,
                        text=label,
                        bg=theme_colors["app_bg"],
                        fg=theme_colors["text_secondary"],
                        font="{Segoe UI} 8"
                    ).pack(side=tk.LEFT)
                
                # Selection indicator - store reference for easy access
                indicator = tk.Frame(card_frame, bg=self.palette.get("accent_color", "#2563eb"), height=3)
                indicator.pack_forget()  # Hide by default
                # Store indicator reference in card_frame for easy access
                card_frame._indicator = indicator
                card_frame._theme_name = theme_name
                
                def on_click(event=None):
                    # Hide all indicators by finding them through stored references
                    for widget in parent_frame.winfo_children():
                        if hasattr(widget, '_indicator'):
                            try:
                                widget._indicator.pack_forget()
                            except:
                                pass
                    
                    # Show this indicator at the bottom of the card
                    try:
                        indicator.pack(fill=tk.X, side=tk.BOTTOM)
                    except:
                        pass
                    selected_theme[0] = theme_name
                    return "break"
                
                # Bind click to all interactive parts
                # Bind to main card frame (this will catch clicks on empty areas)
                card_frame.bind("<Button-1>", on_click)
                card_frame.configure(cursor="hand2")
                
                # Bind to header
                header.bind("<Button-1>", on_click)
                header.configure(cursor="hand2")
                
                # Bind to preview frame
                preview_frame.bind("<Button-1>", on_click)
                preview_frame.configure(cursor="hand2")
                
                # Bind to swatch frame
                swatch_frame.bind("<Button-1>", on_click)
                swatch_frame.configure(cursor="hand2")
                
                # Bind to all child widgets in swatch_frame recursively
                def bind_to_children(parent):
                    for child in parent.winfo_children():
                        child.bind("<Button-1>", on_click)
                        if isinstance(child, tk.Frame):
                            bind_to_children(child)
                
                bind_to_children(swatch_frame)
                
                return card_frame
            
            # Create cards for Light and Dark themes first (from THEMES)
            for theme_name in ["Light", "Dark"]:
                if theme_name in THEMES:
                    display_name = f"{theme_name} Theme"
                    create_theme_card(display_name, THEMES[theme_name], scrollable_frame)
            
            # Create cards for each color theme
            for theme_name, theme_colors in COLOR_THEMES.items():
                create_theme_card(theme_name, theme_colors, scrollable_frame)
            
            # Pack canvas and scrollbar in the content container
            canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            # Button frame - pack at bottom of main_frame, not inside content_container
            button_frame = ttk.Frame(main_frame, style="Techfix.Surface.TFrame")
            button_frame.pack(fill=tk.X, pady=(15, 5), side=tk.BOTTOM)
            
            # Reset to default button
            def reset_to_default():
                try:
                    # Reset to current theme's default colors
                    base_theme = self.theme_name
                    if base_theme in THEMES:
                        self.palette = THEMES[base_theme].copy()
                        # Clear custom colors from settings
                        self._save_custom_colors(None)
                        self._apply_theme(base_theme, initial=True)
                        dialog.destroy()
                        messagebox.showinfo("Theme Reset", f"Colors reset to {base_theme} theme defaults.")
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to reset colors: {e}")
            
            ttk.Button(
                button_frame,
                text="Reset to Default",
                command=reset_to_default,
                style="Techfix.TButton"
            ).pack(side=tk.LEFT, padx=(0, 10))
            
            # Apply button
            def apply_theme():
                try:
                    if not selected_theme[0]:
                        messagebox.showwarning("No Selection", "Please select a color theme first.")
                        return
                    
                    theme_name = selected_theme[0]
                    
                    # If Light Theme or Dark Theme is selected, update theme_name
                    if theme_name == "Light Theme":
                        self.theme_name = "Light"
                        # Clear custom colors to use default Light theme
                        self._save_custom_colors(None)
                        self._apply_theme("Light", initial=True)
                    elif theme_name == "Dark Theme":
                        self.theme_name = "Dark"
                        # Clear custom colors to use default Dark theme
                        self._save_custom_colors(None)
                        self._apply_theme("Dark", initial=True)
                    else:
                        # For other color themes, apply custom colors
                        if theme_name in COLOR_THEMES:
                            theme_colors = COLOR_THEMES[theme_name]
                            self.palette = theme_colors.copy()
                            # Save as custom colors
                            self._save_custom_colors(theme_colors)
                            # Apply the updated theme
                            self._configure_style()
                            self._update_theme_widgets()
                        else:
                            messagebox.showerror("Error", f"Theme '{theme_name}' not found.")
                            return
                    
                    dialog.destroy()
                    messagebox.showinfo("Theme Applied", f"{theme_name} has been applied!")
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to apply theme: {e}")
            
            ttk.Button(
                button_frame,
                text="Apply Theme",
                command=apply_theme,
                style="Techfix.TButton"
            ).pack(side=tk.LEFT, padx=(0, 10))
            
            # Cancel button
            ttk.Button(
                button_frame,
                text="Cancel",
                command=dialog.destroy,
                style="Techfix.TButton"
            ).pack(side=tk.RIGHT)
            
            # Update canvas scroll region
            def update_scroll_region():
                canvas.update_idletasks()
                canvas.configure(scrollregion=canvas.bbox("all"))
            
            scrollable_frame.after(100, update_scroll_region)
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open color theme picker: {e}")

    def _animate_view_transition(self, target: tk.Widget, *, duration_ms: int = 250) -> None:
        try:
            target.update_idletasks()
            w = max(1, target.winfo_width() or 800)
            h = max(1, target.winfo_height() or 600)
            overlay = tk.Canvas(target, highlightthickness=0, bd=0, bg=self.palette.get('surface_bg', '#ffffff'))
            overlay.place(x=0, y=0, width=w, height=h)
            try:
                overlay.lift()
            except Exception:
                pass
            steps = max(10, int(duration_ms / 15))
            dx = w / float(steps)
            rect = overlay.create_rectangle(0, 0, 1, h, fill=self.palette.get('surface_bg', '#ffffff'), outline="")
            i = 0
            def sweep():
                nonlocal i
                i += 1
                x2 = int(min(w, i * dx))
                overlay.coords(rect, 0, 0, x2, h)
                if i < steps:
                    self.after(15, sweep)
                else:
                    try:
                        overlay.place_forget()
                        overlay.destroy()
                    except Exception:
                        pass
            sweep()
        except Exception:
            pass

    def _animate_swipe_to(self, target: tk.Widget, *, direction: str = 'right', duration_ms: int = 250) -> None:
        try:
            # Slide the target frame into view using place, then restore pack
            parent = getattr(self, 'content_area', self)
            target.update_idletasks()
            parent.update_idletasks()
            w = max(1, parent.winfo_width() or target.winfo_width() or 800)
            h = max(1, parent.winfo_height() or target.winfo_height() or 600)
            try:
                target.pack_forget()
            except Exception:
                pass
            start_x = w if direction == 'right' else -w
            target.place(in_=parent, x=start_x, y=0, width=w, height=h)
            try:
                target.lift()
            except Exception:
                pass
            steps = max(10, int(duration_ms / 15))
            dx = w / float(steps)
            i = 0
            def step():
                nonlocal i, start_x
                i += 1
                if direction == 'right':
                    x = int(max(0, start_x - i * dx))
                else:
                    x = int(min(0, start_x + i * dx))
                target.place_configure(x=x)
                if i < steps:
                    self.after(15, step)
                else:
                    try:
                        target.place_forget()
                        target.pack(fill=tk.BOTH, expand=True)
                    except Exception:
                        pass
            step()
        except Exception:
            pass

    def _animate_tab_pulse(self, frame: tk.Widget, *, duration_ms: int = 200) -> None:
        try:
            colors = self.palette
            base = colors.get('surface_bg', '#ffffff')
            accent = colors.get('accent_color', '#2563eb')
            def hex_to_rgb(h: str):
                h = h.lstrip('#')
                return (int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16))
            def rgb_to_hex(r: int, g: int, b: int):
                return f"#{r:02x}{g:02x}{b:02x}"
            r1, g1, b1 = hex_to_rgb(accent)
            r2, g2, b2 = hex_to_rgb(base)
            steps = max(8, int(duration_ms / 15))
            style_name = f"Techfix.Pulse.{id(frame)}.TFrame"
            self.style.configure(style_name, background=accent)
            frame.configure(style=style_name)
            i = 0
            def step():
                nonlocal i
                i += 1
                t = i / float(steps)
                r = int(r1 + (r2 - r1) * t)
                g = int(g1 + (g2 - g1) * t)
                b = int(b1 + (b2 - b1) * t)
                c = rgb_to_hex(r, g, b)
                try:
                    self.style.configure(style_name, background=c)
                except Exception:
                    pass
                if i < steps:
                    self.after(15, step)
                else:
                    try:
                        frame.configure(style="Techfix.Surface.TFrame")
                    except Exception:
                        pass
            step()
        except Exception:
            pass

    def _style_combobox_popdown(self, cb: ttk.Combobox) -> None:
        try:
            colors = self.palette
            pop = cb.tk.call('ttk::combobox::PopdownWindow', cb)
            lb = self.nametowidget(f'{pop}.f.l')
            lb.configure(
                background=colors.get('surface_bg', '#ffffff'),
                foreground=colors.get('text_primary', '#000000'),
                selectbackground=colors.get('accent_color', '#2563eb'),
                selectforeground='#ffffff',
                highlightthickness=0,
                borderwidth=0,
            )
        except Exception:
            pass

    def _style_all_combobox_popdowns(self) -> None:
        try:
            for name in dir(self):
                try:
                    w = getattr(self, name)
                except Exception:
                    continue
                if isinstance(w, ttk.Combobox):
                    try:
                        self._style_combobox_popdown(w)
                    except Exception:
                        pass
        except Exception:
            pass

    def _style_menus(self) -> None:
        try:
            colors = self.palette
            for m in (getattr(self, 'menubar', None), getattr(self, 'file_menu', None), getattr(self, 'view_menu', None), getattr(self, 'help_menu', None)):
                if m is None:
                    continue
                try:
                    m.configure(
                        bg=colors.get('surface_bg', '#ffffff'),
                        fg=colors.get('text_primary', '#000000'),
                        activebackground=colors.get('accent_color', '#2563eb'),
                        activeforeground='#ffffff',
                        tearoff=False,
                    )
                except Exception:
                    pass
        except Exception:
            pass

    # ========== AESTHETIC ENHANCEMENTS ==========
    
    def _create_gradient_canvas(self, parent, width, height, color1, color2, direction='horizontal'):
        """Create a canvas with gradient background."""
        try:
            canvas = tk.Canvas(parent, width=width, height=height, highlightthickness=0, bd=0)
            def hex_to_rgb(h):
                h = h.lstrip('#')
                return (int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16))
            def rgb_to_hex(r, g, b):
                return f"#{r:02x}{g:02x}{b:02x}"
            r1, g1, b1 = hex_to_rgb(color1)
            r2, g2, b2 = hex_to_rgb(color2)
            if direction == 'horizontal':
                for i in range(width):
                    t = i / float(width)
                    r = int(r1 + (r2 - r1) * t)
                    g = int(g1 + (g2 - g1) * t)
                    b = int(b1 + (b2 - b1) * t)
                    c = rgb_to_hex(r, g, b)
                    canvas.create_line(i, 0, i, height, fill=c, width=1)
            else:
                for i in range(height):
                    t = i / float(height)
                    r = int(r1 + (r2 - r1) * t)
                    g = int(g1 + (g2 - g1) * t)
                    b = int(b1 + (b2 - b1) * t)
                    c = rgb_to_hex(r, g, b)
                    canvas.create_line(0, i, width, i, fill=c, width=1)
            return canvas
        except Exception:
            return tk.Canvas(parent, width=width, height=height, highlightthickness=0, bd=0, bg=color1)

    def _create_glassmorphism_overlay(self, parent, alpha=0.85):
        """Create a glassmorphism effect overlay."""
        try:
            colors = self.palette
            overlay = tk.Frame(parent, bg=colors.get('surface_bg', '#ffffff'))
            # Simulate glassmorphism with semi-transparent background
            # Note: Tkinter doesn't support true transparency, so we use a lightened color
            def lighten_color(hex_color, factor=0.15):
                h = hex_color.lstrip('#')
                r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
                r = min(255, int(r + (255 - r) * factor))
                g = min(255, int(g + (255 - g) * factor))
                b = min(255, int(b + (255 - b) * factor))
                return f"#{r:02x}{g:02x}{b:02x}"
            glass_color = lighten_color(colors.get('surface_bg', '#ffffff'), factor=alpha)
            overlay.configure(bg=glass_color)
            return overlay
        except Exception:
            return tk.Frame(parent, bg=self.palette.get('surface_bg', '#ffffff'))

    def _add_hover_animation(self, widget, scale_factor=1.05, glow_color=None):
        """Add hover scale and glow animation to a widget."""
        try:
            if glow_color is None:
                glow_color = self.palette.get('accent_color', '#2563eb')
            
            original_bg = widget.cget('bg') if hasattr(widget, 'cget') else None
            original_relief = widget.cget('relief') if hasattr(widget, 'cget') else None
            
            def on_enter(e):
                try:
                    # Scale effect (simulated with padding/border)
                    widget.configure(relief=tk.RAISED, bd=2)
                    # Glow effect (simulated with highlight)
                    widget.configure(highlightbackground=glow_color, highlightthickness=2)
                except Exception:
                    pass
            
            def on_leave(e):
                try:
                    if original_relief:
                        widget.configure(relief=original_relief, bd=1)
                    else:
                        widget.configure(relief=tk.FLAT, bd=0)
                    widget.configure(highlightthickness=0)
                except Exception:
                    pass
            
            widget.bind('<Enter>', on_enter)
            widget.bind('<Leave>', on_leave)
        except Exception:
            pass

    def _create_loading_spinner(self, parent, size=40):
        """Create an animated loading spinner."""
        try:
            canvas = tk.Canvas(parent, width=size, height=size, highlightthickness=0, bd=0, bg=self.palette.get('surface_bg', '#ffffff'))
            colors = self.palette
            accent = colors.get('accent_color', '#2563eb')
            center = size // 2
            radius = size // 3
            
            spinner_id = None
            angle = 0
            
            def animate():
                nonlocal angle, spinner_id
                canvas.delete('spinner')
                angle += 15
                for i in range(8):
                    a = (angle + i * 45) % 360
                    x = center + radius * 0.7 * (1 if i % 2 == 0 else 0.5) * (1 if a < 180 else -1)
                    y = center + radius * 0.7 * (1 if i % 2 == 0 else 0.5) * (1 if 90 < a < 270 else -1)
                    alpha = 1.0 - (i * 0.1)
                    color = accent
                    canvas.create_oval(x-3, y-3, x+3, y+3, fill=color, outline='', tags='spinner')
                canvas.after(50, animate)
            
            animate()
            return canvas
        except Exception:
            return tk.Label(parent, text="⏳", font=("Segoe UI", 16))

    def _create_progress_bar(self, parent, width=200, height=20):
        """Create an animated progress bar."""
        try:
            frame = tk.Frame(parent, bg=self.palette.get('surface_bg', '#ffffff'))
            canvas = tk.Canvas(frame, width=width, height=height, highlightthickness=0, bd=0, bg=self.palette.get('entry_border', '#d8dee9'))
            canvas.pack()
            
            colors = self.palette
            accent = colors.get('accent_color', '#2563eb')
            
            progress = 0
            def animate():
                nonlocal progress
                canvas.delete('progress')
                progress = (progress + 2) % 100
                w = int((width - 4) * progress / 100)
                canvas.create_rectangle(2, 2, w + 2, height - 2, fill=accent, outline='', tags='progress')
                canvas.after(30, animate)
            
            animate()
            return frame
        except Exception:
            return tk.Label(parent, text="Loading...", font=FONT_BASE)

    def _create_status_badge(self, parent, text, status="info", pulse=False):
        """Create an animated status badge with optional pulse effect."""
        try:
            colors = self.palette
            status_color = STATUS_COLORS.get(status, STATUS_COLORS["info"])
            bg_color = colors.get('surface_bg', '#ffffff')
            
            frame = tk.Frame(parent, bg=bg_color)
            badge = tk.Label(
                frame,
                text=f"{ICONS.get(status, '●')} {text}",
                bg=status_color,
                fg="#ffffff",
                font=("{Segoe UI} 9 bold"),
                padx=8,
                pady=4,
                relief=tk.RAISED,
                bd=1
            )
            badge.pack()
            
            if pulse:
                pulse_alpha = 1.0
                def pulse_animate():
                    nonlocal pulse_alpha
                    pulse_alpha = (pulse_alpha + 0.1) % 1.0
                    try:
                        def hex_to_rgb(h):
                            h = h.lstrip('#')
                            return (int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16))
                        def rgb_to_hex(r, g, b):
                            return f"#{r:02x}{g:02x}{b:02x}"
                        r, g, b = hex_to_rgb(status_color)
                        alpha = 0.7 + 0.3 * abs(pulse_alpha - 0.5) * 2
                        r = int(r * alpha)
                        g = int(g * alpha)
                        b = int(b * alpha)
                        pulse_color = rgb_to_hex(min(255, r), min(255, g), min(255, b))
                        badge.configure(bg=pulse_color)
                    except Exception:
                        pass
                    frame.after(100, pulse_animate)
                pulse_animate()
            
            return frame
        except Exception:
            return tk.Label(parent, text=text, font=FONT_BASE)

    def _create_mini_chart(self, parent, data, width=150, height=60, title=None):
        """Create a mini sparkline chart with optional title."""
        try:
            container = tk.Frame(parent, bg=self.palette.get('surface_bg', '#ffffff'))
            
            if title:
                title_label = tk.Label(
                    container,
                    text=title,
                    bg=self.palette.get('surface_bg', '#ffffff'),
                    fg=self.palette.get('text_secondary', '#4b5563'),
                    font=("{Segoe UI} 8"),
                    anchor=tk.W
                )
                title_label.pack(anchor=tk.W, padx=2, pady=(0, 2))
            
            canvas = tk.Canvas(container, width=width, height=height, highlightthickness=0, bd=0, bg=self.palette.get('surface_bg', '#ffffff'))
            canvas.pack()
            
            if not data or len(data) < 2:
                # Show placeholder
                canvas.create_text(width//2, height//2, text="No data", fill=self.palette.get('text_secondary', '#4b5563'), font=("{Segoe UI} 9"))
                return container
            
            colors = self.palette
            accent = colors.get('accent_color', '#2563eb')
            
            min_val = min(data)
            max_val = max(data)
            range_val = max_val - min_val if max_val != min_val else 1
            
            points = []
            for i, val in enumerate(data):
                x = int((i / (len(data) - 1)) * (width - 20) + 10)
                y = int(height - 10 - ((val - min_val) / range_val) * (height - 20))
                points.append((x, y))
            
            # Draw grid lines
            for i in range(3):
                y_pos = 10 + (height - 20) * i / 2
                canvas.create_line(10, y_pos, width - 10, y_pos, fill=colors.get('entry_border', '#d8dee9'), width=1, dash=(2, 2))
            
            # Draw line
            if len(points) > 1:
                for i in range(len(points) - 1):
                    canvas.create_line(points[i][0], points[i][1], points[i+1][0], points[i+1][1], fill=accent, width=2, smooth=True)
            
            # Draw points
            for x, y in points:
                canvas.create_oval(x-3, y-3, x+3, y+3, fill=accent, outline='', width=1)
            
            # Add min/max labels
            canvas.create_text(5, 10, text=f"{max_val:.0f}", fill=colors.get('text_secondary', '#4b5563'), font=("{Segoe UI} 7"), anchor=tk.W)
            canvas.create_text(5, height - 10, text=f"{min_val:.0f}", fill=colors.get('text_secondary', '#4b5563'), font=("{Segoe UI} 7"), anchor=tk.W)
            
            return container
        except Exception:
            return tk.Label(parent, text="📊", font=("Segoe UI", 20))

    def _enhance_button_style(self, button, rounded=True, shadow=True):
        """Enhance button with rounded corners and shadow effect."""
        try:
            if rounded:
                # Simulate rounded corners with relief
                button.configure(relief=tk.RAISED, bd=2)
            if shadow:
                # Simulate shadow with highlight
                button.configure(highlightbackground=self.palette.get('entry_border', '#d8dee9'), highlightthickness=1)
            # Add hover effect
            self._add_hover_animation(button)
        except Exception:
            pass

    def _enhance_typography(self, widget, text_shadow=False, letter_spacing=None):
        """Enhance typography with text shadow and letter spacing."""
        try:
            # Tkinter has limited typography support, but we can simulate effects
            if text_shadow:
                # Create a shadow effect by layering labels
                pass  # Would require creating multiple label layers
            if letter_spacing:
                # Letter spacing would require custom font or character-by-character rendering
                pass  # Limited support in Tkinter
        except Exception:
            pass

    def _animate_tab_fade(self, target_frame, fade_in=True, duration_ms=300):
        """Enhanced tab transition with fade effect."""
        try:
            # This is a subtle fade effect that works with the swipe animation
            # The actual fade is handled by the swipe animation's overlay
            # This method can be extended for more complex fade effects if needed
            pass
        except Exception:
            pass

    def _setup_keyboard_shortcuts(self) -> None:
        """Setup keyboard shortcuts."""
        try:
            # Common shortcuts
            self.bind('<Control-n>', lambda e: self._nav_to(0))  # New transaction
            self.bind('<Control-s>', lambda e: self._save_current())  # Save
            self.bind('<Control-f>', lambda e: self._show_search_dialog())  # Search
            self.bind('<Control-i>', lambda e: self._show_import_dialog())  # Import
            self.bind('<Control-d>', lambda e: self._show_dashboard())  # Dashboard
            self.bind('<Control-z>', lambda e: self._undo_action())  # Undo
            self.bind('<Control-y>', lambda e: self._redo_action())  # Redo
            self.bind('<Control-q>', lambda e: self._on_close())  # Quit
            self.bind('<F1>', lambda e: self._show_context_help())  # Help
        except Exception:
            pass
    
    def _update_undo_redo_states(self) -> None:
        """Update undo/redo button states."""
        try:
            from . import undo
            # Update menu if it exists
            if hasattr(self, 'edit_menu'):
                self.edit_menu.entryconfig("Undo", state=tk.NORMAL if undo.can_undo() else tk.DISABLED)
                self.edit_menu.entryconfig("Redo", state=tk.NORMAL if undo.can_redo() else tk.DISABLED)
            # Update toolbar buttons if they exist
            if hasattr(self, 'undo_btn'):
                self.undo_btn.configure(state=tk.NORMAL if undo.can_undo() else tk.DISABLED)
            if hasattr(self, 'redo_btn'):
                self.redo_btn.configure(state=tk.NORMAL if undo.can_redo() else tk.DISABLED)
        except Exception:
            pass
    
    def _undo_action(self) -> None:
        """Undo last action."""
        try:
            from . import undo
            result = undo.undo()
            if result:
                self.set_status(f"Undone: {result.get('message', 'Action undone')}", "success")
                self._load_all_views()
                self._update_undo_redo_states()
            else:
                self.set_status("Nothing to undo", "info")
        except Exception as e:
            messagebox.showerror("Error", f"Undo failed: {e}")
    
    def _redo_action(self) -> None:
        """Redo last undone action."""
        try:
            from . import undo
            result = undo.redo()
            if result:
                self.set_status(f"Redone: {result.get('message', 'Action redone')}", "success")
                self._load_all_views()
                self._update_undo_redo_states()
            else:
                self.set_status("Nothing to redo", "info")
        except Exception as e:
            messagebox.showerror("Error", f"Redo failed: {e}")
    
    def _show_search_dialog(self) -> None:
        """Show enhanced search dialog."""
        try:
            from . import search
            
            dialog = tk.Toplevel(self)
            dialog.title("Search - TechFix")
            dialog.transient(self)
            dialog.grab_set()
            dialog.geometry("900x700")
            dialog.configure(bg=self.palette.get("surface_bg", "#ffffff"))
            
            # Center dialog
            dialog.update_idletasks()
            x = (dialog.winfo_screenwidth() // 2) - (900 // 2)
            y = (dialog.winfo_screenheight() // 2) - (700 // 2)
            dialog.geometry(f"900x700+{x}+{y}")
            
            main_frame = ttk.Frame(dialog, style="Techfix.Surface.TFrame", padding=24)
            main_frame.pack(fill=tk.BOTH, expand=True)
            
            # Enhanced header
            header_frame = ttk.Frame(main_frame, style="Techfix.Surface.TFrame")
            header_frame.pack(fill=tk.X, pady=(0, 20))
            
            ttk.Label(
                header_frame,
                text="🔍 Search",
                font=("{Segoe UI Semibold} 22"),
                style="TLabel"
            ).pack(side=tk.LEFT)
            
            # Search box with enhanced styling
            search_container = ttk.Frame(main_frame, style="Techfix.Surface.TFrame")
            search_container.pack(fill=tk.X, pady=(0, 20))
            
            search_box_frame = tk.Frame(
                search_container,
                bg=self.palette.get("app_bg", "#f9fafb"),
                relief=tk.FLAT,
                bd=1,
                highlightbackground=self.palette.get("entry_border", "#e5e7eb"),
                highlightthickness=1
            )
            search_box_frame.pack(fill=tk.X, padx=2, pady=2)
            
            search_inner = tk.Frame(search_box_frame, bg=self.palette.get("app_bg", "#f9fafb"))
            search_inner.pack(fill=tk.X, padx=16, pady=12)
            
            search_icon = tk.Label(
                search_inner,
                text="🔍",
                font=("{Segoe UI} 16"),
                bg=self.palette.get("app_bg", "#f9fafb"),
                fg=self.palette.get("text_secondary", "#6b7280")
            )
            search_icon.pack(side=tk.LEFT, padx=(0, 12))
            
            search_var = tk.StringVar()
            search_entry = tk.Entry(
                search_inner,
                textvariable=search_var,
                font=("{Segoe UI} 12"),
                bg=self.palette.get("app_bg", "#f9fafb"),
                fg=self.palette.get("text_primary", "#111827"),
                relief=tk.FLAT,
                bd=0,
                highlightthickness=0
            )
            search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
            search_entry.insert(0, "Search transactions, accounts, customers, vendors...")
            search_entry.configure(fg=self.palette.get("text_secondary", "#9ca3af"))
            search_entry.focus()
            
            def on_search_focus_in(e):
                if search_entry.get() == "Search transactions, accounts, customers, vendors...":
                    search_entry.delete(0, tk.END)
                    search_entry.configure(fg=self.palette.get("text_primary", "#111827"))
            
            def on_search_focus_out(e):
                if not search_entry.get().strip():
                    search_entry.insert(0, "Search transactions, accounts, customers, vendors...")
                    search_entry.configure(fg=self.palette.get("text_secondary", "#9ca3af"))
            
            search_entry.bind('<FocusIn>', on_search_focus_in)
            search_entry.bind('<FocusOut>', on_search_focus_out)
            
            # Results container with scrollable area
            results_container = ttk.Frame(main_frame, style="Techfix.Surface.TFrame")
            results_container.pack(fill=tk.BOTH, expand=True)
            
            # Canvas for scrollable results
            canvas = tk.Canvas(
                results_container,
                bg=self.palette.get("surface_bg", "#ffffff"),
                highlightthickness=0,
                bd=0
            )
            scrollbar = ttk.Scrollbar(results_container, orient="vertical", command=canvas.yview)
            results_frame = tk.Frame(canvas, bg=self.palette.get("surface_bg", "#ffffff"))
            
            results_frame.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
            )
            
            canvas.create_window((0, 0), window=results_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)
            
            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")
            
            # Mouse wheel scrolling - bind to canvas and dialog, not globally
            def on_mousewheel(event):
                try:
                    if canvas.winfo_exists():
                        canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
                except tk.TclError:
                    pass  # Widget was destroyed
            
            canvas.bind("<MouseWheel>", on_mousewheel)
            dialog.bind("<MouseWheel>", on_mousewheel)
            
            # Cleanup binding when dialog is destroyed
            def cleanup_bindings():
                try:
                    canvas.unbind("<MouseWheel>")
                    dialog.unbind("<MouseWheel>")
                except:
                    pass
            dialog.bind("<Destroy>", lambda e: cleanup_bindings())
            
            def create_result_card(parent, icon, title, subtitle, data_type, item_id, on_click=None):
                """Create a clickable result card."""
                card_frame = tk.Frame(
                    parent,
                    bg=self.palette.get("app_bg", "#f9fafb"),
                    relief=tk.FLAT,
                    bd=1,
                    highlightbackground=self.palette.get("entry_border", "#e5e7eb"),
                    highlightthickness=1,
                    cursor="hand2"
                )
                card_frame.pack(fill=tk.X, pady=4, padx=4)
                
                content_frame = tk.Frame(card_frame, bg=self.palette.get("app_bg", "#f9fafb"))
                content_frame.pack(fill=tk.BOTH, expand=True, padx=16, pady=12)
                
                # Icon and text
                icon_label = tk.Label(
                    content_frame,
                    text=icon,
                    font=("{Segoe UI} 20"),
                    bg=self.palette.get("app_bg", "#f9fafb"),
                    fg=self.palette.get("accent_color", "#2563eb")
                )
                icon_label.pack(side=tk.LEFT, padx=(0, 12))
                
                text_frame = tk.Frame(content_frame, bg=self.palette.get("app_bg", "#f9fafb"))
                text_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
                
                title_label = tk.Label(
                    text_frame,
                    text=title,
                    font=("{Segoe UI Semibold} 13"),
                    bg=self.palette.get("app_bg", "#f9fafb"),
                    fg=self.palette.get("text_primary", "#111827"),
                    anchor=tk.W
                )
                title_label.pack(anchor=tk.W)
                
                subtitle_label = tk.Label(
                    text_frame,
                    text=subtitle,
                    font=("{Segoe UI} 11"),
                    bg=self.palette.get("app_bg", "#f9fafb"),
                    fg=self.palette.get("text_secondary", "#6b7280"),
                    anchor=tk.W
                )
                subtitle_label.pack(anchor=tk.W, pady=(4, 0))
                
                # Hover effect
                def on_enter(e):
                    card_frame.configure(bg=self.palette.get("accent_color", "#2563eb"), highlightbackground=self.palette.get("accent_color", "#2563eb"))
                    content_frame.configure(bg=self.palette.get("accent_color", "#2563eb"))
                    icon_label.configure(bg=self.palette.get("accent_color", "#2563eb"), fg="#ffffff")
                    text_frame.configure(bg=self.palette.get("accent_color", "#2563eb"))
                    title_label.configure(bg=self.palette.get("accent_color", "#2563eb"), fg="#ffffff")
                    subtitle_label.configure(bg=self.palette.get("accent_color", "#2563eb"), fg="#f0f0f0")
                
                def on_leave(e):
                    card_frame.configure(bg=self.palette.get("app_bg", "#f9fafb"), highlightbackground=self.palette.get("entry_border", "#e5e7eb"))
                    content_frame.configure(bg=self.palette.get("app_bg", "#f9fafb"))
                    icon_label.configure(bg=self.palette.get("app_bg", "#f9fafb"), fg=self.palette.get("accent_color", "#2563eb"))
                    text_frame.configure(bg=self.palette.get("app_bg", "#f9fafb"))
                    title_label.configure(bg=self.palette.get("app_bg", "#f9fafb"), fg=self.palette.get("text_primary", "#111827"))
                    subtitle_label.configure(bg=self.palette.get("app_bg", "#f9fafb"), fg=self.palette.get("text_secondary", "#6b7280"))
                
                card_frame.bind('<Enter>', on_enter)
                card_frame.bind('<Leave>', on_leave)
                content_frame.bind('<Enter>', on_enter)
                content_frame.bind('<Leave>', on_leave)
                icon_label.bind('<Enter>', on_enter)
                icon_label.bind('<Leave>', on_leave)
                text_frame.bind('<Enter>', on_enter)
                text_frame.bind('<Leave>', on_leave)
                title_label.bind('<Enter>', on_enter)
                title_label.bind('<Leave>', on_leave)
                subtitle_label.bind('<Enter>', on_enter)
                subtitle_label.bind('<Leave>', on_leave)
                
                if on_click:
                    def handle_click(e):
                        on_click()
                    card_frame.bind('<Button-1>', handle_click)
                    content_frame.bind('<Button-1>', handle_click)
                    icon_label.bind('<Button-1>', handle_click)
                    text_frame.bind('<Button-1>', handle_click)
                    title_label.bind('<Button-1>', handle_click)
                    subtitle_label.bind('<Button-1>', handle_click)
                
                return card_frame
            
            def do_search():
                query = search_entry.get().strip()
                if query == "Search transactions, accounts, customers, vendors...":
                    query = ""
                
                # Clear results
                for widget in results_frame.winfo_children():
                    widget.destroy()
                
                if not query or len(query) < 2:
                    # Show empty state
                    empty_frame = tk.Frame(results_frame, bg=self.palette.get("surface_bg", "#ffffff"))
                    empty_frame.pack(fill=tk.BOTH, expand=True, pady=80)
                    
                    empty_icon = tk.Label(
                        empty_frame,
                        text="🔍",
                        font=("{Segoe UI} 64"),
                        bg=self.palette.get("surface_bg", "#ffffff"),
                        fg=self.palette.get("text_secondary", "#9ca3af")
                    )
                    empty_icon.pack()
                    
                    empty_title = tk.Label(
                        empty_frame,
                        text="Start searching",
                        font=("{Segoe UI Semibold} 18"),
                        bg=self.palette.get("surface_bg", "#ffffff"),
                        fg=self.palette.get("text_primary", "#111827")
                    )
                    empty_title.pack(pady=(16, 8))
                    
                    empty_subtitle = tk.Label(
                        empty_frame,
                        text="Enter at least 2 characters to search across transactions, accounts, customers, and vendors.",
                        font=("{Segoe UI} 11"),
                        bg=self.palette.get("surface_bg", "#ffffff"),
                        fg=self.palette.get("text_secondary", "#6b7280")
                    )
                    empty_subtitle.pack()
                    return
                
                # Perform search
                results = search.global_search(query, limit=20)
                
                total_results = sum(len(v) for v in results.values())
                
                if total_results == 0:
                    # No results state
                    empty_frame = tk.Frame(results_frame, bg=self.palette.get("surface_bg", "#ffffff"))
                    empty_frame.pack(fill=tk.BOTH, expand=True, pady=80)
                    
                    empty_icon = tk.Label(
                        empty_frame,
                        text="🔎",
                        font=("{Segoe UI} 64"),
                        bg=self.palette.get("surface_bg", "#ffffff"),
                        fg=self.palette.get("text_secondary", "#9ca3af")
                    )
                    empty_icon.pack()
                    
                    empty_title = tk.Label(
                        empty_frame,
                        text="No results found",
                        font=("{Segoe UI Semibold} 18"),
                        bg=self.palette.get("surface_bg", "#ffffff"),
                        fg=self.palette.get("text_primary", "#111827")
                    )
                    empty_title.pack(pady=(16, 8))
                    
                    empty_subtitle = tk.Label(
                        empty_frame,
                        text=f"No matches found for '{query}'. Try different keywords.",
                        font=("{Segoe UI} 11"),
                        bg=self.palette.get("surface_bg", "#ffffff"),
                        fg=self.palette.get("text_secondary", "#6b7280")
                    )
                    empty_subtitle.pack()
                    return
                
                # Show results count
                count_frame = tk.Frame(results_frame, bg=self.palette.get("surface_bg", "#ffffff"))
                count_frame.pack(fill=tk.X, pady=(0, 12))
                
                count_label = tk.Label(
                    count_frame,
                    text=f"Found {total_results} result{'s' if total_results != 1 else ''}",
                    font=("{Segoe UI Semibold} 12"),
                    bg=self.palette.get("surface_bg", "#ffffff"),
                    fg=self.palette.get("text_secondary", "#6b7280")
                )
                count_label.pack(side=tk.LEFT)
                
                # Journal Entries
                if results['journal_entries']:
                    section_frame = tk.Frame(results_frame, bg=self.palette.get("surface_bg", "#ffffff"))
                    section_frame.pack(fill=tk.X, pady=(0, 8))
                    
                    section_label = tk.Label(
                        section_frame,
                        text=f"🧾 Journal Entries ({len(results['journal_entries'])})",
                        font=("{Segoe UI Semibold} 14"),
                        bg=self.palette.get("surface_bg", "#ffffff"),
                        fg=self.palette.get("text_primary", "#111827")
                    )
                    section_label.pack(anchor=tk.W, pady=(0, 8))
                    
                    for entry in results['journal_entries']:
                        date_str = entry.get('date', '')[:10] if entry.get('date') else 'No date'
                        desc = entry.get('description', 'No description')[:80]
                        status = entry.get('status', 'posted')
                        entry_id = entry.get('id')
                        
                        def navigate_to_entry(eid=entry_id):
                            dialog.destroy()
                            self._nav_to(0)  # Navigate to Transactions tab
                            # Could add logic to highlight/select the entry
                        
                        create_result_card(
                            section_frame,
                            "🧾",
                            desc,
                            f"{date_str} • Status: {status}",
                            "journal_entry",
                            entry_id,
                            navigate_to_entry
                        )
                
                # Accounts
                if results['accounts']:
                    section_frame = tk.Frame(results_frame, bg=self.palette.get("surface_bg", "#ffffff"))
                    section_frame.pack(fill=tk.X, pady=(0, 8))
                    
                    section_label = tk.Label(
                        section_frame,
                        text=f"📚 Accounts ({len(results['accounts'])})",
                        font=("{Segoe UI Semibold} 14"),
                        bg=self.palette.get("surface_bg", "#ffffff"),
                        fg=self.palette.get("text_primary", "#111827")
                    )
                    section_label.pack(anchor=tk.W, pady=(0, 8))
                    
                    for account in results['accounts']:
                        code = account.get('code', '')
                        name = account.get('name', 'No name')
                        acc_type = account.get('type', '')
                        acc_id = account.get('id')
                        
                        def navigate_to_account(aid=acc_id):
                            dialog.destroy()
                            self._nav_to(2)  # Navigate to Ledger tab
                            # Could add logic to filter by account
                        
                        create_result_card(
                            section_frame,
                            "📚",
                            f"{code} - {name}",
                            f"Type: {acc_type}",
                            "account",
                            acc_id,
                            navigate_to_account
                        )
                
                # Customers
                if results['customers']:
                    section_frame = tk.Frame(results_frame, bg=self.palette.get("surface_bg", "#ffffff"))
                    section_frame.pack(fill=tk.X, pady=(0, 8))
                    
                    section_label = tk.Label(
                        section_frame,
                        text=f"👤 Customers ({len(results['customers'])})",
                        font=("{Segoe UI Semibold} 14"),
                        bg=self.palette.get("surface_bg", "#ffffff"),
                        fg=self.palette.get("text_primary", "#111827")
                    )
                    section_label.pack(anchor=tk.W, pady=(0, 8))
                    
                    for customer in results['customers']:
                        code = customer.get('code', '')
                        name = customer.get('name', 'No name')
                        contact = customer.get('contact', '')
                        email = customer.get('email', '')
                        cust_id = customer.get('id')
                        
                        subtitle = f"Code: {code}"
                        if contact:
                            subtitle += f" • Contact: {contact}"
                        if email:
                            subtitle += f" • {email}"
                        
                        def navigate_to_customer(cid=cust_id):
                            dialog.destroy()
                            # Could navigate to customer details if implemented
                        
                        create_result_card(
                            section_frame,
                            "👤",
                            name,
                            subtitle,
                            "customer",
                            cust_id,
                            navigate_to_customer
                        )
                
                # Vendors
                if results['vendors']:
                    section_frame = tk.Frame(results_frame, bg=self.palette.get("surface_bg", "#ffffff"))
                    section_frame.pack(fill=tk.X, pady=(0, 8))
                    
                    section_label = tk.Label(
                        section_frame,
                        text=f"🏢 Vendors ({len(results['vendors'])})",
                        font=("{Segoe UI Semibold} 14"),
                        bg=self.palette.get("surface_bg", "#ffffff"),
                        fg=self.palette.get("text_primary", "#111827")
                    )
                    section_label.pack(anchor=tk.W, pady=(0, 8))
                    
                    for vendor in results['vendors']:
                        code = vendor.get('code', '')
                        name = vendor.get('name', 'No name')
                        contact = vendor.get('contact', '')
                        email = vendor.get('email', '')
                        vend_id = vendor.get('id')
                        
                        subtitle = f"Code: {code}"
                        if contact:
                            subtitle += f" • Contact: {contact}"
                        if email:
                            subtitle += f" • {email}"
                        
                        def navigate_to_vendor(vid=vend_id):
                            dialog.destroy()
                            # Could navigate to vendor details if implemented
                        
                        create_result_card(
                            section_frame,
                            "🏢",
                            name,
                            subtitle,
                            "vendor",
                            vend_id,
                            navigate_to_vendor
                        )
                
                canvas.update_idletasks()
                canvas.configure(scrollregion=canvas.bbox("all"))
            
            search_entry.bind('<Return>', lambda e: do_search())
            search_entry.bind('<KeyRelease>', lambda e: do_search() if len(search_entry.get().strip()) >= 2 else None)
            
            # Action buttons
            btn_frame = ttk.Frame(main_frame, style="Techfix.Surface.TFrame")
            btn_frame.pack(fill=tk.X, pady=(20, 0))
            
            ttk.Button(
                btn_frame,
                text="🔄 Clear",
                command=lambda: (search_entry.delete(0, tk.END), search_entry.insert(0, "Search transactions, accounts, customers, vendors..."), search_entry.configure(fg=self.palette.get("text_secondary", "#9ca3af")), do_search()),
                style="Techfix.TButton"
            ).pack(side=tk.LEFT, padx=(0, 8))
            
            ttk.Button(
                btn_frame,
                text="Close",
                command=dialog.destroy,
                style="Techfix.TButton"
            ).pack(side=tk.RIGHT)
            
            # Initial empty state
            do_search()
            
        except Exception as e:
            messagebox.showerror("Error", f"Search error: {e}")
    
    def _show_backup_dialog(self) -> None:
        """Show backup dialog."""
        try:
            from . import backup
            
            # Show dialog for backup description
            dialog = tk.Toplevel(self)
            dialog.title("Backup Database")
            dialog.transient(self)
            dialog.grab_set()
            dialog.geometry("500x200")
            dialog.configure(bg=self.palette.get("surface_bg", "#ffffff"))
            
            main_frame = ttk.Frame(dialog, style="Techfix.Surface.TFrame", padding=24)
            main_frame.pack(fill=tk.BOTH, expand=True)
            
            ttk.Label(
                main_frame,
                text="Create Database Backup",
                font=("{Segoe UI Semibold} 16"),
                style="TLabel"
            ).pack(pady=(0, 12))
            
            ttk.Label(
                main_frame,
                text="Enter a description for this backup (optional):",
                style="TLabel"
            ).pack(anchor=tk.W, pady=(0, 8))
            
            desc_var = tk.StringVar()
            desc_entry = ttk.Entry(main_frame, textvariable=desc_var, width=40, style="Techfix.TEntry")
            desc_entry.pack(fill=tk.X, pady=(0, 12))
            desc_entry.focus()
            
            info_label = ttk.Label(
                main_frame,
                text="Backup will include database and settings.",
                style="TLabel",
                foreground=self.palette.get("text_secondary", "#6b7280")
            )
            info_label.pack(anchor=tk.W, pady=(0, 12))
            
            def create_backup():
                description = desc_var.get().strip() or None
                dialog.destroy()
                
                # Show progress
                self.set_status("Creating backup...", "info")
                self.update()
                
                try:
                    backup_path = backup.create_full_backup(description)
                    
                    if backup_path:
                        # Get file size
                        try:
                            size_bytes = backup_path.stat().st_size
                            size_str = f"{size_bytes / (1024*1024):.2f} MB" if size_bytes > 1024*1024 else f"{size_bytes / 1024:.2f} KB"
                        except:
                            size_str = "Unknown"
                        
                        messagebox.showinfo(
                            "Backup Created Successfully",
                            f"Backup created successfully!\n\n"
                            f"File: {backup_path.name}\n"
                            f"Size: {size_str}\n"
                            f"Location: {backup_path.parent}\n\n"
                            f"Your database and settings have been backed up.",
                            parent=self
                        )
                        self.set_status(f"Backup created: {backup_path.name}", "success")
                    else:
                        messagebox.showerror("Backup Failed", "Failed to create backup. Please check the logs for details.", parent=self)
                        self.set_status("Backup failed", "error")
                except Exception as e:
                    logger.error(f"Backup error: {e}", exc_info=True)
                    messagebox.showerror("Backup Error", f"An error occurred while creating backup:\n{e}", parent=self)
                    self.set_status("Backup error", "error")
            
            btn_frame = ttk.Frame(main_frame, style="Techfix.Surface.TFrame")
            btn_frame.pack(fill=tk.X, pady=(12, 0))
            
            ttk.Button(
                btn_frame,
                text="Create Backup",
                command=create_backup,
                style="Techfix.TButton"
            ).pack(side=tk.RIGHT, padx=(8, 0))
            
            ttk.Button(
                btn_frame,
                text="Cancel",
                command=dialog.destroy,
                style="Techfix.TButton"
            ).pack(side=tk.RIGHT)
            
            # Allow Enter key to create backup
            desc_entry.bind('<Return>', lambda e: create_backup())
            
        except Exception as e:
            logger.error(f"Backup dialog error: {e}", exc_info=True)
            messagebox.showerror("Error", f"Backup dialog error: {e}")
    
    def _show_restore_dialog(self) -> None:
        """Show restore dialog with list of available backups."""
        try:
            from . import backup
            
            backups = backup.list_backups()
            if not backups:
                messagebox.showinfo("No Backups", "No backups found. Please create a backup first.", parent=self)
                return
            
            # Create dialog with backup list
            dialog = tk.Toplevel(self)
            dialog.title("Restore Database")
            dialog.transient(self)
            dialog.grab_set()
            dialog.geometry("700x500")
            dialog.configure(bg=self.palette.get("surface_bg", "#ffffff"))
            
            main_frame = ttk.Frame(dialog, style="Techfix.Surface.TFrame", padding=24)
            main_frame.pack(fill=tk.BOTH, expand=True)
            
            ttk.Label(
                main_frame,
                text="Select Backup to Restore",
                font=("{Segoe UI Semibold} 16"),
                style="TLabel"
            ).pack(pady=(0, 12))
            
            # Backup list frame
            list_frame = ttk.Frame(main_frame, style="Techfix.Surface.TFrame")
            list_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 12))
            
            # Treeview for backups
            columns = ("Name", "Type", "Size", "Created")
            backup_tree = ttk.Treeview(list_frame, columns=columns, show="headings", style="Techfix.Treeview", selectmode=tk.BROWSE)
            
            for col in columns:
                backup_tree.heading(col, text=col)
                if col == "Name":
                    backup_tree.column(col, width=300, stretch=True)
                elif col == "Type":
                    backup_tree.column(col, width=100, stretch=False)
                elif col == "Size":
                    backup_tree.column(col, width=120, stretch=False)
                else:
                    backup_tree.column(col, width=180, stretch=False)
            
            scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=backup_tree.yview)
            backup_tree.configure(yscrollcommand=scrollbar.set)
            backup_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            # Populate backup list
            def format_size(size_bytes):
                """Format file size in human-readable format."""
                for unit in ['B', 'KB', 'MB', 'GB']:
                    if size_bytes < 1024.0:
                        return f"{size_bytes:.1f} {unit}"
                    size_bytes /= 1024.0
                return f"{size_bytes:.1f} TB"
            
            selected_backup = None
            path_map: dict[str, str] = {}
            for backup_info in backups:
                created_str = backup_info['created'].strftime("%Y-%m-%d %H:%M:%S")
                size_str = format_size(backup_info['size'])
                type_str = backup_info['type'].title()
                
                item_id = backup_tree.insert('', 'end', values=(
                    backup_info['name'],
                    type_str,
                    size_str,
                    created_str
                ))
                # Store backup path in a side map keyed by item id.
                path_map[item_id] = str(backup_info['path'])
            
            # Select first item by default
            if backups:
                first_item = backup_tree.get_children()[0]
                backup_tree.selection_set(first_item)
                backup_tree.focus(first_item)
            
            def on_backup_select(event):
                selection = backup_tree.selection()
                if selection:
                    item = backup_tree.item(selection[0])
                    nonlocal selected_backup
                    selected_backup = item.get('values', [])
                    # Get path from stored side map
                    selected_backup = path_map.get(selection[0])
                    # Fallback: find by name if map entry missing
                    if not selected_backup:
                        for backup_info in backups:
                            if backup_info['name'] == item['values'][0]:
                                selected_backup = str(backup_info['path'])
                                break
            
            backup_tree.bind('<<TreeviewSelect>>', on_backup_select)
            
            # Buttons
            btn_frame = ttk.Frame(main_frame, style="Techfix.Surface.TFrame")
            btn_frame.pack(fill=tk.X, pady=(12, 0))
            
            def _show_loading(message: str) -> tuple[tk.Toplevel, ttk.Progressbar]:
                """Show a simple modal loading dialog."""
                loading = tk.Toplevel(dialog)
                loading.title("Please wait")
                loading.transient(dialog)
                loading.grab_set()
                loading.resizable(False, False)
                loading.configure(bg=self.palette.get("surface_bg", "#ffffff"))
                loading.protocol("WM_DELETE_WINDOW", lambda: None)  # Disable close

                frame = ttk.Frame(loading, padding=16, style="Techfix.Surface.TFrame")
                frame.pack(fill=tk.BOTH, expand=True)
                ttk.Label(frame, text=message, style="Techfix.TLabel").pack(pady=(0, 12))
                pb = ttk.Progressbar(frame, mode="indeterminate")
                pb.pack(fill=tk.X)
                pb.start(12)
                loading.update_idletasks()
                return loading, pb

            def perform_restore():
                selection = backup_tree.selection()
                if not selection:
                    messagebox.showwarning("No Selection", "Please select a backup to restore.", parent=dialog)
                    return
                
                item = backup_tree.item(selection[0])
                try:
                    backup_path_str = path_map.get(selection[0], "")
                    if not backup_path_str:
                        # Fallback: find by name
                        for backup_info in backups:
                            if backup_info['name'] == item['values'][0]:
                                backup_path_str = str(backup_info['path'])
                                break
                    
                    if not backup_path_str:
                        messagebox.showerror("Error", "Could not determine backup path.", parent=dialog)
                        return
                    
                    backup_path = Path(backup_path_str)
                    
                    if not messagebox.askyesno(
                        "Confirm Restore",
                        f"This will replace the current database with:\n{backup_path.name}\n\n"
                        f"Type: {item['values'][1]}\n"
                        f"Size: {item['values'][2]}\n"
                        f"Created: {item['values'][3]}\n\n"
                        "A backup of your current database will be created first.\n"
                        "Continue?",
                        parent=dialog
                    ):
                        return
                    
                    loading, pb = _show_loading("Restoring backup, please wait...")
                    try:
                        # Close database connections before restore
                        try:
                            if hasattr(self, 'engine') and self.engine:
                                self.engine.close()
                        except:
                            pass
                        
                        # Perform restore
                        if backup_path.suffix == '.zip':
                            success = backup.restore_full_backup(backup_path)
                        else:
                            success = backup.restore_backup(backup_path)
                    finally:
                        pb.stop()
                        loading.destroy()
                    
                    if success:
                        messagebox.showinfo(
                            "Success",
                            "Database restored successfully.\n\n"
                            "Please restart the application to ensure all data is reloaded.",
                            parent=dialog
                        )
                        dialog.destroy()
                    else:
                        messagebox.showerror("Error", "Restore failed. Please check the logs for details.", parent=dialog)
                except Exception as e:
                    logger.error(f"Restore error: {e}", exc_info=True)
                    messagebox.showerror("Error", f"Restore failed: {e}", parent=dialog)
            
            ttk.Button(
                btn_frame,
                text="Restore Selected",
                command=perform_restore,
                style="Techfix.TButton"
            ).pack(side=tk.RIGHT, padx=(8, 0))
            
            ttk.Button(
                btn_frame,
                text="Cancel",
                command=dialog.destroy,
                style="Techfix.TButton"
            ).pack(side=tk.RIGHT)
            
            # Also allow file picker as fallback
            def browse_file():
                backup_path = filedialog.askopenfilename(
                    title="Select Backup File",
                    initialdir=backup.BACKUP_DIR,
                    filetypes=[("Backup files", "*.db *.zip"), ("All files", "*.*")],
                    parent=dialog
                )
                if backup_path:
                    backup_path = Path(backup_path)
                    if messagebox.askyesno(
                        "Confirm Restore",
                        f"Restore from {backup_path.name}?\n\nThis will replace the current database.",
                        parent=dialog
                    ):
                        loading, pb = _show_loading("Restoring backup, please wait...")
                        try:
                            try:
                                if hasattr(self, 'engine') and self.engine:
                                    self.engine.close()
                            except:
                                pass
                            
                            if backup_path.suffix == '.zip':
                                success = backup.restore_full_backup(backup_path)
                            else:
                                success = backup.restore_backup(backup_path)
                        finally:
                            pb.stop()
                            loading.destroy()
                        
                        if success:
                            messagebox.showinfo("Success", "Database restored. Please restart the application to ensure all data is reloaded.", parent=dialog)
                            dialog.destroy()
                        else:
                            messagebox.showerror("Error", "Restore failed", parent=dialog)
            
            ttk.Button(
                btn_frame,
                text="Browse File...",
                command=browse_file,
                style="Techfix.TButton"
            ).pack(side=tk.LEFT)
            
        except Exception as e:
            logger.error(f"Restore dialog error: {e}", exc_info=True)
            messagebox.showerror("Error", f"Restore dialog error: {e}")
    
    def _show_import_dialog(self) -> None:
        """Show import dialog."""
        try:
            from . import import_data
            
            file_path = filedialog.askopenfilename(
                title="Import Transactions",
                filetypes=[
                    ("Excel files", "*.xlsx *.xls"),
                    ("CSV files", "*.csv"),
                    ("All files", "*.*")
                ],
                parent=self
            )
            
            if not file_path:
                return
            
            file_path = Path(file_path)
            
            # Validate file exists
            if not file_path.exists():
                messagebox.showerror("Error", f"File not found: {file_path}", parent=self)
                return
            
            # Check file format
            if file_path.suffix.lower() not in ['.xlsx', '.xls', '.csv']:
                messagebox.showerror("Error", "Unsupported file format. Please select an Excel (.xlsx, .xls) or CSV (.csv) file.", parent=self)
                return
            
            # Check if pandas is available for CSV
            if file_path.suffix.lower() == '.csv':
                if not import_data.PANDAS_AVAILABLE:
                    messagebox.showerror(
                        "Error",
                        "CSV import requires pandas library.\n\n"
                        "Please install it using: pip install pandas",
                        parent=self
                    )
                    return
            
            # Show progress
            self.set_status(f"Importing transactions from {file_path.name}...", "info")
            self.update()
            
            try:
                # Perform import
                if file_path.suffix.lower() in ['.xlsx', '.xls']:
                    success, errors, error_msgs = import_data.import_transactions_from_excel(
                        file_path,
                        period_id=self.current_period_id,
                        default_status="draft"
                    )
                else:  # CSV
                    success, errors, error_msgs = import_data.import_transactions_from_csv(
                        file_path,
                        period_id=self.current_period_id,
                        default_status="draft"
                    )
                
                # Build result message
                if success > 0:
                    msg = f"✅ Successfully imported {success} transaction(s)"
                    if errors > 0:
                        msg += f"\n⚠️ {errors} error(s) occurred"
                        if error_msgs:
                            msg += "\n\nErrors:\n" + "\n".join(error_msgs[:10])
                            if len(error_msgs) > 10:
                                msg += f"\n... and {len(error_msgs) - 10} more error(s)"
                    
                    messagebox.showinfo("Import Complete", msg, parent=self)
                    self.set_status(f"Imported {success} transaction(s) from {file_path.name}", "success")
                    
                    # Refresh views
                    self._load_all_views()
                else:
                    if errors > 0:
                        msg = f"❌ Import failed: {errors} error(s) occurred"
                        if error_msgs:
                            msg += "\n\nErrors:\n" + "\n".join(error_msgs[:10])
                            if len(error_msgs) > 10:
                                msg += f"\n... and {len(error_msgs) - 10} more error(s)"
                        messagebox.showerror("Import Failed", msg, parent=self)
                        self.set_status("Import failed", "error")
                    else:
                        messagebox.showwarning("Import Complete", "No transactions were imported. Please check your file format.", parent=self)
                        self.set_status("No transactions imported", "warning")
                        
            except Exception as import_error:
                logger.error(f"Import error: {import_error}", exc_info=True)
                error_msg = str(import_error)
                if "pandas" in error_msg.lower() or "openpyxl" in error_msg.lower():
                    error_msg = f"Import library error: {error_msg}\n\nPlease ensure required libraries are installed:\npip install pandas openpyxl"
                messagebox.showerror("Import Error", f"An error occurred during import:\n\n{error_msg}", parent=self)
                self.set_status("Import error", "error")
                
        except Exception as e:
            logger.error(f"Import dialog error: {e}", exc_info=True)
            messagebox.showerror("Error", f"Import dialog error: {e}")
    
    def _show_dashboard(self) -> None:
        """Show analytics dashboard."""
        try:
            # Update navigation state to highlight Dashboard button
            self._current_nav_index = -1
            for ind, btn, nav_idx in getattr(self, '_nav_buttons', ()):
                if nav_idx == -1:
                    try:
                        ind.configure(bg=self.palette.get('accent_color'))
                    except Exception:
                        pass
                    try:
                        btn.configure(style='Techfix.Nav.Selected.TButton')
                    except Exception:
                        pass
                else:
                    try:
                        ind.configure(bg=self.palette.get('surface_bg'))
                    except Exception:
                        pass
                    try:
                        btn.configure(style='Techfix.Nav.TButton')
                    except Exception:
                        pass
            
            from . import analytics
            
            dialog = tk.Toplevel(self)
            dialog.title("Dashboard - Analytics")
            dialog.transient(self)
            dialog.geometry("1000x700")
            dialog.configure(bg=self.palette.get("surface_bg", "#ffffff"))
            
            # Center dialog
            dialog.update_idletasks()
            x = (dialog.winfo_screenwidth() // 2) - (1000 // 2)
            y = (dialog.winfo_screenheight() // 2) - (700 // 2)
            dialog.geometry(f"1000x700+{x}+{y}")
            
            main_frame = ttk.Frame(dialog, style="Techfix.Surface.TFrame", padding=24)
            main_frame.pack(fill=tk.BOTH, expand=True)
            
            # Header section
            header_frame = ttk.Frame(main_frame, style="Techfix.Surface.TFrame")
            header_frame.pack(fill=tk.X, pady=(0, 24))
            
            header_title = ttk.Label(
                header_frame,
                text="📊 Dashboard",
                font=("{Segoe UI Semibold} 24"),
                style="TLabel"
            )
            header_title.pack(side=tk.LEFT)
            
            header_subtitle = ttk.Label(
                header_frame,
                text="Financial Overview & Analytics",
                font=("{Segoe UI} 12"),
                foreground=self.palette.get("text_secondary", "#6b7280"),
                style="TLabel"
            )
            header_subtitle.pack(side=tk.LEFT, padx=(12, 0))
            
            metrics = analytics.get_financial_metrics(period_id=self.current_period_id)
            
            # Enhanced metrics cards in a grid
            metrics_container = ttk.Frame(main_frame, style="Techfix.Surface.TFrame")
            metrics_container.pack(fill=tk.X, pady=(0, 24))
            
            # Configure grid columns
            for i in range(4):
                metrics_container.columnconfigure(i, weight=1, uniform="metric")
            
            def create_metric_card(parent, icon, title, value, color="#2563eb", row=0, col=0):
                """Create a styled metric card."""
                card_frame = tk.Frame(
                    parent,
                    bg=self.palette.get("surface_bg", "#ffffff"),
                    relief=tk.FLAT,
                    bd=0
                )
                card_frame.grid(row=row, column=col, padx=8, pady=8, sticky="nsew")
                
                # Card background with subtle border
                card_bg = tk.Frame(
                    card_frame,
                    bg=self.palette.get("app_bg", "#f9fafb"),
                    relief=tk.FLAT,
                    bd=1,
                    highlightbackground=self.palette.get("entry_border", "#e5e7eb"),
                    highlightthickness=1
                )
                card_bg.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)
                
                content_frame = tk.Frame(card_bg, bg=self.palette.get("app_bg", "#f9fafb"))
                content_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
                
                # Icon and title row
                icon_frame = tk.Frame(content_frame, bg=self.palette.get("app_bg", "#f9fafb"))
                icon_frame.pack(fill=tk.X, pady=(0, 12))
                
                icon_label = tk.Label(
                    icon_frame,
                    text=icon,
                    font=("{Segoe UI} 24"),
                    bg=self.palette.get("app_bg", "#f9fafb"),
                    fg=color
                )
                icon_label.pack(side=tk.LEFT)
                
                title_label = tk.Label(
                    icon_frame,
                    text=title,
                    font=("{Segoe UI} 11"),
                    bg=self.palette.get("app_bg", "#f9fafb"),
                    fg=self.palette.get("text_secondary", "#6b7280")
                )
                title_label.pack(side=tk.LEFT, padx=(12, 0))
                
                # Value
                value_label = tk.Label(
                    content_frame,
                    text=value,
                    font=("{Segoe UI Semibold} 20"),
                    bg=self.palette.get("app_bg", "#f9fafb"),
                    fg=self.palette.get("text_primary", "#111827"),
                    anchor=tk.W
                )
                value_label.pack(fill=tk.X)
                
                return card_frame
            
            # Create metric cards
            revenue = metrics.get('total_revenue', 0)
            expenses = metrics.get('total_expenses', 0)
            net_income = metrics.get('net_income', 0)
            txn_count = metrics.get('transaction_count', 0)
            
            create_metric_card(metrics_container, "💰", "Total Revenue", f"₱{revenue:,.2f}", "#10b981", 0, 0)
            create_metric_card(metrics_container, "📉", "Total Expenses", f"₱{expenses:,.2f}", "#ef4444", 0, 1)
            create_metric_card(metrics_container, "📊", "Net Income", f"₱{net_income:,.2f}", "#3b82f6", 0, 2)
            create_metric_card(metrics_container, "🧾", "Transactions", f"{txn_count:,}", "#8b5cf6", 0, 3)
            
            # Revenue trend chart section
            trend_frame = ttk.LabelFrame(main_frame, text="📈 Revenue Trend (Last 30 Days)", style="Techfix.TLabelframe")
            trend_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
            
            trend_inner = ttk.Frame(trend_frame, style="Techfix.Surface.TFrame")
            trend_inner.pack(fill=tk.BOTH, expand=True, padx=16, pady=16)
            
            trend_data = analytics.get_revenue_trend(days=30)
            if trend_data:
                chart = self._create_mini_chart(trend_inner, [d['revenue'] for d in trend_data], width=920, height=250, title=None)
                chart.pack(fill=tk.BOTH, expand=True)
            else:
                empty_chart = tk.Label(
                    trend_inner,
                    text="📊 No revenue data available",
                    font=("{Segoe UI} 14"),
                    bg=self.palette.get("surface_bg", "#ffffff"),
                    fg=self.palette.get("text_secondary", "#6b7280")
                )
                empty_chart.pack(expand=True)
            
            # Action buttons
            btn_frame = ttk.Frame(main_frame, style="Techfix.Surface.TFrame")
            btn_frame.pack(fill=tk.X)
            
            def refresh_dashboard():
                loading = self.show_loading("Please wait... Refreshing dashboard...", dialog)
                try:
                    dialog.update()  # Update UI to show loading indicator
                    dialog.destroy()
                    self._show_dashboard()
                finally:
                    if loading:
                        self.hide_loading(loading)
            
            ttk.Button(
                btn_frame,
                text="🔄 Refresh",
                command=refresh_dashboard,
                style="Techfix.TButton"
            ).pack(side=tk.LEFT, padx=(0, 8))
            
            ttk.Button(
                btn_frame,
                text="Close",
                command=dialog.destroy,
                style="Techfix.TButton"
            ).pack(side=tk.RIGHT)
            
        except Exception as e:
            messagebox.showerror("Error", f"Dashboard error: {e}")
    
    def _show_notifications(self) -> None:
        """Show notifications window."""
        try:
            from . import notifications
            
            dialog = tk.Toplevel(self)
            dialog.title("Notifications - TechFix")
            dialog.transient(self)
            dialog.grab_set()
            dialog.geometry("800x600")
            dialog.configure(bg=self.palette.get("surface_bg", "#ffffff"))
            
            # Center dialog
            dialog.update_idletasks()
            x = (dialog.winfo_screenwidth() // 2) - (800 // 2)
            y = (dialog.winfo_screenheight() // 2) - (600 // 2)
            dialog.geometry(f"800x600+{x}+{y}")
            
            main_frame = ttk.Frame(dialog, style="Techfix.Surface.TFrame", padding=24)
            main_frame.pack(fill=tk.BOTH, expand=True)
            
            # Enhanced header
            header_frame = ttk.Frame(main_frame, style="Techfix.Surface.TFrame")
            header_frame.pack(fill=tk.X, pady=(0, 20))
            
            header_left = ttk.Frame(header_frame, style="Techfix.Surface.TFrame")
            header_left.pack(side=tk.LEFT, fill=tk.Y)
            
            ttk.Label(
                header_left,
                text="📬 Notifications",
                font=("{Segoe UI Semibold} 22"),
                style="TLabel"
            ).pack(side=tk.LEFT)
            
            user_id = self.current_user['id'] if self.current_user else None
            notifs = notifications.get_user_notifications(user_id, unread_only=False, limit=50)
            unread_count = sum(1 for n in notifs if n.get('is_read', 0) == 0)
            
            if unread_count > 0:
                badge_frame = tk.Frame(header_left, bg=self.palette.get("accent_color", "#2563eb"), relief=tk.FLAT)
                badge_frame.pack(side=tk.LEFT, padx=(16, 0))
                
                count_label = tk.Label(
                    badge_frame,
                    text=f"{unread_count} unread",
                    font=("{Segoe UI Semibold} 11"),
                    bg=self.palette.get("accent_color", "#2563eb"),
                    fg="#ffffff",
                    padx=12,
                    pady=6
                )
                count_label.pack()
            else:
                status_label = ttk.Label(
                    header_left,
                    text="All caught up! ✓",
                    font=("{Segoe UI} 12"),
                    foreground="#10b981",
                    style="TLabel"
                )
                status_label.pack(side=tk.LEFT, padx=(16, 0))
            
            # Action buttons in header (top right)
            header_right = ttk.Frame(header_frame, style="Techfix.Surface.TFrame")
            header_right.pack(side=tk.RIGHT, fill=tk.Y)
            
            def mark_all_read():
                if user_id:
                    notifications.mark_all_read(user_id)
                    dialog.destroy()
                    self._show_notifications()  # Refresh
            
            if unread_count > 0:
                ttk.Button(
                    header_right,
                    text="✓ Mark All as Read",
                    command=mark_all_read,
                    style="Techfix.TButton"
                ).pack(side=tk.LEFT, padx=(0, 8))
            
            def refresh_notifications():
                loading = self.show_loading("Please wait... Refreshing notifications...", dialog)
                try:
                    dialog.update()  # Update UI to show loading indicator
                    dialog.destroy()
                    self._show_notifications()
                finally:
                    if loading:
                        self.hide_loading(loading)
            
            ttk.Button(
                header_right,
                text="🔄 Refresh",
                command=refresh_notifications,
                style="Techfix.TButton"
            ).pack(side=tk.LEFT, padx=(0, 8))
            
            ttk.Button(
                header_right,
                text="Close",
                command=dialog.destroy,
                style="Techfix.TButton"
            ).pack(side=tk.LEFT)
            
            # Scrollable frame for notifications
            canvas = tk.Canvas(
                main_frame,
                bg=self.palette.get("surface_bg", "#ffffff"),
                highlightthickness=0,
                bd=0
            )
            scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
            scrollable_frame = tk.Frame(canvas, bg=self.palette.get("surface_bg", "#ffffff"))
            
            scrollable_frame.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
            )
            
            canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)
            
            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")
            
            # Mouse wheel scrolling - bind to canvas and dialog, not globally
            def on_mousewheel(event):
                try:
                    if canvas.winfo_exists():
                        canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
                except tk.TclError:
                    pass  # Widget was destroyed
            
            canvas.bind("<MouseWheel>", on_mousewheel)
            dialog.bind("<MouseWheel>", on_mousewheel)
            
            # Cleanup binding when dialog is destroyed
            def cleanup_bindings():
                try:
                    canvas.unbind("<MouseWheel>")
                    dialog.unbind("<MouseWheel>")
                except:
                    pass
            dialog.bind("<Destroy>", lambda e: cleanup_bindings())
            
            if notifs:
                for idx, notif in enumerate(notifs):
                    is_unread = notif.get('is_read', 0) == 0
                    
                    # Create card frame
                    card_frame = tk.Frame(
                        scrollable_frame,
                        bg=self.palette.get("app_bg", "#f9fafb") if is_unread else self.palette.get("surface_bg", "#ffffff"),
                        relief=tk.FLAT,
                        bd=1,
                        highlightbackground=self.palette.get("accent_color", "#2563eb") if is_unread else self.palette.get("entry_border", "#e5e7eb"),
                        highlightthickness=2 if is_unread else 1
                    )
                    card_frame.pack(fill=tk.X, pady=6, padx=4)
                    
                    # Card content
                    content_frame = tk.Frame(card_frame, bg=card_frame.cget("bg"))
                    content_frame.pack(fill=tk.BOTH, expand=True, padx=16, pady=16)
                    
                    # Top row: icon, title, and action button
                    top_row = tk.Frame(content_frame, bg=card_frame.cget("bg"))
                    top_row.pack(fill=tk.X, pady=(0, 8))
                    
                    # Notification type icon
                    notif_type = notif.get('type', 'info')
                    icon_map = {
                        'info': 'ℹ️',
                        'success': '✅',
                        'warning': '⚠️',
                        'error': '❌'
                    }
                    icon_text = icon_map.get(notif_type, '📌')
                    
                    icon_label = tk.Label(
                        top_row,
                        text=icon_text,
                        font=("{Segoe UI} 20"),
                        bg=card_frame.cget("bg"),
                        fg=self.palette.get("accent_color", "#2563eb") if is_unread else self.palette.get("text_secondary", "#6b7280")
                    )
                    icon_label.pack(side=tk.LEFT, padx=(0, 12))
                    
                    # Title and message frame
                    text_frame = tk.Frame(top_row, bg=card_frame.cget("bg"))
                    text_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
                    
                    title_label = tk.Label(
                        text_frame,
                        text=notif.get('title', 'Notification'),
                        font=("{Segoe UI Semibold} 13"),
                        bg=card_frame.cget("bg"),
                        fg=self.palette.get("text_primary", "#111827"),
                        anchor=tk.W
                    )
                    title_label.pack(anchor=tk.W)
                    
                    message_label = tk.Label(
                        text_frame,
                        text=notif.get('message', ''),
                        font=("{Segoe UI} 11"),
                        bg=card_frame.cget("bg"),
                        fg=self.palette.get("text_secondary", "#6b7280"),
                        anchor=tk.W,
                        wraplength=600,
                        justify=tk.LEFT
                    )
                    message_label.pack(anchor=tk.W, pady=(4, 0))
                    
                    # Bottom row: timestamp and action
                    bottom_row = tk.Frame(content_frame, bg=card_frame.cget("bg"))
                    bottom_row.pack(fill=tk.X, pady=(8, 0))
                    
                    if notif.get('created_at'):
                        time_text = notif.get('created_at', '')[:16].replace('T', ' ')
                        time_label = tk.Label(
                            bottom_row,
                            text=f"🕒 {time_text}",
                            font=("{Segoe UI} 9"),
                            bg=card_frame.cget("bg"),
                            fg=self.palette.get("text_secondary", "#9ca3af")
                        )
                        time_label.pack(side=tk.LEFT)
                    
                    # Mark as read button for unread notifications
                    if is_unread:
                        def mark_read(nid=notif.get('id')):
                            if nid:
                                notifications.mark_notification_read(nid)
                                dialog.destroy()
                                self._show_notifications()  # Refresh
                        
                        read_btn = tk.Button(
                            bottom_row,
                            text="✓ Mark as Read",
                            command=mark_read,
                            font=("{Segoe UI} 10"),
                            bg=self.palette.get("accent_color", "#2563eb"),
                            fg="#ffffff",
                            relief=tk.FLAT,
                            padx=16,
                            pady=6,
                            cursor="hand2",
                            activebackground=self.palette.get("accent_hover", "#1d4ed8"),
                            activeforeground="#ffffff"
                        )
                        read_btn.pack(side=tk.RIGHT)
            else:
                # Enhanced empty state
                empty_frame = tk.Frame(scrollable_frame, bg=self.palette.get("surface_bg", "#ffffff"))
                empty_frame.pack(fill=tk.BOTH, expand=True, pady=80)
                
                empty_icon = tk.Label(
                    empty_frame,
                    text="📭",
                    font=("{Segoe UI} 64"),
                    bg=self.palette.get("surface_bg", "#ffffff"),
                    fg=self.palette.get("text_secondary", "#9ca3af")
                )
                empty_icon.pack()
                
                empty_title = tk.Label(
                    empty_frame,
                    text="No notifications",
                    font=("{Segoe UI Semibold} 18"),
                    bg=self.palette.get("surface_bg", "#ffffff"),
                    fg=self.palette.get("text_primary", "#111827")
                )
                empty_title.pack(pady=(16, 8))
                
                empty_subtitle = tk.Label(
                    empty_frame,
                    text="You're all caught up! New notifications will appear here.",
                    font=("{Segoe UI} 11"),
                    bg=self.palette.get("surface_bg", "#ffffff"),
                    fg=self.palette.get("text_secondary", "#6b7280")
                )
                empty_subtitle.pack()
            
        except Exception as e:
            logger.error(f"Notifications error: {e}", exc_info=True)
            messagebox.showerror("Error", f"Notifications error: {e}")
    
    def _is_user_admin(self) -> bool:
        """Check if current user is admin by verifying role from database."""
        if not self.current_user or not self.current_user.get('id'):
            logger.warning("Admin check failed: No current user or user ID")
            return False
        
        try:
            from . import db
            user_id = self.current_user.get('id')
            if not user_id:
                logger.warning("Admin check failed: User ID is None or empty")
                return False
            
            conn = db.get_connection()
            try:
                # Verify role directly from database - must be active and have Admin role
                cur = conn.execute("""
                    SELECT r.name as role_name, u.is_active
                    FROM users u
                    LEFT JOIN roles r ON u.role_id = r.id
                    WHERE u.id = ? AND u.is_active = 1
                """, (user_id,))
                row = cur.fetchone()
                if not row:
                    logger.warning(f"Admin check failed: User {user_id} not found or inactive")
                    return False
                
                role_name = row['role_name']
                if not role_name:
                    logger.warning(f"Admin check failed: User {user_id} has no role assigned")
                    return False
                
                # Case-insensitive comparison - database stores "Admin" but we check for "admin"
                role_name_lower = str(role_name).strip().lower()
                is_admin = role_name_lower == 'admin'
                
                if not is_admin:
                    logger.debug(f"Admin check failed: User {user_id} has role '{role_name}' (not Admin)")
                
                return is_admin
            finally:
                conn.close()
        except Exception as e:
            logger.error(f"Error checking admin status: {e}", exc_info=True)
            return False
    
    def _has_permission(self, permission: str) -> bool:
        """Check if current user has a specific permission."""
        if not self.current_user or not self.current_user.get('role_id'):
            return False
        
        try:
            from . import auth
            role_id = self.current_user.get('role_id')
            return auth.has_permission(role_id, permission)
        except Exception as e:
            logger.error(f"Error checking permission '{permission}': {e}", exc_info=True)
            return False
    
    def _show_admin_panel(self) -> None:
        """Show admin panel for user and system management."""
        try:
            # Check if user is admin - verify from database for security
            if not self.current_user:
                messagebox.showerror("Access Denied", "You must be logged in to access the admin panel.")
                return
            
            # Verify admin status from database (not just cached user info)
            if not self._is_user_admin():
                messagebox.showerror("Access Denied", "Admin access required. Only users with Admin role can access this panel.")
                return
            
            from . import auth, db
            
            dialog = tk.Toplevel(self)
            dialog.title("Admin Panel - TechFix")
            dialog.transient(self)
            dialog.grab_set()
            dialog.resizable(True, True)  # Allow resizing
            
            # Set initial size - reduced height significantly to ensure close button is visible
            width, height = 1100, 650
            dialog.configure(bg=self.palette.get("surface_bg", "#ffffff"))
            
            # Center dialog and ensure it fits on screen with proper margins
            dialog.update_idletasks()
            screen_width = dialog.winfo_screenwidth()
            screen_height = dialog.winfo_screenheight()
            
            # Ensure dialog fits on screen with generous margin for title bar (50px top, 50px bottom)
            max_height = screen_height - 100  # 50px top + 50px bottom margin
            if height > max_height:
                height = max_height
            if width > screen_width - 50:
                width = screen_width - 50
            
            # Calculate position - ensure title bar is always visible
            x = (screen_width // 2) - (width // 2)
            # Position with at least 50px from top to ensure title bar is fully visible
            y = max(50, (screen_height // 2) - (height // 2))
            
            # Double-check: if dialog would extend beyond screen, adjust
            if y + height > screen_height - 20:
                y = screen_height - height - 20
            
            dialog.geometry(f"{width}x{height}+{x}+{y}")
            dialog.minsize(900, 550)  # Set minimum size
            
            main_frame = ttk.Frame(dialog, style="Techfix.Surface.TFrame", padding=24)
            main_frame.pack(fill=tk.BOTH, expand=True)
            
            # Header
            header_frame = ttk.Frame(main_frame, style="Techfix.Surface.TFrame")
            header_frame.pack(fill=tk.X, pady=(0, 20))
            
            ttk.Label(
                header_frame,
                text="⚙️ Admin Panel",
                font=("{Segoe UI Semibold} 22"),
                style="TLabel"
            ).pack(side=tk.LEFT)
            
            # Tab notebook for different admin sections
            notebook = ttk.Notebook(main_frame)
            notebook.pack(fill=tk.BOTH, expand=True)
            
            # Users Tab
            users_frame = ttk.Frame(notebook, style="Techfix.Surface.TFrame", padding=16)
            notebook.add(users_frame, text="👥 Users")
            
            # Users header with actions
            users_header = ttk.Frame(users_frame, style="Techfix.Surface.TFrame")
            users_header.pack(fill=tk.X, pady=(0, 12))
            
            ttk.Label(
                users_header,
                text="User Management",
                font=("{Segoe UI Semibold} 16"),
                style="TLabel"
            ).pack(side=tk.LEFT)
            
            btn_frame = ttk.Frame(users_header, style="Techfix.Surface.TFrame")
            btn_frame.pack(side=tk.RIGHT)
            
            def refresh_users():
                loading = self.show_loading("Please wait... Loading users...", dialog)
                try:
                    dialog.update()  # Update UI to show loading indicator
                    # Clear existing items
                    for item in users_tree.get_children():
                        users_tree.delete(item)
                    
                    # Load users
                    conn = db.get_connection()
                    try:
                        cur = conn.execute("""
                            SELECT u.id, u.username, u.full_name, u.is_active, 
                                   r.name as role_name, c.name as company_name,
                                   u.created_at, u.last_login
                            FROM users u
                            LEFT JOIN roles r ON u.role_id = r.id
                            LEFT JOIN companies c ON u.company_id = c.id
                            ORDER BY u.username
                        """)
                        for row in cur.fetchall():
                            status = "Active" if row['is_active'] else "Inactive"
                            role = row['role_name'] or "No Role"
                            company = row['company_name'] or "No Company"
                            created = row['created_at'][:10] if row['created_at'] else "N/A"
                            last_login = row['last_login'][:16] if row['last_login'] else "Never"
                            
                            users_tree.insert('', 'end', values=(
                                row['id'],
                                row['username'],
                                row['full_name'] or '',
                                role,
                                company,
                                status,
                                created,
                                last_login
                            ))
                    except Exception as e:
                        logger.error(f"Error loading users: {e}", exc_info=True)
                        messagebox.showerror("Error", f"Failed to load users: {e}")
                    finally:
                        conn.close()
                finally:
                    if loading:
                        self.hide_loading(loading)
            
            ttk.Button(
                btn_frame,
                text="➕ New User",
                command=lambda: self._show_create_user_dialog(dialog, refresh_users),
                style="Techfix.TButton"
            ).pack(side=tk.LEFT, padx=(0, 8))
            
            ttk.Button(
                btn_frame,
                text="🔄 Refresh",
                command=refresh_users,
                style="Techfix.TButton"
            ).pack(side=tk.LEFT, padx=(0, 8))
            
            # Users treeview
            users_tree_frame = ttk.Frame(users_frame, style="Techfix.Surface.TFrame")
            users_tree_frame.pack(fill=tk.BOTH, expand=True)
            
            columns = ("ID", "Username", "Full Name", "Role", "Company", "Status", "Created", "Last Login")
            users_tree = ttk.Treeview(users_tree_frame, columns=columns, show="headings", style="Techfix.Treeview")
            
            for col in columns:
                users_tree.heading(col, text=col)
                if col == "ID":
                    users_tree.column(col, width=50, stretch=False)
                elif col in ("Status", "Role", "Company"):
                    users_tree.column(col, width=100, stretch=False)
                elif col == "Username":
                    users_tree.column(col, width=120, stretch=False)
                elif col in ("Created", "Last Login"):
                    users_tree.column(col, width=120, stretch=False)
                else:
                    users_tree.column(col, width=150, stretch=True)
            
            scrollbar_users = ttk.Scrollbar(users_tree_frame, orient=tk.VERTICAL, command=users_tree.yview)
            users_tree.configure(yscrollcommand=scrollbar_users.set)
            users_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar_users.pack(side=tk.RIGHT, fill=tk.Y)
            
            # User actions frame
            user_actions_frame = ttk.Frame(users_frame, style="Techfix.Surface.TFrame")
            user_actions_frame.pack(fill=tk.X, pady=(12, 0))
            
            def edit_selected_user():
                selection = users_tree.selection()
                if not selection:
                    messagebox.showwarning("No Selection", "Please select a user to edit.")
                    return
                item = users_tree.item(selection[0])
                user_id = item['values'][0]
                self._show_edit_user_dialog(dialog, user_id, refresh_users)
            
            def delete_selected_user():
                # Security check: Verify admin status before allowing delete
                if not self._is_user_admin():
                    messagebox.showerror("Access Denied", "Admin access required.")
                    return
                
                selection = users_tree.selection()
                if not selection:
                    messagebox.showwarning("No Selection", "Please select a user to delete.")
                    return
                item = users_tree.item(selection[0])
                user_id = item['values'][0]
                username = item['values'][1]
                
                if user_id == self.current_user['id']:
                    messagebox.showerror("Error", "You cannot delete your own account.")
                    return
                
                if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete user '{username}'? This action cannot be undone."):
                    conn = db.get_connection()
                    try:
                        conn.execute("DELETE FROM users WHERE id = ?", (user_id,))
                        conn.commit()
                        messagebox.showinfo("Success", f"User '{username}' deleted successfully.")
                        refresh_users()
                    except Exception as e:
                        messagebox.showerror("Error", f"Failed to delete user: {e}")
                    finally:
                        conn.close()
            
            def toggle_user_status():
                # Security check: Verify admin status before allowing status change
                if not self._is_user_admin():
                    messagebox.showerror("Access Denied", "Admin access required.")
                    return
                
                selection = users_tree.selection()
                if not selection:
                    messagebox.showwarning("No Selection", "Please select a user.")
                    return
                item = users_tree.item(selection[0])
                user_id = item['values'][0]
                username = item['values'][1]
                current_status = item['values'][5]
                new_status = 0 if current_status == "Active" else 1
                
                conn = db.get_connection()
                try:
                    conn.execute("UPDATE users SET is_active = ? WHERE id = ?", (new_status, user_id))
                    conn.commit()
                    messagebox.showinfo("Success", f"User '{username}' status updated.")
                    refresh_users()
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to update user status: {e}")
                finally:
                    conn.close()
            
            def reset_user_password():
                # Security check: Verify admin status before allowing password reset
                if not self._is_user_admin():
                    messagebox.showerror("Access Denied", "Admin access required.")
                    return
                
                selection = users_tree.selection()
                if not selection:
                    messagebox.showwarning("No Selection", "Please select a user.")
                    return
                item = users_tree.item(selection[0])
                user_id = item['values'][0]
                username = item['values'][1]
                
                new_password = simpledialog.askstring(
                    "Reset Password",
                    f"Enter new password for '{username}':",
                    show='*',
                    parent=dialog
                )
                
                if new_password:
                    if auth.reset_password(user_id, new_password):
                        messagebox.showinfo("Success", f"Password reset for user '{username}'.")
                    else:
                        messagebox.showerror("Error", "Failed to reset password.")
            
            ttk.Button(
                user_actions_frame,
                text="✏️ Edit",
                command=edit_selected_user,
                style="Techfix.TButton"
            ).pack(side=tk.LEFT, padx=(0, 8))
            
            ttk.Button(
                user_actions_frame,
                text="🔑 Reset Password",
                command=reset_user_password,
                style="Techfix.TButton"
            ).pack(side=tk.LEFT, padx=(0, 8))
            
            ttk.Button(
                user_actions_frame,
                text="🔄 Toggle Status",
                command=toggle_user_status,
                style="Techfix.TButton"
            ).pack(side=tk.LEFT, padx=(0, 8))
            
            ttk.Button(
                user_actions_frame,
                text="🗑️ Delete",
                command=delete_selected_user,
                style="Techfix.TButton"
            ).pack(side=tk.LEFT, padx=(0, 8))
            
            # Roles Tab
            roles_frame = ttk.Frame(notebook, style="Techfix.Surface.TFrame", padding=16)
            notebook.add(roles_frame, text="👤 Roles")
            
            roles_header = ttk.Frame(roles_frame, style="Techfix.Surface.TFrame")
            roles_header.pack(fill=tk.X, pady=(0, 12))
            
            ttk.Label(
                roles_header,
                text="Role Management",
                font=("{Segoe UI Semibold} 16"),
                style="TLabel"
            ).pack(side=tk.LEFT)
            
            def refresh_roles():
                loading = self.show_loading("Please wait... Loading roles...", dialog)
                try:
                    dialog.update()  # Update UI to show loading indicator
                    for item in roles_tree.get_children():
                        roles_tree.delete(item)
                    
                    conn = db.get_connection()
                    try:
                        cur = conn.execute("SELECT id, name, description FROM roles ORDER BY name")
                        for row in cur.fetchall():
                            roles_tree.insert('', 'end', values=(
                                row['id'],
                                row['name'],
                                row['description'] or ''
                            ))
                    except Exception as e:
                        messagebox.showerror("Error", f"Failed to load roles: {e}")
                    finally:
                        conn.close()
                finally:
                    if loading:
                        self.hide_loading(loading)
            
            roles_btn_frame = ttk.Frame(roles_header, style="Techfix.Surface.TFrame")
            roles_btn_frame.pack(side=tk.RIGHT)
            
            ttk.Button(
                roles_btn_frame,
                text="➕ New Role",
                command=lambda: self._show_create_role_dialog(dialog, refresh_roles),
                style="Techfix.TButton"
            ).pack(side=tk.LEFT, padx=(0, 8))
            
            ttk.Button(
                roles_btn_frame,
                text="🔄 Refresh",
                command=refresh_roles,
                style="Techfix.TButton"
            ).pack(side=tk.LEFT)
            
            roles_tree_frame = ttk.Frame(roles_frame, style="Techfix.Surface.TFrame")
            roles_tree_frame.pack(fill=tk.BOTH, expand=True)
            
            roles_columns = ("ID", "Name", "Description")
            roles_tree = ttk.Treeview(roles_tree_frame, columns=roles_columns, show="headings", style="Techfix.Treeview")
            
            for col in roles_columns:
                roles_tree.heading(col, text=col)
                if col == "ID":
                    roles_tree.column(col, width=50, stretch=False)
                else:
                    roles_tree.column(col, width=200, stretch=True)
            
            scrollbar_roles = ttk.Scrollbar(roles_tree_frame, orient=tk.VERTICAL, command=roles_tree.yview)
            roles_tree.configure(yscrollcommand=scrollbar_roles.set)
            roles_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar_roles.pack(side=tk.RIGHT, fill=tk.Y)
            
            # Double-click to edit
            def on_role_double_click(event):
                selection = roles_tree.selection()
                if selection:
                    item = roles_tree.item(selection[0])
                    role_id = item['values'][0]
                    self._show_edit_role_dialog(dialog, role_id, refresh_roles)
            
            roles_tree.bind('<Double-1>', on_role_double_click)
            
            # Role actions frame
            role_actions_frame = ttk.Frame(roles_frame, style="Techfix.Surface.TFrame")
            role_actions_frame.pack(fill=tk.X, pady=(12, 0))
            
            def edit_selected_role():
                # Security check: Verify admin status
                if not self._is_user_admin():
                    messagebox.showerror("Access Denied", "Admin access required.")
                    return
                
                selection = roles_tree.selection()
                if not selection:
                    messagebox.showwarning("No Selection", "Please select a role to edit.")
                    return
                item = roles_tree.item(selection[0])
                role_id = item['values'][0]
                self._show_edit_role_dialog(dialog, role_id, refresh_roles)
            
            def delete_selected_role():
                # Security check: Verify admin status
                if not self._is_user_admin():
                    messagebox.showerror("Access Denied", "Admin access required.")
                    return
                
                selection = roles_tree.selection()
                if not selection:
                    messagebox.showwarning("No Selection", "Please select a role to delete.")
                    return
                item = roles_tree.item(selection[0])
                role_id = item['values'][0]
                role_name = item['values'][1]
                
                # Check if role is in use
                conn = db.get_connection()
                try:
                    user_count = conn.execute("SELECT COUNT(*) FROM users WHERE role_id = ?", (role_id,)).fetchone()[0]
                    if user_count > 0:
                        messagebox.showerror(
                            "Cannot Delete",
                            f"Cannot delete role '{role_name}' because it is assigned to {user_count} user(s).\nPlease reassign users to another role first."
                        )
                        return
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to check role usage: {e}")
                finally:
                    conn.close()
                
                if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete role '{role_name}'? This action cannot be undone."):
                    conn = db.get_connection()
                    try:
                        conn.execute("DELETE FROM roles WHERE id = ?", (role_id,))
                        conn.commit()
                        messagebox.showinfo("Success", f"Role '{role_name}' deleted successfully.")
                        refresh_roles()
                    except Exception as e:
                        messagebox.showerror("Error", f"Failed to delete role: {e}")
                    finally:
                        conn.close()
            
            ttk.Button(
                role_actions_frame,
                text="✏️ Edit",
                command=edit_selected_role,
                style="Techfix.TButton"
            ).pack(side=tk.LEFT, padx=(0, 8))
            
            ttk.Button(
                role_actions_frame,
                text="🗑️ Delete",
                command=delete_selected_role,
                style="Techfix.TButton"
            ).pack(side=tk.LEFT, padx=(0, 8))
            
            # System Info Tab
            system_frame = ttk.Frame(notebook, style="Techfix.Surface.TFrame", padding=16)
            notebook.add(system_frame, text="ℹ️ System Info")
            
            ttk.Label(
                system_frame,
                text="System Information",
                font=("{Segoe UI Semibold} 16"),
                style="TLabel"
            ).pack(anchor=tk.W, pady=(0, 12))
            
            info_text = tk.Text(
                system_frame,
                wrap=tk.WORD,
                font=("{Segoe UI} 11"),
                bg=self.palette.get("surface_bg", "#ffffff"),
                fg=self.palette.get("text_primary", "#111827"),
                relief=tk.FLAT,
                padx=12,
                pady=12
            )
            info_text.pack(fill=tk.BOTH, expand=True)
            
            conn = db.get_connection()
            try:
                # Get system stats
                user_count = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
                active_users = conn.execute("SELECT COUNT(*) FROM users WHERE is_active = 1").fetchone()[0]
                role_count = conn.execute("SELECT COUNT(*) FROM roles").fetchone()[0]
                company_count = conn.execute("SELECT COUNT(*) FROM companies").fetchone()[0]
                entry_count = conn.execute("SELECT COUNT(*) FROM journal_entries").fetchone()[0]
                account_count = conn.execute("SELECT COUNT(*) FROM accounts").fetchone()[0]
                
                info_text.insert('1.0', f"""System Statistics:

Users: {user_count} total ({active_users} active)
Roles: {role_count}
Companies: {company_count}
Journal Entries: {entry_count}
Accounts: {account_count}

Database: techfix.sqlite3
Current User: {self.current_user.get('username', 'N/A')}
Role: {self.current_user.get('role_name', 'N/A')}
""")
                info_text.configure(state=tk.DISABLED)
            except Exception as e:
                info_text.insert('1.0', f"Error loading system information: {e}")
                info_text.configure(state=tk.DISABLED)
            finally:
                conn.close()
            
            # Load initial data
            refresh_users()
            refresh_roles()
            
            # Close button
            btn_frame_bottom = ttk.Frame(main_frame, style="Techfix.Surface.TFrame")
            btn_frame_bottom.pack(fill=tk.X, pady=(20, 0))
            
            ttk.Button(
                btn_frame_bottom,
                text="Close",
                command=dialog.destroy,
                style="Techfix.TButton"
            ).pack(side=tk.RIGHT)
            
            # Final positioning check to ensure title bar is visible
            dialog.update_idletasks()
            dialog.update()
            
            # Verify dialog position - ensure title bar is visible
            current_y = dialog.winfo_y()
            if current_y < 30:  # If too close to top (title bar is ~30px)
                screen_height = dialog.winfo_screenheight()
                current_height = dialog.winfo_height()
                # Reposition with proper top margin
                new_y = 30
                # Make sure dialog doesn't go off bottom of screen
                if new_y + current_height > screen_height - 20:
                    new_y = screen_height - current_height - 20
                current_x = dialog.winfo_x()
                dialog.geometry(f"{dialog.winfo_width()}x{current_height}+{current_x}+{new_y}")
                dialog.update_idletasks()
            
            dialog.focus_force()
            dialog.lift()
            
        except Exception as e:
            logger.error(f"Admin panel error: {e}", exc_info=True)
            messagebox.showerror("Error", f"Admin panel error: {e}")
    
    def _show_create_user_dialog(self, parent, refresh_callback):
        """Show dialog to create a new user."""
        # Security check: Only admins can create users
        if not self._is_user_admin():
            messagebox.showerror("Access Denied", "Admin access required to create users.")
            return
        
        try:
            from . import auth, db
            
            dialog = tk.Toplevel(parent)
            dialog.title("Create New User")
            dialog.transient(parent)
            dialog.grab_set()
            dialog.geometry("500x400")
            dialog.configure(bg=self.palette.get("surface_bg", "#ffffff"))
            
            main_frame = ttk.Frame(dialog, style="Techfix.Surface.TFrame", padding=24)
            main_frame.pack(fill=tk.BOTH, expand=True)
            
            ttk.Label(
                main_frame,
                text="Create New User",
                font=("{Segoe UI Semibold} 18"),
                style="TLabel"
            ).pack(pady=(0, 20))
            
            # Form fields
            fields_frame = ttk.Frame(main_frame, style="Techfix.Surface.TFrame")
            fields_frame.pack(fill=tk.BOTH, expand=True)
            
            ttk.Label(fields_frame, text="Username:", style="TLabel").grid(row=0, column=0, sticky=tk.W, pady=8, padx=(0, 12))
            username_var = tk.StringVar()
            username_entry = ttk.Entry(fields_frame, textvariable=username_var, width=30, style="Techfix.TEntry")
            username_entry.grid(row=0, column=1, sticky=tk.W, pady=8)
            username_entry.focus()
            
            ttk.Label(fields_frame, text="Full Name:", style="TLabel").grid(row=1, column=0, sticky=tk.W, pady=8, padx=(0, 12))
            fullname_var = tk.StringVar()
            ttk.Entry(fields_frame, textvariable=fullname_var, width=30, style="Techfix.TEntry").grid(row=1, column=1, sticky=tk.W, pady=8)
            
            ttk.Label(fields_frame, text="Password:", style="TLabel").grid(row=2, column=0, sticky=tk.W, pady=8, padx=(0, 12))
            password_var = tk.StringVar()
            ttk.Entry(fields_frame, textvariable=password_var, show='*', width=30, style="Techfix.TEntry").grid(row=2, column=1, sticky=tk.W, pady=8)
            
            ttk.Label(fields_frame, text="Role:", style="TLabel").grid(row=3, column=0, sticky=tk.W, pady=8, padx=(0, 12))
            role_var = tk.StringVar()
            role_combo = ttk.Combobox(fields_frame, textvariable=role_var, state="readonly", width=27, style="Techfix.TCombobox")
            role_combo.grid(row=3, column=1, sticky=tk.W, pady=8)
            
            # Load roles
            conn = db.get_connection()
            try:
                cur = conn.execute("SELECT name FROM roles ORDER BY name")
                roles = [row['name'] for row in cur.fetchall()]
                role_combo['values'] = roles
                if roles:
                    role_combo.current(0)
            finally:
                conn.close()
            
            ttk.Label(fields_frame, text="Active:", style="TLabel").grid(row=4, column=0, sticky=tk.W, pady=8, padx=(0, 12))
            active_var = tk.BooleanVar(value=True)
            ttk.Checkbutton(fields_frame, variable=active_var, style="Techfix.TCheckbutton").grid(row=4, column=1, sticky=tk.W, pady=8)
            
            def create_user():
                username = username_var.get().strip()
                full_name = fullname_var.get().strip()
                password = password_var.get()
                role_name = role_var.get()
                is_active = 1 if active_var.get() else 0
                
                if not username:
                    messagebox.showerror("Error", "Username is required.", parent=dialog)
                    return
                
                if not password:
                    messagebox.showerror("Error", "Password is required.", parent=dialog)
                    return
                
                conn = db.get_connection()
                try:
                    # Check if username exists
                    existing = conn.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
                    if existing:
                        messagebox.showerror("Error", f"Username '{username}' already exists.", parent=dialog)
                        return
                    
                    # Get role ID
                    role_row = conn.execute("SELECT id FROM roles WHERE name = ?", (role_name,)).fetchone()
                    if not role_row:
                        messagebox.showerror("Error", "Invalid role selected.", parent=dialog)
                        return
                    role_id = role_row['id']
                    
                    # Get default company
                    company_row = conn.execute("SELECT id FROM companies WHERE code = 'DEFAULT'").fetchone()
                    company_id = company_row['id'] if company_row else None
                    
                    # Hash password
                    password_hash = auth.hash_password(password)
                    
                    # Create user
                    conn.execute("""
                        INSERT INTO users (username, full_name, role_id, company_id, is_active, password_hash)
                        VALUES (?, ?, ?, ?, ?, ?)
                    """, (username, full_name or None, role_id, company_id, is_active, password_hash))
                    conn.commit()
                    
                    messagebox.showinfo("Success", f"User '{username}' created successfully.", parent=dialog)
                    dialog.destroy()
                    refresh_callback()
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to create user: {e}", parent=dialog)
                finally:
                    conn.close()
            
            btn_frame = ttk.Frame(main_frame, style="Techfix.Surface.TFrame")
            btn_frame.pack(fill=tk.X, pady=(20, 0))
            
            ttk.Button(
                btn_frame,
                text="Create",
                command=create_user,
                style="Techfix.TButton"
            ).pack(side=tk.RIGHT, padx=(8, 0))
            
            ttk.Button(
                btn_frame,
                text="Cancel",
                command=dialog.destroy,
                style="Techfix.TButton"
            ).pack(side=tk.RIGHT)
            
        except Exception as e:
            messagebox.showerror("Error", f"Error creating user dialog: {e}")
    
    def _show_edit_user_dialog(self, parent, user_id, refresh_callback):
        """Show dialog to edit an existing user."""
        # Security check: Only admins can edit users
        if not self._is_user_admin():
            messagebox.showerror("Access Denied", "Admin access required to edit users.")
            return
        
        try:
            from . import db
            
            dialog = tk.Toplevel(parent)
            dialog.title("Edit User")
            dialog.transient(parent)
            dialog.grab_set()
            dialog.geometry("500x400")
            dialog.configure(bg=self.palette.get("surface_bg", "#ffffff"))
            
            main_frame = ttk.Frame(dialog, style="Techfix.Surface.TFrame", padding=24)
            main_frame.pack(fill=tk.BOTH, expand=True)
            
            ttk.Label(
                main_frame,
                text="Edit User",
                font=("{Segoe UI Semibold} 18"),
                style="TLabel"
            ).pack(pady=(0, 20))
            
            # Load user data
            conn = db.get_connection()
            try:
                cur = conn.execute("""
                    SELECT username, full_name, role_id, is_active
                    FROM users WHERE id = ?
                """, (user_id,))
                user_data = cur.fetchone()
                
                if not user_data:
                    messagebox.showerror("Error", "User not found.", parent=dialog)
                    dialog.destroy()
                    return
            finally:
                conn.close()
            
            # Form fields
            fields_frame = ttk.Frame(main_frame, style="Techfix.Surface.TFrame")
            fields_frame.pack(fill=tk.BOTH, expand=True)
            
            ttk.Label(fields_frame, text="Username:", style="TLabel").grid(row=0, column=0, sticky=tk.W, pady=8, padx=(0, 12))
            username_var = tk.StringVar(value=user_data['username'])
            username_entry = ttk.Entry(fields_frame, textvariable=username_var, width=30, style="Techfix.TEntry", state='readonly')
            username_entry.grid(row=0, column=1, sticky=tk.W, pady=8)
            
            ttk.Label(fields_frame, text="Full Name:", style="TLabel").grid(row=1, column=0, sticky=tk.W, pady=8, padx=(0, 12))
            fullname_var = tk.StringVar(value=user_data['full_name'] or '')
            ttk.Entry(fields_frame, textvariable=fullname_var, width=30, style="Techfix.TEntry").grid(row=1, column=1, sticky=tk.W, pady=8)
            
            ttk.Label(fields_frame, text="Role:", style="TLabel").grid(row=2, column=0, sticky=tk.W, pady=8, padx=(0, 12))
            role_var = tk.StringVar()
            role_combo = ttk.Combobox(fields_frame, textvariable=role_var, state="readonly", width=27, style="Techfix.TCombobox")
            role_combo.grid(row=2, column=1, sticky=tk.W, pady=8)
            
            # Load roles and set current
            conn = db.get_connection()
            try:
                cur = conn.execute("SELECT id, name FROM roles ORDER BY name")
                roles = []
                current_role_id = user_data['role_id']
                current_role_name = None
                for row in cur.fetchall():
                    roles.append(row['name'])
                    if row['id'] == current_role_id:
                        current_role_name = row['name']
                role_combo['values'] = roles
                if current_role_name:
                    role_var.set(current_role_name)
            finally:
                conn.close()
            
            ttk.Label(fields_frame, text="Active:", style="TLabel").grid(row=3, column=0, sticky=tk.W, pady=8, padx=(0, 12))
            active_var = tk.BooleanVar(value=bool(user_data['is_active']))
            ttk.Checkbutton(fields_frame, variable=active_var, style="Techfix.TCheckbutton").grid(row=3, column=1, sticky=tk.W, pady=8)
            
            def update_user():
                full_name = fullname_var.get().strip()
                role_name = role_var.get()
                is_active = 1 if active_var.get() else 0
                
                conn = db.get_connection()
                try:
                    # Get role ID
                    role_row = conn.execute("SELECT id FROM roles WHERE name = ?", (role_name,)).fetchone()
                    if not role_row:
                        messagebox.showerror("Error", "Invalid role selected.", parent=dialog)
                        return
                    role_id = role_row['id']
                    
                    # Update user
                    conn.execute("""
                        UPDATE users 
                        SET full_name = ?, role_id = ?, is_active = ?
                        WHERE id = ?
                    """, (full_name or None, role_id, is_active, user_id))
                    conn.commit()
                    
                    messagebox.showinfo("Success", "User updated successfully.", parent=dialog)
                    dialog.destroy()
                    refresh_callback()
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to update user: {e}", parent=dialog)
                finally:
                    conn.close()
            
            btn_frame = ttk.Frame(main_frame, style="Techfix.Surface.TFrame")
            btn_frame.pack(fill=tk.X, pady=(20, 0))
            
            ttk.Button(
                btn_frame,
                text="Update",
                command=update_user,
                style="Techfix.TButton"
            ).pack(side=tk.RIGHT, padx=(8, 0))
            
            ttk.Button(
                btn_frame,
                text="Cancel",
                command=dialog.destroy,
                style="Techfix.TButton"
            ).pack(side=tk.RIGHT)
            
        except Exception as e:
            messagebox.showerror("Error", f"Error editing user: {e}")
    
    def _show_create_role_dialog(self, parent, refresh_callback):
        """Show dialog to create a new role."""
        # Security check: Only admins can create roles
        if not self._is_user_admin():
            messagebox.showerror("Access Denied", "Admin access required to create roles.")
            return
        
        try:
            from . import db
            
            dialog = tk.Toplevel(parent)
            dialog.title("Create New Role")
            dialog.transient(parent)
            dialog.grab_set()
            dialog.geometry("500x250")
            dialog.configure(bg=self.palette.get("surface_bg", "#ffffff"))
            
            main_frame = ttk.Frame(dialog, style="Techfix.Surface.TFrame", padding=24)
            main_frame.pack(fill=tk.BOTH, expand=True)
            
            ttk.Label(
                main_frame,
                text="Create New Role",
                font=("{Segoe UI Semibold} 18"),
                style="TLabel"
            ).pack(pady=(0, 20))
            
            # Form fields
            fields_frame = ttk.Frame(main_frame, style="Techfix.Surface.TFrame")
            fields_frame.pack(fill=tk.BOTH, expand=True)
            
            ttk.Label(fields_frame, text="Role Name:", style="TLabel").grid(row=0, column=0, sticky=tk.W, pady=8, padx=(0, 12))
            name_var = tk.StringVar()
            name_entry = ttk.Entry(fields_frame, textvariable=name_var, width=30, style="Techfix.TEntry")
            name_entry.grid(row=0, column=1, sticky=tk.W, pady=8)
            name_entry.focus()
            
            ttk.Label(fields_frame, text="Description:", style="TLabel").grid(row=1, column=0, sticky=tk.W, pady=8, padx=(0, 12))
            desc_var = tk.StringVar()
            desc_entry = ttk.Entry(fields_frame, textvariable=desc_var, width=30, style="Techfix.TEntry")
            desc_entry.grid(row=1, column=1, sticky=tk.W, pady=8)
            
            def create_role():
                name = name_var.get().strip()
                description = desc_var.get().strip()
                
                if not name:
                    messagebox.showerror("Error", "Role name is required.", parent=dialog)
                    return
                
                conn = db.get_connection()
                try:
                    # Check if role name exists
                    existing = conn.execute("SELECT id FROM roles WHERE name = ?", (name,)).fetchone()
                    if existing:
                        messagebox.showerror("Error", f"Role '{name}' already exists.", parent=dialog)
                        return
                    
                    # Create role
                    conn.execute("""
                        INSERT INTO roles (name, description)
                        VALUES (?, ?)
                    """, (name, description or None))
                    conn.commit()
                    
                    messagebox.showinfo("Success", f"Role '{name}' created successfully.", parent=dialog)
                    dialog.destroy()
                    refresh_callback()
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to create role: {e}", parent=dialog)
                finally:
                    conn.close()
            
            btn_frame = ttk.Frame(main_frame, style="Techfix.Surface.TFrame")
            btn_frame.pack(fill=tk.X, pady=(20, 0))
            
            ttk.Button(
                btn_frame,
                text="Create",
                command=create_role,
                style="Techfix.TButton"
            ).pack(side=tk.RIGHT, padx=(8, 0))
            
            ttk.Button(
                btn_frame,
                text="Cancel",
                command=dialog.destroy,
                style="Techfix.TButton"
            ).pack(side=tk.RIGHT)
            
        except Exception as e:
            messagebox.showerror("Error", f"Error creating role dialog: {e}")
    
    def _show_edit_role_dialog(self, parent, role_id, refresh_callback):
        """Show dialog to edit an existing role."""
        # Security check: Only admins can edit roles
        if not self._is_user_admin():
            messagebox.showerror("Access Denied", "Admin access required to edit roles.")
            return
        
        try:
            from . import db
            
            dialog = tk.Toplevel(parent)
            dialog.title("Edit Role")
            dialog.transient(parent)
            dialog.grab_set()
            dialog.geometry("500x250")
            dialog.configure(bg=self.palette.get("surface_bg", "#ffffff"))
            
            main_frame = ttk.Frame(dialog, style="Techfix.Surface.TFrame", padding=24)
            main_frame.pack(fill=tk.BOTH, expand=True)
            
            ttk.Label(
                main_frame,
                text="Edit Role",
                font=("{Segoe UI Semibold} 18"),
                style="TLabel"
            ).pack(pady=(0, 20))
            
            # Load role data
            conn = db.get_connection()
            try:
                cur = conn.execute("""
                    SELECT name, description
                    FROM roles WHERE id = ?
                """, (role_id,))
                role_data = cur.fetchone()
                
                if not role_data:
                    messagebox.showerror("Error", "Role not found.", parent=dialog)
                    dialog.destroy()
                    return
            finally:
                conn.close()
            
            # Form fields
            fields_frame = ttk.Frame(main_frame, style="Techfix.Surface.TFrame")
            fields_frame.pack(fill=tk.BOTH, expand=True)
            
            ttk.Label(fields_frame, text="Role Name:", style="TLabel").grid(row=0, column=0, sticky=tk.W, pady=8, padx=(0, 12))
            name_var = tk.StringVar(value=role_data['name'])
            name_entry = ttk.Entry(fields_frame, textvariable=name_var, width=30, style="Techfix.TEntry", state='readonly')
            name_entry.grid(row=0, column=1, sticky=tk.W, pady=8)
            
            ttk.Label(fields_frame, text="Description:", style="TLabel").grid(row=1, column=0, sticky=tk.W, pady=8, padx=(0, 12))
            desc_var = tk.StringVar(value=role_data['description'] or '')
            desc_entry = ttk.Entry(fields_frame, textvariable=desc_var, width=30, style="Techfix.TEntry")
            desc_entry.grid(row=1, column=1, sticky=tk.W, pady=8)
            desc_entry.focus()
            
            def update_role():
                description = desc_var.get().strip()
                
                conn = db.get_connection()
                try:
                    # Update role
                    conn.execute("""
                        UPDATE roles 
                        SET description = ?
                        WHERE id = ?
                    """, (description or None, role_id))
                    conn.commit()
                    
                    messagebox.showinfo("Success", "Role updated successfully.", parent=dialog)
                    dialog.destroy()
                    refresh_callback()
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to update role: {e}", parent=dialog)
                finally:
                    conn.close()
            
            btn_frame = ttk.Frame(main_frame, style="Techfix.Surface.TFrame")
            btn_frame.pack(fill=tk.X, pady=(20, 0))
            
            ttk.Button(
                btn_frame,
                text="Update",
                command=update_role,
                style="Techfix.TButton"
            ).pack(side=tk.RIGHT, padx=(8, 0))
            
            ttk.Button(
                btn_frame,
                text="Cancel",
                command=dialog.destroy,
                style="Techfix.TButton"
            ).pack(side=tk.RIGHT)
            
        except Exception as e:
            messagebox.showerror("Error", f"Error editing role: {e}")
    
    def _show_user_menu(self) -> None:
        """Show enhanced user menu popup with options."""
        if not self.auth_module or not self.current_user:
            # User is not logged in - show login option
            menu = tk.Menu(self, tearoff=0)
            menu.add_command(label="Login", command=lambda: self._show_login() and self._load_all_views())
            try:
                menu.tk_popup(self.winfo_pointerx(), self.winfo_pointery())
            except Exception:
                pass
            return
        
        # Create custom popup window for logged-in users
        popup = tk.Toplevel(self)
        popup.overrideredirect(True)  # Remove window decorations
        popup.attributes('-topmost', True)
        
        # Get user information
        username = self.current_user.get('username', 'User')
        full_name = self.current_user.get('full_name')
        role_name = self.current_user.get('role_name', 'User')
        
        # Get theme colors
        surface_bg = self.palette.get("surface_bg", "#ffffff")
        accent_color = self.palette.get("accent_color", "#2563eb")
        text_primary = self.palette.get("text_primary", "#1f2937")
        text_secondary = self.palette.get("text_secondary", "#4b5563")
        hover_bg = "#f3f4f6"
        
        # Main container with shadow effect
        container = tk.Frame(popup, bg="#e5e7eb", relief=tk.FLAT, bd=0)
        container.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)
        
        # Inner frame with white background
        inner_frame = tk.Frame(container, bg=surface_bg, relief=tk.FLAT, bd=0)
        inner_frame.pack(fill=tk.BOTH, expand=True)
        
        # User info section
        user_info_frame = tk.Frame(inner_frame, bg=accent_color)
        user_info_frame.pack(fill=tk.X)
        
        user_content = tk.Frame(user_info_frame, bg=accent_color)
        user_content.pack(fill=tk.X, padx=16, pady=14)
        
        # User icon and name
        user_icon = tk.Label(
            user_content,
            text="👤",
            bg=accent_color,
            fg="#ffffff",
            font=("{Segoe UI} 20")
        )
        user_icon.pack(side=tk.LEFT, padx=(0, 12))
        
        name_frame = tk.Frame(user_content, bg=accent_color)
        name_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        username_label = tk.Label(
            name_frame,
            text=full_name if full_name else username,
            bg=accent_color,
            fg="#ffffff",
            font=("{Segoe UI Semibold} 13"),
            anchor="w"
        )
        username_label.pack(fill=tk.X)
        
        role_label = tk.Label(
            name_frame,
            text=f"@{username} • {role_name}",
            bg=accent_color,
            fg="#dbeafe",
            font=("{Segoe UI} 10"),
            anchor="w"
        )
        role_label.pack(fill=tk.X, pady=(2, 0))
        
        # Separator
        sep1 = tk.Frame(inner_frame, bg="#e5e7eb", height=1)
        sep1.pack(fill=tk.X)
        
        # Menu items container
        menu_frame = tk.Frame(inner_frame, bg=surface_bg)
        menu_frame.pack(fill=tk.X, padx=4, pady=4)
        
        # Change Password button
        def create_menu_item(parent, icon, text, command=None):
            item_frame = tk.Frame(parent, bg=surface_bg, cursor="hand2")
            item_frame.pack(fill=tk.X, pady=2)
            
            def on_enter(e):
                item_frame.configure(bg=hover_bg)
                icon_label.configure(bg=hover_bg)
                text_label.configure(bg=hover_bg)
            
            def on_leave(e):
                item_frame.configure(bg=surface_bg)
                icon_label.configure(bg=surface_bg)
                text_label.configure(bg=surface_bg)
            
            def on_click(e):
                if command:
                    popup.destroy()
                    command()
            
            item_frame.bind('<Enter>', on_enter)
            item_frame.bind('<Leave>', on_leave)
            item_frame.bind('<Button-1>', on_click)
            
            icon_label = tk.Label(
                item_frame,
                text=icon,
                bg=surface_bg,
                fg=text_secondary,
                font=("{Segoe UI} 14"),
                width=3,
                anchor="w"
            )
            icon_label.pack(side=tk.LEFT, padx=(12, 8), pady=10)
            icon_label.bind('<Enter>', on_enter)
            icon_label.bind('<Leave>', on_leave)
            icon_label.bind('<Button-1>', on_click)
            
            text_label = tk.Label(
                item_frame,
                text=text,
                bg=surface_bg,
                fg=text_primary,
                font=("{Segoe UI} 11"),
                anchor="w"
            )
            text_label.pack(side=tk.LEFT, fill=tk.X, expand=True, pady=10)
            text_label.bind('<Enter>', on_enter)
            text_label.bind('<Leave>', on_leave)
            text_label.bind('<Button-1>', on_click)
            
            return item_frame
        
        # Add menu items
        create_menu_item(menu_frame, "🔑", "Change Password...", self._show_change_password)
        
        # Add Admin Panel option for admin users
        if self._is_user_admin():
            create_menu_item(menu_frame, "⚙️", "Admin Panel...", self._show_admin_panel)
        
        # Separator
        sep2 = tk.Frame(inner_frame, bg="#e5e7eb", height=1)
        sep2.pack(fill=tk.X, pady=(4, 0))
        
        # Logout button
        logout_frame = tk.Frame(inner_frame, bg=surface_bg, cursor="hand2")
        logout_frame.pack(fill=tk.X, padx=4, pady=(4, 4))
        
        def on_logout_enter(e):
            logout_frame.configure(bg="#fef2f2")
            logout_icon.configure(bg="#fef2f2", fg="#dc2626")
            logout_text.configure(bg="#fef2f2", fg="#dc2626")
        
        def on_logout_leave(e):
            logout_frame.configure(bg=surface_bg)
            logout_icon.configure(bg=surface_bg, fg="#dc2626")
            logout_text.configure(bg=surface_bg, fg="#dc2626")
        
        def on_logout_click(e):
            popup.destroy()
            self._logout()
        
        logout_frame.bind('<Enter>', on_logout_enter)
        logout_frame.bind('<Leave>', on_logout_leave)
        logout_frame.bind('<Button-1>', on_logout_click)
        
        logout_icon = tk.Label(
            logout_frame,
            text="🚪",
            bg=surface_bg,
            fg="#dc2626",
            font=("{Segoe UI} 14"),
            width=3,
            anchor="w"
        )
        logout_icon.pack(side=tk.LEFT, padx=(12, 8), pady=10)
        logout_icon.bind('<Enter>', on_logout_enter)
        logout_icon.bind('<Leave>', on_logout_leave)
        logout_icon.bind('<Button-1>', on_logout_click)
        
        logout_text = tk.Label(
            logout_frame,
            text="Logout",
            bg=surface_bg,
            fg="#dc2626",
            font=("{Segoe UI Semibold} 11"),
            anchor="w"
        )
        logout_text.pack(side=tk.LEFT, fill=tk.X, expand=True, pady=10)
        logout_text.bind('<Enter>', on_logout_enter)
        logout_text.bind('<Leave>', on_logout_leave)
        logout_text.bind('<Button-1>', on_logout_click)
        
        # Calculate position near the user button
        popup.update_idletasks()
        width = 280
        height = popup.winfo_reqheight()
        
        try:
            # Get button position
            btn_x = self.user_menu_btn.winfo_rootx()
            btn_y = self.user_menu_btn.winfo_rooty()
            btn_height = self.user_menu_btn.winfo_height()
            
            # Position popup below the button, aligned to the right
            x = btn_x + self.user_menu_btn.winfo_width() - width
            y = btn_y + btn_height + 4
            
            # Ensure popup stays on screen
            screen_width = self.winfo_screenwidth()
            screen_height = self.winfo_screenheight()
            
            if x + width > screen_width:
                x = screen_width - width - 10
            if x < 0:
                x = 10
            if y + height > screen_height:
                y = btn_y - height - 4
            
            popup.geometry(f"{width}x{height}+{x}+{y}")
        except Exception:
            # Fallback to cursor position
            popup.geometry(f"{width}x{height}+{self.winfo_pointerx()}+{self.winfo_pointery()}")
        
        # Close popup function
        def close_popup():
            try:
                if popup.winfo_exists():
                    try:
                        popup.grab_release()
                    except:
                        pass
                    popup.destroy()
            except:
                pass
        
        def on_escape(e):
            close_popup()
        
        popup.bind('<Escape>', on_escape)
        
        # Update to show popup
        popup.update_idletasks()
        popup.focus_set()
        
        # Use grab_set to capture all events - clicking outside will close popup
        try:
            popup.grab_set()
            # Release grab when popup is destroyed
            popup.bind('<Destroy>', lambda e: popup.grab_release() if popup.winfo_exists() else None)
        except:
            pass
    
    def _show_change_password(self) -> None:
        """Show change password dialog."""
        if not self.auth_module or not self.current_user:
            return
        
        dialog = tk.Toplevel(self)
        dialog.title("Change Password - TechFix Solutions")
        dialog.transient(self)
        dialog.grab_set()
        dialog.resizable(False, False)
        
        # Get theme colors
        bg_color = self.palette.get("app_bg", "#f5f7fb")
        surface_bg = self.palette.get("surface_bg", "#ffffff")
        accent_color = self.palette.get("accent_color", "#2563eb")
        accent_hover = self.palette.get("accent_hover", "#1d4ed8")
        text_primary = self.palette.get("text_primary", "#1f2937")
        text_secondary = self.palette.get("text_secondary", "#4b5563")
        entry_border = self.palette.get("entry_border", "#d8dee9")
        
        dialog.configure(bg=bg_color)
        
        # Set initial size - increased to ensure all content is visible
        width, height = 480, 750
        x = (dialog.winfo_screenwidth() // 2) - (width // 2)
        y = (dialog.winfo_screenheight() // 2) - (height // 2)
        dialog.geometry(f"{width}x{height}+{x}+{y}")
        dialog.minsize(width, height)
        dialog.maxsize(width, height)
        
        # Main container with padding
        container = tk.Frame(dialog, bg=bg_color)
        container.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Card-style panel
        card_frame = tk.Frame(container, bg=surface_bg, relief=tk.FLAT, bd=0)
        card_frame.pack(fill=tk.BOTH, expand=True)
        
        # Inner padding for card content - reduced vertical padding
        content_frame = tk.Frame(card_frame, bg=surface_bg)
        content_frame.pack(fill=tk.BOTH, expand=True, padx=50, pady=25)
        
        # Use grid for better control
        content_frame.grid_rowconfigure(0, weight=0)  # Header - no expand
        content_frame.grid_rowconfigure(1, weight=0)  # Form - no expand  
        content_frame.grid_rowconfigure(2, weight=1)  # Spacer - expands
        content_frame.grid_rowconfigure(3, weight=0)  # Buttons - no expand, always at bottom
        
        # Header section - reduced spacing
        header_frame = tk.Frame(content_frame, bg=surface_bg)
        header_frame.grid(row=0, column=0, sticky="ew", pady=(0, 20))
        
        # Icon - smaller
        icon_label = tk.Label(
            header_frame,
            text="🔑",
            font=("{Segoe UI} 40"),
            bg=surface_bg,
            fg=accent_color
        )
        icon_label.pack(pady=(0, 10))
        
        # Title
        title_label = tk.Label(
            header_frame,
            text="Change Password",
            font=("{Segoe UI Semibold} 22"),
            bg=surface_bg,
            fg=text_primary
        )
        title_label.pack(pady=(0, 5))
        
        subtitle = tk.Label(
            header_frame,
            text="Update your account password",
            bg=surface_bg,
            fg=text_secondary,
            font=("{Segoe UI} 10")
        )
        subtitle.pack()
        
        # Form section - don't expand, just take needed space
        form_frame = tk.Frame(content_frame, bg=surface_bg)
        form_frame.grid(row=1, column=0, sticky="ew", pady=(0, 15))
        
        # Old Password field
        old_pwd_frame = tk.Frame(form_frame, bg=surface_bg)
        old_pwd_frame.pack(fill=tk.X, pady=(0, 15))
        
        old_pwd_label = tk.Label(
            old_pwd_frame,
            text="Old Password",
            bg=surface_bg,
            fg=text_primary,
            font=("{Segoe UI Semibold} 11")
        )
        old_pwd_label.pack(anchor=tk.W, pady=(0, 10))
        
        old_pwd_var = tk.StringVar()
        old_pwd_entry_frame = tk.Frame(old_pwd_frame, bg=entry_border, bd=1, relief=tk.SOLID)
        old_pwd_entry_frame.pack(fill=tk.X)
        
        old_pwd_inner = tk.Frame(old_pwd_entry_frame, bg=surface_bg)
        old_pwd_inner.pack(fill=tk.BOTH, expand=True, padx=14, pady=14)
        
        old_pwd_icon = tk.Label(
            old_pwd_inner,
            text="🔒",
            bg=surface_bg,
            fg=text_secondary,
            font=("{Segoe UI} 14")
        )
        old_pwd_icon.pack(side=tk.LEFT, padx=(0, 10))
        
        old_pwd_entry = tk.Entry(
            old_pwd_inner,
            textvariable=old_pwd_var,
            font=("{Segoe UI} 12"),
            relief=tk.FLAT,
            bd=0,
            bg=surface_bg,
            fg=text_primary,
            show="*",
            insertbackground=accent_color,
            highlightthickness=0
        )
        old_pwd_entry.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        old_pwd_entry.focus()
        
        def on_old_pwd_focus_in(e):
            old_pwd_entry_frame.configure(bg=accent_color, bd=2)
            old_pwd_icon.config(fg=accent_color)
        
        def on_old_pwd_focus_out(e):
            old_pwd_entry_frame.configure(bg=entry_border, bd=1)
            old_pwd_icon.config(fg=text_secondary)
        
        old_pwd_entry.bind('<FocusIn>', on_old_pwd_focus_in)
        old_pwd_entry.bind('<FocusOut>', on_old_pwd_focus_out)
        
        # New Password field
        new_pwd_frame = tk.Frame(form_frame, bg=surface_bg)
        new_pwd_frame.pack(fill=tk.X, pady=(0, 15))
        
        new_pwd_label = tk.Label(
            new_pwd_frame,
            text="New Password",
            bg=surface_bg,
            fg=text_primary,
            font=("{Segoe UI Semibold} 11")
        )
        new_pwd_label.pack(anchor=tk.W, pady=(0, 10))
        
        new_pwd_var = tk.StringVar()
        new_pwd_entry_frame = tk.Frame(new_pwd_frame, bg=entry_border, bd=1, relief=tk.SOLID)
        new_pwd_entry_frame.pack(fill=tk.X)
        
        new_pwd_inner = tk.Frame(new_pwd_entry_frame, bg=surface_bg)
        new_pwd_inner.pack(fill=tk.BOTH, expand=True, padx=14, pady=14)
        
        new_pwd_icon = tk.Label(
            new_pwd_inner,
            text="🔒",
            bg=surface_bg,
            fg=text_secondary,
            font=("{Segoe UI} 14")
        )
        new_pwd_icon.pack(side=tk.LEFT, padx=(0, 10))
        
        new_pwd_entry = tk.Entry(
            new_pwd_inner,
            textvariable=new_pwd_var,
            font=("{Segoe UI} 12"),
            relief=tk.FLAT,
            bd=0,
            bg=surface_bg,
            fg=text_primary,
            show="*",
            insertbackground=accent_color,
            highlightthickness=0
        )
        new_pwd_entry.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        def on_new_pwd_focus_in(e):
            new_pwd_entry_frame.configure(bg=accent_color, bd=2)
            new_pwd_icon.config(fg=accent_color)
        
        def on_new_pwd_focus_out(e):
            new_pwd_entry_frame.configure(bg=entry_border, bd=1)
            new_pwd_icon.config(fg=text_secondary)
        
        new_pwd_entry.bind('<FocusIn>', on_new_pwd_focus_in)
        new_pwd_entry.bind('<FocusOut>', on_new_pwd_focus_out)
        
        # Confirm Password field
        confirm_pwd_frame = tk.Frame(form_frame, bg=surface_bg)
        confirm_pwd_frame.pack(fill=tk.X, pady=(0, 15))
        
        confirm_pwd_label = tk.Label(
            confirm_pwd_frame,
            text="Confirm New Password",
            bg=surface_bg,
            fg=text_primary,
            font=("{Segoe UI Semibold} 11")
        )
        confirm_pwd_label.pack(anchor=tk.W, pady=(0, 10))
        
        confirm_pwd_var = tk.StringVar()
        confirm_pwd_entry_frame = tk.Frame(confirm_pwd_frame, bg=entry_border, bd=1, relief=tk.SOLID)
        confirm_pwd_entry_frame.pack(fill=tk.X)
        
        confirm_pwd_inner = tk.Frame(confirm_pwd_entry_frame, bg=surface_bg)
        confirm_pwd_inner.pack(fill=tk.BOTH, expand=True, padx=14, pady=14)
        
        confirm_pwd_icon = tk.Label(
            confirm_pwd_inner,
            text="🔒",
            bg=surface_bg,
            fg=text_secondary,
            font=("{Segoe UI} 14")
        )
        confirm_pwd_icon.pack(side=tk.LEFT, padx=(0, 10))
        
        confirm_pwd_entry = tk.Entry(
            confirm_pwd_inner,
            textvariable=confirm_pwd_var,
            font=("{Segoe UI} 12"),
            relief=tk.FLAT,
            bd=0,
            bg=surface_bg,
            fg=text_primary,
            show="*",
            insertbackground=accent_color,
            highlightthickness=0
        )
        confirm_pwd_entry.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        def on_confirm_pwd_focus_in(e):
            confirm_pwd_entry_frame.configure(bg=accent_color, bd=2)
            confirm_pwd_icon.config(fg=accent_color)
        
        def on_confirm_pwd_focus_out(e):
            confirm_pwd_entry_frame.configure(bg=entry_border, bd=1)
            confirm_pwd_icon.config(fg=text_secondary)
        
        confirm_pwd_entry.bind('<FocusIn>', on_confirm_pwd_focus_in)
        confirm_pwd_entry.bind('<FocusOut>', on_confirm_pwd_focus_out)
        
        # Error message label
        error_frame = tk.Frame(form_frame, bg=surface_bg, height=35)
        error_frame.pack(fill=tk.X, pady=(0, 8))
        error_frame.pack_propagate(False)
        
        error_label = tk.Label(
            error_frame,
            text="",
            bg=surface_bg,
            fg="#ef4444",
            font=("{Segoe UI} 10"),
            wraplength=350,
            justify=tk.LEFT,
            anchor="nw"
        )
        error_label.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        def show_error(message: str):
            error_label.config(text=message)
        
        def hide_error():
            error_label.config(text="")
        
        # Spacer to push buttons to bottom
        spacer = tk.Frame(content_frame, bg=surface_bg, height=1)
        spacer.grid(row=2, column=0, sticky="nsew")
        
        # Buttons frame - ensure it's always visible with proper spacing
        btn_frame = tk.Frame(content_frame, bg=surface_bg)
        btn_frame.grid(row=3, column=0, sticky="ew", pady=(15, 10))
        
        # Configure column
        content_frame.grid_columnconfigure(0, weight=1)
        btn_frame.grid_columnconfigure(0, weight=1)
        btn_frame.grid_columnconfigure(1, weight=1)
        
        def do_change():
            hide_error()
            
            old_pwd = old_pwd_var.get().strip()
            new_pwd = new_pwd_var.get().strip()
            confirm_pwd = confirm_pwd_var.get().strip()
            
            if not old_pwd:
                show_error("Please enter your old password")
                old_pwd_entry.focus()
                return
            
            if not new_pwd:
                show_error("Please enter a new password")
                new_pwd_entry.focus()
                return
            
            if len(new_pwd) < 6:
                show_error("New password must be at least 6 characters long")
                new_pwd_entry.focus()
                return
            
            if new_pwd != confirm_pwd:
                show_error("New passwords do not match")
                confirm_pwd_entry.focus()
                return
            
            if self.auth_module.change_password(self.current_user['id'], old_pwd, new_pwd):
                messagebox.showinfo("Success", "Password changed successfully", parent=dialog)
                dialog.destroy()
            else:
                show_error("Password change failed. Please check your old password.")
                old_pwd_entry.focus()
                old_pwd_entry.select_range(0, tk.END)
        
        def on_cancel():
            dialog.destroy()
        
        # Cancel button
        cancel_btn = tk.Button(
            btn_frame,
            text="Cancel",
            command=on_cancel,
            bg="#6b7280",
            fg="white",
            padx=35,
            pady=14,
            font=("{Segoe UI Semibold} 11"),
            cursor="hand2",
            relief=tk.FLAT,
            bd=0,
            activebackground="#4b5563",
            activeforeground="white",
            width=12
        )
        cancel_btn.grid(row=0, column=0, padx=(0, 12), sticky="w")
        
        # Change Password button
        change_btn = tk.Button(
            btn_frame,
            text="Change Password",
            command=do_change,
            bg=accent_color,
            fg="white",
            padx=35,
            pady=14,
            font=("{Segoe UI Semibold} 11"),
            cursor="hand2",
            relief=tk.FLAT,
            bd=0,
            activebackground=accent_hover,
            activeforeground="white"
        )
        change_btn.grid(row=0, column=1, sticky="ew")
        
        def on_change_enter(e):
            change_btn.config(bg=accent_hover)
        
        def on_change_leave(e):
            change_btn.config(bg=accent_color)
        
        change_btn.bind('<Enter>', on_change_enter)
        change_btn.bind('<Leave>', on_change_leave)
        
        # Bind Enter key
        old_pwd_entry.bind('<Return>', lambda e: new_pwd_entry.focus())
        new_pwd_entry.bind('<Return>', lambda e: confirm_pwd_entry.focus())
        confirm_pwd_entry.bind('<Return>', lambda e: do_change())
        
        # Handle window close
        dialog.protocol('WM_DELETE_WINDOW', on_cancel)
        
        # Force update to ensure all widgets are rendered
        dialog.update_idletasks()
        dialog.update()
        
        # Ensure minimum height for all content to be visible
        dialog.minsize(480, 750)
        dialog.maxsize(480, 750)
        
        # Final update and focus
        dialog.update_idletasks()
        dialog.focus_force()
        dialog.lift()
    
    def _logout(self) -> None:
        """Logout current user."""
        if self.auth_module and self.session_token:
            self.auth_module.invalidate_session(self.session_token)
        self.current_user = None
        self.session_token = None
        if self._show_login():
            self._load_all_views()
        else:
            self.destroy()
    
    def _show_shortcuts_help(self) -> None:
        """Show keyboard shortcuts help."""
        shortcuts = """
Keyboard Shortcuts:

Navigation:
  Ctrl+1-0    Navigate to tabs
  F11          Toggle fullscreen
  Escape       Exit fullscreen

Actions:
  Ctrl+N       New Transaction
  Ctrl+S       Save
  Ctrl+F       Search
  Ctrl+I       Import Data
  Ctrl+D       Dashboard
  Ctrl+Z       Undo
  Ctrl+Y       Redo
  Ctrl+Q       Quit

Help:
  F1           Context Help
        """
        messagebox.showinfo("Keyboard Shortcuts", shortcuts.strip())
    
    def _show_context_help(self) -> None:
        """Show context-sensitive help."""
        messagebox.showinfo("Help", "Press F1 for help\nSee 'How to Use?' tab for detailed documentation")
    
    def _save_current(self) -> None:
        """Save current form (if applicable)."""
        # This would save the current transaction form if open
        self.set_status("Save functionality", "info")
    
    def _on_window_resize(self, event=None):
        try:
            if hasattr(self, '_resize_after_id') and self._resize_after_id:
                try:
                    self.after_cancel(self._resize_after_id)
                except Exception:
                    pass
            def _apply():
                self._apply_resize_layout()
                self._resize_after_id = None
            self._resize_after_id = self.after(120, _apply)
        except Exception:
            pass

    def _apply_resize_layout(self) -> None:
        if hasattr(self, 'main_frame') and hasattr(self.main_frame, 'winfo_children'):
            self.main_frame.update_idletasks()
        try:
            # --- Aspect-ratio-aware layout tweaks ---------------------------------
            # Use current window size to adjust proportions so the UI feels
            # balanced on wide desktop monitors vs. shorter laptop screens.
            try:
                win_w = max(self.winfo_width(), 1)
                win_h = max(self.winfo_height(), 1)
                aspect = win_w / win_h
            except Exception:
                win_w = win_h = 1
                aspect = 1.6  # sensible default

            # Sidebar width responds to aspect ratio:
            # - On very wide screens, keep it relatively slim.
            # - On taller / squarer laptop-style windows, allow a bit more width.
            if hasattr(self, 'sidebar'):
                if aspect >= 1.8:
                    # Very wide (big external monitor)
                    base_collapsed = 80
                    base_expanded = 210
                elif aspect >= 1.4:
                    # Typical laptop / standard monitor
                    base_collapsed = 90
                    base_expanded = 230
                else:
                    # Tall window (portrait / split-screen)
                    base_collapsed = 100
                    base_expanded = 250

                self.sidebar_collapsed_width = base_collapsed
                self.sidebar_expanded_width = base_expanded

                target_width = self.sidebar_expanded_width if getattr(self, "_sidebar_expanded", False) else self.sidebar_collapsed_width
                try:
                    self.sidebar.configure(width=target_width)
                    self.sidebar.update_idletasks()
                except Exception:
                    pass

            # Optionally tweak the "Global Action" button styling for very wide monitors.
            try:
                if hasattr(self, "global_action_btn"):
                    if win_w >= 1400:
                        self.global_action_btn.configure(style="Techfix.TButton")
                    else:
                        self.global_action_btn.configure(style="Techfix.TButton")
            except Exception:
                pass

            # --- Existing responsive column behavior ------------------------------
            if hasattr(self, 'cycle_tree'):
                total = self.cycle_tree.winfo_width() or 1
                col_defs = {
                    'step': int(total * 0.08),
                    'name': int(total * 0.32),
                    'status': int(total * 0.18),
                    'note': int(total * 0.30),
                    'updated': int(total * 0.12),
                }
                for c, w in col_defs.items():
                    try:
                        self.cycle_tree.column(c, width=max(60, w))
                    except Exception:
                        pass
            if hasattr(self, 'notebook'):
                self.notebook.update_idletasks()
            if hasattr(self, 'closing_preview_tree'):
                total = self.closing_preview_tree.winfo_width() or 1
                widths = {
                    'code': int(total * 0.15),
                    'name': int(total * 0.35),
                    'action': int(total * 0.30),
                    'amount': int(total * 0.20),
                }
                for c, w in widths.items():
                    try:
                        self.closing_preview_tree.column(c, width=max(80, w))
                    except Exception:
                        pass
            if hasattr(self, 'txn_recent_tree'):
                total = self.txn_recent_tree.winfo_width() or 1
                col_defs = {
                    'date': int(total * 0.12),
                    'reference': int(total * 0.10),
                    'description': int(total * 0.42),
                    'debit': int(total * 0.12),
                    'credit': int(total * 0.12),
                    'account': int(total * 0.12),
                }
                for c, w in col_defs.items():
                    try:
                        self.txn_recent_tree.column(c, width=max(80, w), stretch=(c == 'description'))
                    except Exception:
                        pass
            if hasattr(self, 'journal_tree'):
                total = self.journal_tree.winfo_width() or 1
                col_defs = {
                    'date': int(total * 0.12),
                    'reference': int(total * 0.10),
                    'description': int(total * 0.42),
                    'debit': int(total * 0.12),
                    'credit': int(total * 0.12),
                    'account': int(total * 0.12),
                }
                for c, w in col_defs.items():
                    try:
                        self.journal_tree.column(c, width=max(80, w), stretch=(c == 'description'))
                    except Exception:
                        pass
            
            # --- Responsive button positioning for transaction form ---
            # Ensure buttons stay visible regardless of aspect ratio
            if hasattr(self, 'action_frame') and hasattr(self, 'btn_container') and hasattr(self, 'btn_recent'):
                try:
                    # Get the available width of the action frame
                    self.action_frame.update_idletasks()
                    frame_width = self.action_frame.winfo_width() or 500
                    
                    # Always keep all buttons in one horizontal row
                    self.btn_recent.grid_configure(row=0, column=0)
                    self.btn_clear.grid_configure(row=0, column=1)
                    self.btn_draft.grid_configure(row=0, column=2)
                    self.btn_post.grid_configure(row=0, column=3)
                    self.btn_container.grid_configure(sticky="e")
                except Exception:
                    pass
            
            # --- Responsive scan buttons layout ---
            # Wrap scan buttons to multiple rows when space is tight
            if hasattr(self, 'scan_row') and hasattr(self, 'scan_btn'):
                try:
                    self.scan_row.update_idletasks()
                    scan_row_width = self.scan_row.winfo_width() or 400
                    
                    # Estimate width needed for all buttons in one row
                    # Scan (10) + Scan Image (12) + Enter Manually (14) + padding ≈ 350px
                    estimated_scan_width = 350
                    
                    if scan_row_width >= estimated_scan_width:
                        # Enough space - keep all buttons in one row
                        self.scan_btn.grid_configure(row=0, column=0)
                        self.scan_img_btn.grid_configure(row=0, column=1)
                        self.manual_entry_btn.grid_configure(row=0, column=2)
                    else:
                        # Tight space - wrap to two rows
                        self.scan_btn.grid_configure(row=0, column=0)
                        self.scan_img_btn.grid_configure(row=1, column=0)
                        self.manual_entry_btn.grid_configure(row=1, column=1)
                except Exception:
                    pass
            
            # --- Options container layout ---
            # Always keep all options in one horizontal row
            if hasattr(self, 'options_container') and hasattr(self, 'adjust_cb'):
                try:
                    # Always keep all in one row
                    self.adjust_cb.grid_configure(row=0, column=0)
                    self.reverse_cb.grid_configure(row=0, column=1)
                    self.reverse_label.grid_configure(row=0, column=2)
                    self.txn_reverse_date.grid_configure(row=0, column=3)
                except Exception:
                    pass
            
            # --- Responsive Journal tab toolbar layout ---
            if hasattr(self, 'journal_toolbar') and hasattr(self, 'journal_left_buttons'):
                try:
                    self.journal_toolbar.update_idletasks()
                    toolbar_width = self.journal_toolbar.winfo_width() or 800
                    
                    # Estimate width needed: Refresh + Export + Paging + Filters ≈ 600px
                    estimated_toolbar_width = 600
                    
                    if toolbar_width >= estimated_toolbar_width:
                        # Enough space - keep all in one row
                        self.journal_left_buttons.grid_configure(row=0, column=0)
                        self.journal_filter_frame.grid_configure(row=0, column=2)
                        # Reset filter widgets to row 0
                        if hasattr(self, 'journal_filter_label'):
                            self.journal_filter_label.grid_configure(row=0, column=0)
                            self.journal_date_from.grid_configure(row=0, column=1)
                            if hasattr(self, 'journal_to_label'):
                                self.journal_to_label.grid_configure(row=0, column=2)
                            self.journal_date_to.grid_configure(row=0, column=3)
                            self.journal_account_filter.grid_configure(row=0, column=4)
                    else:
                        # Tight space - wrap filters to second row
                        self.journal_left_buttons.grid_configure(row=0, column=0)
                        self.journal_filter_frame.grid_configure(row=1, column=0, columnspan=3, sticky="w")
                        # Move filter widgets to row 1
                        if hasattr(self, 'journal_filter_label'):
                            self.journal_filter_label.grid_configure(row=1, column=0)
                            self.journal_date_from.grid_configure(row=1, column=1)
                            if hasattr(self, 'journal_to_label'):
                                self.journal_to_label.grid_configure(row=1, column=2)
                            self.journal_date_to.grid_configure(row=1, column=3)
                            self.journal_account_filter.grid_configure(row=1, column=4)
                except Exception:
                    pass
            
            # --- Responsive Ledger tab toolbar layout ---
            if hasattr(self, 'ledger_toolbar') and hasattr(self, 'ledger_left_buttons'):
                try:
                    self.ledger_toolbar.update_idletasks()
                    toolbar_width = self.ledger_toolbar.winfo_width() or 800
                    
                    # Estimate width needed: Refresh + Post + Export + Paging + Filter ≈ 650px
                    estimated_toolbar_width = 650
                    
                    if toolbar_width >= estimated_toolbar_width:
                        # Enough space - keep all in one row
                        self.ledger_left_buttons.grid_configure(row=0, column=0)
                        self.ledger_filter_frame.grid_configure(row=0, column=2)
                        # Reset filter widgets to row 0
                        if hasattr(self, 'ledger_account_label'):
                            self.ledger_account_label.grid_configure(row=0, column=0)
                            self.ledger_account_filter.grid_configure(row=0, column=1)
                    else:
                        # Tight space - wrap filters to second row
                        self.ledger_left_buttons.grid_configure(row=0, column=0)
                        self.ledger_filter_frame.grid_configure(row=1, column=0, columnspan=3, sticky="w")
                        # Move filter widgets to row 1
                        if hasattr(self, 'ledger_account_label'):
                            self.ledger_account_label.grid_configure(row=1, column=0)
                            self.ledger_account_filter.grid_configure(row=1, column=1)
                except Exception:
                    pass
            
            # --- Responsive Financial Statements tab buttons layout ---
            # Keep all buttons in a single horizontal row
            if hasattr(self, 'fs_btn_frame') and hasattr(self, 'fs_run_btn'):
                try:
                    # Always keep all buttons in one horizontal row
                    self.fs_preset_box.grid_configure(row=0, column=0)
                    self.fs_run_btn.grid_configure(row=0, column=1)
                    self.fs_export_xls_btn.grid_configure(row=0, column=2)
                    self.fs_export_txt_btn.grid_configure(row=0, column=3)
                except Exception:
                    pass
        except Exception:
            pass

    def _read_settings_file(self) -> dict:
        try:
            settings_path = db.DB_DIR / "settings.json"
            if settings_path.exists():
                data = json.loads(settings_path.read_text(encoding="utf-8"))
                if isinstance(data, dict):
                    return data
        except Exception:
            pass
        return {}

    def _write_settings_file(self, data: dict) -> None:
        try:
            settings_path = db.DB_DIR / "settings.json"
            settings_path.parent.mkdir(parents=True, exist_ok=True)
            settings_path.write_text(json.dumps(data, indent=2), encoding="utf-8")
        except Exception:
            pass

    def _load_window_settings(self):
        try:
            data = self._read_settings_file()
            geom = data.get("geometry")
            full = data.get("fullscreen")
            theme = data.get("theme")
            if isinstance(geom, str) and geom:
                self.geometry(geom)
            if isinstance(full, bool):
                self.attributes('-fullscreen', full)
            if isinstance(theme, str) and theme in THEMES:
                self.theme_name = theme
                self.palette = THEMES[theme].copy()  # Use copy to avoid modifying original
                # Load custom colors if available
                custom_colors = self._load_custom_colors()
                if custom_colors:
                    self.palette.update(custom_colors)
            # Load saved email accounts (used by export dialog)
            email_accounts = data.get("email_accounts")
            self.email_accounts = email_accounts if isinstance(email_accounts, list) else []
        except Exception:
            self.email_accounts = []

    def _save_window_settings(self):
        try:
            data = self._read_settings_file()
            data.update(
                {
                    "geometry": self.winfo_geometry(),
                    "fullscreen": bool(self.attributes('-fullscreen')),
                    "theme": self.theme_name,
                    "email_accounts": getattr(self, "email_accounts", []),
                }
            )
            self._write_settings_file(data)
        except Exception:
            pass

    def _save_email_account(self, account: dict, remember_password: bool) -> None:
        """Persist a recently used email account (optionally storing password)."""
        try:
            # Ensure container exists
            if not hasattr(self, "email_accounts") or not isinstance(self.email_accounts, list):
                self.email_accounts = []
            acct = {
                "recipient": account.get("recipient", "").strip(),
                "smtp_server": account.get("smtp_server", "").strip(),
                "smtp_port": account.get("smtp_port", "").strip() or "587",
                "smtp_user": account.get("smtp_user", "").strip(),
            }
            if remember_password:
                acct["smtp_password"] = account.get("smtp_password", "")
            else:
                acct["smtp_password"] = ""

            # Deduplicate by smtp_user + server + port
            key = (acct["smtp_user"], acct["smtp_server"], acct["smtp_port"])
            filtered: list = []
            for item in self.email_accounts:
                if not isinstance(item, dict):
                    continue
                if (item.get("smtp_user"), item.get("smtp_server"), item.get("smtp_port")) == key:
                    continue
                filtered.append(item)
            # Prepend latest and cap list length
            filtered.insert(0, acct)
            self.email_accounts = filtered[:5]
            self._save_window_settings()
        except Exception:
            pass

    def _on_close(self):
        try:
            self._save_window_settings()
        except Exception:
            pass
        self.destroy()

    def _on_global_action(self) -> None:
        """
        Open a small chooser window that lets the user pick a target
        aspect ratio (e.g., laptop vs. desktop monitor). Choosing an option
        resizes this main window to that preset.
        """
        try:
            # Reuse the chooser window if it already exists.
            if hasattr(self, "_aspect_chooser") and self._aspect_chooser is not None:
                try:
                    self._aspect_chooser.deiconify()
                    self._aspect_chooser.lift()
                    self._aspect_chooser.focus_set()
                    return
                except Exception:
                    self._aspect_chooser = None

            chooser = tk.Toplevel(self)
            chooser.title("Choose Aspect Ratio")
            chooser.resizable(False, False)
            chooser.transient(self)
            try:
                chooser.grab_set()
            except Exception:
                pass

            self._aspect_chooser = chooser

            frame = ttk.Frame(chooser, padding=12)
            frame.pack(fill=tk.BOTH, expand=True)

            ttk.Label(
                frame,
                text="Pick a window size / aspect ratio profile:",
                style="Techfix.AppBar.TLabel",
            ).pack(anchor="w", pady=(0, 8))

            # Preset options: (label, width, height)
            presets = [
                ("Laptop 16:9 (1366 x 768)", 1366, 768),
                ("Desktop 16:9 (1920 x 1080)", 1920, 1080),
                ("Square-ish 4:3 (1200 x 900)", 1200, 900),
                ("Tall / Coding Split (1100 x 1200)", 1100, 1200),
            ]

            for text, w, h in presets:
                ttk.Button(
                    frame,
                    text=text,
                    style="Techfix.TButton",
                    command=lambda ww=w, hh=h, win=chooser: self._apply_aspect_preset(ww, hh, win),
                ).pack(fill=tk.X, pady=3)

            ttk.Button(
                frame,
                text="Close",
                style="Techfix.TButton",
                command=chooser.destroy,
            ).pack(fill=tk.X, pady=(10, 0))

            # Roughly center the chooser over the main window.
            try:
                self.update_idletasks()
                chooser.update_idletasks()
                main_x = self.winfo_rootx()
                main_y = self.winfo_rooty()
                main_w = self.winfo_width()
                main_h = self.winfo_height()
                ch_w = chooser.winfo_width()
                ch_h = chooser.winfo_height()
                x = max(0, main_x + (main_w - ch_w) // 2)
                y = max(0, main_y + (main_h - ch_h) // 2)
                chooser.geometry(f"+{x}+{y}")
            except Exception:
                pass
        except Exception:
            pass

    def _apply_aspect_preset(self, target_width: int, target_height: int, chooser: Optional[tk.Toplevel] = None) -> None:
        """
        Resize the main window to a given width/height, roughly centered
        on the current screen, and close the chooser if provided.
        """
        try:
            # Clamp to screen size so we don't place the window off-screen.
            screen_w = max(self.winfo_screenwidth(), 1)
            screen_h = max(self.winfo_screenheight(), 1)

            w = min(target_width, screen_w)
            h = min(target_height, screen_h)

            x = max(0, (screen_w - w) // 2)
            y = max(0, (screen_h - h) // 2)

            # Ensure normal state so geometry is respected.
            try:
                self.state("normal")
            except Exception:
                pass

            self.geometry(f"{w}x{h}+{x}+{y}")

            if chooser is not None:
                try:
                    chooser.destroy()
                except Exception:
                    pass
        except Exception:
            pass

    def set_status(self, text: str, status_type: str = "info") -> None:
        """Set status message with optional status badge."""
        try:
            if hasattr(self, 'status_var'):
                self.status_var.set(str(text))
            
            # Update header status area with badge if available
            if hasattr(self, 'header_status_area'):
                try:
                    # Clear existing badges
                    for widget in self.header_status_area.winfo_children():
                        widget.destroy()
                    
                    # Create new status badge
                    if text:
                        badge = self._create_status_badge(
                            self.header_status_area,
                            text,
                            status=status_type,
                            pulse=(status_type == "warning" or status_type == "error")
                        )
                        badge.pack(side=tk.RIGHT, padx=4)
                except Exception:
                    pass
        except Exception:
            pass
    
    def show_loading(self, message: str = "Please wait...", parent=None):
        """Show a loading spinner with message."""
        import time
        try:
            if parent is None:
                parent = self
            
            # Create loading overlay
            loading_frame = self._create_glassmorphism_overlay(parent)
            loading_frame.place(x=0, y=0, relwidth=1, relheight=1)
            loading_frame.lift()
            
            # Create loading content with rounded corners effect
            content_frame = tk.Frame(
                loading_frame, 
                bg=self.palette.get('surface_bg', '#ffffff'),
                relief=tk.FLAT,
                bd=0
            )
            content_frame.place(relx=0.5, rely=0.5, anchor=tk.CENTER)
            
            # Add padding frame for better visual appearance
            padding_frame = tk.Frame(content_frame, bg=self.palette.get('surface_bg', '#ffffff'))
            padding_frame.pack(padx=30, pady=30)
            
            # Create animated spinner
            spinner = self._create_loading_spinner(padding_frame, size=50)
            spinner.pack(pady=(0, 15))
            
            # Loading message label
            label = tk.Label(
                padding_frame,
                text=message,
                bg=self.palette.get('surface_bg', '#ffffff'),
                fg=self.palette.get('text_primary', '#1f2937'),
                font=("{Segoe UI} 12")
            )
            label.pack()
            
            # Store timestamp when loading was shown (minimum 2 seconds display)
            loading_frame._loading_active = True
            loading_frame._content_frame = content_frame
            loading_frame._show_time = time.time()
            return loading_frame
        except Exception:
            return None
    
    def hide_loading(self, loading_frame):
        """Hide loading spinner after ensuring minimum 2 second display time."""
        import time
        try:
            if loading_frame and hasattr(loading_frame, '_loading_active'):
                # Calculate elapsed time since loading was shown
                if hasattr(loading_frame, '_show_time'):
                    elapsed = time.time() - loading_frame._show_time
                    min_display_time = 2.0  # Minimum 2 seconds
                    
                    if elapsed < min_display_time:
                        # Schedule hide after remaining time
                        remaining = min_display_time - elapsed
                        loading_frame.after(int(remaining * 1000), lambda: self._actually_hide_loading(loading_frame))
                        return
                
                # If enough time has passed, hide immediately
                self._actually_hide_loading(loading_frame)
        except Exception:
            pass
    
    def _actually_hide_loading(self, loading_frame):
        """Actually hide and destroy the loading frame."""
        try:
            if loading_frame and hasattr(loading_frame, '_loading_active'):
                loading_frame.place_forget()
                loading_frame.destroy()
        except Exception:
            pass

    # --- Centralized error / logging helpers ---------------------------------

    def _handle_exception(self, context: str, exc: BaseException) -> None:
        """
        Log an exception with context and surface a short message in the status bar.

        Call this in addition to any user‑facing messagebox for operations where
        we want better diagnostics without changing existing UX.
        """
        try:
            logger.exception("TechFixApp error in %s", context, exc_info=exc)
        except Exception:
            # Never allow logging failures to propagate into the UI.
            pass
        try:
            self.set_status(f"{context}: {exc}")
        except Exception:
            pass

    def _report_callback_exception(self, exc_type, exc_value, exc_traceback) -> None:  # type: ignore[override]
        """
        Tkinter hook: called for exceptions raised in event callbacks.
        Route these through the logger and show a generic dialog so the app
        doesn't silently swallow errors.
        """
        try:
            logger.error(
                "Unhandled Tk callback exception",
                exc_info=(exc_type, exc_value, exc_traceback),
            )
        except Exception:
            pass
        try:
            messagebox.showerror(
                "Unexpected Error",
                "An unexpected error occurred in the UI.\n\n"
                "Details have been logged; please check the log output for more information.",
            )
        except Exception:
            # As a last resort, ignore if even messagebox fails.
            pass

    def _on_period_change(self, event=None):
        selected = self.period_var.get()
        if not selected:
            return
        for period_id, name in self.periods:
            if name == selected:
                # snapshot previous period form
                try:
                    if self._last_period_id:
                        snap = self._snapshot_txn_form()
                        if snap:
                            self.period_form_cache[int(self._last_period_id)] = snap
                except Exception:
                    pass
                self.current_period_id = period_id
                try:
                    self.engine.set_active_period(period_id)
                except Exception:
                    pass
                self._refresh_cycle_and_views()
                # apply cached form or prefill from last entry
                try:
                    snap = self.period_form_cache.get(int(period_id))
                    if snap:
                        self._apply_txn_form(snap)
                    else:
                        self._prefill_txn_from_last_entry()
                except Exception:
                    pass
                self._last_period_id = period_id
                break

    def _load_periods(self):
        try:
            rows = self.engine.list_periods()
            items = []
            for r in rows:
                name = r["name"] if isinstance(r, sqlite3.Row) else (r[1] if len(r) > 1 else "")
                pid = r["id"] if isinstance(r, sqlite3.Row) else r[0]
                items.append((pid, name))
            self.periods = items
            if hasattr(self, 'period_combo'):
                current = self.period_combo.get()
                self.period_combo['values'] = [p[1] for p in self.periods]
                if self.periods and not current:
                    self.current_period_id = self.periods[0][0]
                    self.period_var.set(self.periods[0][1])
                    try:
                        self.engine.set_active_period(self.current_period_id)
                    except Exception:
                        pass
        except Exception:
            pass

    def _prompt_new_period(self):
        """Show dialog to create a new accounting period."""
        # Create a simple dialog to get start and end dates
        class PeriodDialog(tk.Toplevel):
            def __init__(self, parent):
                super().__init__(parent)
                self.title("New Accounting Period")
                self.parent = parent
                self.result = None
                try:
                    self.configure(bg=parent.palette.get('surface_bg', '#ffffff'))
                except Exception:
                    pass
                
                # Make dialog modal
                self.transient(parent)
                self.grab_set()
                
                # Set focus
                self.focus_set()
                
                # Add widgets
                ttk.Label(self, text="Start Date (YYYY-MM-DD):", style="Techfix.TLabel").grid(row=0, column=0, padx=5, pady=5)
                self.start_entry = ttk.Entry(self, style="Techfix.TEntry")
                self.start_entry.grid(row=0, column=1, padx=5, pady=5)
                ttk.Button(self, text="📅", command=lambda: self._pick_date(self.start_entry), style="Techfix.Theme.TButton", width=3).grid(row=0, column=2, padx=2)
                
                ttk.Label(self, text="End Date (YYYY-MM-DD):", style="Techfix.TLabel").grid(row=1, column=0, padx=5, pady=5)
                self.end_entry = ttk.Entry(self, style="Techfix.TEntry")
                self.end_entry.grid(row=1, column=1, padx=5, pady=5)
                ttk.Button(self, text="📅", command=lambda: self._pick_date(self.end_entry), style="Techfix.Theme.TButton", width=3).grid(row=1, column=2, padx=2)
                
                btn_frame = ttk.Frame(self, style="Techfix.Surface.TFrame")
                btn_frame.grid(row=2, column=0, columnspan=3, pady=10)
                
                ttk.Button(btn_frame, text="Create", command=self.on_ok, style="Techfix.TButton").pack(side=tk.LEFT, padx=5)
                ttk.Button(btn_frame, text="Cancel", command=self.on_cancel, style="Techfix.Theme.TButton").pack(side=tk.LEFT, padx=5)
                
                # Center the dialog
                self.update_idletasks()
                width = self.winfo_width()
                height = self.winfo_height()
                x = (self.winfo_screenwidth() // 2) - (width // 2)
                y = (self.winfo_screenheight() // 2) - (height // 2)
                self.geometry(f'{width}x{height}+{x}+{y}')

            def _pick_date(self, entry_widget):
                class DatePicker(tk.Toplevel):
                    def __init__(self, parent, callback=None, initial_date=None):
                        super().__init__(parent)
                        self.transient(parent)
                        self.grab_set()
                        self.callback = callback
                        self.title("Pick Date")
                        self.resizable(False, False)

                        today = datetime.now().date()
                        if initial_date:
                            try:
                                dt = datetime.strptime(initial_date, '%Y-%m-%d').date()
                            except Exception:
                                dt = today
                        else:
                            dt = today

                        self.year = dt.year
                        self.month = dt.month

                        header = ttk.Frame(self)
                        header.pack(fill=tk.X, padx=8, pady=6)
                        ttk.Button(header, text="◀", width=3, command=self._prev_month).pack(side=tk.LEFT)
                        self.title_lbl = ttk.Label(header, text="", anchor=tk.CENTER)
                        self.title_lbl.pack(side=tk.LEFT, expand=True)
                        ttk.Button(header, text="▶", width=3, command=self._next_month).pack(side=tk.RIGHT)

                        self.cal_frame = ttk.Frame(self)
                        self.cal_frame.pack(padx=8, pady=(0,8))
                        self._draw_calendar()

                    def _draw_calendar(self):
                        for child in self.cal_frame.winfo_children():
                            child.destroy()
                        self.title_lbl.config(text=f"{calendar.month_name[self.month]} {self.year}")
                        wdays = ['Mo','Tu','We','Th','Fr','Sa','Su']
                        for c, wd in enumerate(wdays):
                            ttk.Label(self.cal_frame, text=wd, width=3, anchor='center').grid(row=0, column=c)
                        m = calendar.monthcalendar(self.year, self.month)
                        for r, week in enumerate(m, start=1):
                            for c, day in enumerate(week):
                                if day == 0:
                                    ttk.Label(self.cal_frame, text='', width=3).grid(row=r, column=c, padx=1, pady=1)
                                else:
                                    b = ttk.Button(self.cal_frame, text=str(day), width=3, command=lambda d=day: self._select_day(d))
                                    b.grid(row=r, column=c, padx=1, pady=1)

                    def _prev_month(self):
                        self.month -= 1
                        if self.month < 1:
                            self.month = 12
                            self.year -= 1
                        self._draw_calendar()

                    def _next_month(self):
                        self.month += 1
                        if self.month > 12:
                            self.month = 1
                            self.year += 1
                        self._draw_calendar()

                    def _select_day(self, day):
                        try:
                            d = datetime(self.year, self.month, day).strftime('%Y-%m-%d')
                            if callable(self.callback):
                                self.callback(d)
                        except Exception:
                            pass
                        finally:
                            try:
                                self.grab_release()
                            except Exception:
                                pass
                            self.destroy()

                cur = entry_widget.get().strip()
                def _on_date(d):
                    entry_widget.delete(0, tk.END)
                    entry_widget.insert(0, d)
                dp = DatePicker(self, callback=_on_date, initial_date=cur)
                self.wait_window(dp)
                
            def on_ok(self):
                start_date = self.start_entry.get()
                end_date = self.end_entry.get()
                # Auto-compute end date if omitted (last day of start month)
                if start_date and not end_date:
                    try:
                        s = datetime.strptime(start_date, '%Y-%m-%d')
                        # move to first day of next month then step back one day
                        if s.month == 12:
                            nxt = datetime(s.year + 1, 1, 1)
                        else:
                            nxt = datetime(s.year, s.month + 1, 1)
                        end_date = (nxt - timedelta(days=1)).strftime('%Y-%m-%d')
                        self.end_entry.delete(0, tk.END)
                        self.end_entry.insert(0, end_date)
                    except Exception:
                        pass
                # Basic validation
                try:
                    start = datetime.strptime(start_date, '%Y-%m-%d')
                    end = datetime.strptime(end_date, '%Y-%m-%d')
                    
                    if start >= end:
                        messagebox.showerror("Error", "End date must be after start date")
                        return
                        
                    self.result = (start_date, end_date)
                    self.destroy()
                except ValueError:
                    messagebox.showerror("Error", "Please enter dates in YYYY-MM-DD format")
            
            def on_cancel(self):
                self.destroy()
        
        # Show the dialog and process result
        dialog = PeriodDialog(self)
        self.wait_window(dialog)
        
        if dialog.result:
            start_date, end_date = dialog.result
            try:
                name = start_date[:7]
                try:
                    pid = self.engine.create_period(name=name, start_date=start_date, end_date=end_date, make_current=True)
                    self.current_period_id = pid
                    messagebox.showinfo("Success", "New accounting period created successfully")
                    # Refresh period list, select the new period, and load views immediately
                    self._load_periods()
                    if hasattr(self, 'period_combo'):
                        self.period_var.set(name)
                        try:
                            self.period_combo.set(name)
                        except Exception:
                            pass
                    self._on_period_change()
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to create period: {str(e)}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to create period: {str(e)}")

    def _load_cycle_status(self):
        try:
            if not hasattr(self, 'cycle_tree'):
                return
            if not self.engine.current_period_id:
                self.engine.refresh_current_period()
            rows = self.engine.get_cycle_status()
            for item in self.cycle_tree.get_children():
                self.cycle_tree.delete(item)
            for r in rows:
                vals = (r['step'], r['step_name'], r['status'], r['note'] if 'note' in r else '', r['updated_at'])
                tag = r['status']
                self.cycle_tree.insert('', 'end', values=vals, tags=(tag,))
            try:
                self.cycle_tree.tag_configure('completed', background=self.palette.get('tab_selected_bg', '#e0ecff'))
                self.cycle_tree.tag_configure('in_progress', background=self.palette.get('table_stripe', '#eef2ff'))
                self.cycle_tree.tag_configure('pending', background=self.palette.get('surface_bg', '#ffffff'))
            except Exception:
                pass
            pass
        except Exception as e:
            print(f"Error loading cycle status: {e}")

    def _refresh_cycle_and_views(self) -> None:
        """Refresh cycle and all views with loading indicator."""
        loading = self.show_loading("Please wait... Refreshing data...")
        try:
            self.update()  # Update UI to show loading indicator
            self.engine.refresh_current_period()
        except Exception:
            pass
        try:
            self._load_cycle_status()
        except Exception:
            pass
        try:
            self._load_all_views()
        except Exception:
            pass
        finally:
            if loading:
                self.hide_loading(loading)

    def _load_15_entries_example(self) -> None:
        """Load 15 example entries for testing purposes."""
        try:
            # Confirm with user
            result = messagebox.askyesno(
                "Load Example Entries",
                "This will create 15 sample journal entries in the current period.\n\n"
                "This includes:\n"
                "- 10 regular transactions\n"
                "- 3 adjusting entries\n"
                "- 2 closing entries\n\n"
                "Continue?",
                icon='question'
            )
            if not result:
                return

            # Helper function to get account ID by name
            def get_account_id(name: str) -> int:
                account = db.get_account_by_name(name, self.engine.conn)
                if account is None:
                    raise ValueError(f"Account '{name}' not found in database")
                return account['id']

            # Get account IDs
            accounts = {}
            account_names = [
                'Cash', "Owner's Capital", 'Supplies', 'Office Equipment',
                'Accounts Payable', 'Accounts Receivable', 'Service Revenue',
                'Utilities Expense', 'Rent Expense', "Owner's Drawings",
                'Supplies Expense', 'Depreciation Expense', 'Accumulated Depreciation',
                'Utilities Payable', 'Salaries & Wages'
            ]
            
            for name in account_names:
                accounts[name] = get_account_id(name)

            # Helper function to post entries
            def post(date_str: str, description: str, lines: List[tuple], is_adjusting: bool = False, is_closing: bool = False) -> int:
                journal_lines = [JournalLine(account_id=acc_id, debit=debit, credit=credit) 
                                for acc_id, debit, credit in lines]
                return self.engine.record_entry(
                    date_str,
                    description,
                    journal_lines,
                    is_adjusting=is_adjusting,
                    is_closing=is_closing,
                    status='posted'
                )

            # Set up dates
            year = date.today().year
            base_date = date(year, 1, 1)

            # Entry 1: Owner investment
            post(
                (base_date + timedelta(days=0)).isoformat(),
                'Owner investment',
                [(accounts['Cash'], 150000.0, 0.0), (accounts["Owner's Capital"], 0.0, 150000.0)]
            )

            # Entry 2: Purchase supplies for cash
            post(
                (base_date + timedelta(days=2)).isoformat(),
                'Purchase supplies (cash)',
                [(accounts['Supplies'], 8000.0, 0.0), (accounts['Cash'], 0.0, 8000.0)]
            )

            # Entry 3: Purchase equipment on account
            post(
                (base_date + timedelta(days=5)).isoformat(),
                'Purchase office equipment on account',
                [(accounts['Office Equipment'], 50000.0, 0.0), (accounts['Accounts Payable'], 0.0, 50000.0)]
            )

            # Entry 4: Service revenue (cash)
            post(
                (base_date + timedelta(days=10)).isoformat(),
                'Service revenue (cash)',
                [(accounts['Cash'], 25000.0, 0.0), (accounts['Service Revenue'], 0.0, 25000.0)]
            )

            # Entry 5: Service revenue (billed)
            post(
                (base_date + timedelta(days=12)).isoformat(),
                'Service revenue (billed)',
                [(accounts['Accounts Receivable'], 35000.0, 0.0), (accounts['Service Revenue'], 0.0, 35000.0)]
            )

            # Entry 6: Pay rent expense
            post(
                (base_date + timedelta(days=15)).isoformat(),
                'Paid rent expense',
                [(accounts['Rent Expense'], 12000.0, 0.0), (accounts['Cash'], 0.0, 12000.0)]
            )

            # Entry 7: Pay utilities expense
            post(
                (base_date + timedelta(days=18)).isoformat(),
                'Paid utilities expense',
                [(accounts['Utilities Expense'], 5000.0, 0.0), (accounts['Cash'], 0.0, 5000.0)]
            )

            # Entry 8: Pay salaries
            post(
                (base_date + timedelta(days=20)).isoformat(),
                'Paid salaries',
                [(accounts['Salaries & Wages'], 20000.0, 0.0), (accounts['Cash'], 0.0, 20000.0)]
            )

            # Entry 9: Received collection from AR
            post(
                (base_date + timedelta(days=22)).isoformat(),
                'Received collection from accounts receivable',
                [(accounts['Cash'], 20000.0, 0.0), (accounts['Accounts Receivable'], 0.0, 20000.0)]
            )

            # Entry 10: Paid accounts payable
            post(
                (base_date + timedelta(days=25)).isoformat(),
                'Paid accounts payable',
                [(accounts['Accounts Payable'], 30000.0, 0.0), (accounts['Cash'], 0.0, 30000.0)]
            )

            # Entry 11: Adjust supplies used
            remaining_supplies = 3000.0
            supplies_used = 8000.0 - remaining_supplies
            post(
                (base_date + timedelta(days=28)).isoformat(),
                'Adjust supplies used',
                [(accounts['Supplies Expense'], supplies_used, 0.0), (accounts['Supplies'], 0.0, supplies_used)],
                is_adjusting=True
            )

            # Entry 12: Record depreciation
            depreciation_amount = 1000.0
            post(
                (base_date + timedelta(days=28)).isoformat(),
                'Record depreciation expense',
                [(accounts['Depreciation Expense'], depreciation_amount, 0.0), 
                 (accounts['Accumulated Depreciation'], 0.0, depreciation_amount)],
                is_adjusting=True
            )

            # Entry 13: Accrue utilities expense
            accrued_utilities = 2000.0
            post(
                (base_date + timedelta(days=28)).isoformat(),
                'Accrue utilities expense',
                [(accounts['Utilities Expense'], accrued_utilities, 0.0), 
                 (accounts['Utilities Payable'], 0.0, accrued_utilities)],
                is_adjusting=True
            )

            # Entries 14-15: Closing entries
            closing_date = (base_date + timedelta(days=31)).isoformat()
            closing_entry_ids = self.engine.make_closing_entries(closing_date)

            # Refresh views
            self._refresh_cycle_and_views()

            messagebox.showinfo(
                "Success",
                f"Successfully loaded 15 example entries!\n\n"
                f"- 10 regular transactions\n"
                f"- 3 adjusting entries\n"
                f"- {len(closing_entry_ids)} closing entries\n\n"
                f"All entries have been posted to the current period."
            )

        except Exception as e:
            logger.exception("Error loading 15 entries example")
            messagebox.showerror("Error", f"Failed to load example entries:\n{str(e)}")

    def _snapshot_txn_form(self) -> Dict[str, str]:
        snap: Dict[str, str] = {}
        try:
            if hasattr(self, 'txn_date'):
                snap['date'] = self.txn_date.get().strip()
            if hasattr(self, 'txn_desc'):
                snap['desc'] = self.txn_desc.get().strip()
            if hasattr(self, 'debit_acct'):
                snap['debit_acct'] = self.debit_acct.get().strip()
            if hasattr(self, 'credit_acct'):
                snap['credit_acct'] = self.credit_acct.get().strip()
            if hasattr(self, 'debit_amt'):
                snap['debit_amt'] = self.debit_amt.get().strip()
            if hasattr(self, 'credit_amt'):
                snap['credit_amt'] = self.credit_amt.get().strip()
            if hasattr(self, 'txn_doc_ref'):
                snap['doc_ref'] = self.txn_doc_ref.get().strip()
            if hasattr(self, 'txn_external_ref'):
                snap['ext_ref'] = self.txn_external_ref.get().strip()
            if hasattr(self, 'txn_source_type'):
                snap['source_type'] = self.txn_source_type.get().strip()
            if hasattr(self, 'txn_memo'):
                snap['memo'] = self.txn_memo.get('1.0', tk.END).strip()
            if hasattr(self, 'txn_reverse_date'):
                snap['reverse_on'] = self.txn_reverse_date.get().strip()
        except Exception:
            pass
        return snap

    def _apply_txn_form(self, snap: Dict[str, str]) -> None:
        try:
            if 'date' in snap and hasattr(self, 'txn_date'):
                self.txn_date.delete(0, tk.END); self.txn_date.insert(0, snap.get('date', ''))
            if 'desc' in snap and hasattr(self, 'txn_desc'):
                self.txn_desc.delete(0, tk.END); self.txn_desc.insert(0, snap.get('desc', ''))
            
            if 'debit_amt' in snap and hasattr(self, 'debit_amt'):
                self.debit_amt.delete(0, tk.END); self.debit_amt.insert(0, snap.get('debit_amt', ''))
            if 'credit_amt' in snap and hasattr(self, 'credit_amt'):
                self.credit_amt.delete(0, tk.END); self.credit_amt.insert(0, snap.get('credit_amt', ''))
            if 'doc_ref' in snap and hasattr(self, 'txn_doc_ref'):
                self.txn_doc_ref.delete(0, tk.END); self.txn_doc_ref.insert(0, snap.get('doc_ref', ''))
            if 'ext_ref' in snap and hasattr(self, 'txn_external_ref'):
                self.txn_external_ref.delete(0, tk.END); self.txn_external_ref.insert(0, snap.get('ext_ref', ''))
            if 'source_type' in snap and hasattr(self, 'txn_source_type'):
                self.txn_source_type.set(snap.get('source_type', ''))
            if 'memo' in snap and hasattr(self, 'txn_memo'):
                self.txn_memo.delete('1.0', tk.END); self.txn_memo.insert('1.0', snap.get('memo', ''))
            if 'reverse_on' in snap and hasattr(self, 'txn_reverse_date'):
                self.txn_reverse_date.delete(0, tk.END); self.txn_reverse_date.insert(0, snap.get('reverse_on', ''))
        except Exception:
            pass

    def _prefill_txn_from_last_entry(self) -> None:
        try:
            pid = int(self.current_period_id or 0)
            if not pid:
                return
            row = self.engine.conn.execute(
                "SELECT id, date, description FROM journal_entries WHERE period_id=? ORDER BY id DESC LIMIT 1",
                (pid,)
            ).fetchone()
            if not row:
                return
            try:
                if hasattr(self, 'txn_date'):
                    self.txn_date.delete(0, tk.END); self.txn_date.insert(0, row['date'])
                if hasattr(self, 'txn_desc'):
                    self.txn_desc.delete(0, tk.END); self.txn_desc.insert(0, row['description'] or '')
            except Exception:
                pass
            lines = self.engine.conn.execute(
                """
                SELECT jl.debit, jl.credit, a.code, a.name
                FROM journal_lines jl
                JOIN accounts a ON a.id = jl.account_id
                WHERE jl.entry_id=?
                ORDER BY jl.id
                """,
                (int(row['id']),)
            ).fetchall()
            # Do not prefill accounts/amounts from last entry to avoid unintended defaults
        except Exception:
            pass

    def _render_cycle_list(self, rows: List[sqlite3.Row]) -> None:
        return

    def _show_login(self) -> bool:
        """Show login dialog. Returns True if login successful."""
        # This method is kept for backward compatibility but now uses the standalone login dialog
        if not self.auth_module:
            return True  # No auth required
        
        # Check if main window is visible (re-login scenario)
        window_was_visible = self.winfo_viewable()
        
        # Hide main window during login
        if window_was_visible:
            self.withdraw()
        
        try:
            from . import login_dialog
            login_result = login_dialog.show_login_dialog(self.auth_module, self.palette, parent=self)
            if login_result:
                self.current_user = login_result.get("user")
                self.session_token = login_result.get("session_token")
                # Show main window again if it was visible before
                if window_was_visible:
                    self.deiconify()
                return True
            # Login cancelled - show window again if it was visible
            if window_was_visible:
                self.deiconify()
            return False
        except Exception as e:
            logger.error(f"Login dialog error: {e}", exc_info=True)
            # Show window again if it was visible
            if window_was_visible:
                self.deiconify()
            return False
    
    def _build_ui(self) -> None:
        # Main container with responsive layout
        self.main_frame = ttk.Frame(self, style="Techfix.App.TFrame")
        self.main_frame.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)
        self.main_frame.grid_rowconfigure(0, weight=1)
        self.main_frame.grid_columnconfigure(0, weight=1)

        container = ttk.Frame(self.main_frame, style="Techfix.App.TFrame")
        container.pack(fill=tk.BOTH, expand=True)

        # Sidebar (left) + content (right) layout for modern look
        # Start collapsed and expand on click for a modern "auto-hide" experience.
        # Increase collapsed width so icons have more breathing room.
        self.sidebar_collapsed_width = 90
        self.sidebar_expanded_width = 220
        # Track current sidebar state so we don't constantly re-layout on every event.
        self._sidebar_expanded: bool = False
        sidebar = ttk.Frame(
            container,
            style="Techfix.Surface.TFrame",
            width=self.sidebar_collapsed_width,
        )
        sidebar.pack(side=tk.LEFT, fill=tk.Y, padx=(12, 6), pady=12)
        sidebar.pack_propagate(False)
        self.sidebar = sidebar

        right_wrap = ttk.Frame(container, style="Techfix.App.TFrame")
        right_wrap.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(6, 12), pady=12)

        # Header goes in right content area with gradient effect
        header_container = tk.Frame(right_wrap, bg=self.palette["app_bg"])
        header_container.pack(fill=tk.X, padx=0, pady=(0, 8))
        
        # Create gradient header
        header = tk.Frame(header_container, bg=self.palette["accent_color"], height=60)
        header.pack(fill=tk.X)
        header.pack_propagate(False)  # Prevent frame from shrinking
        self.header_frame = header
        
        # Header left side: Title and subtitle (create first, before gradient)
        header_left = tk.Frame(header, bg=self.palette["accent_color"])
        header_left.pack(side=tk.LEFT, padx=18, pady=12)
        header_left.lift()  # Ensure it's on top
        
        title_label = tk.Label(
            header_left,
            text="TechFix Solutions",
            bg=self.palette["accent_color"],
            fg="#ffffff",
            font="{Segoe UI Semibold} 16"
        )
        title_label.pack(side=tk.LEFT, padx=(0, 12))
        title_label.lift()
        
        subtitle_label = tk.Label(
            header_left,
            text="Integrated accounting workspace",
            bg=self.palette["accent_color"],
            fg=self.palette.get("subtitle_fg", "#dbeafe"),
            font=FONT_BASE
        )
        subtitle_label.pack(side=tk.LEFT)
        subtitle_label.lift()
        
        # Header right side: Search, Notifications, User menu, Status
        header_right = tk.Frame(header, bg=self.palette["accent_color"])
        header_right.pack(side=tk.RIGHT, padx=18, pady=12)
        header_right.lift()  # Ensure it's on top
        
        # When packing to RIGHT, pack in reverse order (rightmost first)
        # Status indicator area (rightmost)
        status_area = tk.Frame(header_right, bg=self.palette["accent_color"])
        status_area.pack(side=tk.RIGHT, padx=(12, 0))
        self.header_status_area = status_area
        
        # User menu button (second from right)
        username = "User"
        is_admin = False
        if self.auth_module and self.current_user:
            username = self.current_user.get('username', 'User')
            # Verify admin status from database for security
            is_admin = self._is_user_admin()
        
        # Button text: show "admin" for admin users, otherwise show username
        btn_text = "👤 admin" if is_admin else f"👤 {username}"
        # Always show user menu (which includes admin panel option for admins)
        btn_command = self._show_user_menu
        
        self.user_menu_btn = tk.Button(
            header_right,
            text=btn_text,
            command=btn_command,
            bg="#ffffff",
            fg=self.palette["accent_color"],
            relief=tk.RAISED,
            font="{Segoe UI Semibold} 10",
            cursor="hand2",
            padx=12,
            pady=6,
            bd=1
        )
        self.user_menu_btn.pack(side=tk.RIGHT, padx=(0, 8))
        
        # Notifications button (third from right)
        self.notif_btn = tk.Button(
            header_right,
            text="🔔 Notifications",
            command=self._show_notifications,
            bg="#ffffff",
            fg=self.palette["accent_color"],
            relief=tk.RAISED,
            font="{Segoe UI Semibold} 10",
            cursor="hand2",
            padx=12,
            pady=6,
            bd=1
        )
        self.notif_btn.pack(side=tk.RIGHT, padx=(0, 8))
        
        # Search icon button (opens search dialog)
        search_btn = tk.Button(
            header_right,
            text=ICONS.get("search", "🔍"),
            command=self._show_search_dialog,
            bg="#ffffff",
            fg=self.palette["accent_color"],
            relief=tk.RAISED,
            font="{Segoe UI Semibold} 12",
            cursor="hand2",
            padx=12,
            pady=6,
            bd=1
        )
        search_btn.pack(side=tk.RIGHT, padx=(0, 8))
        
        # Force update to ensure widgets are rendered
        header.update_idletasks()
        
        # Store references for later use
        self.header_left_frame = header_left
        self.header_right_frame = header_right
        
        # Note: Gradient canvas temporarily disabled to ensure widgets are visible
        # The solid background color is sufficient for now

        toolbar = ttk.Frame(right_wrap, style="Techfix.App.TFrame")
        toolbar.pack(fill=tk.X, padx=12, pady=(0, 12))
        self.toolbar_frame = toolbar

        ttk.Label(toolbar, text="Period", style="Techfix.AppBar.TLabel").pack(side=tk.LEFT, padx=(0, 6))
        self.period_var = tk.StringVar()
        self.period_combo = ttk.Combobox(
            toolbar,
            textvariable=self.period_var,
            state="readonly",
            width=14,
            style="Techfix.TCombobox",
        )
        self.period_combo.pack(side=tk.LEFT, padx=(0, 12))
        self.period_combo.bind("<<ComboboxSelected>>", self._on_period_change)

        new_period_btn = ttk.Button(
            toolbar,
            text=f"{ICONS.get('add', '+')} New Period",
            command=self._prompt_new_period,
            style="Techfix.Enhanced.TButton",
        )
        new_period_btn.pack(side=tk.LEFT, padx=(0, 12))
        # Add hover animation
        try:
            self._add_hover_animation(new_period_btn)
        except Exception:
            pass

        refresh_btn = ttk.Button(
            toolbar,
            text=f"{ICONS.get('refresh', '🔄')} Refresh Cycle",
            command=self._refresh_cycle_and_views,
            style="Techfix.Enhanced.TButton",
        )
        refresh_btn.pack(side=tk.LEFT, padx=(0, 12))
        # Add hover animation
        try:
            self._add_hover_animation(refresh_btn)
        except Exception:
            pass

        ttk.Button(
            toolbar,
            text="Load 15 Entries Example",
            command=self._load_15_entries_example,
            style="Techfix.TButton",
        ).pack(side=tk.LEFT, padx=(0, 12))
        
        # Undo/Redo buttons
        undo_btn = ttk.Button(
            toolbar,
            text="↶ Undo",
            command=self._undo_action,
            style="Techfix.TButton",
        )
        undo_btn.pack(side=tk.LEFT, padx=(0, 6))
        self.undo_btn = undo_btn
        
        redo_btn = ttk.Button(
            toolbar,
            text="↷ Redo",
            command=self._redo_action,
            style="Techfix.TButton",
        )
        redo_btn.pack(side=tk.LEFT, padx=(0, 12))
        self.redo_btn = redo_btn

        # Color theme button
        theme_frame = ttk.Frame(toolbar, style="Techfix.App.TFrame")
        theme_frame.pack(side=tk.RIGHT, padx=(0, 12))
        
        self.colors_btn = ttk.Button(
            theme_frame,
            text="Colors",
            command=self._open_color_theme_picker,
            style="Techfix.Theme.TButton",
            width=10
        )
        self.colors_btn.pack(side=tk.LEFT)

        cycle_frame = ttk.Labelframe(right_wrap, text="Accounting Cycle Status", style="Techfix.TLabelframe")
        cycle_frame.pack(fill=tk.X, padx=12, pady=(0, 12))
        self.cycle_frame = cycle_frame

        cycle_top = ttk.Frame(cycle_frame, style="Techfix.Surface.TFrame")
        cycle_top.pack(fill=tk.X, padx=4, pady=(4, 0))
        ttk.Label(cycle_top, text="Track progress through the 10-step cycle.", style="Techfix.TLabel").pack(side=tk.LEFT)

        actions = ttk.Frame(cycle_frame, style="Techfix.Surface.TFrame")
        actions.pack(fill=tk.X, padx=4, pady=(4, 4))

        ttk.Button(
            actions,
            text="Mark Selected Completed",
            command=lambda: self._update_cycle_step_status("completed"),
            style="Techfix.TButton",
        ).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(
            actions,
            text="Mark Selected In Progress",
            command=lambda: self._update_cycle_step_status("in_progress"),
            style="Techfix.TButton",
        ).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(
            actions,
            text="Reset to Pending",
            command=lambda: self._update_cycle_step_status("pending"),
            style="Techfix.TButton",
        ).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(
            actions,
            text="Reset All to Pending",
            command=self._reset_all_cycle_to_pending,
            style="Techfix.TButton",
        ).pack(side=tk.LEFT, padx=(0, 8))

        cols = ("step", "name", "status", "note", "updated")
        self.cycle_tree = ttk.Treeview(
            cycle_frame,
            columns=cols,
            show="headings",
            style="Techfix.Treeview",
            selectmode="browse",
            height=6,
        )
        self.cycle_tree_scroll = ttk.Scrollbar(cycle_frame, orient=tk.VERTICAL, command=self.cycle_tree.yview)
        self.cycle_tree.configure(yscrollcommand=self.cycle_tree_scroll.set)
        for c in cols:
            width = 80 if c == "step" else 140
            if c == "updated":
                width = 160
            self.cycle_tree.heading(c, text=c.title(), anchor="w")
            self.cycle_tree.column(c, width=width if c != "note" else 260, stretch=(c == "note"))
        self.cycle_tree.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(4, 0), pady=(0, 6))
        self.cycle_tree_scroll.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 4), pady=(0, 6))

    

        notebook_wrap = ttk.Frame(right_wrap, style="Techfix.App.TFrame")
        notebook_wrap.pack(fill=tk.BOTH, expand=True, padx=12, pady=(0, 12))

        # Create a single content area and individual tab frames (we'll show/hide them)
        self.content_area = ttk.Frame(notebook_wrap, style="Techfix.App.TFrame")
        self.content_area.pack(fill=tk.BOTH, expand=True)

        self.tab_txn = ttk.Frame(self.content_area, style="Techfix.Surface.TFrame")
        self.tab_journal = ttk.Frame(self.content_area, style="Techfix.Surface.TFrame")
        self.tab_ledger = ttk.Frame(self.content_area, style="Techfix.Surface.TFrame")
        self.tab_trial = ttk.Frame(self.content_area, style="Techfix.Surface.TFrame")
        self.tab_adjust = ttk.Frame(self.content_area, style="Techfix.Surface.TFrame")
        self.tab_fs = ttk.Frame(self.content_area, style="Techfix.Surface.TFrame")
        self.tab_closing = ttk.Frame(self.content_area, style="Techfix.Surface.TFrame")
        self.tab_postclosing = ttk.Frame(self.content_area, style="Techfix.Surface.TFrame")
        self.tab_export = ttk.Frame(self.content_area, style="Techfix.Surface.TFrame")
        self.tab_audit = ttk.Frame(self.content_area, style="Techfix.Surface.TFrame")
        self.tab_help = ttk.Frame(self.content_area, style="Techfix.Surface.TFrame")

        self._build_transactions_tab()
        self._build_journal_tab()
        self._build_ledger_tab()
        self._build_trial_tab()
        self._build_adjust_tab()
        self._build_fs_tab()
        self._build_closing_tab()
        self._build_postclosing_tab()
        self._build_export_tab()
        self._build_audit_tab()
        self._build_help_tab()

        # Build simple sidebar navigation (shows/hides content frames)
        # Compact header area inside the sidebar (just the toggle button)
        profile = ttk.Frame(sidebar, style="Techfix.Surface.TFrame")
        profile.pack(fill=tk.X, pady=(4, 4), padx=6)

        # Sidebar toggle button (click to expand/collapse)
        toggle_row = ttk.Frame(profile, style="Techfix.Surface.TFrame")
        toggle_row.pack(fill=tk.X, padx=6, pady=(0, 2))
        self.sidebar_toggle_btn = ttk.Button(
            toggle_row,
            text="☰",
            style="Techfix.Nav.TButton",
            command=self._toggle_sidebar,
        )
        # Match other nav buttons: full-width, centered content
        self.sidebar_toggle_btn.pack(side=tk.LEFT, fill=tk.X, expand=True)
        # Keep a placeholder brand canvas attribute (used by theme code),
        # but do not pack it so it doesn't take extra vertical space.
        self.sidebar_brand_canvas = tk.Canvas(
            profile,
            height=0,
            highlightthickness=0,
            bd=0,
            bg=self.palette.get('surface_bg'),
        )
        try:
            self.sidebar_brand_canvas.bind('<Configure>', lambda e: self._draw_sidebar_brand())
        except Exception:
            pass
        try:
            self._draw_sidebar_brand()
        except Exception:
            pass

        # Factory to create sidebar nav rows (indicator + button)
        def make_nav(text: str, index: int, emoji: str = ""):
            if not hasattr(self, '_nav_buttons'):
                self._nav_buttons = []

            row = ttk.Frame(sidebar, style="Techfix.Surface.TFrame")
            row.pack(fill=tk.X, pady=2, padx=6)

            # Left accent indicator (thin bar)
            indicator = tk.Frame(row, width=6, bg=self.palette["surface_bg"], highlightthickness=0, bd=0, relief=tk.FLAT)
            indicator.pack(side=tk.LEFT, fill=tk.Y)

            # Button expands to fill remaining space
            # Store emoji/label on the button so we can toggle between
            # icon-only (collapsed) and icon+label (expanded) sidebar states.
            # Special handling for Dashboard (index -1)
            if index == -1:
                cmd = lambda: self._show_dashboard()
            else:
                cmd = lambda i=index: self._nav_to(i)
            
            btn = ttk.Button(
                row,
                text=emoji or text,
                style="Techfix.Nav.TButton",
                command=cmd,
            )
            # Custom attributes for hover-collapse behavior
            btn._nav_emoji = emoji
            btn._nav_label = text
            btn.pack(side=tk.LEFT, fill=tk.X, expand=True)

            # Hover effects: change indicator on enter/leave when not selected
            def on_enter(e, ind=indicator):
                try:
                    ind.configure(bg=self.palette.get('tab_active_bg'))
                except Exception:
                    pass

            def on_leave(e, ind=indicator, b=btn, idx=index):
                try:
                    # keep selected indicator if currently selected
                    if getattr(self, '_current_nav_index', None) == idx:
                        ind.configure(bg=self.palette.get('accent_color'))
                    else:
                        ind.configure(bg=self.palette.get('surface_bg'))
                except Exception:
                    pass

            btn.bind('<Enter>', on_enter)
            btn.bind('<Leave>', on_leave)

            # Store nav_index with each button for proper highlighting
            self._nav_buttons.append((indicator, btn, index))
            return (indicator, btn)

        # Add navigation buttons (icons via emoji for simplicity)
        make_nav("Dashboard", -1, "📊")  # Special index for dashboard
        make_nav("Transactions", 0, "🧾")
        make_nav("Journal", 1, "📓")
        make_nav("Ledger", 2, "📚")
        make_nav("Trial Balance", 3, "🧮")
        make_nav("Adjustments", 4, "⚙️")
        make_nav("Fin. Statements", 5, "💰")
        make_nav("Closing", 6, "🔒")
        make_nav("Post-Closing", 7, "📈")
        make_nav("Im/Export", 8, "⬇️")
        make_nav("Audit Log", 9, "🧪")
        make_nav("How to Use?", 10, "❓")

        # Keep an ordered list of the content frames to show/hide
        self._tab_frames = [
            self.tab_txn,
            self.tab_journal,
            self.tab_ledger,
            self.tab_trial,
            self.tab_adjust,
            self.tab_fs,
            self.tab_closing,
            self.tab_postclosing,
            self.tab_export,
            self.tab_audit,
            self.tab_help,
        ]

        exit_wrap = ttk.Frame(sidebar, style="Techfix.Surface.TFrame")
        exit_wrap.pack(side=tk.BOTTOM, fill=tk.X, padx=6, pady=(12, 6))

        # A global action button that stays visible on all window sizes (PC, laptop, etc.)
        # because it is anchored to the bottom of the sidebar and stretches horizontally.
        self.global_action_btn = ttk.Button(
            exit_wrap,
            text="Global Action",
            command=self._on_global_action,
            style="Techfix.TButton",
        )
        self.global_action_btn.pack(fill=tk.X, pady=(0, 4))

        ttk.Button(
            exit_wrap,
            text="Exit",
            command=self._on_close,
            style="Techfix.Danger.TButton",
        ).pack(fill=tk.X)

        # Global keyboard shortcuts for navigation between major tabs
        try:
            # Ctrl+1 shows dashboard
            self.bind("<Control-Key-1>", lambda e: self._show_dashboard())
            # Ctrl+2..Ctrl+0 jump between sidebar tabs (Transactions through Help)
            # Map Ctrl+2 to index 0 (Transactions), Ctrl+3 to index 1, etc.
            for idx in range(min(len(self._tab_frames), 9)):  # Map up to 9 tabs (indices 0-8)
                digit = (idx + 2) % 10  # Maps: 0->2, 1->3, ..., 7->9, 8->0
                seq = f"<Control-Key-{digit}>"
                self.bind(seq, lambda e, i=idx: self._nav_to(i))
        except Exception:
            pass
        
        # Additional keyboard shortcuts
        self._setup_keyboard_shortcuts()

    # --- Sidebar toggle behavior (click-to-expand/collapse) ---------------------

    def _toggle_sidebar(self) -> None:
        """
        Toggle sidebar between collapsed and expanded states on button click.
        """
        try:
            if getattr(self, "_sidebar_expanded", False):
                self._collapse_sidebar()
            else:
                self._expand_sidebar()
        except Exception:
            pass

    def _expand_sidebar(self) -> None:
        """
        Expand the left sidebar (used by the click-to-toggle button).
        Also restore full text labels for nav buttons.
        """
        try:
            if not hasattr(self, "sidebar"):
                return

            # Avoid redundant layout work if we're already expanded.
            if getattr(self, "_sidebar_expanded", False):
                return

            target_width = getattr(self, "sidebar_expanded_width", 220)
            self.sidebar.configure(width=target_width)
            self.sidebar.update_idletasks()

            # Restore full labels on navigation buttons when expanded
            for ind, btn, nav_idx in getattr(self, "_nav_buttons", []):
                try:
                    emoji = getattr(btn, "_nav_emoji", "")
                    label = getattr(btn, "_nav_label", "")
                    if emoji and label:
                        btn.configure(text=f"{emoji}  {label}")
                    elif label:
                        btn.configure(text=label)
                except Exception:
                    continue

            # Mark state so we don't expand repeatedly.
            self._sidebar_expanded = True
        except Exception:
            # Avoid breaking the app if anything goes wrong
            pass

    def _collapse_sidebar(self) -> None:
        """
        Collapse the left sidebar (used by the click-to-toggle button).
        Show icon-only (emoji) nav buttons to save space.
        """
        try:
            if not hasattr(self, "sidebar"):
                return

            # If we're already collapsed, do nothing to avoid extra layout work.
            if not getattr(self, "_sidebar_expanded", False):
                return

            target_width = getattr(self, "sidebar_collapsed_width", 60)
            self.sidebar.configure(width=target_width)
            self.sidebar.update_idletasks()

            # Show only icons on navigation buttons when collapsed
            for ind, btn, nav_idx in getattr(self, "_nav_buttons", []):
                try:
                    emoji = getattr(btn, "_nav_emoji", "")
                    label = getattr(btn, "_nav_label", "")
                    if emoji:
                        btn.configure(text=emoji)
                    elif label:
                        # If no emoji, keep a very short label
                        btn.configure(text=label[:3] + "…" if len(label) > 4 else label)
                except Exception:
                    continue

            # Mark state so we don't collapse repeatedly.
            self._sidebar_expanded = False
        except Exception:
            # Avoid breaking the app if anything goes wrong
            pass

    def _build_menubar(self) -> None:
        self.menubar = tk.Menu(self)
        
        # File menu
        self.file_menu = tk.Menu(self.menubar, tearoff=0)
        self.file_menu.add_command(label="New Transaction", command=lambda: self._nav_to(0), accelerator="Ctrl+N")
        self.file_menu.add_separator()
        self.file_menu.add_command(label="Import Data...", command=self._show_import_dialog, accelerator="Ctrl+I")
        self.file_menu.add_separator()
        self.file_menu.add_command(label="Backup Database...", command=self._show_backup_dialog)
        self.file_menu.add_command(label="Restore Database...", command=self._show_restore_dialog)
        self.file_menu.add_separator()
        self.file_menu.add_command(label="Exit", command=self._on_close, accelerator="Ctrl+Q")
        self.menubar.add_cascade(label="File", menu=self.file_menu)

        # Edit menu
        self.edit_menu = tk.Menu(self.menubar, tearoff=0)
        self.edit_menu.add_command(label="Undo", command=self._undo_action, accelerator="Ctrl+Z", state=tk.DISABLED)
        self.edit_menu.add_command(label="Redo", command=self._redo_action, accelerator="Ctrl+Y", state=tk.DISABLED)
        self.edit_menu.add_separator()
        self.edit_menu.add_command(label="Search...", command=self._show_search_dialog, accelerator="Ctrl+F")
        self.menubar.add_cascade(label="Edit", menu=self.edit_menu)

        # View menu
        self.view_menu = tk.Menu(self.menubar, tearoff=0)
        self.view_menu.add_command(label="Light Theme", command=lambda: self._apply_theme("Light"))
        self.view_menu.add_command(label="Dark Theme", command=lambda: self._apply_theme("Dark"))
        self.view_menu.add_separator()
        self.view_menu.add_command(label="Customize Colors...", command=self._open_color_picker)
        self.view_menu.add_separator()
        self.view_menu.add_command(label="Dashboard", command=self._show_dashboard, accelerator="Ctrl+D")
        self.view_menu.add_separator()
        self.view_menu.add_command(label="Toggle Fullscreen", command=lambda: self.attributes('-fullscreen', not self.attributes('-fullscreen')), accelerator="F11")
        self.menubar.add_cascade(label="View", menu=self.view_menu)

        # Tools menu
        self.tools_menu = tk.Menu(self.menubar, tearoff=0)
        self.tools_menu.add_command(label="Notifications", command=self._show_notifications)
        if self.auth_module:
            self.tools_menu.add_separator()
            self.tools_menu.add_command(label="Change Password...", command=self._show_change_password)
            self.tools_menu.add_command(label="Logout", command=self._logout)
        self.menubar.add_cascade(label="Tools", menu=self.tools_menu)

        # Help menu
        self.help_menu = tk.Menu(self.menubar, tearoff=0)
        self.help_menu.add_command(label="Keyboard Shortcuts", command=self._show_shortcuts_help)
        self.help_menu.add_command(label="About", command=lambda: messagebox.showinfo("About TechFix", "TechFix Accounting App\nVersion 1.0"))
        self.menubar.add_cascade(label="Help", menu=self.help_menu)

        self.config(menu=self.menubar)
        self._style_menus()
        
        # Update undo/redo menu states
        self._update_undo_redo_states()

    def _update_cycle_step_status(self, status: str) -> None:
        try:
            if not hasattr(self, 'cycle_tree'):
                return
            sel = self.cycle_tree.selection()
            if not sel:
                return
            item = sel[0]
            vals = self.cycle_tree.item(item, 'values')
            step = int(vals[0])
            note = None
            try:
                note = simpledialog.askstring("Update Step", "Note", parent=self) or None
            except Exception:
                note = None
            self.engine.set_cycle_step_status(step, status, note)
            self._load_cycle_status()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update cycle step: {e}")

    def _nav_to(self, index: int) -> None:
        """Navigate to notebook tab by index if possible."""
        try:
            # Hide all frames
            if not hasattr(self, '_tab_frames'):
                return
            for f in getattr(self, '_tab_frames'):
                try:
                    f.pack_forget()
                except Exception:
                    pass

            # Bound index
            if index < 0:
                index = 0
            if index >= len(self._tab_frames):
                index = 0

            sel = self._tab_frames[index]
            try:
                cur = getattr(self, '_current_nav_index', None)
                direction = 'right'
                try:
                    if isinstance(cur, int) and index < cur:
                        direction = 'left'
                except Exception:
                    pass
                # Enhanced: Use fade + swipe animation
                self._animate_swipe_to(sel, direction=direction)
                self._animate_tab_fade(sel, fade_in=True, duration_ms=250)
            except Exception:
                try:
                    sel.pack(fill=tk.BOTH, expand=True)
                except Exception:
                    pass

            # Update nav button visuals (indicator and selected style)
            try:
                self._current_nav_index = index
                for ind, btn, nav_idx in getattr(self, '_nav_buttons', ()):
                    if nav_idx == index:
                        try:
                            ind.configure(bg=self.palette.get('accent_color'))
                        except Exception:
                            pass
                        try:
                            btn.configure(style='Techfix.Nav.Selected.TButton')
                        except Exception:
                            pass
                    else:
                        try:
                            ind.configure(bg=self.palette.get('surface_bg'))
                        except Exception:
                            pass
                        try:
                            btn.configure(style='Techfix.Nav.TButton')
                        except Exception:
                            pass
            except Exception:
                pass
        except Exception:
            pass

    def _post_to_ledger_action(self) -> None:
        try:
            self.engine.set_cycle_step_status(3, "completed", "Ledger posted")
            self._load_cycle_status()
            messagebox.showinfo("Posted", "Ledger posting marked completed")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to post to ledger: {e}")

    def _reset_all_cycle_to_pending(self) -> None:
        """Reset all 10 accounting cycle steps to 'pending' for the active period."""
        try:
            if not self.engine.current_period_id:
                messagebox.showwarning("No Period", "No active accounting period selected.")
                return
            if not messagebox.askyesno("Confirm Reset", "Reset all accounting cycle steps to 'pending' for the current period?"):
                return
            pid = self.engine.current_period_id
            conn = self.engine.conn
            # Use constant list length from db module when available
            try:
                total_steps = len(db.ACCOUNTING_CYCLE_STEPS)
            except Exception:
                total_steps = 10
            for s in range(1, total_steps + 1):
                db.set_cycle_step_status(pid, s, 'pending', note='Reset to pending (user)', conn=conn)
            self._load_cycle_status()
            messagebox.showinfo("Reset Complete", "All cycle steps have been reset to pending.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to reset cycle steps: {e}")

    def _browse_source_document(self):
        try:
            filetypes = [
                ('All supported files', '*.pdf;*.jpg;*.jpeg;*.png;*.doc;*.docx;*.xls;*.xlsx;*.json;*.csv;*.txt'),
                ('PDF files', '*.pdf'),
                ('Image files', '*.jpg;*.jpeg;*.png'),
                ('Word documents', '*.doc;*.docx'),
                ('Excel files', '*.xls;*.xlsx'),
                ('Data files', '*.json;*.csv;*.txt'),
                ('All files', '*.*')
            ]
            initial_dir = getattr(self, 'doc_library_dir', None)
            filename = filedialog.askopenfilename(
                title="Select Source Document",
                filetypes=filetypes,
                defaultextension=".pdf",
                initialdir=initial_dir if initial_dir else None
            )
            if filename:
                # Update the StringVar which is bound to the entry field
                if hasattr(self, 'txn_attachment_path'):
                    self.txn_attachment_path.set(filename)
                # Also update the entry field directly if needed
                if hasattr(self, 'txn_attachment_display'):
                    try:
                        self.txn_attachment_display.configure(state='normal')
                        self.txn_attachment_display.delete(0, tk.END)
                        self.txn_attachment_display.insert(0, filename)
                        self.txn_attachment_display.configure(state='readonly')
                    except Exception:
                        pass
                try:
                    if hasattr(self, 'txn_prefill_status'):
                        self.txn_prefill_status.configure(
                            text="Document attached - Remember to click 'Record & Post' (not 'Save Draft') to include in balance sheet"
                        )
                except Exception:
                    pass
            try:
                import os
                self._audit('document_selected', {'file': filename, 'exists': os.path.exists(filename), 'readable': os.access(filename, os.R_OK)})
                if 'SampleSourceDocs' in filename:
                    sj = os.path.splitext(filename)[0] + '.json'
                    self._audit('sample_docs_sidecar', {'file': filename, 'sidecar': sj, 'sidecar_exists': os.path.exists(sj), 'sidecar_readable': os.access(sj, os.R_OK)})
                try:
                    self._append_recent_document(filename)
                    if hasattr(self, 'doc_recent_cb'):
                        self.doc_recent_cb.configure(values=getattr(self, '_recent_docs', []))
                except Exception:
                    pass
            except Exception:
                pass
            try:
                # Keep existing assignments unless new document provides accounts explicitly
                self._accounts_prefilled = False
            except Exception:
                pass
            try:
                self._prefill_date_from_source_document(filename)
            except Exception as e:
                self._audit('document_prefill_error', {'file': filename, 'stage': 'date', 'error': str(e)})
            try:
                self._prefill_amounts_from_source_document(filename)
            except Exception as e:
                self._audit('document_prefill_error', {'file': filename, 'stage': 'amounts', 'error': str(e)})
            try:
                self._prefill_from_source_document(filename)
            except Exception as e:
                self._audit('document_prefill_error', {'file': filename, 'stage': 'structured', 'error': str(e)})
            try:
                ok_accounts = self._validate_accounts_assigned()
                ok_amounts = self._validate_amounts_present()
                self._audit('document_prefill_summary', {'file': filename, 'accounts_ok': ok_accounts, 'amounts_ok': ok_amounts, 'debit': (self.debit_amt.get().strip() if hasattr(self, 'debit_amt') else None), 'credit': (self.credit_amt.get().strip() if hasattr(self, 'credit_amt') else None), 'debit_acct': (self.debit_acct.get().strip() if hasattr(self, 'debit_acct') else None), 'credit_acct': (self.credit_acct.get().strip() if hasattr(self, 'credit_acct') else None)})
            except Exception:
                pass
            try:
                self._load_document_preview(filename)
            except Exception:
                self._audit('document_preview_error', {'file': filename})
            # Update button states after prefilling
            try:
                self._update_post_buttons_enabled()
            except Exception:
                pass
        except Exception as e:
            # Show error if browse fails
            try:
                messagebox.showerror("Error", f"Failed to browse for document: {e}")
            except Exception:
                pass
            
    def _scan_source_document(self) -> None:
        try:
            self.txn_prefill_status.configure(text="Opening camera…")
        except Exception:
            pass
        try:
            import threading
            import time
            try:
                import cv2
            except Exception:
                ok = False
                try:
                    ok = self._attempt_install_packages(["opencv-python"])
                except Exception:
                    ok = False
                if ok:
                    try:
                        import cv2
                    except Exception:
                        cv2 = None  # type: ignore
                else:
                    cv2 = None  # type: ignore
                if cv2 is None:
                    messagebox.showerror("Scan", "Camera library not available. Install opencv-python")
                    try:
                        self.txn_prefill_status.configure(text="Scan error: camera library missing")
                    except Exception:
                        pass
                    return
            # Try to import pyzbar, but don't fail if it's not available
            # OpenCV QR detector can work as fallback
            zbar_decode = None  # type: ignore
            try:
                from pyzbar.pyzbar import decode as zbar_decode
            except (ImportError, FileNotFoundError, OSError) as e:
                # pyzbar may fail due to missing DLLs on Windows (libiconv.dll, libzbar-64.dll)
                # This is OK - OpenCV QR detector will be used as fallback
                try:
                    logger.debug(f"pyzbar not available (will use OpenCV fallback): {e}")
                except Exception:
                    pass
                zbar_decode = None  # type: ignore
            except Exception:
                # Other errors - try to install, but don't block if it fails
                ok = False
                try:
                    ok = self._attempt_install_packages(["pyzbar"])
                except Exception:
                    ok = False
                if ok:
                    try:
                        from pyzbar.pyzbar import decode as zbar_decode
                    except Exception:
                        zbar_decode = None  # type: ignore
                else:
                    zbar_decode = None  # type: ignore
            try:
                from PIL import Image, ImageTk
            except Exception:
                ok = False
                try:
                    ok = self._attempt_install_packages(["Pillow"])
                except Exception:
                    ok = False
                if ok:
                    try:
                        from PIL import Image, ImageTk
                    except Exception:
                        messagebox.showerror("Scan", "Image toolkit not available. Install Pillow")
                        try:
                            self.txn_prefill_status.configure(text="Scan error: Pillow missing")
                        except Exception:
                            pass
                        return
                else:
                    messagebox.showerror("Scan", "Image toolkit not available. Install Pillow")
                    try:
                        self.txn_prefill_status.configure(text="Scan error: Pillow missing")
                    except Exception:
                        pass
                    return

            class ScanWindow(tk.Toplevel):
                def __init__(self, parent):
                    super().__init__(parent)
                    self.title("Scan Source Document")
                    self.geometry("800x600")
                    self.resizable(False, False)
                    self.protocol('WM_DELETE_WINDOW', self._on_close)
                    self._running = True
                    
                    # Apply theme colors
                    try:
                        palette = parent.palette
                        self.configure(bg=palette.get("app_bg", "#f5f7fb"))
                    except Exception:
                        palette = THEMES.get("Light", {})
                        self.configure(bg=palette.get("app_bg", "#f5f7fb"))
                    
                    # Main container
                    main_container = ttk.Frame(self, style="Techfix.Surface.TFrame")
                    main_container.pack(fill=tk.BOTH, expand=True, padx=12, pady=12)
                    
                    # Header section with title and status
                    header_frame = ttk.Frame(main_container, style="Techfix.Surface.TFrame")
                    header_frame.pack(fill=tk.X, pady=(0, 12))
                    
                    # Title
                    title_label = tk.Label(
                        header_frame,
                        text="Scan QR Code or Barcode",
                        bg=palette.get("surface_bg", "#ffffff"),
                        fg=palette.get("text_primary", "#1f2937"),
                        font=("{Segoe UI Semibold} 16")
                    )
                    title_label.pack(anchor=tk.W, pady=(0, 4))
                    
                    # Status label with better styling
                    status_frame = ttk.Frame(header_frame, style="Techfix.Surface.TFrame")
                    status_frame.pack(fill=tk.X, pady=(4, 0))
                    
                    self.status = tk.Label(
                        status_frame,
                        text="Point camera at QR code or barcode…",
                        bg=palette.get("surface_bg", "#ffffff"),
                        fg=palette.get("text_secondary", "#4b5563"),
                        font=FONT_BASE
                    )
                    self.status.pack(side=tk.LEFT)
                    
                    # Close button
                    close_btn = ttk.Button(
                        status_frame,
                        text="Close",
                        command=self._on_close,
                        style="Techfix.Theme.TButton",
                        width=10
                    )
                    close_btn.pack(side=tk.RIGHT)
                    
                    # Preview frame with border
                    preview_frame = ttk.Frame(main_container, style="Techfix.Surface.TFrame")
                    preview_frame.pack(fill=tk.BOTH, expand=True)
                    
                    # Create a frame for the preview with border styling
                    preview_border = tk.Frame(
                        preview_frame,
                        bg=palette.get("entry_border", "#d8dee9"),
                        highlightthickness=0
                    )
                    preview_border.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)
                    
                    # Preview label with centered content
                    self.preview = tk.Label(
                        preview_border,
                        text="",
                        bg=palette.get("surface_bg", "#ffffff"),
                        fg=palette.get("text_primary", "#1f2937"),
                        anchor=tk.CENTER
                    )
                    self.preview.pack(fill=tk.BOTH, expand=True)
                    self.preview_border = preview_border
                    self.palette = palette  # Store palette for later use
                    
                    # Instructions at bottom
                    instructions_frame = ttk.Frame(main_container, style="Techfix.Surface.TFrame")
                    instructions_frame.pack(fill=tk.X, pady=(12, 0))
                    
                    instructions_text = (
                        "• Position the QR code or barcode within the frame\n"
                        "• Ensure good lighting and hold steady\n"
                        "• The window will close automatically when a code is detected"
                    )
                    instructions_label = tk.Label(
                        instructions_frame,
                        text=instructions_text,
                        bg=palette.get("surface_bg", "#ffffff"),
                        fg=palette.get("text_secondary", "#6b7280"),
                        font=("{Segoe UI} 9"),
                        justify=tk.LEFT,
                        anchor=tk.W
                    )
                    instructions_label.pack(anchor=tk.W, fill=tk.X)
                    
                    # Center window on screen
                    self.update_idletasks()
                    width = self.winfo_width()
                    height = self.winfo_height()
                    x = (self.winfo_screenwidth() // 2) - (width // 2)
                    y = (self.winfo_screenheight() // 2) - (height // 2)
                    self.geometry(f'{width}x{height}+{x}+{y}')
                    
                    self.cap = cv2.VideoCapture(0)
                    if not self.cap or not self.cap.isOpened():
                        self.status.configure(
                            text="✗ Camera access failed",
                            fg="#dc2626",
                            bg=palette.get("surface_bg", "#ffffff")
                        )
                        messagebox.showerror("Scan", "Cannot access camera. Check permissions/device.")
                        self._running = False
                        return
                    self._frame_loop()
                def _frame_loop(self):
                    if not self._running:
                        return
                    ret, frame = self.cap.read()
                    if ret:
                        try:
                            import numpy as np  # optional if present
                        except Exception:
                            np = None
                        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                        payload = None
                        codes = []
                        if zbar_decode:
                            try:
                                codes = zbar_decode(gray)
                            except Exception as e:
                                # pyzbar may fail due to missing DLLs on Windows
                                # This is OK - we'll use OpenCV as fallback
                                codes = []
                                try:
                                    logger.debug(f"pyzbar decode failed (will use OpenCV fallback): {e}")
                                except Exception:
                                    pass
                            if codes:
                                try:
                                    payload = codes[0].data.decode('utf-8', errors='replace')
                                except Exception:
                                    payload = None
                        if payload is None:
                            try:
                                # Try QR code detection
                                qr_detector = cv2.QRCodeDetector()
                                data, points, _ = qr_detector.detectAndDecode(gray)
                                if points is not None and data:
                                    payload = data
                                else:
                                    # Try barcode detection
                                    try:
                                        barcode_detector = cv2.barcode_BarcodeDetector()
                                        retval, decoded_info, decoded_type = barcode_detector.detectAndDecode(gray)
                                        if retval and decoded_info and len(decoded_info) > 0:
                                            barcode_data = decoded_info[0] if isinstance(decoded_info, list) else str(decoded_info)
                                            # For camera scan, barcode data is used directly (no file lookup)
                                            payload = barcode_data
                                    except AttributeError:
                                        # barcode_BarcodeDetector not available
                                        pass
                                    except Exception:
                                        pass
                            except Exception:
                                pass
                        if payload:
                            try:
                                palette = self.master.palette
                            except Exception:
                                palette = THEMES.get("Light", {})
                            self.status.configure(
                                text="✓ Code detected – processing…",
                                fg="#059669",
                                bg=palette.get("surface_bg", "#ffffff")
                            )
                            self._running = False
                            # Validate payload before applying
                            try:
                                data = self.master._parse_scanned_payload(payload)
                                if data:
                                    # Valid data - apply it and close window
                                    self._apply_payload(payload)
                                    self._cleanup()
                                    return
                                else:
                                    # Invalid data - show error but still close window
                                    try:
                                        self.master.txn_prefill_status.configure(text="Scan error: invalid data format")
                                    except Exception:
                                        pass
                                    try:
                                        palette = self.master.palette
                                    except Exception:
                                        palette = THEMES.get("Light", {})
                                    self.status.configure(
                                        text="✗ Invalid code format",
                                        fg="#dc2626",
                                        bg=palette.get("surface_bg", "#ffffff")
                                    )
                                    self.update()
                                    messagebox.showerror("Scan", "Invalid code format. Expected JSON or key=value pairs.")
                                    self._cleanup()
                                    return
                            except Exception:
                                # Error parsing - close window anyway
                                self._cleanup()
                                return
                        rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                        img = Image.fromarray(rgb)
                        
                        # Calculate size to fit in preview area
                        try:
                            preview_width = self.preview.winfo_width() if self.preview.winfo_width() > 1 else 760
                            preview_height = self.preview.winfo_height() if self.preview.winfo_height() > 1 else 420
                        except Exception:
                            preview_width = 760
                            preview_height = 420
                        
                        # Maintain aspect ratio and fit to preview
                        img_width, img_height = img.size
                        scale_w = preview_width / img_width
                        scale_h = preview_height / img_height
                        scale = min(scale_w, scale_h)
                        
                        new_width = int(img_width * scale)
                        new_height = int(img_height * scale)
                        img = img.resize((new_width, new_height), Image.Resampling.LANCZOS)
                        
                        imgtk = ImageTk.PhotoImage(image=img)
                        self.preview.configure(image=imgtk)
                        self.preview.imgtk = imgtk
                    try:
                        self.after(30, self._frame_loop)
                    except Exception:
                        pass
                def _apply_payload(self, text: str):
                    try:
                        data = self.master._parse_scanned_payload(text)
                    except Exception:
                        data = None
                    if not data:
                        # This should not happen as we validate before calling this
                        try:
                            self.master.txn_prefill_status.configure(text="Scan error: invalid data format")
                        except Exception:
                            pass
                        messagebox.showerror("Scan", "Invalid code format. Expected JSON or key=value pairs.")
                        return
                    try:
                        # For camera scan, we don't have a filename, so pass empty string
                        self.master._apply_scanned_data(data, '')
                        self.master.txn_prefill_status.configure(text="Scan successful – fields populated")
                    except Exception as e:
                        try:
                            logger.debug(f"Error applying scanned data: {e}", exc_info=True)
                        except Exception:
                            pass
                        messagebox.showerror("Scan", f"Failed to apply scanned data: {str(e)}")
                def _cleanup(self):
                    try:
                        if self.cap:
                            self.cap.release()
                    except Exception:
                        pass
                    try:
                        self.destroy()
                    except Exception:
                        pass
                def _on_close(self):
                    self._running = False
                    self._cleanup()

            ScanWindow(self)
        except Exception:
            try:
                self.txn_prefill_status.configure(text="Scan failed")
            except Exception:
                pass
    def _attempt_install_packages(self, pkgs: list[str]) -> bool:
        try:
            exe = sys.executable
            for p in pkgs:
                try:
                    subprocess.check_call([exe, '-m', 'pip', 'install', p])
                except Exception:
                    return False
            return True
        except Exception:
            return False

    def _parse_scanned_payload(self, text: str) -> dict | None:
        try:
            s = text.strip()
            if not s:
                return None
            if s.startswith('{') and s.endswith('}'):
                try:
                    d = json.loads(s)
                except Exception:
                    d = None
            else:
                d = {}
                parts = [p for p in s.replace('|', '&').split('&') if p]
                for p in parts:
                    if '=' in p:
                        k, v = p.split('=', 1)
                        d[k.strip()] = v.strip()
            if not isinstance(d, dict):
                return None
            # Normalize keys
            m = {}
            def get(*keys):
                for k in keys:
                    if k in d:
                        return d[k]
                return None
            m['date'] = get('date')
            m['source_type'] = get('source_type','source')
            m['document_ref'] = get('document_ref','doc_no','doc','reference')
            m['external_ref'] = get('external_ref','ext_ref')
            m['description'] = get('description','desc')
            m['debit_amount'] = get('debit_amount','debit','amount')
            m['credit_amount'] = get('credit_amount','credit','amount')
            m['debit_account'] = get('debit_account','debit_acct','debit_account_name')
            m['credit_account'] = get('credit_account','credit_acct','credit_account_name')
            m['memo'] = get('memo','note')
            # Remove empty
            m = {k:v for k,v in m.items() if v not in (None,'')}
            return m or None
        except Exception:
            return None

    def _apply_scanned_data(self, data: dict, filename: str = '') -> None:
        try:
            # Use existing prefill and auto-entry helpers
            def _set_entry(w, v):
                try:
                    if w and v is not None:
                        w.delete(0, tk.END); w.insert(0, str(v))
                        return True
                except Exception:
                    pass
                return False
            if hasattr(self, 'txn_date') and data.get('date'):
                _set_entry(self.txn_date, data.get('date'))
            if hasattr(self, 'txn_desc') and data.get('description'):
                _set_entry(self.txn_desc, data.get('description'))
            if hasattr(self, 'txn_doc_ref') and data.get('document_ref'):
                _set_entry(self.txn_doc_ref, data.get('document_ref'))
            if hasattr(self, 'txn_external_ref') and data.get('external_ref'):
                _set_entry(self.txn_external_ref, data.get('external_ref'))
            if hasattr(self, 'txn_source_type') and data.get('source_type'):
                try:
                    self.txn_source_type.set(data.get('source_type'))
                except Exception:
                    pass
            # Remind user to post the transaction (not save as draft)
            try:
                if hasattr(self, 'txn_prefill_status'):
                    current_text = self.txn_prefill_status.cget('text')
                    if 'successful' in current_text.lower() or 'scanned' in current_text.lower():
                        # Add reminder about posting
                        self.txn_prefill_status.configure(
                            text=f"{current_text} - Remember to click 'Record & Post' (not 'Save Draft') to include in balance sheet"
                        )
            except Exception:
                pass
            # Automatically check adjusting entry checkbox if source_type is "Adjust" or description contains "Adjusting entry"
            if hasattr(self, 'txn_is_adjust'):
                try:
                    source_type = data.get('source_type', '').strip()
                    description = data.get('description', '').strip()
                    is_adjusting = (
                        source_type.lower() == 'adjust' or
                        'adjusting entry' in description.lower()
                    )
                    if is_adjusting:
                        self.txn_is_adjust.set(1)
                except Exception:
                    pass
            if hasattr(self, 'txn_memo') and data.get('memo'):
                try:
                    self.txn_memo.delete('1.0', tk.END); self.txn_memo.insert('1.0', data.get('memo'))
                except Exception:
                    pass
            # Try to get accounts directly from data first (like JSON files)
            dd = None
            cc = None
            if data.get('debit_account'):
                try:
                    dd = self._match_account_display(data.get('debit_account'))
                    # If exact match failed, try with stripped/cleaned version
                    if not dd:
                        cleaned = str(data.get('debit_account')).strip()
                        if cleaned:
                            dd = self._match_account_display(cleaned)
                except Exception as e:
                    try:
                        logger.debug(f"Error matching debit account: {e}", exc_info=True)
                    except Exception:
                        pass
            if data.get('credit_account'):
                try:
                    cc = self._match_account_display(data.get('credit_account'))
                    # If exact match failed, try with stripped/cleaned version
                    if not cc:
                        cleaned = str(data.get('credit_account')).strip()
                        if cleaned:
                            cc = self._match_account_display(cleaned)
                except Exception as e:
                    try:
                        logger.debug(f"Error matching credit account: {e}", exc_info=True)
                    except Exception:
                        pass
            
            # Always try auto-entry suggestions to fill in missing accounts
            # Use filename if provided, otherwise try to construct a meaningful filename from data
            auto_entry_filename = filename
            if not auto_entry_filename:
                # Try to construct filename from source_type and document_ref for better inference
                source_type = data.get('source_type', '')
                doc_ref = data.get('document_ref', '')
                if source_type or doc_ref:
                    parts = []
                    if source_type:
                        parts.append(source_type)
                    if doc_ref:
                        parts.append(doc_ref)
                    auto_entry_filename = '_'.join(parts) if parts else ''
            
            sugg = {}
            try:
                sugg = self._auto_entry_from_data(data, auto_entry_filename)
            except Exception as e:
                try:
                    logger.debug(f"Error in auto-entry from data: {e}", exc_info=True)
                except Exception:
                    pass
                sugg = {}
            
            # Use auto-entry suggestions to fill in missing accounts (even if one account was already matched)
            # Note: sugg accounts are already matched by _default_accounts_for_source and _infer_accounts_from_context
            if not dd:
                dd = sugg.get('debit_account_display')
            if not cc:
                cc = sugg.get('credit_account_display')
            
            # Set accounts if we have them (even if only one is found)
            if dd or cc:
                try:
                    self._set_accounts(dd, cc)
                    # Force UI update after setting accounts
                    self.update_idletasks()
                    # Also trigger account changed events to ensure UI is updated
                    if dd:
                        try:
                            self._on_account_changed('debit')
                        except Exception:
                            pass
                    if cc:
                        try:
                            self._on_account_changed('credit')
                        except Exception:
                            pass
                except Exception as e:
                    try:
                        logger.debug(f"Error setting accounts: {e}", exc_info=True)
                    except Exception:
                        pass
            elif data.get('debit_account') or data.get('credit_account'):
                # If we have account data but matching failed, log it for debugging
                try:
                    logger.debug(f"Account matching failed - debit: {data.get('debit_account')}, credit: {data.get('credit_account')}, sugg: {sugg}")
                except Exception:
                    pass
            if hasattr(self, 'debit_amt') and (data.get('debit_amount') is not None or (sugg and sugg.get('debit_amount') is not None)):
                v = data.get('debit_amount', sugg.get('debit_amount'))
                try:
                    self.debit_amt.delete(0, tk.END)
                    self.debit_amt.insert(0, f"{float(v):.2f}")
                    # Trigger update after setting amount
                    self._update_post_buttons_enabled()
                except Exception:
                    pass
            if hasattr(self, 'credit_amt') and (data.get('credit_amount') is not None or (sugg and sugg.get('credit_amount') is not None)):
                v = data.get('credit_amount', sugg.get('credit_amount'))
                try:
                    self.credit_amt.delete(0, tk.END)
                    self.credit_amt.insert(0, f"{float(v):.2f}")
                    # Trigger update after setting amount
                    self._update_post_buttons_enabled()
                except Exception:
                    pass
            try:
                ok_accounts = self._validate_accounts_assigned()
                ok_amounts = self._validate_amounts_present()
                msg = "Prefilled from scan" + (" (ok)" if (ok_accounts and ok_amounts) else " (missing)")
                self.txn_prefill_status.configure(text=msg)
            except Exception:
                pass
            # Update button states after applying ALL data - do this last to ensure everything is set
            try:
                # Force update after a short delay to ensure UI has updated
                self.after(10, self._update_post_buttons_enabled)
                # Also update immediately
                self._update_post_buttons_enabled()
            except Exception:
                pass
        except Exception:
            try:
                self.txn_prefill_status.configure(text="Scan apply failed")
            except Exception:
                pass

    def _scan_from_image_file(self) -> None:
        try:
            from tkinter import filedialog
            import os
            # Update status
            try:
                if hasattr(self, 'txn_prefill_status'):
                    self.txn_prefill_status.configure(text="Selecting image to scan…")
            except Exception:
                pass
            path = filedialog.askopenfilename(title="Select image to scan", filetypes=[('Images','*.png;*.jpg;*.jpeg;*.bmp;*.gif')])
            if not path:
                return
            # Update status
            try:
                if hasattr(self, 'txn_prefill_status'):
                    self.txn_prefill_status.configure(text="Scanning image…")
            except Exception:
                pass
            # Try to import pyzbar, but don't fail if it's not available
            # OpenCV QR detector can work as fallback
            zbar_decode = None  # type: ignore
            try:
                from pyzbar.pyzbar import decode as zbar_decode
            except (ImportError, FileNotFoundError, OSError) as e:
                # pyzbar may fail due to missing DLLs on Windows
                # This is OK - OpenCV QR detector will be used as fallback
                try:
                    logger.debug(f"pyzbar not available (will use OpenCV fallback): {e}")
                except Exception:
                    pass
                zbar_decode = None  # type: ignore
            except Exception:
                zbar_decode = None  # type: ignore
            try:
                from PIL import Image
            except Exception:
                messagebox.showerror("Scan", "Image toolkit not available")
                return
            payload = None
            if zbar_decode:
                try:
                    img = Image.open(path)
                    res = zbar_decode(img)
                    if res:
                        try:
                            payload = res[0].data.decode('utf-8', errors='replace')
                        except Exception:
                            payload = None
                except Exception as e:
                    # pyzbar may fail due to missing DLLs on Windows
                    # This is OK - we'll use OpenCV as fallback
                    payload = None
                    try:
                        logger.debug(f"pyzbar decode failed (will use OpenCV fallback): {e}")
                    except Exception:
                        pass
            if payload is None:
                try:
                    import cv2
                    im = cv2.imread(path)
                    if im is not None and im.size > 0:
                        # Try QR code detection first
                        qr_detector = cv2.QRCodeDetector()
                        data, points, _ = qr_detector.detectAndDecode(im)
                        if points is not None and data and len(str(data).strip()) > 0:
                            payload = str(data).strip()
                        else:
                            # Try QR detection with grayscale for better detection
                            try:
                                gray = cv2.cvtColor(im, cv2.COLOR_BGR2GRAY)
                                data, points, _ = qr_detector.detectAndDecode(gray)
                                if points is not None and data and len(str(data).strip()) > 0:
                                    payload = str(data).strip()
                            except Exception:
                                pass
                        
                        # If QR code not found, try barcode detection
                        # Note: OpenCV barcode detector often fails with Code128 barcodes,
                        # so we rely on the .txt file fallback below
                        if payload is None:
                            try:
                                # OpenCV barcode detector (available in OpenCV 4.5+)
                                barcode_detector = cv2.barcode_BarcodeDetector()
                                retval, decoded_info, decoded_type = barcode_detector.detectAndDecode(im)
                                if retval and decoded_info and len(decoded_info) > 0:
                                    # decoded_info is a list, get first result
                                    barcode_data = decoded_info[0] if isinstance(decoded_info, list) else str(decoded_info)
                                    # Barcode may contain just doc_ref, try to find corresponding .txt file
                                    looked_up = self._lookup_barcode_data(barcode_data, path)
                                    payload = looked_up if looked_up else barcode_data
                                else:
                                    # Try barcode detection with grayscale
                                    try:
                                        gray = cv2.cvtColor(im, cv2.COLOR_BGR2GRAY)
                                        retval, decoded_info, decoded_type = barcode_detector.detectAndDecode(gray)
                                        if retval and decoded_info and len(decoded_info) > 0:
                                            barcode_data = decoded_info[0] if isinstance(decoded_info, list) else str(decoded_info)
                                            looked_up = self._lookup_barcode_data(barcode_data, path)
                                            payload = looked_up if looked_up else barcode_data
                                    except Exception:
                                        pass
                            except AttributeError:
                                # barcode_BarcodeDetector not available in this OpenCV version
                                pass
                            except Exception as e:
                                # Log barcode detection error but continue
                                try:
                                    logger.debug(f"OpenCV barcode detection error: {e}", exc_info=True)
                                except Exception:
                                    pass
                    else:
                        # Image file couldn't be read
                        try:
                            logger.debug(f"OpenCV could not read image file: {path}")
                        except Exception:
                            pass
                except ImportError:
                    # OpenCV not available, will be handled below
                    pass
                except Exception as e:
                    # Log error for debugging but continue
                    try:
                        logger.debug(f"OpenCV detection error: {e}", exc_info=True)
                    except Exception:
                        pass
                    pass
            if not payload:
                # If no code detected, try to read corresponding .txt file (for barcode images)
                # This is important for barcode images where the barcode contains just a doc_ref
                # or when OpenCV barcode detector fails to detect the barcode
                try:
                    import os
                    import re
                    from pathlib import Path
                    image_path_obj = Path(path)
                    
                    # Try direct .txt file with same name (e.g., txn_1_code128.png -> txn_1_code128.txt)
                    txt_path = image_path_obj.with_suffix('.txt')
                    if txt_path.exists():
                        try:
                            with open(txt_path, 'r', encoding='utf-8') as f:
                                content = f.read().strip()
                                if content:
                                    payload = content
                                    # Update status
                                    try:
                                        if hasattr(self, 'txn_prefill_status'):
                                            self.txn_prefill_status.configure(text="Barcode not detected, but found corresponding data file")
                                    except Exception:
                                        pass
                        except Exception as e:
                            try:
                                logger.debug(f"Error reading .txt file: {e}", exc_info=True)
                            except Exception:
                                pass
                    
                    # Also try looking for files with similar patterns in the same directory
                    # This handles cases where the filename doesn't match exactly
                    if not payload:
                        parent_dir = image_path_obj.parent
                        if parent_dir.exists():
                            img_name = image_path_obj.stem.lower()
                            
                            # Extract transaction number from image filename
                            # Patterns: txn_1_code128, txn_1, barcode_00001, etc.
                            img_txn_match = re.search(r'txn[_\s]*(\d+)', img_name)
                            if not img_txn_match:
                                img_txn_match = re.search(r'(\d+)', img_name)
                            
                            # Try to find matching .txt file
                            for txt_file in parent_dir.glob('*.txt'):
                                try:
                                    txt_name = txt_file.stem.lower()
                                    
                                    # Check if it's a code128.txt file
                                    if 'code128' in txt_name:
                                        # Extract transaction number from txt filename
                                        txt_txn_match = re.search(r'txn[_\s]*(\d+)', txt_name)
                                        if not txt_txn_match:
                                            txt_txn_match = re.search(r'(\d+)', txt_name)
                                        
                                        # If transaction numbers match
                                        if img_txn_match and txt_txn_match and img_txn_match.group(1) == txt_txn_match.group(1):
                                            with open(txt_file, 'r', encoding='utf-8') as f:
                                                content = f.read().strip()
                                                if content and (content.startswith('{') or '=' in content):
                                                    payload = content
                                                    try:
                                                        if hasattr(self, 'txn_prefill_status'):
                                                            self.txn_prefill_status.configure(text="Found matching transaction data file")
                                                    except Exception:
                                                        pass
                                                    break
                                except Exception:
                                    continue
                except Exception as e:
                    try:
                        logger.debug(f"Error in .txt file fallback: {e}", exc_info=True)
                    except Exception:
                        pass
            
            if not payload:
                # Check for missing libraries first
                missing_libs = []
                try:
                    from pyzbar.pyzbar import decode
                except Exception:
                    missing_libs.append("pyzbar")
                try:
                    import cv2
                except Exception:
                    missing_libs.append("opencv-python")
                
                if missing_libs:
                    # Only show library error if critical libraries are missing
                    if "opencv-python" in missing_libs:
                        messagebox.showerror(
                            "Scanning Libraries Not Available",
                            f"Required library not available: opencv-python\n\n"
                            "Please install it with: pip install opencv-python"
                        )
                        try:
                            if hasattr(self, 'txn_prefill_status'):
                                self.txn_prefill_status.configure(text="Scan error: opencv-python not installed")
                        except Exception:
                            pass
                        return
                    else:
                        # Only pyzbar missing (optional for barcodes)
                        msg = "No code detected in image. The image may not contain a QR code or barcode, or it may be too small/blurry.\n\nNote: Barcode scanning requires pyzbar library."
                else:
                    # Libraries are available but no code detected
                    msg = "No QR code or barcode detected in the selected image.\n\nThe image may be:\n• Too small or low resolution\n• Blurry or out of focus\n• Not containing a QR code or barcode\n• In an unsupported format"
                
                # Show error and let user try again or use manual entry
                response = messagebox.askyesno(
                    "Scan Failed", 
                    msg + "\n\nWould you like to try entering the data manually instead?"
                )
                if response:
                    self._manual_data_entry()
                else:
                    # Update status to show scan failed, but allow user to try again
                    try:
                        if hasattr(self, 'txn_prefill_status'):
                            self.txn_prefill_status.configure(text="Scan failed - no code detected. Try another image or use manual entry.")
                    except Exception:
                        pass
                return
            if not payload:
                # This should have been handled above, but double-check
                try:
                    if hasattr(self, 'txn_prefill_status'):
                        self.txn_prefill_status.configure(text="Scan failed - no data found")
                except Exception:
                    pass
                return
            
            try:
                data = self._parse_scanned_payload(payload)
            except Exception as e:
                try:
                    logger.debug(f"Error parsing scanned payload: {e}", exc_info=True)
                except Exception:
                    pass
                data = None
            if not data:
                messagebox.showerror("Scan", f"Invalid code data in image.\n\nPayload: {payload[:100] if payload else 'None'}")
                try:
                    if hasattr(self, 'txn_prefill_status'):
                        self.txn_prefill_status.configure(text="Scan failed - invalid data format")
                except Exception:
                    pass
                return
            try:
                # Pass the image file path to help with auto-entry inference
                self._apply_scanned_data(data, path)
                if hasattr(self, 'txn_prefill_status'):
                    self.txn_prefill_status.configure(text="Scan from image successful")
            except Exception as e:
                try:
                    logger.debug(f"Error applying scanned data: {e}", exc_info=True)
                except Exception:
                    pass
                messagebox.showerror("Scan", f"Failed to apply data from image: {str(e)}")
        except Exception:
            try:
                self.txn_prefill_status.configure(text="Scan from image failed")
            except Exception:
                pass
    
    def _lookup_barcode_data(self, barcode_data: str, image_path: str) -> str | None:
        """
        Look up full transaction data from barcode.
        If barcode contains just a document reference, try to find corresponding .txt file.
        """
        try:
            import os
            from pathlib import Path
            
            # If barcode data looks like JSON or key=value, return it directly
            if (barcode_data.startswith('{') and barcode_data.endswith('}')) or '=' in barcode_data:
                return barcode_data
            
            # Otherwise, it's likely just a document reference
            # Try to find corresponding .txt file in the same directory
            image_path_obj = Path(image_path)
            txt_path = image_path_obj.with_suffix('.txt')
            
            if txt_path.exists():
                try:
                    with open(txt_path, 'r', encoding='utf-8') as f:
                        content = f.read().strip()
                        if content:
                            return content
                except Exception:
                    pass
            
            # Also try looking for files with the barcode data in the name
            # (for mock_codes directory structure: txn_X_code128.txt)
            parent_dir = image_path_obj.parent
            if parent_dir.exists():
                # Try pattern: *{barcode_data}*.txt or *code128.txt
                for txt_file in parent_dir.glob('*code128.txt'):
                    try:
                        with open(txt_file, 'r', encoding='utf-8') as f:
                            content = f.read().strip()
                            # Check if this file's data contains the barcode as document_ref
                            if content and barcode_data in content:
                                return content
                    except Exception:
                        continue
        except Exception:
            pass
        return None
    
    def _manual_data_entry(self) -> None:
        """Allow user to manually paste/enter QR code or barcode data."""
        # Create dialog window matching app styling
        dialog = tk.Toplevel(self)
        dialog.title("Enter Code Data Manually")
        dialog.transient(self)
        dialog.grab_set()
        dialog.focus_set()
        
        # Configure dialog background
        try:
            dialog.configure(bg=self.palette.get('surface_bg', '#ffffff'))
        except Exception:
            pass
        
        # Main container frame
        main_frame = ttk.Frame(dialog, style="Techfix.Surface.TFrame")
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Instructions section
        instructions_frame = ttk.Frame(main_frame, style="Techfix.Surface.TFrame")
        instructions_frame.pack(fill=tk.X, pady=(0, 15))
        
        ttk.Label(
            instructions_frame,
            text="Enter Code Data",
            style="Techfix.TLabelframe.Label"
        ).pack(anchor=tk.W, pady=(0, 8))
        
        instructions_text = tk.Text(
            instructions_frame,
            height=4,
            wrap=tk.WORD,
            bg=self.palette.get("surface_bg", "#ffffff"),
            fg=self.palette.get("text_secondary", "#4b5563"),
            font=FONT_BASE,
            relief=tk.FLAT,
            bd=0,
            highlightthickness=0,
            padx=0,
            pady=0
        )
        instructions_text.pack(fill=tk.X)
        instructions_text.insert('1.0', 
            "Paste the data from your QR code or barcode here.\n\n"
            "Supported formats:\n"
            "• JSON: {\"date\":\"2024-01-01\",\"source_type\":\"Receipt\",...}\n"
            "• Key=Value: date=2024-01-01&source=Receipt&doc=12345&amount=99.99"
        )
        instructions_text.config(state=tk.DISABLED)
        
        # Data entry section
        data_frame = ttk.Frame(main_frame, style="Techfix.Surface.TFrame")
        data_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 15))
        
        ttk.Label(
            data_frame,
            text="Data:",
            style="Techfix.TLabel"
        ).pack(anchor=tk.W, pady=(0, 5))
        
        # Text entry with scrollbar
        text_container = ttk.Frame(data_frame, style="Techfix.Surface.TFrame")
        text_container.pack(fill=tk.BOTH, expand=True)
        
        data_text = tk.Text(
            text_container,
            wrap=tk.WORD,
            bg=self.palette.get("surface_bg", "#ffffff"),
            fg=self.palette.get("text_primary", "#1f2937"),
            font=FONT_MONO,
            relief=tk.SOLID,
            bd=1,
            highlightthickness=1,
            highlightbackground=self.palette.get("entry_border", "#d8dee9"),
            highlightcolor=self.palette.get("accent_color", "#2563eb"),
            padx=8,
            pady=8,
            insertbackground=self.palette.get("accent_color", "#2563eb")
        )
        scrollbar = ttk.Scrollbar(text_container, orient=tk.VERTICAL, command=data_text.yview)
        data_text.configure(yscrollcommand=scrollbar.set)
        
        data_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        data_text.focus_set()
        
        # Button frame
        button_frame = ttk.Frame(main_frame, style="Techfix.Surface.TFrame")
        button_frame.pack(fill=tk.X)
        
        def process_data():
            payload = data_text.get('1.0', tk.END).strip()
            if not payload:
                messagebox.showwarning("No Data", "Please enter some data.")
                return
            
            try:
                data = self._parse_scanned_payload(payload)
            except Exception:
                data = None
            
            if not data:
                messagebox.showerror("Invalid Data", 
                    "Could not parse the data.\n\n"
                    "Make sure it's valid JSON or key=value format.\n\n"
                    "Example JSON: {\"date\":\"2024-01-01\",\"source_type\":\"Receipt\"}\n"
                    "Example key=value: date=2024-01-01&source=Receipt&doc=12345"
                )
                return
            
            dialog.destroy()
            try:
                # Apply the data (no filename for manual entry)
                self._apply_scanned_data(data, '')
                # Force update button states after a short delay to ensure all UI updates are complete
                self.after(100, self._update_post_buttons_enabled)
                # Also update immediately
                self._update_post_buttons_enabled()
                if hasattr(self, 'txn_prefill_status'):
                    # Check if everything was set correctly
                    accounts_ok = self._validate_accounts_assigned() if hasattr(self, '_validate_accounts_assigned') else False
                    amounts_ok = self._validate_amounts_present() if hasattr(self, '_validate_amounts_present') else False
                    if accounts_ok and amounts_ok:
                        self.txn_prefill_status.configure(text="Manual data entry successful - ready to record")
                    else:
                        missing = []
                        if not accounts_ok:
                            missing.append("accounts")
                        if not amounts_ok:
                            missing.append("amounts")
                        self.txn_prefill_status.configure(text=f"Manual data entry - missing: {', '.join(missing)}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to apply data: {e}")
                if hasattr(self, 'txn_prefill_status'):
                    self.txn_prefill_status.configure(text="Manual data entry failed")
        
        def load_from_file():
            file_path = filedialog.askopenfilename(
                title="Select text file with code data",
                filetypes=[('Text files', '*.txt'), ('All files', '*.*')]
            )
            if file_path:
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        content = f.read().strip()
                        data_text.delete('1.0', tk.END)
                        data_text.insert('1.0', content)
                except Exception as e:
                    messagebox.showerror("Error", f"Could not read file: {e}")
        
        ttk.Button(
            button_frame,
            text="Load from File",
            command=load_from_file,
            style="Techfix.Theme.TButton"
        ).pack(side=tk.LEFT, padx=(0, 5))
        
        ttk.Button(
            button_frame,
            text="Cancel",
            command=dialog.destroy,
            style="Techfix.Theme.TButton"
        ).pack(side=tk.RIGHT, padx=(5, 0))
        
        ttk.Button(
            button_frame,
            text="Process Data",
            command=process_data,
            style="Techfix.TButton"
        ).pack(side=tk.RIGHT)
        
        # Set initial size and center dialog
        dialog.update_idletasks()
        dialog.geometry("700x500")
        width = dialog.winfo_width()
        height = dialog.winfo_height()
        x = (dialog.winfo_screenwidth() // 2) - (width // 2)
        y = (dialog.winfo_screenheight() // 2) - (height // 2)
        dialog.geometry(f'{width}x{height}+{x}+{y}')
    
    def _prefill_from_source_document(self, filename: str) -> None:
        try:
            import os, json, csv
            data = None
            updated = []
            # Normalize the filename path
            filename = os.path.normpath(os.path.abspath(filename))
            side_json = os.path.splitext(filename)[0] + ".json"
            # Log for debugging
            self._audit('prefill_sidecar_check', {'file': filename, 'sidecar': side_json, 'exists': os.path.exists(side_json)})
            if os.path.exists(side_json):
                try:
                    with open(side_json, "r", encoding="utf-8") as f:
                        data = json.load(f)
                    self._audit('prefill_json_loaded', {'file': side_json, 'keys': list(data.keys()) if isinstance(data, dict) else 'not_dict'})
                except Exception as e:
                    self._audit('prefill_json_read_error', {'file': side_json, 'error': str(e)})
            elif filename.lower().endswith(".json"):
                with open(filename, "r", encoding="utf-8") as f:
                    data = json.load(f)
            elif filename.lower().endswith(".csv"):
                with open(filename, newline="", encoding="utf-8") as f:
                    r = csv.DictReader(f)
                    data = next(r, None)
            if not data:
                base = os.path.basename(filename)
                name, _ = os.path.splitext(base)
                parts = name.split("_")
                if len(parts) >= 2:
                    date = parts[0] if len(parts[0]) == 10 else None
                    typ = parts[1].title()
                    docno = parts[2] if len(parts) >= 3 else None
                    desc_rest = " ".join(parts[3:]).replace("-", " ") if len(parts) > 3 else ""
                    # Try to locate a nearby JSON by doc number if sidecar missing
                    if not data and docno:
                        try:
                            dirpath = os.path.dirname(filename)
                            for fn in os.listdir(dirpath):
                                if fn.lower().endswith('.json') and docno in fn:
                                    with open(os.path.join(dirpath, fn), 'r', encoding='utf-8') as f:
                                        data = json.load(f)
                                    break
                        except Exception:
                            pass
                    if hasattr(self, "txn_source_type"):
                        self.txn_source_type.set(typ if typ in ["Invoice", "Receipt", "Bank", "Adjust", "Payroll", "Other"] else "")
                        updated.append("Source")
                    # Check if adjusting entry based on source type from filename
                    if hasattr(self, 'txn_is_adjust') and typ == "Adjust":
                        try:
                            self.txn_is_adjust.set(1)
                        except Exception:
                            pass
                    if hasattr(self, "txn_doc_ref") and docno:
                        try:
                            self.txn_doc_ref.delete(0, tk.END); self.txn_doc_ref.insert(0, docno)
                            updated.append("Document Ref")
                        except Exception:
                            pass
                    if hasattr(self, "txn_desc"):
                        try:
                            dval = (typ + (" " + docno if docno else "") + (" " + desc_rest if desc_rest else "")).strip()
                            self.txn_desc.delete(0, tk.END); self.txn_desc.insert(0, dval)
                            updated.append("Description")
                            # Also check adjusting entry based on description
                            if hasattr(self, 'txn_is_adjust') and 'adjusting entry' in dval.lower():
                                try:
                                    self.txn_is_adjust.set(1)
                                except Exception:
                                    pass
                        except Exception:
                            pass
                    if hasattr(self, "txn_date") and date:
                        try:
                            self.txn_date.delete(0, tk.END); self.txn_date.insert(0, date)
                            updated.append("Date")
                        except Exception:
                            pass
                    # If we loaded data from nearby JSON, continue below to fill structured fields
                    if data:
                        pass
                    else:
                        try:
                            pseudo = {
                                'source_type': typ,
                                'document_ref': docno,
                                'description': (typ + (" " + docno if docno else "") + (" " + desc_rest if desc_rest else "")).strip(),
                            }
                            sugg = self._auto_entry_from_data(pseudo, filename)
                            dd = sugg.get('debit_account_display')
                            cc = sugg.get('credit_account_display')
                            if dd or cc:
                                self._set_accounts(dd, cc)
                                updated.append("Accounts")
                            if hasattr(self, 'txn_desc') and sugg.get('description'):
                                _set_entry(self.txn_desc, sugg.get('description'))
                                updated.append("Description")
                            # continue with status update below
                        except Exception:
                            pass
                elif not data:
                    try:
                        messagebox.showwarning("Prefill", "No structured data found in selected document. Use a matching .json/.csv or filename pattern: YYYY-MM-DD_Type_DocNo_Description")
                    except Exception:
                        pass
                    return
            def _set_entry(w, v):
                try:
                    if w and v is not None:
                        w.delete(0, tk.END); w.insert(0, str(v))
                        return True
                except Exception:
                    pass
                return False
            # Only process data if we have a dict
            if not isinstance(data, dict):
                return
            if hasattr(self, "txn_date") and data.get("date"):
                if _set_entry(self.txn_date, data.get("date")):
                    updated.append("Date")
            if hasattr(self, "txn_desc") and data.get("description"):
                if _set_entry(self.txn_desc, data.get("description")):
                    updated.append("Description")
            if hasattr(self, "txn_doc_ref") and data.get("document_ref"):
                if _set_entry(self.txn_doc_ref, data.get("document_ref")):
                    updated.append("Document Ref")
            if hasattr(self, "txn_external_ref") and data.get("external_ref"):
                if _set_entry(self.txn_external_ref, data.get("external_ref")):
                    updated.append("External Ref")
            if hasattr(self, "txn_source_type") and data.get("source_type"):
                try:
                    self.txn_source_type.set(data.get("source_type"))
                    updated.append("Source")
                except Exception:
                    pass
            # Automatically check adjusting entry checkbox if source_type is "Adjust" or description contains "Adjusting entry"
            if hasattr(self, 'txn_is_adjust'):
                try:
                    source_type = data.get('source_type', '').strip()
                    description = data.get('description', '').strip()
                    is_adjusting = (
                        source_type.lower() == 'adjust' or
                        'adjusting entry' in description.lower()
                    )
                    if is_adjusting:
                        self.txn_is_adjust.set(1)
                except Exception:
                    pass
            if hasattr(self, "txn_memo") and data.get("memo"):
                try:
                    self.txn_memo.delete("1.0", tk.END); self.txn_memo.insert("1.0", data.get("memo"))
                    updated.append("Memo")
                except Exception:
                    pass
            try:
                if isinstance(data, dict):
                    # Try direct account matching first
                    accounts_set = self._assign_accounts_from_data(data)
                    
                    # Get current account values to see what's missing
                    current_dd = None
                    current_cc = None
                    try:
                        if hasattr(self, 'debit_acct_var'):
                            current_dd = self.debit_acct_var.get().strip()
                        elif hasattr(self, 'debit_acct'):
                            current_dd = self.debit_acct.get().strip()
                        if hasattr(self, 'credit_acct_var'):
                            current_cc = self.credit_acct_var.get().strip()
                        elif hasattr(self, 'credit_acct'):
                            current_cc = self.credit_acct.get().strip()
                    except Exception:
                        pass
                    
                    # Always try auto-entry to fill in missing accounts and other fields
                    sugg = self._auto_entry_from_data(data, filename)
                    if sugg:
                        # Use auto-entry suggestions to fill in missing accounts
                        dd = current_dd if current_dd else sugg.get('debit_account_display')
                        cc = current_cc if current_cc else sugg.get('credit_account_display')
                        if dd or cc:
                            # Only update if we have new account information
                            if (dd and not current_dd) or (cc and not current_cc):
                                self._set_accounts(dd, cc)
                                if not accounts_set or (dd and not current_dd) or (cc and not current_cc):
                                    updated.append("Accounts")
                        # Always try to set amounts and other fields from auto-entry
                        if hasattr(self, 'debit_amt') and sugg.get('debit_amount') is not None:
                            _set_entry(self.debit_amt, f"{float(sugg.get('debit_amount')):.2f}")
                            updated.append("Debit Amount")
                        if hasattr(self, 'credit_amt') and sugg.get('credit_amount') is not None:
                            _set_entry(self.credit_amt, f"{float(sugg.get('credit_amount')):.2f}")
                            updated.append("Credit Amount")
                        if hasattr(self, 'txn_desc') and sugg.get('description'):
                            _set_entry(self.txn_desc, sugg.get('description'))
                            updated.append("Description")
                        if hasattr(self, 'txn_source_type') and sugg.get('source_type'):
                            self.txn_source_type.set(sugg.get('source_type'))
                        if hasattr(self, 'txn_doc_ref') and sugg.get('document_ref'):
                            _set_entry(self.txn_doc_ref, sugg.get('document_ref'))
            except Exception:
                pass
            if hasattr(self, "debit_amt") and data.get("debit_amount") is not None:
                try:
                    if _set_entry(self.debit_amt, f"{float(data.get('debit_amount')):.2f}"):
                        updated.append("Debit Amount")
                except Exception:
                    pass
            if hasattr(self, "credit_amt") and data.get("credit_amount") is not None:
                try:
                    if _set_entry(self.credit_amt, f"{float(data.get('credit_amount')):.2f}"):
                        updated.append("Credit Amount")
                except Exception:
                    pass
            try:
                assigned_ok = self._validate_accounts_assigned()
                amounts_ok = self._validate_amounts_present()
                msg = f"Prefilled: {', '.join(updated)}" if updated else "No structured data found"
                if hasattr(self, 'txn_prefill_status'):
                    self.txn_prefill_status.configure(text=(msg + (" (ok)" if (assigned_ok and amounts_ok) else " (missing)")))
                self._audit('auto_entry_validation', {'file': filename, 'accounts_ok': assigned_ok, 'amounts_ok': amounts_ok})
                if not assigned_ok:
                    messagebox.showwarning("Accounts", "Accounts not set from document. Please select Debit and Credit accounts.")
            except Exception:
                pass
        except Exception as e:
            # Log the error for debugging
            try:
                self._audit('prefill_from_source_document_error', {'file': filename, 'error': str(e)})
            except Exception:
                pass

    def _prefill_date_from_source_document(self, filename: str) -> None:
        try:
            import os, json, csv
            date_val = None
            side_json = os.path.splitext(filename)[0] + ".json"
            if os.path.exists(side_json):
                with open(side_json, "r", encoding="utf-8") as f:
                    d = json.load(f)
                    date_val = d.get("date")
            elif filename.lower().endswith(".json"):
                with open(filename, "r", encoding="utf-8") as f:
                    d = json.load(f)
                    date_val = d.get("date")
            elif filename.lower().endswith(".csv"):
                with open(filename, newline="", encoding="utf-8") as f:
                    r = csv.DictReader(f)
                    row = next(r, None)
                    if row:
                        date_val = row.get("date")
            if not date_val:
                base = os.path.basename(filename)
                name, _ = os.path.splitext(base)
                parts = name.split("_")
                if parts and len(parts[0]) == 10:
                    date_val = parts[0]
            if hasattr(self, "txn_date"):
                cur = self.txn_date.get().strip()
                if (not cur) and date_val:
                    try:
                        self.txn_date.delete(0, tk.END); self.txn_date.insert(0, date_val)
                    except Exception:
                        pass
                    try:
                        if hasattr(self, 'txn_prefill_status'):
                            self.txn_prefill_status.configure(text="Date set from document")
                    except Exception:
                        pass
        except Exception:
            pass

    def _prefill_amounts_from_source_document(self, filename: str) -> None:
        try:
            import os, json, csv
            def parse_num(v):
                try:
                    import re
                    s = str(v).strip()
                    neg = False
                    if s.startswith('(') and s.endswith(')'):
                        neg = True
                        s = s[1:-1]
                    s = re.sub(r"[^0-9.\-]", "", s)
                    if not s:
                        return None
                    val = float(s)
                    if neg:
                        val = -abs(val)
                    return val
                except Exception:
                    return None
            def extract_amounts(data_obj):
                debit_val = None
                credit_val = None
                default_amt = None
                if isinstance(data_obj, dict):
                    keys = set(k.lower() for k in data_obj.keys())
                    # direct keys
                    for k in ("debit_amount","debit"): 
                        if k in keys:
                            dv = parse_num(data_obj.get(k))
                            if dv is not None:
                                debit_val = dv
                    for k in ("credit_amount","credit"):
                        if k in keys:
                            cv = parse_num(data_obj.get(k))
                            if cv is not None:
                                credit_val = cv
                    # total-like keys
                    for k in ("amount","total","grand_total","net_amount","gross_amount","subtotal","payment_amount"):
                        if k in keys and default_amt is None:
                            default_amt = parse_num(data_obj.get(k))
                    # nested lists
                    for lk in ("lines","items","entries","details"):
                        if lk in keys and isinstance(data_obj.get(lk), list):
                            dv_sum = 0.0; cv_sum = 0.0; any_d=False; any_c=False
                            for it in data_obj.get(lk):
                                if isinstance(it, dict):
                                    dv = parse_num(it.get("debit"))
                                    cv = parse_num(it.get("credit"))
                                    amt = parse_num(it.get("amount"))
                                    if dv is not None:
                                        dv_sum += dv; any_d=True
                                    if cv is not None:
                                        cv_sum += cv; any_c=True
                                    if amt is not None and default_amt is None:
                                        default_amt = amt
                            if any_d and debit_val is None:
                                debit_val = dv_sum
                            if any_c and credit_val is None:
                                credit_val = cv_sum
                return debit_val, credit_val, default_amt
            debit_val = None
            credit_val = None
            data = None
            side_json = os.path.splitext(filename)[0] + ".json"
            if os.path.exists(side_json):
                with open(side_json, "r", encoding="utf-8") as f:
                    data = json.load(f)
            elif filename.lower().endswith(".json"):
                with open(filename, "r", encoding="utf-8") as f:
                    data = json.load(f)
            elif filename.lower().endswith(".csv"):
                with open(filename, newline="", encoding="utf-8") as f:
                    r = csv.DictReader(f)
                    data = next(r, None)
            if isinstance(data, dict):
                dv, cv, def_amt = extract_amounts(data)
                debit_val = dv
                credit_val = cv
                if debit_val is None and credit_val is None and def_amt is not None:
                    debit_val = def_amt
                    credit_val = def_amt
            # Apply to form if found
            set_any = False
            if hasattr(self, 'debit_amt') and debit_val is not None:
                try:
                    self.debit_amt.delete(0, tk.END); self.debit_amt.insert(0, f"{debit_val:.2f}")
                    set_any = True
                except Exception:
                    pass
            if hasattr(self, 'credit_amt') and credit_val is not None:
                try:
                    self.credit_amt.delete(0, tk.END); self.credit_amt.insert(0, f"{credit_val:.2f}")
                    set_any = True
                except Exception:
                    pass
            if set_any:
                try:
                    if hasattr(self, 'txn_prefill_status'):
                        self.txn_prefill_status.configure(text="Amounts set from document")
                    self._audit('document_amounts_set', {'file': filename, 'debit': (self.debit_amt.get().strip() if hasattr(self, 'debit_amt') else None), 'credit': (self.credit_amt.get().strip() if hasattr(self, 'credit_amt') else None)})
                except Exception:
                    pass
            else:
                self._audit('document_amounts_missing', {'file': filename})
        except Exception:
            self._audit('document_prefill_error', {'file': filename, 'stage': 'amounts_exception'})
            
    def _clear_transaction_form(self):
        """Reset all fields in the transaction form to their default values."""
        # Clear entry fields
        if hasattr(self, 'txn_date'):
            try:
                self.txn_date.delete(0, tk.END)
                self.txn_date.insert(0, datetime.now().strftime('%Y-%m-%d'))
            except Exception:
                pass
        # Clear document/reference fields (some names changed over time)
        if hasattr(self, 'txn_doc_ref'):
            try:
                self.txn_doc_ref.delete(0, tk.END)
            except Exception:
                pass
        if hasattr(self, 'txn_external_ref'):
            try:
                self.txn_external_ref.delete(0, tk.END)
            except Exception:
                pass
        if hasattr(self, 'txn_desc'):
            try:
                self.txn_desc.delete(0, tk.END)
            except Exception:
                pass
        if hasattr(self, 'txn_memo'):
            try:
                self.txn_memo.delete('1.0', tk.END)
            except Exception:
                pass
        if hasattr(self, 'txn_attachment_path'):
            try:
                self.txn_attachment_path.set('')
            except Exception:
                pass
        if hasattr(self, 'txn_source_type'):
            try:
                self.txn_source_type.set('')
            except Exception:
                pass
        if hasattr(self, 'debit_acct'):
            try:
                self.debit_acct.set('')
            except Exception:
                pass
        if hasattr(self, 'credit_acct'):
            try:
                self.credit_acct.set('')
            except Exception:
                pass
        if hasattr(self, 'debit_amt'):
            try:
                self.debit_amt.delete(0, tk.END)
            except Exception:
                pass
        if hasattr(self, 'credit_amt'):
            try:
                self.credit_amt.delete(0, tk.END)
            except Exception:
                pass
        if hasattr(self, 'txn_reverse_date'):
            try:
                self.txn_reverse_date.delete(0, tk.END)
            except Exception:
                pass
        if hasattr(self, 'txn_is_adjust'):
            try:
                self.txn_is_adjust.set(0)
            except Exception:
                pass
        if hasattr(self, 'txn_schedule_reverse'):
            try:
                self.txn_schedule_reverse.set(0)
            except Exception:
                pass
        try:
            self._accounts_prefilled = False
        except Exception:
            pass
        # Clear prefill status under Document
        try:
            if hasattr(self, 'txn_prefill_status'):
                self.txn_prefill_status.configure(text='')
        except Exception:
            pass
        # Clear document state
        try:
            self.current_document_path = None
            if hasattr(self, 'doc_recent_var'):
                try:
                    self.doc_recent_var.set('')
                except Exception:
                    pass
        except Exception:
            pass
            
        # Clear any journal entry lines if they exist
        if hasattr(self, 'journal_entries') and hasattr(self.journal_entries, 'delete'):
            self.journal_entries.delete(0, tk.END)
            
        # Reset the transaction ID if it exists
        if hasattr(self, 'current_transaction_id'):
            self.current_transaction_id = None
            
        # Set focus to the first field (txn_date Entry)
        if hasattr(self, 'txn_date'):
            try:
                self.txn_date.focus_set()
            except Exception:
                pass

    def _set_txn_date_today(self) -> None:
        """Set the transaction date field to today's date."""
        try:
            if hasattr(self, 'txn_date'):
                self.txn_date.delete(0, tk.END)
                self.txn_date.insert(0, datetime.now().strftime('%Y-%m-%d'))
        except Exception:
            pass

    def _pick_txn_date(self) -> None:
        """Open a small date picker dialog and set the txn_date when picked."""
        try:
            class DatePicker(tk.Toplevel):
                def __init__(self, parent, callback=None, initial_date=None):
                    super().__init__(parent)
                    self.transient(parent)
                    self.grab_set()
                    self.callback = callback
                    self.title("Pick Date")
                    self.resizable(False, False)

                    today = datetime.now().date()
                    if initial_date:
                        try:
                            dt = datetime.strptime(initial_date, '%Y-%m-%d').date()
                        except Exception:
                            dt = today
                    else:
                        dt = today

                    self.year = dt.year
                    self.month = dt.month

                    header = ttk.Frame(self)
                    header.pack(fill=tk.X, padx=8, pady=6)
                    ttk.Button(header, text="◀", width=3, command=self._prev_month).pack(side=tk.LEFT)
                    self.title_lbl = ttk.Label(header, text="", anchor=tk.CENTER)
                    self.title_lbl.pack(side=tk.LEFT, expand=True)
                    ttk.Button(header, text="▶", width=3, command=self._next_month).pack(side=tk.RIGHT)

                    self.cal_frame = ttk.Frame(self)
                    self.cal_frame.pack(padx=8, pady=(0,8))

                    self._draw_calendar()

                def _draw_calendar(self):
                    for child in self.cal_frame.winfo_children():
                        child.destroy()
                    self.title_lbl.config(text=f"{calendar.month_name[self.month]} {self.year}")
                    wdays = ['Mo','Tu','We','Th','Fr','Sa','Su']
                    for c, wd in enumerate(wdays):
                        ttk.Label(self.cal_frame, text=wd, width=3, anchor='center').grid(row=0, column=c)
                    m = calendar.monthcalendar(self.year, self.month)
                    for r, week in enumerate(m, start=1):
                        for c, day in enumerate(week):
                            if day == 0:
                                ttk.Label(self.cal_frame, text='', width=3).grid(row=r, column=c, padx=1, pady=1)
                            else:
                                b = ttk.Button(self.cal_frame, text=str(day), width=3, command=lambda d=day: self._select_day(d))
                                b.grid(row=r, column=c, padx=1, pady=1)

                def _prev_month(self):
                    self.month -= 1
                    if self.month < 1:
                        self.month = 12
                        self.year -= 1
                    self._draw_calendar()

                def _next_month(self):
                    self.month += 1
                    if self.month > 12:
                        self.month = 1
                        self.year += 1
                    self._draw_calendar()

                def _select_day(self, day):
                    try:
                        d = datetime(self.year, self.month, day).strftime('%Y-%m-%d')
                        if callable(self.callback):
                            self.callback(d)
                    except Exception:
                        pass
                    finally:
                        try:
                            self.grab_release()
                        except Exception:
                            pass
                        self.destroy()

            def _on_date_chosen(d):
                try:
                    if hasattr(self, 'txn_date'):
                        self.txn_date.delete(0, tk.END)
                        self.txn_date.insert(0, d)
                except Exception:
                    pass

            # open the picker
            cur = None
            try:
                cur = self.txn_date.get().strip()
            except Exception:
                cur = None
            dp = DatePicker(self, callback=_on_date_chosen, initial_date=cur)
            self.wait_window(dp)
        except Exception:
            pass

    def _build_scrollable_tab(self, parent: ttk.Frame) -> ttk.Frame:
        """Create a scrollable frame within a tab"""
        # Create a container frame that will hold the canvas and scrollbar
        container = ttk.Frame(parent, style="Techfix.Surface.TFrame")
        container.pack(fill=tk.BOTH, expand=True)

        # Create canvas with scrollbar
        canvas = tk.Canvas(container, bd=0, highlightthickness=0, bg=self.palette["surface_bg"])
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)

        # Create a frame inside the canvas that will hold our content
        scrollable_frame = ttk.Frame(canvas, style="Techfix.Surface.TFrame")

        # Configure the canvas
        canvas.configure(yscrollcommand=scrollbar.set)

        # Pack the scrollbar and canvas
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        # Create window in canvas for the scrollable frame
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw", tags=("scrollable_frame",))

        # Configure the scroll region when the frame changes size
        def _on_frame_configure(event):
            # Update the scroll region to encompass the inner frame
            canvas.configure(scrollregion=canvas.bbox("all"))
            # Make sure the frame fills the canvas width
            canvas.itemconfig("scrollable_frame", width=event.width)

        def _on_mousewheel(event):
            # Handle mousewheel scrolling
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            return "break"

        # Bind the configure event to handle resizing
        scrollable_frame.bind("<Configure>", _on_frame_configure)

        # Bind mousewheel events to the canvas and all its children
        def bind_mousewheel(widget):
            widget.bind("<MouseWheel>", _on_mousewheel)
            widget.bind("<Button-4>", lambda e: canvas.yview_scroll(-1, "units") if canvas.winfo_containing(e.x_root, e.y_root) else None)
            widget.bind("<Button-5>", lambda e: canvas.yview_scroll(1, "units") if canvas.winfo_containing(e.x_root, e.y_root) else None)

            for child in widget.winfo_children():
                bind_mousewheel(child)

        bind_mousewheel(scrollable_frame)

        # Bind mousewheel for Windows
        canvas.bind("<MouseWheel>", _on_mousewheel)

        # Bind mousewheel for Linux (button 4/5)
        canvas.bind("<Button-4>", lambda e: canvas.yview_scroll(-1, "units") if canvas.winfo_containing(e.x_root, e.y_root) in (canvas, scrollbar) else None)
        canvas.bind("<Button-5>", lambda e: canvas.yview_scroll(1, "units") if canvas.winfo_containing(e.x_root, e.y_root) in (canvas, scrollbar) else None)

        # Store canvas for theme updates
        if not hasattr(self, '_scroll_canvases'):
            self._scroll_canvases = []
        self._scroll_canvases.append(canvas)

        # Make sure entry widgets don't capture scroll events
        def _on_entry_mousewheel(event):
            return "break"

        # Apply to all existing entry widgets
        for child in scrollable_frame.winfo_children():
            if isinstance(child, (ttk.Entry, ttk.Combobox, tk.Text, ttk.Spinbox)):
                child.bind("<MouseWheel>", _on_entry_mousewheel)

        return scrollable_frame

    def _build_transactions_tab(self) -> None:
        frame = self.tab_txn
        
        # Single column layout - form takes full width
        main_container = ttk.Frame(frame, style="Techfix.Surface.TFrame")
        main_container.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        # Configure single column to take full width
        main_container.columnconfigure(0, weight=1)
        main_container.rowconfigure(0, weight=1)

        left_panel = ttk.Frame(main_container, style="Techfix.Surface.TFrame")
        left_panel.grid(row=0, column=0, sticky="nsew")
        
        # Create scrollable container for the form
        scroll_container = ttk.Frame(left_panel, style="Techfix.Surface.TFrame")
        scroll_container.pack(fill=tk.BOTH, expand=True)
        
        # Create canvas with scrollbar
        canvas = tk.Canvas(scroll_container, bd=0, highlightthickness=0, bg=self.palette["surface_bg"])
        scrollbar = ttk.Scrollbar(scroll_container, orient="vertical", command=canvas.yview)
        
        # Form frame - inside scrollable canvas
        form = ttk.LabelFrame(canvas, text="New Transaction", style="Techfix.TLabelframe")
        
        # Configure the canvas
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Pack the scrollbar and canvas
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        
        # Create window in canvas for the form
        canvas_window = canvas.create_window((0, 0), window=form, anchor="nw", tags=("form_frame",))
        
        # Configure the scroll region when the form changes size
        def _on_form_configure(event):
            # Update the scroll region to encompass the form
            canvas.configure(scrollregion=canvas.bbox("all"))
            # Make sure the form fills the canvas width
            canvas_width = event.width
            canvas.itemconfig("form_frame", width=canvas_width)
        
        def _on_canvas_configure(event):
            # Update form width when canvas resizes
            canvas_width = event.width
            canvas.itemconfig("form_frame", width=canvas_width)
        
        def _on_mousewheel(event):
            # Handle mousewheel scrolling
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            return "break"
        
        # Bind events
        form.bind("<Configure>", _on_form_configure)
        canvas.bind("<Configure>", _on_canvas_configure)
        canvas.bind("<MouseWheel>", _on_mousewheel)
        
        # Bind mousewheel to canvas and form children
        def bind_mousewheel(widget):
            widget.bind("<MouseWheel>", _on_mousewheel)
            for child in widget.winfo_children():
                bind_mousewheel(child)
        
        bind_mousewheel(canvas)
        
        # Store canvas reference for theme updates
        if not hasattr(self, '_scroll_canvases'):
            self._scroll_canvases = []
        self._scroll_canvases.append(canvas)

        # Column weight strategy: better balanced distribution
        # Labels: compact, Input fields: expand proportionally
        form.columnconfigure(0, weight=0, minsize=100)  # Label column 1
        form.columnconfigure(1, weight=1, minsize=180)  # Input column 1
        form.columnconfigure(2, weight=0, minsize=100)  # Label column 2
        form.columnconfigure(3, weight=2, minsize=200)  # Input column 2 (wider for description)
        form.columnconfigure(4, weight=0, minsize=0)    # Spacer (not used)
        
        # Configure form grid row weights for better vertical distribution
        form.rowconfigure(0, weight=0)  # Date/Description row
        form.rowconfigure(1, weight=0)  # Debit row
        form.rowconfigure(2, weight=0)  # Credit row
        form.rowconfigure(3, weight=0)  # Doc references row
        form.rowconfigure(4, weight=0)  # Source/Attachment row
        form.rowconfigure(5, weight=0)  # Memo row (fixed height)
        form.rowconfigure(6, weight=0)  # Options row
        form.rowconfigure(7, weight=0)  # Action buttons row
        
        # Date and Description row with better spacing
        ttk.Label(form, text="Date:").grid(row=0, column=0, sticky="w", padx=(4, 2), pady=(6, 4))
        # Date entry with picker and Today button
        date_frame = ttk.Frame(form, style="Techfix.Surface.TFrame")
        date_frame.grid(row=0, column=1, sticky="w", padx=2, pady=(6, 4))
        self.txn_date = ttk.Entry(date_frame, style="Techfix.TEntry", width=12)
        self.txn_date.pack(side=tk.LEFT, padx=(0, 2))
        ttk.Button(date_frame, text="📅", width=3, command=lambda: self._pick_txn_date(), style="Techfix.TButton").pack(side=tk.LEFT, padx=(2,0))
        ttk.Button(date_frame, text="Today", width=6, command=lambda: self._set_txn_date_today(), style="Techfix.TButton").pack(side=tk.LEFT, padx=(4,0))

        ttk.Label(form, text="Description:").grid(row=0, column=2, sticky="w", padx=(8, 2), pady=(6, 4))
        self.txn_desc = ttk.Entry(form, style="Techfix.TEntry")
        self.txn_desc.grid(row=0, column=3, columnspan=2, sticky="we", padx=2, pady=(6, 4))

        accounts = db.get_accounts()
        account_names = [f"{a['code']} - {a['name']}" for a in accounts]
        self.account_id_by_display = {f"{a['code']} - {a['name']}": a["id"] for a in accounts}
        for a in accounts:
            try:
                self.account_id_by_display[str(a['code'])] = a['id']
            except Exception:
                pass
            try:
                self.account_id_by_display[str(a['name'])] = a['id']
                self.account_id_by_display[str(a['name']).lower()] = a['id']
            except Exception:
                pass

        # Debit line with better spacing
        ttk.Label(form, text="Debit Account:").grid(row=1, column=0, sticky="w", padx=(4, 2), pady=4)
        self.debit_acct_var = tk.StringVar(value="")
        self.debit_acct = ttk.Combobox(form, values=[''] + account_names, textvariable=self.debit_acct_var, style="Techfix.TCombobox")
        self.debit_acct.grid(row=1, column=1, sticky="we", padx=2, pady=4)
        try:
            self.debit_acct.set('')
        except Exception:
            pass
        try:
            self.debit_acct.bind('<<ComboboxSelected>>', lambda e: self._on_account_changed('debit'))
        except Exception:
            pass
        
        ttk.Label(form, text="Amount:").grid(row=1, column=2, sticky="e", padx=(8, 2), pady=4)
        self.debit_amt = ttk.Entry(form, style="Techfix.TEntry")
        self.debit_amt.grid(row=1, column=3, sticky="we", padx=2, pady=4)
        try:
            self.debit_amt.bind('<KeyRelease>', lambda e: self._update_post_buttons_enabled())
        except Exception:
            pass

        # Credit line with better spacing
        ttk.Label(form, text="Credit Account:").grid(row=2, column=0, sticky="w", padx=(4, 2), pady=4)
        self.credit_acct_var = tk.StringVar(value="")
        self.credit_acct = ttk.Combobox(form, values=[''] + account_names, textvariable=self.credit_acct_var, style="Techfix.TCombobox")
        self.credit_acct.grid(row=2, column=1, sticky="we", padx=2, pady=4)
        try:
            self.credit_acct.set('')
        except Exception:
            pass
        try:
            self.credit_acct.bind('<<ComboboxSelected>>', lambda e: self._on_account_changed('credit'))
        except Exception:
            pass
        
        ttk.Label(form, text="Amount:").grid(row=2, column=2, sticky="e", padx=(8, 2), pady=4)
        self.credit_amt = ttk.Entry(form, style="Techfix.TEntry")
        self.credit_amt.grid(row=2, column=3, sticky="we", padx=2, pady=4)
        try:
            self.credit_amt.bind('<KeyRelease>', lambda e: self._update_post_buttons_enabled())
        except Exception:
            pass

        # Document references row with better spacing
        ttk.Label(form, text="Doc #:").grid(row=3, column=0, sticky="w", padx=(4, 2), pady=4)
        self.txn_doc_ref = ttk.Entry(form, style="Techfix.TEntry")
        self.txn_doc_ref.grid(row=3, column=1, sticky="we", padx=2, pady=4)

        ttk.Label(form, text="External Ref:").grid(row=3, column=2, sticky="e", padx=(8, 2), pady=4)
        self.txn_external_ref = ttk.Entry(form, style="Techfix.TEntry")
        self.txn_external_ref.grid(row=3, column=3, sticky="we", padx=2, pady=4)

        # Source type row
        ttk.Label(form, text="Source:").grid(row=4, column=0, sticky="w", padx=(4, 2), pady=(6, 4))
        self.txn_source_type = ttk.Combobox(
            form,
            values=["", "Invoice", "Receipt", "Bank", "Adjust", "Payroll", "Other"],
            state="readonly",
            style="Techfix.TCombobox",
            width=15
        )
        self.txn_source_type.set("")
        self.txn_source_type.grid(row=4, column=1, sticky="we", padx=2, pady=(6, 4))

        # Document attachment section - split into two rows for better layout
        ttk.Label(form, text="Document:").grid(row=4, column=2, sticky="e", padx=(8, 2), pady=(6, 4))
        self.txn_attachment_path = tk.StringVar(value="")
        
        # Create a frame for the entry and buttons
        attach_frame = ttk.Frame(form, style="Techfix.Surface.TFrame")
        attach_frame.grid(row=4, column=3, columnspan=2, sticky="we", padx=2, pady=(6, 4))
        
        # Configure the frame to handle resizing
        attach_frame.columnconfigure(0, weight=1)
        attach_frame.rowconfigure(0, weight=0)  # Entry row
        attach_frame.rowconfigure(1, weight=0)  # Scan buttons row
        attach_frame.rowconfigure(2, weight=0)  # Status label row - fixed height
        
        # First row: Entry field and Browse button
        entry_row = ttk.Frame(attach_frame, style="Techfix.Surface.TFrame")
        entry_row.grid(row=0, column=0, sticky="we")
        entry_row.columnconfigure(0, weight=1)
        
        self.txn_attachment_display = ttk.Entry(
            entry_row, 
            textvariable=self.txn_attachment_path, 
            state="readonly", 
            style="Techfix.TEntry"
        )
        self.txn_attachment_display.grid(row=0, column=0, sticky="we", padx=(0, 4))
        
        browse_btn = ttk.Button(
            entry_row, 
            text="Browse...", 
            command=self._browse_source_document, 
            style="Techfix.TButton",
            width=10
        )
        browse_btn.grid(row=0, column=1, sticky="e")
        
        # Second row: Scan buttons - use grid for better responsive layout
        scan_row = ttk.Frame(attach_frame, style="Techfix.Surface.TFrame")
        scan_row.grid(row=1, column=0, sticky="we", pady=(4, 0))
        
        # Configure scan_row to allow wrapping
        scan_row.columnconfigure(0, weight=1, minsize=0)
        scan_row.columnconfigure(1, weight=0)
        scan_row.columnconfigure(2, weight=0)
        scan_row.columnconfigure(3, weight=0)
        scan_row.rowconfigure(0, weight=0)
        scan_row.rowconfigure(1, weight=0)  # Second row for wrapping
        
        scan_btn = ttk.Button(
            scan_row,
            text="Scan",
            command=self._scan_source_document,
            style="Techfix.Theme.TButton",
            width=10
        )
        scan_btn.grid(row=0, column=0, padx=(0, 2), pady=2, sticky="w")
        
        scan_img_btn = ttk.Button(
            scan_row,
            text="Scan Image",
            command=self._scan_from_image_file,
            style="Techfix.Theme.TButton",
            width=12
        )
        scan_img_btn.grid(row=0, column=1, padx=(2, 4), pady=2, sticky="w")
        
        manual_entry_btn = ttk.Button(
            scan_row,
            text="Enter Manually",
            command=self._manual_data_entry,
            style="Techfix.Theme.TButton",
            width=14
        )
        manual_entry_btn.grid(row=0, column=2, padx=(4, 0), pady=2, sticky="w")
        
        # Store references for responsive layout
        self.scan_row = scan_row
        self.scan_btn = scan_btn
        self.scan_img_btn = scan_img_btn
        self.manual_entry_btn = manual_entry_btn
        
        # Status label below buttons - wrap text to prevent stretching
        self.txn_prefill_status = ttk.Label(
            attach_frame, 
            text="", 
            style="Techfix.TLabel",
            wraplength=400  # Prevent label from stretching too wide
        )
        self.txn_prefill_status.grid(row=2, column=0, sticky="w", pady=(4, 0))


        # Memo field with better spacing
        ttk.Label(form, text="Memo:").grid(row=5, column=0, sticky="nw", padx=(4, 2), pady=(6, 2))
        
        # Create a frame to hold the text widget and scrollbar
        memo_frame = ttk.Frame(form, style="Techfix.TFrame")
        memo_frame.grid(row=5, column=1, columnspan=4, sticky="nsew", padx=2, pady=(6, 2))
        
        # Configure grid weights for memo frame
        memo_frame.columnconfigure(0, weight=1)
        memo_frame.rowconfigure(0, weight=1)
        
        # Text widget with fixed height
        self.txn_memo = tk.Text(
            memo_frame,
            height=4,
            bg=self.palette["surface_bg"],
            fg=self.palette["text_primary"],
            bd=1,
            relief=tk.SOLID,
            highlightthickness=1,
            highlightbackground=self.palette.get("entry_border", "#d8dee9"),
            highlightcolor=self.palette.get("accent_color", "#2563eb"),
            wrap=tk.WORD,
            font=("Segoe UI", 9),
            padx=6,
            pady=4
        )
        self.txn_memo.grid(row=0, column=0, sticky="nsew")
        
        # Add scrollbar for memo field
        scrollbar = ttk.Scrollbar(
            memo_frame,
            orient=tk.VERTICAL,
            command=self.txn_memo.yview
        )
        scrollbar.grid(row=0, column=1, sticky="ns")
        self.txn_memo.config(yscrollcommand=scrollbar.set)

        # Options frame - all items in one horizontal row
        # Combined options and action buttons frame - aligned on same row
        combined_frame = ttk.Frame(form, style="Techfix.Surface.TFrame")
        combined_frame.grid(row=6, column=0, columnspan=5, sticky="ew", padx=2, pady=(6, 6))
        
        # Configure combined frame columns - options on left, buttons on right
        combined_frame.columnconfigure(0, weight=1)  # Options container
        combined_frame.columnconfigure(1, weight=0)  # Button container
        combined_frame.rowconfigure(0, weight=0)
        
        # Options container - all items in one horizontal row
        options_container = ttk.Frame(combined_frame, style="Techfix.Surface.TFrame")
        options_container.grid(row=0, column=0, sticky="w", padx=0, pady=2)
        
        # Configure options container columns for horizontal layout
        options_container.columnconfigure(0, weight=0)
        options_container.columnconfigure(1, weight=0)
        options_container.columnconfigure(2, weight=0)
        options_container.columnconfigure(3, weight=0)
        options_container.rowconfigure(0, weight=0)
        
        self.txn_is_adjust = tk.IntVar(value=0)
        self.adjust_cb = ttk.Checkbutton(
            options_container, 
            text="Adjusting Entry", 
            variable=self.txn_is_adjust, 
            style="Techfix.TCheckbutton"
        )
        self.adjust_cb.grid(row=0, column=0, padx=(0, 16), pady=2, sticky="w")

        self.txn_schedule_reverse = tk.IntVar(value=0)
        self.reverse_cb = ttk.Checkbutton(
            options_container,
            text="Schedule Reversal",
            variable=self.txn_schedule_reverse,
            style="Techfix.TCheckbutton",
            command=lambda: self._on_schedule_reverse_toggle()
        )
        self.reverse_cb.grid(row=0, column=1, padx=(0, 8), pady=2, sticky="w")

        self.reverse_label = ttk.Label(options_container, text="Reverse Date:")
        self.reverse_label.grid(row=0, column=2, padx=(0, 4), pady=2, sticky="w")
        self.txn_reverse_date = ttk.Entry(options_container, width=12, style="Techfix.TEntry")
        self.txn_reverse_date.grid(row=0, column=3, pady=2, sticky="w")
        
        # Store references for responsive layout
        self.options_container = options_container
        try:
            # Disabled until Schedule Reversal is checked
            self.txn_reverse_date.configure(state='disabled')
        except Exception:
            pass

        # Action buttons - responsive layout that adapts to available space
        action_frame = ttk.Frame(combined_frame, style="Techfix.Surface.TFrame")
        action_frame.grid(row=0, column=1, sticky="e", padx=0, pady=2)
        
        # Configure action frame - buttons container directly
        action_frame.columnconfigure(0, weight=0)
        action_frame.rowconfigure(0, weight=0)
        
        # Button container - configured for responsive wrapping to multiple rows
        btn_container = ttk.Frame(action_frame, style="Techfix.Surface.TFrame")
        btn_container.grid(row=0, column=0, sticky="e")
        
        # Configure button container for horizontal layout (single row)
        btn_container.columnconfigure(0, weight=0)
        btn_container.columnconfigure(1, weight=0)
        btn_container.columnconfigure(2, weight=0)
        btn_container.columnconfigure(3, weight=0)
        btn_container.rowconfigure(0, weight=0)
        
        # Common button style with fixed width
        button_style = {
            'style': 'Techfix.TButton',
            'width': 14,  # Slightly wider for better text fit
            'padding': (10, 4)  # More padding for better clickability
        }
        
        # Use grid for buttons to ensure they stay visible and don't stretch
        self.btn_recent = ttk.Button(btn_container, text="Recent Transactions", command=self._open_recent_transactions_window, **button_style)
        self.btn_recent.grid(row=0, column=0, padx=(0, 6), pady=2, sticky="")
        
        self.btn_clear = ttk.Button(btn_container, text="Clear Form", command=self._clear_transaction_form, **button_style)
        self.btn_clear.grid(row=0, column=1, padx=6, pady=2, sticky="")
        
        self.btn_draft = ttk.Button(btn_container, text="Save Draft", command=lambda: self._record_transaction("draft"), **button_style)
        self.btn_draft.grid(row=0, column=2, padx=6, pady=2, sticky="")
        
        self.btn_post = ttk.Button(btn_container, text="Record & Post", command=lambda: self._record_transaction("posted"), **button_style)
        self.btn_post.grid(row=0, column=3, padx=(6, 0), pady=2, sticky="")
        
        # Store reference for responsive layout updates
        self.action_frame = action_frame
        self.btn_container = btn_container
        try:
            self._update_post_buttons_enabled()
        except Exception:
            pass


        # Keyboard shortcuts in the Transactions tab
        try:
            # Ctrl+Enter posts, Ctrl+Shift+Enter saves draft (when focus is within txn form)
            form.bind_all("<Control-Return>", lambda e: self._record_transaction("posted"))
            form.bind_all("<Control-KP_Enter>", lambda e: self._record_transaction("posted"))
            form.bind_all("<Control-Shift-Return>", lambda e: self._record_transaction("draft"))
            form.bind_all("<Control-Shift-KP_Enter>", lambda e: self._record_transaction("draft"))
        except Exception:
            pass

    def _open_recent_transactions_window(self) -> None:
        """Open Recent Transactions in a separate window."""
        try:
            # Check if window already exists
            if hasattr(self, '_recent_txn_window') and self._recent_txn_window.winfo_exists():
                try:
                    self._recent_txn_window.lift()
                    self._recent_txn_window.focus()
                    return
                except Exception:
                    pass
            
            # Create new window
            window = tk.Toplevel(self)
            window.title("Recent Transactions")
            window.geometry("1000x600")
            window.configure(bg=self.palette["surface_bg"])
            
            # Make window appear on top and focus it
            window.lift()
            window.focus_force()
            window.attributes('-topmost', True)
            window.after_idle(lambda: window.attributes('-topmost', False))
            
            # Store reference
            self._recent_txn_window = window
            
            # Main container
            main_frame = ttk.Frame(window, style="Techfix.Surface.TFrame")
            main_frame.pack(fill=tk.BOTH, expand=True, padx=12, pady=12)
            main_frame.columnconfigure(0, weight=1)
            main_frame.rowconfigure(1, weight=1)
            
            # Controls frame
            controls = ttk.Frame(main_frame, style="Techfix.Surface.TFrame")
            controls.grid(row=0, column=0, sticky="ew", padx=0, pady=(0, 6))
            controls.columnconfigure(0, weight=1)  # Spacer
            
            # Left side buttons
            left_buttons = ttk.Frame(controls, style="Techfix.Surface.TFrame")
            left_buttons.pack(side=tk.LEFT)
            def refresh_recent_transactions():
                loading = self.show_loading("Please wait... Loading transactions...", window)
                try:
                    window.update()  # Update UI to show loading indicator
                    self._load_recent_transactions_window()
                finally:
                    if loading:
                        self.hide_loading(loading)
            
            ttk.Button(left_buttons, text="Refresh", command=refresh_recent_transactions, style="Techfix.TButton").pack(side=tk.LEFT, padx=(0, 6))
            # Disable delete button for viewers
            delete_btn_state = 'normal' if self._has_permission('edit') else 'disabled'
            delete_btn = ttk.Button(left_buttons, text="Delete Selected", command=self._delete_selected_transaction_window, style="Techfix.Danger.TButton", state=delete_btn_state)
            delete_btn.pack(side=tk.LEFT)
            
            # Close button on the right
            def on_close():
                try:
                    if hasattr(self, '_recent_txn_window'):
                        del self._recent_txn_window
                except Exception:
                    pass
                window.destroy()
            
            ttk.Button(controls, text="Close", command=on_close, style="Techfix.TButton").pack(side=tk.RIGHT)
            
            # Tree frame
            tree_frame = ttk.Frame(main_frame, style="Techfix.Surface.TFrame")
            tree_frame.grid(row=1, column=0, sticky="nsew", padx=0, pady=0)
            tree_frame.columnconfigure(0, weight=1)
            tree_frame.rowconfigure(0, weight=1)
            
            # Create treeview
            cols = ("date", "reference", "description", "debit", "credit", "account")
            txn_tree = ttk.Treeview(
                tree_frame,
                columns=cols,
                show="headings",
                style="Techfix.Treeview",
                selectmode="browse",
            )
            
            # Headings
            txn_tree.heading("date", text="Date")
            txn_tree.heading("reference", text="Ref")
            txn_tree.heading("description", text="Description")
            txn_tree.heading("debit", text="Debit")
            txn_tree.heading("credit", text="Credit")
            txn_tree.heading("account", text="Account")
            
            # Column sizing - description expands
            txn_tree.column("date", width=100, anchor=tk.W, stretch=False)
            txn_tree.column("reference", width=80, anchor=tk.W, stretch=False)
            txn_tree.column("description", width=400, anchor=tk.W, stretch=True)
            txn_tree.column("debit", width=100, anchor=tk.E, stretch=False)
            txn_tree.column("credit", width=100, anchor=tk.E, stretch=False)
            txn_tree.column("account", width=200, anchor=tk.W, stretch=False)
            
            # Scrollbars
            vsb = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=txn_tree.yview)
            hsb = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=txn_tree.xview)
            txn_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
            
            # Layout
            txn_tree.grid(row=0, column=0, sticky="nsew")
            vsb.grid(row=0, column=1, sticky="ns")
            hsb.grid(row=1, column=0, columnspan=2, sticky="ew")
            
            # Store tree reference for this window
            window.txn_tree = txn_tree
            
            # Load transactions
            self._load_recent_transactions_window(tree=txn_tree)
            
            # Handle window close via X button
            window.protocol("WM_DELETE_WINDOW", on_close)
            
        except Exception as e:
            try:
                messagebox.showerror("Error", f"Failed to open Recent Transactions window: {e}")
            except Exception:
                pass

    def _load_recent_transactions_window(self, tree=None, limit: int = 50) -> None:
        """Load recent transactions into the window's treeview."""
        try:
            if tree is None:
                # Try to get tree from window
                if (hasattr(self, '_recent_txn_window') and 
                    self._recent_txn_window.winfo_exists() and 
                    hasattr(self._recent_txn_window, 'txn_tree')):
                    tree = self._recent_txn_window.txn_tree
                else:
                    return
            
            # Clear existing items
            for it in tree.get_children():
                tree.delete(it)
            
            conn = self.engine.conn
            pid = self.engine.current_period_id
            clause = "WHERE je.period_id = ?" if pid else ""
            params = ([int(pid)] if pid else []) + [limit]
            sql = (
                """
                SELECT je.id AS entry_id, je.date AS date, je.description AS description, je.status AS status
                FROM journal_entries je
                """
                + (clause or "")
                + """
                ORDER BY date(je.date) DESC, je.id DESC
                LIMIT ?
                """
            )
            cur = conn.execute(sql, params)
            entries = cur.fetchall()
            for e in entries:
                eid = e['entry_id'] if 'entry_id' in e.keys() else e['id']
                date = e['date'] if 'date' in e.keys() else ''
                desc = e['description'] if 'description' in e.keys() else ''
                # Aggregate debit/credit and pick a representative account name (first line)
                try:
                    rsum = conn.execute(
                        """
                        SELECT COALESCE(SUM(jl.debit),0) AS debit, COALESCE(SUM(jl.credit),0) AS credit
                        FROM journal_lines jl
                        WHERE jl.entry_id=?
                        """,
                        (eid,)
                    ).fetchone()
                    debit = float(rsum['debit'] if 'debit' in rsum.keys() else 0) if rsum else 0.0
                    credit = float(rsum['credit'] if 'credit' in rsum.keys() else 0) if rsum else 0.0
                    racct = conn.execute(
                        """
                        SELECT a.name
                        FROM journal_lines jl
                        JOIN accounts a ON a.id = jl.account_id
                        WHERE jl.entry_id=?
                        ORDER BY jl.id
                        LIMIT 1
                        """,
                        (eid,)
                    ).fetchone()
                    acct = (racct['name'] if racct and 'name' in racct.keys() else '')
                except Exception:
                    debit = credit = 0.0
                    acct = ''
                
                # Get reference
                ref = conn.execute("SELECT document_ref FROM journal_entries WHERE id=?", (eid,)).fetchone()
                ref_str = (ref['document_ref'] if ref and 'document_ref' in ref.keys() else '') or ''
                
                # Format amounts
                debit_str = f"{debit:,.2f}" if debit > 0 else ""
                credit_str = f"{credit:,.2f}" if credit > 0 else ""
                
                tree.insert("", "end", iid=f"entry_{eid}", values=(date, ref_str, desc, debit_str, credit_str, acct))
        except Exception as e:
            import traceback
            logging.error(f"Error loading recent transactions: {e}")
            logging.error(traceback.format_exc())

    def _delete_selected_transaction_window(self) -> None:
        """Delete selected transaction from the window."""
        # Check permission - viewers cannot delete transactions
        if not self._has_permission('edit'):
            messagebox.showerror(
                'Access Denied',
                'You do not have permission to delete transactions. '
                'Only users with edit permission (Accountant, Manager, Admin) can delete transactions.'
            )
            return
        
        try:
            # Check if window exists and is still valid
            if not hasattr(self, '_recent_txn_window'):
                messagebox.showwarning("Delete", "Recent Transactions window is not open.")
                return
            
            window = self._recent_txn_window
            try:
                # Check if window still exists
                if not window.winfo_exists():
                    messagebox.showwarning("Delete", "Recent Transactions window has been closed.")
                    delattr(self, '_recent_txn_window')
                    return
            except tk.TclError:
                messagebox.showwarning("Delete", "Recent Transactions window has been closed.")
                if hasattr(self, '_recent_txn_window'):
                    delattr(self, '_recent_txn_window')
                return
            
            if not hasattr(window, 'txn_tree'):
                messagebox.showerror("Delete", "Transaction list not found in window.")
                return
            
            tree = window.txn_tree
            sel = tree.selection()
            if not sel:
                messagebox.showinfo("Delete", "Please select a transaction to delete.")
                return
            
            item = sel[0]
            entry_id = None
            try:
                # Extract entry_id from iid (item itself is the iid in Treeview)
                # The iid format is "entry_{entry_id}" as set in _load_recent_transactions_window
                if item and isinstance(item, str) and item.startswith('entry_'):
                    entry_id = int(item.split('_')[1])
                else:
                    # Fallback: try to get iid from tree item
                    try:
                        iid = tree.item(item, 'iid')
                        if iid and iid.startswith('entry_'):
                            entry_id = int(iid.split('_')[1])
                    except Exception:
                        pass
                    
                    # If still no entry_id, try to find by matching description and date
                    if not entry_id:
                        vals = tree.item(item, 'values')
                        if len(vals) >= 3:
                            date = vals[0] if vals[0] else ''
                            desc = vals[2] if vals[2] else ''
                            if date and desc:
                                conn = self.engine.conn
                                cur = conn.execute(
                                    "SELECT id FROM journal_entries WHERE date=? AND description=? LIMIT 1",
                                    (date, desc)
                                )
                                result = cur.fetchone()
                                if result:
                                    entry_id = result['id']
            except Exception as e:
                logger.exception("Error extracting entry_id")
                messagebox.showerror("Delete", f"Could not determine the selected transaction ID: {e}")
                return
            
            if not entry_id:
                messagebox.showerror("Delete", "Could not determine the selected transaction ID.")
                return
            
            # Confirm deletion
            if not messagebox.askyesno("Confirm Delete", f"Delete transaction #{entry_id}?\n\nThis action cannot be undone."):
                return
            
            # Delete from database
            conn = self.engine.conn
            try:
                # Delete journal entry (journal_lines will cascade delete due to foreign key)
                conn.execute("DELETE FROM journal_entries WHERE id=?", (entry_id,))
                conn.commit()
                messagebox.showinfo("Success", f"Transaction #{entry_id} has been deleted.")
            except Exception as e:
                logger.exception("Error deleting transaction")
                messagebox.showerror("Delete Failed", f"Error deleting transaction: {str(e)}")
                conn.rollback()
                return
            
            # Reload transactions in window
            try:
                self._load_recent_transactions_window(tree=tree)
            except Exception as e:
                logger.exception("Error reloading transactions window")
            
            # Refresh other views that might show this transaction
            try:
                self._load_journal_entries()
            except Exception:
                pass
            try:
                self._load_recent_transactions()
            except Exception:
                pass
            try:
                self._refresh_cycle_and_views()
            except Exception:
                pass
                
        except Exception as e:
            logger.exception("Unexpected error in _delete_selected_transaction_window")
            messagebox.showerror("Error", f"An unexpected error occurred: {str(e)}")

    def _delete_selected_transaction(self) -> None:
        # Check permission - viewers cannot delete transactions
        if not self._has_permission('edit'):
            messagebox.showerror(
                'Access Denied',
                'You do not have permission to delete transactions. '
                'Only users with edit permission (Accountant, Manager, Admin) can delete transactions.'
            )
            return
        
        try:
            if not hasattr(self, 'txn_recent_tree'):
                return
            sel = self.txn_recent_tree.selection()
            if not sel:
                messagebox.showinfo("Delete", "Select a transaction to delete.")
                return
            item = sel[0]
            vals = self.txn_recent_tree.item(item, 'values')
            try:
                entry_id = int(vals[1]) if len(vals) > 1 and vals[1] else None
            except Exception:
                entry_id = None
            if not entry_id:
                messagebox.showerror("Delete", "Could not determine the selected transaction ID.")
                return
            if not messagebox.askyesno("Confirm Delete", f"Delete transaction #{entry_id}? This cannot be undone."):
                return
            conn = self.engine.conn
            try:
                conn.execute("DELETE FROM journal_entries WHERE id=?", (entry_id,))
                conn.commit()
            except Exception as e:
                messagebox.showerror("Delete Failed", f"Error deleting transaction: {e}")
                return
            try:
                self._load_recent_transactions()
            except Exception:
                pass
            try:
                self._load_journal_entries()
            except Exception:
                pass
        except Exception:
            pass


    def _build_adjust_tab(self) -> None:
        frame = self.tab_adjust

        # Root content area for the Adjustments tab
        content = ttk.Frame(frame, style="Techfix.Surface.TFrame")
        # Outer padding kept modest so controls stay large but avoid huge empty borders
        content.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
        content.grid_rowconfigure(0, weight=1)
        # Give the requests panel a little more width than the common adjustments panel
        content.grid_columnconfigure(0, weight=1)
        content.grid_columnconfigure(1, weight=2)

        # Common Adjustments section (left)
        f = ttk.Labelframe(
            content,
            text="Common Adjustments",
            style="Techfix.TLabelframe",
            padding=(10, 8)
        )
        # A bit of horizontal space between the two main sections
        f.grid(row=0, column=0, sticky="nsew", padx=(0, 8), pady=(0, 0))

        # Configure columns - make entry fields expand
        f.columnconfigure(0, weight=0)  # Labels
        f.columnconfigure(1, weight=3)  # Entry fields (3x weight)
        f.columnconfigure(2, weight=1)  # Buttons (1x weight)

        # Configure rows with a little vertical breathing room
        for i in range(4):
            f.rowconfigure(i, weight=0, pad=4)

        # Date input row - using grid for better control
        ttk.Label(f, text="Date:").grid(row=0, column=0, sticky="e", padx=4, pady=4)
        self.adjust_date = ttk.Entry(f, style="Techfix.TEntry")
        self.adjust_date.grid(row=0, column=1, sticky="we", padx=4, pady=4)
        btns = ttk.Frame(f, style="Techfix.Surface.TFrame")
        btns.grid(row=0, column=2, sticky="ew", padx=4, pady=4)
        ttk.Button(btns, text="📅", command=self._pick_adjust_date, style="Techfix.TButton", width=3).pack(side=tk.LEFT, padx=(0, 4))
        ttk.Button(btns, text="Today", command=lambda: self._set_entry_today(self.adjust_date), style="Techfix.TButton").pack(side=tk.LEFT)

        # Adjustment controls in a grid
        row = 1

        # Row 1: Supplies
        ttk.Label(f, text="Supplies:").grid(row=row, column=0, sticky="e", padx=4, pady=4)
        self.sup_remaining = ttk.Entry(f, style="Techfix.TEntry")
        self.sup_remaining.grid(row=row, column=1, sticky="we", padx=4, pady=4)
        ttk.Button(f, text="Adjust", command=self._do_adjust_supplies,
                  style="Techfix.TButton").grid(row=row, column=2, padx=4, pady=4, sticky="ew")

        # Row 2: Prepaid Rent
        row += 1
        ttk.Label(f, text="Prepaid Rent:").grid(row=row, column=0, sticky="e", padx=4, pady=4)
        self.prepaid_amt = ttk.Entry(f, style="Techfix.TEntry")
        self.prepaid_amt.grid(row=row, column=1, sticky="we", padx=4, pady=4)
        ttk.Button(f, text="Amortize", command=self._do_amortize_prepaid,
                  style="Techfix.TButton").grid(row=row, column=2, padx=4, pady=4, sticky="ew")

        # Row 3: Depreciation
        row += 1
        ttk.Label(f, text="Depreciation:").grid(row=row, column=0, sticky="e", padx=4, pady=4)
        self.depr_amt = ttk.Entry(f, style="Techfix.TEntry")
        self.depr_amt.grid(row=row, column=1, sticky="we", padx=4, pady=4)
        ttk.Button(f, text="Calculate", command=self._do_depreciate,
                  style="Techfix.TButton").grid(row=row, column=2, padx=4, pady=4, sticky="ew")

        # Configure column weights for the labelframe
        f.columnconfigure(1, weight=1)  # Entry fields expand

        # Adjustment Requests section (right)
        queue = ttk.Labelframe(
            content,
            text="Adjustment Requests & Approvals",
            style="Techfix.TLabelframe",
            padding=(10, 8)
        )
        queue.grid(row=0, column=1, sticky="nsew", padx=(8, 0), pady=(0, 0))

        # Input form frame
        queue_inputs = ttk.Frame(queue, style="Techfix.Surface.TFrame")
        queue_inputs.pack(fill=tk.X, padx=4, pady=(0, 6))

        # Description row
        desc_frame = ttk.Frame(queue_inputs, style="Techfix.Surface.TFrame")
        desc_frame.pack(fill=tk.X, pady=(0, 6))
        ttk.Label(desc_frame, text="Description:").pack(side=tk.LEFT, padx=(0, 6))
        self.adjust_desc = ttk.Entry(desc_frame, style="Techfix.TEntry")
        self.adjust_desc.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        ttk.Label(desc_frame, text="By:").pack(side=tk.LEFT, padx=(0, 6))
        self.adjust_requested_by = ttk.Entry(desc_frame, style="Techfix.TEntry", width=15)
        self.adjust_requested_by.pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(desc_frame, text="Add Request", command=self._queue_adjustment_request,
                  style="Techfix.TButton").pack(side=tk.RIGHT)

        # Notes row
        notes_frame = ttk.Frame(queue_inputs, style="Techfix.Surface.TFrame")
        notes_frame.pack(fill=tk.X, pady=(0, 4))
        ttk.Label(notes_frame, text="Notes:").pack(side=tk.LEFT, padx=(0, 6))
        self.adjust_notes = ttk.Entry(notes_frame, style="Techfix.TEntry")
        self.adjust_notes.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 6))

        # Treeview frame (use grid so scrollbars attach reliably)
        tree_frame = ttk.Frame(queue, style="Techfix.Surface.TFrame")
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)

        # Treeview for adjustments
        cols = ("id", "description", "requested_on", "status", "notes")
        self.adjust_tree = ttk.Treeview(
            tree_frame,
            columns=cols,
            show="headings",
            style="Techfix.Treeview",
            # Slightly shorter so bottom buttons remain visible on smaller screens
            height=10,
            selectmode="extended"
        )

        # Configure columns with better distribution and stretching
        col_config = {
            "id": {"width": 40, "minwidth": 40, "stretch": False, "anchor": "center"},
            "description": {"width": 300, "minwidth": 150, "stretch": True, "anchor": "w"},
            "requested_on": {"width": 150, "minwidth": 100, "stretch": False, "anchor": "center"},
            "status": {"width": 100, "minwidth": 80, "stretch": False, "anchor": "center"},
            "notes": {"width": 300, "minwidth": 150, "stretch": True, "anchor": "w"}
        }

        # Configure columns
        for c in cols:
            self.adjust_tree.heading(c, text=c.replace("_", " ").title(), anchor="center")
            self.adjust_tree.column(
                c,
                width=col_config[c]["width"],
                minwidth=col_config[c]["minwidth"],
                stretch=col_config[c]["stretch"],
                anchor=col_config[c]["anchor"]
            )

        # Add scrollbars and grid them so layout is stable
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.adjust_tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.adjust_tree.xview)
        self.adjust_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        # grid layout for tree + scrollbars
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        self.adjust_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, columnspan=2, sticky="ew")

        # Ensure mouse wheel scrolls the treeview
        try:
            self.adjust_tree.bind("<MouseWheel>", lambda e: self.adjust_tree.yview_scroll(int(-1 * (e.delta / 120)), "units"))
        except Exception:
            pass

        # Action buttons frame
        btn_frame = ttk.Frame(queue, style="Techfix.Surface.TFrame")
        btn_frame.pack(fill=tk.X, padx=4, pady=(4, 0))

        # Add buttons with pack
        buttons = [
            ("Refresh", self._load_adjustments),
            ("Approve", lambda: self._mark_adjustment_status("approved")),
            ("Post", lambda: self._mark_adjustment_status("posted")),
            ("Draft", lambda: self._mark_adjustment_status("draft"))
        ]

        for text, cmd in buttons:
            btn = ttk.Button(btn_frame, text=text, command=cmd, style="Techfix.TButton")
            btn.pack(side=tk.LEFT, padx=2, pady=2, expand=True, fill=tk.X)

    # Adjustment action handlers
    def _do_adjust_supplies(self) -> None:
        """Create an adjusting journal entry to record supplies used."""
        amt_text = (self.sup_remaining.get() if hasattr(self, 'sup_remaining') else '').strip()
        if not amt_text:
            messagebox.showinfo("No Amount", "Please enter the supplies amount to adjust.")
            return
        try:
            amt = float(amt_text)
        except Exception:
            messagebox.showerror("Invalid Amount", "Please enter a numeric amount for supplies.")
            return

        # Find account ids
        supplies = db.get_account_by_name('Supplies')
        supplies_exp = db.get_account_by_name('Supplies Expense')
        if not supplies or not supplies_exp:
            messagebox.showerror("Missing Accounts", "Required accounts not found: Supplies / Supplies Expense")
            return

        try:
            date_str = (self.adjust_date.get().strip() if hasattr(self, 'adjust_date') else '') or datetime.utcnow().date().isoformat()
            entry_id = db.insert_journal_entry(
                date=date_str,
                description=f"Adjust supplies: used {amt:.2f}",
                lines=[(supplies_exp['id'], amt, 0.0), (supplies['id'], 0.0, amt)],
                is_adjusting=1,
                conn=self.engine.conn,
            )
            messagebox.showinfo("Adjusted", f"Created adjusting entry {entry_id} for supplies ({amt:.2f})")
            self._refresh_after_post()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to create adjusting entry: {e}")

    def _do_amortize_prepaid(self) -> None:
        """Amortize prepaid rent into rent expense."""
        amt_text = (self.prepaid_amt.get() if hasattr(self, 'prepaid_amt') else '').strip()
        if not amt_text:
            messagebox.showinfo("No Amount", "Please enter the amount to amortize.")
            return
        try:
            amt = float(amt_text)
        except Exception:
            messagebox.showerror("Invalid Amount", "Please enter a numeric amount for amortization.")
            return

        prepaid = db.get_account_by_name('Prepaid Rent')
        rent_exp = db.get_account_by_name('Rent Expense')
        if not prepaid or not rent_exp:
            messagebox.showerror("Missing Accounts", "Required accounts not found: Prepaid Rent / Rent Expense")
            return

        try:
            date_str = (self.adjust_date.get().strip() if hasattr(self, 'adjust_date') else '') or datetime.utcnow().date().isoformat()
            entry_id = db.insert_journal_entry(
                date=date_str,
                description=f"Amortize prepaid rent: {amt:.2f}",
                lines=[(rent_exp['id'], amt, 0.0), (prepaid['id'], 0.0, amt)],
                is_adjusting=1,
                conn=self.engine.conn,
            )
            messagebox.showinfo("Amortized", f"Created amortization entry {entry_id} ({amt:.2f})")
            self._refresh_after_post()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to create amortization entry: {e}")

    def _do_depreciate(self) -> None:
        """Record depreciation expense for equipment."""
        amt_text = (self.depr_amt.get() if hasattr(self, 'depr_amt') else '').strip()
        if not amt_text:
            messagebox.showinfo("No Amount", "Please enter the depreciation amount.")
            return
        try:
            amt = float(amt_text)
        except Exception:
            messagebox.showerror("Invalid Amount", "Please enter a numeric amount for depreciation.")
            return

        depr_exp = db.get_account_by_name('Depreciation Expense')
        acc_depr = db.get_account_by_name('Accumulated Depreciation - Equipment')
        if not depr_exp or not acc_depr:
            messagebox.showerror("Missing Accounts", "Required accounts not found: Depreciation Expense / Accumulated Depreciation - Equipment")
            return

        try:
            date_str = (self.adjust_date.get().strip() if hasattr(self, 'adjust_date') else '') or datetime.utcnow().date().isoformat()
            entry_id = db.insert_journal_entry(
                date=date_str,
                description=f"Record depreciation: {amt:.2f}",
                lines=[(depr_exp['id'], amt, 0.0), (acc_depr['id'], 0.0, amt)],
                is_adjusting=1,
                conn=self.engine.conn,
            )
            messagebox.showinfo("Depreciated", f"Created depreciation entry {entry_id} ({amt:.2f})")
            self._refresh_after_post()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to create depreciation entry: {e}")

    def _pick_adjust_date(self) -> None:
        try:
            class DatePicker(tk.Toplevel):
                def __init__(self, parent, callback=None, initial_date=None):
                    super().__init__(parent)
                    self.transient(parent)
                    self.grab_set()
                    self.callback = callback
                    self.title("Pick Date")
                    self.resizable(False, False)

                    today = datetime.now().date()
                    if initial_date:
                        try:
                            dt = datetime.strptime(initial_date, '%Y-%m-%d').date()
                        except Exception:
                            dt = today
                    else:
                        dt = today

                    self.year = dt.year
                    self.month = dt.month

                    header = ttk.Frame(self)
                    header.pack(fill=tk.X, padx=8, pady=6)
                    ttk.Button(header, text="◀", width=3, command=self._prev_month).pack(side=tk.LEFT)
                    self.title_lbl = ttk.Label(header, text="", anchor=tk.CENTER)
                    self.title_lbl.pack(side=tk.LEFT, expand=True)
                    ttk.Button(header, text="▶", width=3, command=self._next_month).pack(side=tk.RIGHT)

                    self.cal_frame = ttk.Frame(self)
                    self.cal_frame.pack(padx=8, pady=(0,8))
                    self._draw_calendar()

                def _draw_calendar(self):
                    for child in self.cal_frame.winfo_children():
                        child.destroy()
                    self.title_lbl.config(text=f"{calendar.month_name[self.month]} {self.year}")
                    wdays = ['Mo','Tu','We','Th','Fr','Sa','Su']
                    for c, wd in enumerate(wdays):
                        ttk.Label(self.cal_frame, text=wd, width=3, anchor='center').grid(row=0, column=c)
                    m = calendar.monthcalendar(self.year, self.month)
                    for r, week in enumerate(m, start=1):
                        for c, day in enumerate(week):
                            if day == 0:
                                ttk.Label(self.cal_frame, text='', width=3).grid(row=r, column=c, padx=1, pady=1)
                            else:
                                b = ttk.Button(self.cal_frame, text=str(day), width=3, command=lambda d=day: self._select_day(d))
                                b.grid(row=r, column=c, padx=1, pady=1)

                def _prev_month(self):
                    self.month -= 1
                    if self.month < 1:
                        self.month = 12
                        self.year -= 1
                    self._draw_calendar()

                def _next_month(self):
                    self.month += 1
                    if self.month > 12:
                        self.month = 1
                        self.year += 1
                    self._draw_calendar()

                def _select_day(self, day):
                    try:
                        d = datetime(self.year, self.month, day).strftime('%Y-%m-%d')
                        if callable(self.callback):
                            self.callback(d)
                    except Exception:
                        pass
                    finally:
                        try:
                            self.grab_release()
                        except Exception:
                            pass
                        self.destroy()

            def _on_pick(d):
                try:
                    self.adjust_date.delete(0, tk.END)
                    self.adjust_date.insert(0, d)
                except Exception:
                    pass

            cur = None
            try:
                cur = self.adjust_date.get().strip()
            except Exception:
                cur = None
            dp = DatePicker(self, callback=_on_pick, initial_date=cur)
            self.wait_window(dp)
        except Exception:
            pass

    def _set_entry_today(self, entry: ttk.Entry) -> None:
        try:
            entry.delete(0, tk.END)
            entry.insert(0, datetime.utcnow().date().isoformat())
        except Exception:
            pass

    def _queue_adjustment_request(self) -> None:
        """Add an adjustment request to the queue."""
        desc = (self.adjust_desc.get() if hasattr(self, 'adjust_desc') else '').strip()
        by = (self.adjust_requested_by.get() if hasattr(self, 'adjust_requested_by') else '').strip()
        notes = (self.adjust_notes.get() if hasattr(self, 'adjust_notes') else '').strip()
        if not desc:
            messagebox.showinfo("Missing Description", "Please enter a description for the adjustment request.")
            return
        try:
            aid = db.create_adjustment_request(self.current_period_id or db.get_current_period()['id'], desc, requested_by=by or None, notes=notes or None, conn=self.engine.conn)
            messagebox.showinfo("Requested", f"Adjustment request {aid} created.")
            self._load_adjustments()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to create adjustment request: {e}")

    def _load_adjustments(self) -> None:
        """Load adjustment requests into the treeview."""
        if hasattr(self, 'adjust_tree'):
            for item in self.adjust_tree.get_children():
                self.adjust_tree.delete(item)
        try:
            rows = db.list_adjustment_requests(self.current_period_id or db.get_current_period()['id'], conn=self.engine.conn)
            for r in rows:
                self.adjust_tree.insert('', 'end', values=(r['id'], r['description'], r['requested_on'], r['status'], r['notes'] or ''))
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load adjustments: {e}")

    def _mark_adjustment_status(self, status: str) -> None:
        """Update status of selected adjustment requests."""
        sel = self.adjust_tree.selection()
        if not sel:
            messagebox.showinfo("Select Request", "Please select one or more adjustment requests to update.")
            return
        try:
            for item in sel:
                vals = self.adjust_tree.item(item, 'values')
                adj_id = int(vals[0])
                db.update_adjustment_status(adj_id, status, conn=self.engine.conn)
            messagebox.showinfo("Updated", f"Updated {len(sel)} request(s) to '{status}'.")
            self._load_adjustments()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update adjustment(s): {e}")

    def _build_journal_tab(self) -> None:
        """Build the journal entries tab with a list of all journal entries."""
        frame = self.tab_journal

        # Create a frame for the toolbar - use grid for responsive layout
        toolbar = ttk.Frame(frame, style="Techfix.Surface.TFrame")
        toolbar.pack(fill=tk.X, padx=4, pady=4)
        
        # Configure toolbar grid for responsive wrapping
        toolbar.columnconfigure(0, weight=0)  # Left buttons
        toolbar.columnconfigure(1, weight=1)  # Spacer
        toolbar.columnconfigure(2, weight=0)  # Right filters
        toolbar.rowconfigure(0, weight=0)
        toolbar.rowconfigure(1, weight=0)  # Second row for wrapping

        # Left side buttons container
        left_buttons = ttk.Frame(toolbar, style="Techfix.Surface.TFrame")
        left_buttons.grid(row=0, column=0, sticky="w", padx=2)
        left_buttons.columnconfigure(0, weight=0)
        left_buttons.columnconfigure(1, weight=0)
        left_buttons.columnconfigure(2, weight=0)
        left_buttons.columnconfigure(3, weight=0)
        left_buttons.rowconfigure(0, weight=0)
        left_buttons.rowconfigure(1, weight=0)  # Second row for wrapping

        # Add refresh button
        refresh_btn = ttk.Button(
            left_buttons,
            text="Refresh",
            command=self._load_journal_entries,
            style="Techfix.TButton",
        )
        refresh_btn.grid(row=0, column=0, padx=2, pady=2, sticky="w")
        
        # Export to Excel button for journal
        export_btn = ttk.Button(
            left_buttons,
            text="Export to Excel",
            command=lambda: self._export_tree_to_excel(
                self.journal_tree,
                default_name=f"journal_{self.fs_date_to.get() if hasattr(self, 'fs_date_to') else ''}.xlsx",
            ),
            style="Techfix.TButton",
        )
        export_btn.grid(row=0, column=1, padx=2, pady=2, sticky="w")

        # Simple paging controls
        paging_frame = ttk.Frame(left_buttons, style="Techfix.Surface.TFrame")
        paging_frame.grid(row=0, column=2, padx=6, pady=2, sticky="w")
        self.journal_page_label = ttk.Label(paging_frame, text="Page 1", style="Techfix.TLabel")
        self.journal_page_label.pack(side=tk.LEFT, padx=(0, 4))
        prev_btn = ttk.Button(
            paging_frame,
            text="◀ Prev",
            style="Techfix.TButton",
            command=lambda: self._change_journal_page(-1),
        )
        next_btn = ttk.Button(
            paging_frame,
            text="Next ▶",
            style="Techfix.TButton",
            command=lambda: self._change_journal_page(1),
        )
        prev_btn.pack(side=tk.LEFT, padx=2)
        next_btn.pack(side=tk.LEFT, padx=2)

        # Add filter controls - use grid for responsive layout
        filter_frame = ttk.Frame(toolbar, style="Techfix.Surface.TFrame")
        filter_frame.grid(row=0, column=2, sticky="e", padx=4)
        filter_frame.columnconfigure(0, weight=0)
        filter_frame.columnconfigure(1, weight=0)
        filter_frame.columnconfigure(2, weight=0)
        filter_frame.columnconfigure(3, weight=0)
        filter_frame.columnconfigure(4, weight=0)
        filter_frame.rowconfigure(0, weight=0)
        filter_frame.rowconfigure(1, weight=0)  # Second row for wrapping
        
        # Store references for responsive layout
        self.journal_toolbar = toolbar
        self.journal_left_buttons = left_buttons
        self.journal_refresh_btn = refresh_btn
        self.journal_export_btn = export_btn
        self.journal_filter_frame = filter_frame

        filter_label = ttk.Label(filter_frame, text="Filter by:", style="Techfix.TLabel")
        filter_label.grid(row=0, column=0, padx=4, pady=2, sticky="w")

        # Date range filter
        self.journal_date_from = ttk.Entry(filter_frame, width=10, style="Techfix.TEntry")
        self.journal_date_from.grid(row=0, column=1, padx=2, pady=2, sticky="w")
        to_label = ttk.Label(filter_frame, text="to", style="Techfix.TLabel")
        to_label.grid(row=0, column=2, padx=2, pady=2, sticky="w")
        self.journal_date_to = ttk.Entry(filter_frame, width=10, style="Techfix.TEntry")
        self.journal_date_to.grid(row=0, column=3, padx=2, pady=2, sticky="w")

        # Account filter
        self.journal_account_filter = ttk.Combobox(
            filter_frame,
            width=20,
            state="readonly",
            style="Techfix.TCombobox",
        )
        self.journal_account_filter.grid(row=0, column=4, padx=4, pady=2, sticky="w")
        
        # Store filter widget references for responsive layout
        self.journal_filter_label = filter_label
        self.journal_to_label = to_label
        try:
            accs = db.get_accounts(conn=self.engine.conn)
            names = ["All"] + [f"{a['code']} - {a['name']}" for a in accs]
            self.journal_account_filter["values"] = names
            try:
                self.journal_account_filter.set("All")
            except Exception:
                pass
        except Exception:
            pass
        self.journal_account_filter.bind("<<ComboboxSelected>>", lambda e: self._load_journal_entries())

        # Create the treeview for journal entries
        columns = ("date", "reference", "description", "debit", "credit", "account")
        self.journal_tree = ttk.Treeview(
            frame,
            columns=columns,
            show="headings",
            selectmode="browse",
            style="Techfix.Treeview",
        )

        # Configure columns
        self.journal_tree.heading("date", text="Date")
        self.journal_tree.heading("reference", text="Reference")
        self.journal_tree.heading("description", text="Description")
        self.journal_tree.heading("debit", text="Debit")
        self.journal_tree.heading("credit", text="Credit")
        self.journal_tree.heading("account", text="Account")

        # Set column widths
        self.journal_tree.column("date", width=100, anchor=tk.W)
        self.journal_tree.column("reference", width=100, anchor=tk.W)
        self.journal_tree.column("description", width=250, anchor=tk.W)
        self.journal_tree.column("debit", width=100, anchor=tk.E)
        self.journal_tree.column("credit", width=100, anchor=tk.E)
        self.journal_tree.column("account", width=200, anchor=tk.W)

        # Add scrollbar
        scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=self.journal_tree.yview)
        self.journal_tree.configure(yscrollcommand=scrollbar.set)

        # Pack the treeview and scrollbar
        self.journal_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=4, pady=4)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y, pady=4)

        # Bind double-click event
        self.journal_tree.bind("<Double-1>", self._on_journal_entry_double_click)

        self._load_journal_entries()

    def _load_journal_entries(self):
        try:
            # Basic in‑memory pagination: fetch a chunk and remember if there is more.
            if not hasattr(self, "_journal_page"):
                self._journal_page = 0
            page_size = getattr(self, "_journal_page_size", 500)

            # Always clear existing rows before loading a page so we don't duplicate content.
            for item in self.journal_tree.get_children():
                self.journal_tree.delete(item)

            offset = self._journal_page * page_size

            rows = db.fetch_journal(period_id=self.engine.current_period_id, conn=self.engine.conn)
            # Optional account filtering from dropdown
            sel = ''
            try:
                sel = self.journal_account_filter.get().strip()
            except Exception:
                sel = ''
            if sel and sel.lower() != 'all':
                try:
                    code, name_match = sel.split(' - ', 1)
                except Exception:
                    code, name_match = sel, ''
                rows = [r for r in rows if ((('code' in r.keys()) and r['code'] == code) or r['name'] == name_match)]
            # Optional date range filter (from/to)
            from_s = to_s = ''
            try:
                from_s = self.journal_date_from.get().strip()
            except Exception:
                from_s = ''
            try:
                to_s = self.journal_date_to.get().strip()
            except Exception:
                to_s = ''
            if from_s or to_s:
                from_dt = None
                to_dt = None
                try:
                    if from_s:
                        from_dt = datetime.strptime(from_s, '%Y-%m-%d').date()
                except Exception:
                    from_dt = None
                try:
                    if to_s:
                        to_dt = datetime.strptime(to_s, '%Y-%m-%d').date()
                except Exception:
                    to_dt = None
                def _in_range(dstr: str) -> bool:
                    try:
                        d = datetime.strptime(dstr, '%Y-%m-%d').date()
                    except Exception:
                        return False
                    if from_dt and d < from_dt:
                        return False
                    if to_dt and d > to_dt:
                        return False
                    return True
                rows = [r for r in rows if _in_range(r['date'])]
            # Apply offset/limit after filters
            sliced = rows[offset : offset + page_size + 1]
            has_more = len(sliced) > page_size
            rows = sliced[:page_size]
            setattr(self, "_journal_has_more", has_more)

            current_entry = None
            total_debit = 0.0
            total_credit = 0.0
            for r in rows:
                eid = r["entry_id"]
                date = r["date"]
                desc = r["description"]
                try:
                    doc_ref = r["document_ref"] if "document_ref" in r.keys() else None
                except Exception:
                    doc_ref = None
                try:
                    ext_ref = r["external_ref"] if "external_ref" in r.keys() else None
                except Exception:
                    ext_ref = None
                ref = (str(doc_ref).strip() if doc_ref else "") or (str(ext_ref).strip() if ext_ref else "")
                acct = r["name"]
                debit = r["debit"]
                credit = r["credit"]
                try:
                    total_debit += float(debit or 0)
                except Exception:
                    pass
                try:
                    total_credit += float(credit or 0)
                except Exception:
                    pass
                # Insert a header row (first line for an entry) with the date and description.
                # Let Tk generate item IDs automatically to avoid duplicate-iid errors when
                # reloading or paging.
                if current_entry != eid:
                    self.journal_tree.insert(
                        '',
                        'end',
                        values=(
                            date,
                            ref,
                            desc,
                            f"{debit:,.2f}" if debit else "",
                            f"{credit:,.2f}" if credit else "",
                            acct,
                        ),
                    )
                    current_entry = eid
                else:
                    # Subsequent lines for the same entry should show blanks for date/description
                    self.journal_tree.insert(
                        '',
                        'end',
                        values=(
                            "",
                            "",
                            "",
                            f"{debit:,.2f}" if debit else "",
                            f"{credit:,.2f}" if credit else "",
                            acct,
                        ),
                    )
            # Insert totals row
            try:
                self.journal_tree.insert(
                    '',
                    'end',
                    values=("", "", "Totals:", f"{total_debit:,.2f}", f"{total_credit:,.2f}", ""),
                    tags=('totals',),
                )
                # Make totals row visually distinct and bold
                self.journal_tree.tag_configure(
                    'totals',
                    background=self.palette.get('tab_selected_bg', '#e0ecff'),
                    foreground=self.palette.get('text_primary', '#000000'),
                    font=FONT_BOLD,
                )
            except Exception:
                pass
            # Update page label if present
            try:
                if hasattr(self, "journal_page_label"):
                    self.journal_page_label.config(text=f"Page {self._journal_page + 1}")
            except Exception:
                pass
        except Exception as e:
            try:
                self._handle_exception("load_journal_entries", e)
            except Exception:
                pass
            messagebox.showerror("Error", f"Failed to load journal entries: {str(e)}")

    def _change_journal_page(self, delta: int) -> None:
        """Move forward/backward through journal pages and reload."""
        try:
            if not hasattr(self, "_journal_page"):
                self._journal_page = 0
            page_size = getattr(self, "_journal_page_size", 500)
            if page_size <= 0:
                page_size = 500
            # Bound page index at 0 and don't advance past "no more data"
            new_page = max(0, self._journal_page + int(delta))
            if delta > 0 and not getattr(self, "_journal_has_more", False):
                return
            self._journal_page = new_page
            # Do not clear existing in this call; _load_journal_entries handles it via _journal_reset
            self._load_journal_entries()
        except Exception as e:
            try:
                self._handle_exception("change_journal_page", e)
            except Exception:
                pass

    def _load_recent_transactions(self, limit: int = 50) -> None:
        """Populate the Recent Transactions tree with the most recent entries, including drafts without lines."""
        try:
            if hasattr(self, 'txn_recent_tree'):
                for it in self.txn_recent_tree.get_children():
                    self.txn_recent_tree.delete(it)
            conn = self.engine.conn
            pid = self.engine.current_period_id
            clause = "WHERE je.period_id = ?" if pid else ""
            params = ([int(pid)] if pid else []) + [limit]
            sql = (
                """
                SELECT je.id AS entry_id, je.date AS date, je.description AS description, je.status AS status
                FROM journal_entries je
                """
                + (clause or "")
                + """
                ORDER BY date(je.date) DESC, je.id DESC
                LIMIT ?
                """
            )
            cur = conn.execute(sql, params)
            entries = cur.fetchall()
            for e in entries:
                eid = e['entry_id'] if 'entry_id' in e.keys() else e['id']
                date = e['date'] if 'date' in e.keys() else ''
                desc = e['description'] if 'description' in e.keys() else ''
                # Aggregate debit/credit and pick a representative account name (first line)
                try:
                    rsum = conn.execute(
                        """
                        SELECT COALESCE(SUM(jl.debit),0) AS debit, COALESCE(SUM(jl.credit),0) AS credit
                        FROM journal_lines jl
                        WHERE jl.entry_id=?
                        """,
                        (eid,)
                    ).fetchone()
                    debit = float(rsum['debit'] if 'debit' in rsum.keys() else 0) if rsum else 0.0
                    credit = float(rsum['credit'] if 'credit' in rsum.keys() else 0) if rsum else 0.0
                    racct = conn.execute(
                        """
                        SELECT a.name
                        FROM journal_lines jl
                        JOIN accounts a ON a.id = jl.account_id
                        WHERE jl.entry_id=?
                        ORDER BY jl.id
                        LIMIT 1
                        """,
                        (eid,)
                    ).fetchone()
                    acct = (racct['name'] if racct and 'name' in racct.keys() else '')
                except Exception:
                    debit = 0.0
                    credit = 0.0
                    acct = ''
                self.txn_recent_tree.insert(
                    '', 'end',
                    values=(date, eid, desc, f"{debit:,.2f}" if debit else '', f"{credit:,.2f}" if credit else '', acct),
                    tags=('row',)
                )
            try:
                self.txn_recent_tree.tag_configure('row', background=self.palette.get('surface_bg', '#ffffff'))
            except Exception:
                pass
            try:
                self.txn_recent_tree.yview_moveto(0)
            except Exception:
                pass
        except Exception:
            pass
    
    def _on_journal_entry_double_click(self, event):
        """Handle double-click on a journal entry"""
        sel = self.journal_tree.selection()
        if not sel:
            return
        item = sel[0]
        # Try to extract the entry id from the item's IID which we set when populating the tree
        iid = str(item)
        entry_id = None
        try:
            if iid.startswith('je-'):
                # IID format: je-<entry_id>-line-<line_id>
                parts = iid.split('-')
                if len(parts) >= 3:
                    entry_id = int(parts[1])
        except Exception:
            entry_id = None

        # Fallback: if we couldn't parse an entry_id from the IID, try reading the visible date column
        if entry_id is None:
            values = self.journal_tree.item(item, 'values')
            if values:
                # If the first displayed column contains a date, this is a header row; otherwise do nothing
                maybe_date = values[0]
                try:
                    # If the date value exists, try to find the corresponding entry by matching date+description
                    if maybe_date:
                        desc = values[2] if len(values) > 2 else None
                        # Query DB for an entry with this date and description (pick the latest match)
                        cur = self.engine.conn.execute("SELECT id FROM journal_entries WHERE date=? AND description=? ORDER BY id DESC LIMIT 1", (maybe_date, desc))
                        row = cur.fetchone()
                        if row:
                            entry_id = int(row['id'])
                except Exception:
                    entry_id = None

        if entry_id:
            try:
                self._load_transaction_for_editing(entry_id)
                self.notebook.select(self.tab_txn)
            except Exception as e:
                messagebox.showerror('Error', f'Failed to open transaction: {e}')
    
    def _load_transaction_for_editing(self, txn_id):
        """Load a transaction into the form for editing"""
        # This method would be implemented to load the selected transaction
        # into the transaction form for editing
        pass

    def _build_ledger_tab(self) -> None:
        """Build the ledger tab with a list of all accounts and their balances."""
        frame = self.tab_ledger

        # Create a frame for the toolbar
        toolbar = ttk.Frame(frame, style="Techfix.Surface.TFrame")
        toolbar.pack(fill=tk.X, padx=4, pady=4)

        # Configure toolbar grid for responsive layout
        toolbar.columnconfigure(0, weight=0)  # Left buttons
        toolbar.columnconfigure(1, weight=1)  # Spacer
        toolbar.columnconfigure(2, weight=0)  # Right filters
        toolbar.rowconfigure(0, weight=0)
        toolbar.rowconfigure(1, weight=0)  # Second row for wrapping

        # Left side buttons container
        left_buttons = ttk.Frame(toolbar, style="Techfix.Surface.TFrame")
        left_buttons.grid(row=0, column=0, sticky="w", padx=2)
        left_buttons.columnconfigure(0, weight=0)
        left_buttons.columnconfigure(1, weight=0)
        left_buttons.columnconfigure(2, weight=0)
        left_buttons.columnconfigure(3, weight=0)
        left_buttons.rowconfigure(0, weight=0)
        left_buttons.rowconfigure(1, weight=0)  # Second row for wrapping

        # Add refresh button
        refresh_btn = ttk.Button(
            left_buttons,
            text="Refresh",
            command=self._load_ledger_entries,
            style="Techfix.TButton",
        )
        refresh_btn.grid(row=0, column=0, padx=2, pady=2, sticky="w")
        
        post_btn = ttk.Button(
            left_buttons,
            text="Post to Ledger",
            command=self._post_to_ledger_action,
            style="Techfix.TButton",
        )
        post_btn.grid(row=0, column=1, padx=6, pady=2, sticky="w")
        
        # Export ledger to Excel
        export_btn = ttk.Button(
            left_buttons,
            text="Export to Excel",
            command=lambda: self._export_tree_to_excel(
                self.ledger_tree,
                default_name=f"ledger_{self.fs_date_to.get() if hasattr(self, 'fs_date_to') else ''}.xlsx",
            ),
            style="Techfix.TButton",
        )
        export_btn.grid(row=0, column=2, padx=6, pady=2, sticky="w")

        # Simple paging controls
        paging_frame = ttk.Frame(left_buttons, style="Techfix.Surface.TFrame")
        paging_frame.grid(row=0, column=3, padx=6, pady=2, sticky="w")
        self.ledger_page_label = ttk.Label(paging_frame, text="Page 1", style="Techfix.TLabel")
        self.ledger_page_label.pack(side=tk.LEFT, padx=(0, 4))
        prev_btn = ttk.Button(
            paging_frame,
            text="◀ Prev",
            style="Techfix.TButton",
            command=lambda: self._change_ledger_page(-1),
        )
        next_btn = ttk.Button(
            paging_frame,
            text="Next ▶",
            style="Techfix.TButton",
            command=lambda: self._change_ledger_page(1),
        )
        prev_btn.pack(side=tk.LEFT, padx=2)
        next_btn.pack(side=tk.LEFT, padx=2)

        # Add account filter - use grid for responsive layout
        filter_frame = ttk.Frame(toolbar, style="Techfix.Surface.TFrame")
        filter_frame.grid(row=0, column=2, sticky="e", padx=4)
        filter_frame.columnconfigure(0, weight=0)
        filter_frame.columnconfigure(1, weight=0)
        filter_frame.rowconfigure(0, weight=0)
        filter_frame.rowconfigure(1, weight=0)  # Second row for wrapping
        
        # Store references for responsive layout
        self.ledger_toolbar = toolbar
        self.ledger_left_buttons = left_buttons
        self.ledger_refresh_btn = refresh_btn
        self.ledger_post_btn = post_btn
        self.ledger_export_btn = export_btn
        self.ledger_filter_frame = filter_frame

        account_label = ttk.Label(filter_frame, text="Account:", style="Techfix.TLabel")
        account_label.grid(row=0, column=0, padx=4, pady=2, sticky="w")
        self.ledger_account_filter = ttk.Combobox(
            filter_frame,
            width=30,
            state="readonly",
            style="Techfix.TCombobox",
        )
        self.ledger_account_filter.grid(row=0, column=1, padx=4, pady=2, sticky="w")
        
        # Store filter widget reference
        self.ledger_account_label = account_label
        self.ledger_account_filter.bind("<<ComboboxSelected>>", lambda e: self._load_ledger_entries())
        try:
            accs = db.get_accounts(conn=self.engine.conn)
            names = ["All"] + [f"{a['code']} - {a['name']}" for a in accs]
            self.ledger_account_filter["values"] = names
            try:
                self.ledger_account_filter.set("All")
            except Exception:
                pass
        except Exception:
            pass

        # Create the treeview for ledger entries
        columns = ("account", "date", "description", "debit", "credit", "balance")
        self.ledger_tree = ttk.Treeview(
            frame,
            columns=columns,
            show="headings",
            selectmode="browse",
            style="Techfix.Treeview",
        )

        # Configure columns
        self.ledger_tree.heading("account", text="Account", anchor=tk.W)
        self.ledger_tree.heading("date", text="Date", anchor=tk.W)
        self.ledger_tree.heading("description", text="Description", anchor=tk.W)
        self.ledger_tree.heading("debit", text="Debit", anchor=tk.E)
        self.ledger_tree.heading("credit", text="Credit", anchor=tk.E)
        self.ledger_tree.heading("balance", text="Balance", anchor=tk.E)

        # Set column widths
        self.ledger_tree.column("account", width=150, anchor=tk.W)
        self.ledger_tree.column("date", width=100, anchor=tk.W)
        self.ledger_tree.column("description", width=250, anchor=tk.W)
        self.ledger_tree.column("debit", width=120, anchor=tk.E)
        self.ledger_tree.column("credit", width=120, anchor=tk.E)
        self.ledger_tree.column("balance", width=120, anchor=tk.E)

        # Add scrollbar
        scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=self.ledger_tree.yview)
        self.ledger_tree.configure(yscrollcommand=scrollbar.set)

        # Pack the treeview and scrollbar
        self.ledger_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=4, pady=4)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y, pady=4)

        # Bind double-click event
        self.ledger_tree.bind("<Double-1>", self._on_ledger_entry_double_click)

        # Load initial data
        self._load_ledger_entries()

    def _load_ledger_entries(self):
        try:
            if not hasattr(self, "_ledger_page"):
                self._ledger_page = 0
            page_size = getattr(self, "_ledger_page_size", 500)

            # Always clear existing rows before loading a page so we don't duplicate content.
            for item in self.ledger_tree.get_children():
                self.ledger_tree.delete(item)

            offset = self._ledger_page * page_size

            # Fetch individual ledger transactions (not trial balance)
            rows = db.fetch_ledger(period_id=self.engine.current_period_id, conn=self.engine.conn)
            
            # Apply account filter if specified
            sel = ''
            try:
                sel = self.ledger_account_filter.get().strip()
            except Exception:
                sel = ''
            if sel and sel.lower() != 'all':
                try:
                    code, name_match = sel.split(' - ', 1)
                except Exception:
                    code, name_match = sel, ''
                rows = [r for r in rows if (('code' in r.keys() and r['code'] == code) or ('name' in r.keys() and r['name'] == name_match))]

            # Apply paging after filters
            sliced = rows[offset : offset + page_size + 1]
            has_more = len(sliced) > page_size
            rows = sliced[:page_size]
            setattr(self, "_ledger_has_more", has_more)
            
            # Calculate running balances per account
            account_balances = {}  # account_id -> running balance
            total_debit = 0.0
            total_credit = 0.0
            
            for r in rows:
                # sqlite3.Row objects use dictionary-style access, not .get()
                account_id = r['account_id'] if 'account_id' in r.keys() else None
                code = r['code'] if 'code' in r.keys() else ''
                name = r['name'] if 'name' in r.keys() else ''
                date = r['date'] if 'date' in r.keys() else ''
                description = r['description'] if 'description' in r.keys() else ''
                debit = float(r['debit'] if 'debit' in r.keys() and r['debit'] is not None else 0)
                credit = float(r['credit'] if 'credit' in r.keys() and r['credit'] is not None else 0)
                
                # Initialize account balance if first transaction for this account
                if account_id not in account_balances:
                    account_balances[account_id] = 0.0
                
                # Update running balance
                account_balances[account_id] += (debit - credit)
                running_balance = account_balances[account_id]
                
                # Format account display
                account_display = f"{code} - {name}" if code else name
                
                # Format balance with Dr/Cr indicator
                if abs(running_balance) < 0.01:
                    balance_display = "0.00"
                elif running_balance > 0:
                    balance_display = f"{running_balance:,.2f} Dr"
                else:
                    balance_display = f"{abs(running_balance):,.2f} Cr"
                
                # Insert row
                self.ledger_tree.insert(
                    '',
                    'end',
                    values=(
                        account_display,
                        date,
                        description,
                        f"{debit:,.2f}" if debit else '',
                        f"{credit:,.2f}" if credit else '',
                        balance_display
                    )
                )
                
                total_debit += debit
                total_credit += credit
            
            # Totals row
            try:
                self.ledger_tree.insert(
                    '',
                    'end',
                    values=(
                        "Totals",
                        "",
                        "",
                        f"{total_debit:,.2f}" if total_debit else '',
                        f"{total_credit:,.2f}" if total_credit else '',
                        "",
                    ),
                    tags=('totals',),
                )
                # Make totals row visually distinct and bold
                self.ledger_tree.tag_configure(
                    'totals',
                    background=self.palette.get('tab_selected_bg', '#e0ecff'),
                    foreground=self.palette.get('text_primary', '#000000'),
                    font=FONT_BOLD,
                )
            except Exception:
                pass
            # Update page label if present
            try:
                if hasattr(self, "ledger_page_label"):
                    self.ledger_page_label.config(text=f"Page {self._ledger_page + 1}")
            except Exception:
                pass
        except Exception as e:
            try:
                self._handle_exception("load_ledger_entries", e)
            except Exception:
                pass
            messagebox.showerror("Error", f"Failed to load ledger entries: {str(e)}")

    def _change_ledger_page(self, delta: int) -> None:
        """Move forward/backward through ledger pages and reload."""
        try:
            if not hasattr(self, "_ledger_page"):
                self._ledger_page = 0
            page_size = getattr(self, "_ledger_page_size", 500)
            if page_size <= 0:
                page_size = 500
            new_page = max(0, self._ledger_page + int(delta))
            if delta > 0 and not getattr(self, "_ledger_has_more", False):
                return
            self._ledger_page = new_page
            self._load_ledger_entries()
        except Exception as e:
            try:
                self._handle_exception("change_ledger_page", e)
            except Exception:
                pass

    def _resolve_account_id(self, sel: str) -> Optional[int]:
        try:
            s = (sel or '').strip()
            if not s:
                return None
            v = self.account_id_by_display.get(s) or self.account_id_by_display.get(s.lower())
            if v:
                return int(v)
            if ' - ' in s:
                left = s.split(' - ', 1)[0].strip()
                v = self.account_id_by_display.get(left)
                if v:
                    return int(v)
            accs = db.get_accounts(conn=self.engine.conn)
            for a in accs:
                if str(a['code']) == s or str(a['name']).lower() == s.lower() or f"{a['code']} - {a['name']}" == s:
                    return int(a['id'])
            return None
        except Exception:
            return None

    def _infer_accounts_from_context(self, desc: str, source: Optional[str]) -> tuple[str | None, str | None]:
        try:
            d = (desc or '').strip().lower()
            s = (source or '').strip().lower()
            pairs: list[tuple[str, tuple[str, str]]] = [
                ('rent', ('403 - Rent Expense', '101 - Cash')),
                ('utilities', ('404 - Utilities Expense', '101 - Cash')),
                ('supplies adjustment', ('405 - Supplies Expense', '124 - Supplies')),
                ('adjust', ('405 - Supplies Expense', '124 - Supplies')),
                ('accrual adjustment', ('405 - Supplies Expense', '124 - Supplies')),
                ('payroll', ('402 - Salaries & Wages', '101 - Cash')),
                ('withdrawal', ("302 - Owner's Drawings", '101 - Cash')),
                ('owner', ("302 - Owner's Drawings", '101 - Cash')),
                ('deposit', ('101 - Cash', '401 - Service Revenue')),
                ('invoice', ('102 - Accounts Receivable', '401 - Service Revenue')),
                ('sales', ('101 - Cash', '401 - Service Revenue')),
            ]
            match = None
            for k, v in pairs:
                if k in d or k == s:
                    match = v
                    break
            if not match:
                return (None, None)
            # Use flexible account matching instead of exact display string matching
            da = self._match_account_display(match[0]) if match[0] else None
            ca = self._match_account_display(match[1]) if match[1] else None
            # If matching failed, try the original display strings as fallback
            if not da and match[0]:
                try:
                    accs = db.get_accounts(conn=self.engine.conn)
                    displays = {f"{a['code']} - {a['name']}" for a in accs}
                    if match[0] in displays:
                        da = match[0]
                except Exception:
                    pass
            if not ca and match[1]:
                try:
                    accs = db.get_accounts(conn=self.engine.conn)
                    displays = {f"{a['code']} - {a['name']}" for a in accs}
                    if match[1] in displays:
                        ca = match[1]
                except Exception:
                    pass
            return (da, ca)
        except Exception:
            return (None, None)

    def _match_account_display(self, text: Optional[str]) -> Optional[str]:
        try:
            if not text:
                return None
            text_str = str(text).strip()
            if not text_str:
                return None
            text_lower = text_str.lower()
            
            # Common account name aliases/mappings
            # Maps scanned names to actual account names in chart of accounts
            account_aliases = {
                "service income": "Service Revenue",
                "service revenue": "Service Revenue",
                "owners capital": "Owner's Capital",
                "owner capital": "Owner's Capital",
                "owners drawings": "Owner's Drawings",
                "owner drawings": "Owner's Drawings",
                "salaries and wages": "Salaries & Wages",
                "salaries & wages": "Salaries & Wages",
            }
            
            # Check if text matches an alias
            if text_lower in account_aliases:
                text_str = account_aliases[text_lower]
                text_lower = text_str.lower()
            
            # Extract account name from "CODE - Name" format if present
            text_name = text_str
            text_code = None
            if ' - ' in text_str:
                parts = text_str.split(' - ', 1)
                if len(parts) == 2:
                    text_code = parts[0].strip()
                    text_name = parts[1].strip()
            text_name_lower = text_name.lower()
            
            accs = db.get_accounts()
            for a in accs:
                disp = f"{a['code']} - {a['name']}"
                code = str(a.get('code', '')).strip()
                name = str(a.get('name', '')).strip()
                name_lower = name.lower()
                
                # Exact matches (highest priority)
                if text_str == disp or text_str == code or text_lower == name_lower:
                    return disp
                
                # Match by extracted account name (if format was "CODE - Name")
                if text_name and text_name_lower == name_lower:
                    return disp
                
                # Check alias mapping for the account name
                if name_lower in account_aliases:
                    if text_name_lower == account_aliases[name_lower].lower():
                        return disp
                
                # Partial matches (if exact match failed)
                # Check if account name is contained in text or vice versa (case-insensitive)
                if text_name_lower in name_lower or name_lower in text_name_lower:
                    return disp
                
                # Check if text matches code (even if partial)
                if code and text_str in code:
                    return disp
                if text_code and code and text_code == code:
                    return disp
            return None
        except Exception:
            return None

    def _set_accounts(self, debit_disp: Optional[str], credit_disp: Optional[str]) -> None:
        try:
            if debit_disp and hasattr(self, 'debit_acct'):
                # Update StringVar first (this is what the combobox reads from)
                if hasattr(self, 'debit_acct_var'):
                    self.debit_acct_var.set(debit_disp)
                # Then update the combobox itself
                try:
                    self.debit_acct.set(debit_disp)
                except Exception:
                    # If set fails, try to add to values first
                    try:
                        current_values = list(self.debit_acct['values'])
                        if debit_disp not in current_values:
                            self.debit_acct['values'] = tuple(list(current_values) + [debit_disp])
                        self.debit_acct.set(debit_disp)
                    except Exception:
                        pass
                self._accounts_prefilled = True
                # Trigger the account changed event to update UI
                try:
                    self._on_account_changed('debit')
                except Exception:
                    pass
            if credit_disp and hasattr(self, 'credit_acct'):
                # Update StringVar first (this is what the combobox reads from)
                if hasattr(self, 'credit_acct_var'):
                    self.credit_acct_var.set(credit_disp)
                # Then update the combobox itself
                try:
                    self.credit_acct.set(credit_disp)
                except Exception:
                    # If set fails, try to add to values first
                    try:
                        current_values = list(self.credit_acct['values'])
                        if credit_disp not in current_values:
                            self.credit_acct['values'] = tuple(list(current_values) + [credit_disp])
                        self.credit_acct.set(credit_disp)
                    except Exception:
                        pass
                self._accounts_prefilled = True
                # Trigger the account changed event to update UI
                try:
                    self._on_account_changed('credit')
                except Exception:
                    pass
            # Update button states after setting accounts - with a small delay to ensure UI updates
            try:
                self.after(50, self._update_post_buttons_enabled)
                self._update_post_buttons_enabled()
            except Exception:
                pass
        except Exception:
            pass

    def _assign_accounts_from_data(self, data_obj: dict) -> bool:
        try:
            da_raw = data_obj.get('debit_account')
            ca_raw = data_obj.get('credit_account')
            dd = self._match_account_display(da_raw)
            cc = self._match_account_display(ca_raw)
            if dd or cc:
                self._set_accounts(dd, cc)
                return True
            return False
        except Exception:
            return False

    def _validate_accounts_assigned(self) -> bool:
        try:
            da = self.debit_acct.get().strip() if hasattr(self, 'debit_acct') else ''
            ca = self.credit_acct.get().strip() if hasattr(self, 'credit_acct') else ''
            ok = bool(da) and bool(ca)
            # Do not show account not set status in the UI
            return ok
        except Exception:
            return False

    def _on_account_changed(self, which: str) -> None:
        try:
            self._accounts_modified_manually = True
            # Do not show account set status in the UI
        except Exception:
            pass

    def _audit(self, action: str, details: dict) -> None:
        try:
            db.log_audit(action=action, details=json.dumps(details), user='system', conn=self.engine.conn)
        except Exception:
            pass

    def _validate_amounts_present(self) -> bool:
        try:
            d = self.debit_amt.get().strip() if hasattr(self, 'debit_amt') else ''
            c = self.credit_amt.get().strip() if hasattr(self, 'credit_amt') else ''
            ok = bool(d) and bool(c)
            return ok
        except Exception:
            return False

    def _validate_transaction_ready(self) -> tuple[bool, list[str]]:
        issues: list[str] = []
        try:
            dd = self.debit_acct.get().strip() if hasattr(self, 'debit_acct') else ''
            cc = self.credit_acct.get().strip() if hasattr(self, 'credit_acct') else ''
            if not dd:
                issues.append('Missing debit account')
            if not cc:
                issues.append('Missing credit account')
            d = self.debit_amt.get().strip() if hasattr(self, 'debit_amt') else ''
            c = self.credit_amt.get().strip() if hasattr(self, 'credit_amt') else ''
            if not d:
                issues.append('Missing debit amount')
            if not c:
                issues.append('Missing credit amount')
            ok = len(issues) == 0
            return ok, issues
        except Exception:
            return False, ['Validation error']

    def _auto_entry_from_data(self, data: dict, filename: str) -> dict:
        try:
            import os
            sugg: dict = {}
            date = data.get('date')
            desc = data.get('description')
            docno = data.get('document_ref') or data.get('doc_no') or None
            st = data.get('source_type') or None
            if not st:
                base = os.path.basename(filename)
                name, _ = os.path.splitext(base)
                parts = name.split('_')
                st = parts[1].title() if len(parts) > 1 else None
            cat = (data.get('category') or data.get('expense_type') or data.get('purpose') or '').strip().lower()
            dv_raw = data.get('debit_amount') or data.get('debit') or data.get('amount') or None
            cv_raw = data.get('credit_amount') or data.get('credit') or data.get('amount') or None
            def fmt_amt(v):
                try:
                    import re
                    s = str(v).strip()
                    neg = False
                    if s.startswith('(') and s.endswith(')'):
                        neg = True
                        s = s[1:-1]
                    s = re.sub(r"[^0-9.\-]", "", s)
                    if not s:
                        return None
                    val = float(s)
                    if neg:
                        val = -abs(val)
                    return round(val, 2)
                except Exception:
                    return None
            dv = fmt_amt(dv_raw)
            cv = fmt_amt(cv_raw)
            if dv is None and cv is not None:
                dv = cv
            if cv is None and dv is not None:
                cv = dv
            da_disp = None
            ca_disp = None
            if st:
                da_src, ca_src = self._default_accounts_for_source(st)
                da_disp = da_src or da_disp
                ca_disp = ca_src or ca_disp
            if not da_disp or not ca_disp:
                k = (cat or (desc or '')).lower()
                da_ctx, ca_ctx = self._infer_accounts_from_context(k, st)
                da_disp = da_disp or da_ctx
                ca_disp = ca_disp or ca_ctx
            if not desc:
                parts = []
                if st:
                    parts.append(str(st))
                if docno:
                    parts.append(str(docno))
                meta = data.get('purpose') or data.get('vendor') or data.get('payee') or ''
                if meta:
                    parts.append(str(meta))
                desc = ' '.join([p for p in parts if p]).strip() or 'Transaction'
            sugg['date'] = date
            sugg['description'] = desc
            sugg['debit_account_display'] = da_disp
            sugg['credit_account_display'] = ca_disp
            sugg['debit_amount'] = dv
            sugg['credit_amount'] = cv
            sugg['document_ref'] = docno
            sugg['source_type'] = st
            return sugg
        except Exception:
            return {}

    def _update_post_buttons_enabled(self) -> None:
        try:
            # Check if user has permission to create transactions
            has_create_permission = self._has_permission('create')
            
            # Try to get from StringVar first, then fallback to combobox
            if hasattr(self, 'debit_acct_var'):
                dd = self.debit_acct_var.get().strip()
            elif hasattr(self, 'debit_acct'):
                dd = self.debit_acct.get().strip()
            else:
                dd = ''
            
            if hasattr(self, 'credit_acct_var'):
                cc = self.credit_acct_var.get().strip()
            elif hasattr(self, 'credit_acct'):
                cc = self.credit_acct.get().strip()
            else:
                cc = ''
            
            d = self.debit_amt.get().strip() if hasattr(self, 'debit_amt') else ''
            c = self.credit_amt.get().strip() if hasattr(self, 'credit_amt') else ''
            
            # Buttons are enabled only if user has permission AND form is filled
            enabled = has_create_permission and bool(dd) and bool(cc) and bool(d) and bool(c)
            state = 'normal' if enabled else 'disabled'
            
            if hasattr(self, 'btn_post'):
                try:
                    self.btn_post.configure(state=state)
                except Exception:
                    pass
            if hasattr(self, 'btn_draft'):
                try:
                    # Draft button also needs permission check
                    self.btn_draft.configure(state=state)
                except Exception:
                    pass
        except Exception as e:
            # Debug: log the error but don't break
            try:
                self._audit('button_enable_error', {'error': str(e)})
            except Exception:
                pass

    def _load_document_preview(self, filename: str) -> None:
        """Store document path for external opening."""
        try:
            import os
            self.current_document_path = filename
            if not os.path.exists(filename):
                messagebox.showerror("Document", "Selected document does not exist")
                self._audit('document_missing', {'file': filename})
                return
            if not os.access(filename, os.R_OK):
                messagebox.showerror("Document", "You do not have permission to read this document")
                self._audit('document_permission_denied', {'file': filename})
                return
            self._audit('document_selected', {'file': filename})
        except Exception:
            self._audit('document_preview_error', {'file': filename, 'stage': 'exception'})

    def _set_view_mode(self, mode: str) -> None:
        try:
            # Mode is advisory; for text we keep continuous, for PDF we set slider range
            self._audit('viewer_mode', {'mode': mode})
        except Exception:
            pass

    def _goto_page(self, page: int) -> None:
        """No-op: document viewer removed."""
        pass

    def _build_pdf_thumbnails(self, doc) -> None:
        try:
            if hasattr(self, 'doc_thumbs'):
                for it in self.doc_thumbs.get_children():
                    self.doc_thumbs.delete(it)
                for i in range(min(100, doc.page_count)):
                    self.doc_thumbs.insert('', 'end', values=(f"Page {i+1}"), tags=(str(i+1),))
        except Exception:
            pass

    def _on_thumb_select(self) -> None:
        try:
            sel = self.doc_thumbs.selection()
            if not sel:
                return
            iid = sel[0]
            tags = self.doc_thumbs.item(iid, 'tags')
            if tags:
                p = int(tags[0])
                self.page_var.set(p)
                self._goto_page(p)
        except Exception:
            pass

    def _bookmark_add(self) -> None:
        try:
            path = getattr(self, 'current_document_path', '')
            mark = f"{os.path.basename(path)} @ {self.page_var.get()}"
            self.bookmarks.insert('', 'end', values=(mark,))
            self._audit('viewer_bookmark_add', {'file': path, 'mark': mark})
        except Exception:
            pass

    def _bookmark_remove(self) -> None:
        try:
            for it in self.bookmarks.selection():
                self.bookmarks.delete(it)
        except Exception:
            pass

    def _share_document(self) -> None:
        try:
            path = getattr(self, 'current_document_path', '')
            if path:
                self.clipboard_clear()
                self.clipboard_append(path)
                messagebox.showinfo('Share', 'Document path copied to clipboard')
        except Exception:
            pass

    def _print_document(self) -> None:
        try:
            import os
            path = getattr(self, 'current_document_path', '')
            if path and os.path.exists(path):
                try:
                    os.startfile(path, 'print')
                except Exception:
                    messagebox.showwarning('Print', 'Printing not available for this file')
        except Exception:
            pass

    def _open_document_external(self) -> None:
        """Open the selected document in the system's default application."""
        try:
            import os
            path = getattr(self, 'current_document_path', None) or (self.txn_attachment_path.get().strip() if hasattr(self, 'txn_attachment_path') else None)
            if path and os.path.exists(path):
                os.startfile(path)
                self._audit('document_open_external', {'file': path})
            else:
                messagebox.showwarning('Document', 'No document selected')
        except Exception:
            pass

    def _on_schedule_reverse_toggle(self) -> None:
        try:
            schedule = bool(self.txn_schedule_reverse.get()) if hasattr(self, 'txn_schedule_reverse') else False
            if schedule:
                self.txn_reverse_date.configure(state='normal')
                try:
                    if hasattr(self, 'txn_date'):
                        d = self.txn_date.get().strip()
                        from datetime import datetime, timedelta
                        dt = datetime.strptime(d, '%Y-%m-%d') if d else datetime.now()
                        self.txn_reverse_date.delete(0, tk.END)
                        self.txn_reverse_date.insert(0, (dt + timedelta(days=30)).strftime('%Y-%m-%d'))
                except Exception:
                    pass
            else:
                self.txn_reverse_date.delete(0, tk.END)
                self.txn_reverse_date.configure(state='disabled')
        except Exception:
            pass

    # Source document picker enhancements
    def _init_document_picker(self) -> None:
        try:
            import os, json
            # Default library folder
            default_dir = os.path.join(str(db.DB_DIR), 'SampleSourceDocs')
            if os.path.exists(default_dir):
                self.doc_library_dir = default_dir
            else:
                try:
                    from pathlib import Path
                    self.doc_library_dir = str(Path.home())
                except Exception:
                    self.doc_library_dir = None
            self._recent_docs = []
            # Load recent list
            fp = os.path.join(str(db.DB_DIR), 'recent_docs.json')
            if os.path.exists(fp):
                with open(fp, 'r', encoding='utf-8') as f:
                    self._recent_docs = json.load(f) or []
            if hasattr(self, 'doc_recent_cb'):
                self.doc_recent_cb.configure(values=self._recent_docs)
        except Exception:
            pass

    def _save_recent_docs(self) -> None:
        try:
            import json, os
            fp = os.path.join(str(db.DB_DIR), 'recent_docs.json')
            with open(fp, 'w', encoding='utf-8') as f:
                json.dump(self._recent_docs[:20], f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def _append_recent_document(self, path: str) -> None:
        try:
            if not hasattr(self, '_recent_docs'):
                self._recent_docs = []
            # Move to front if exists
            self._recent_docs = [p for p in self._recent_docs if p != path]
            self._recent_docs.insert(0, path)
            self._recent_docs = self._recent_docs[:20]
            self._save_recent_docs()
        except Exception:
            pass

    def _pick_recent_document(self) -> None:
        try:
            path = self.doc_recent_var.get().strip() if hasattr(self, 'doc_recent_var') else ''
            if not path:
                return
            self.txn_attachment_path.set(path)
            try:
                self._prefill_date_from_source_document(path)
                self._prefill_amounts_from_source_document(path)
                self._prefill_from_source_document(path)
            except Exception:
                pass
            try:
                self._load_document_preview(path)
            except Exception:
                pass
        except Exception:
            pass

    def _open_picker_folder(self) -> None:
        try:
            import os
            d = getattr(self, 'doc_library_dir', None)
            if d and os.path.exists(d):
                os.startfile(d)
        except Exception:
            pass

    def _clear_document_selection(self) -> None:
        try:
            if hasattr(self, 'txn_attachment_path'):
                self.txn_attachment_path.set('')
            if hasattr(self, 'doc_recent_var'):
                self.doc_recent_var.set('')
            if hasattr(self, 'txn_prefill_status'):
                self.txn_prefill_status.configure(text='')
            self.current_document_path = None
        except Exception:
            pass

    # Document Library helpers
    def _init_document_library(self) -> None:
        try:
            import os
            default_dir = os.path.join(str(db.DB_DIR), 'SampleSourceDocs')
            self.doc_library_dir = default_dir
            self._refresh_document_library(force=True)
        except Exception:
            pass

    def _choose_document_library_folder(self) -> None:
        try:
            from tkinter import filedialog
            d = filedialog.askdirectory(initialdir=self.doc_library_dir if hasattr(self, 'doc_library_dir') else None)
            if d:
                self.doc_library_dir = d
                self._refresh_document_library(force=True)
        except Exception:
            pass

    def _get_doc_metadata(self, file_path: str) -> dict:
        import os, json
        base = os.path.basename(file_path)
        name, ext = os.path.splitext(base)
        parts = name.split('_')
        date = parts[0] if (parts and len(parts[0]) == 10) else ''
        typ = parts[1].title() if len(parts) > 1 else ext.lstrip('.').upper()
        status = 'Unverified'
        sidecar = os.path.splitext(file_path)[0] + '.json'
        try:
            if os.path.exists(sidecar):
                with open(sidecar, 'r', encoding='utf-8') as f:
                    d = json.load(f)
                required = ['debit_account','credit_account','debit_amount','credit_amount']
                ok = all(k in d.keys() for k in required)
                status = 'Verified' if ok else 'Incomplete'
        except Exception:
            status = 'Error'
        return {'name': base, 'date': date, 'type': typ, 'status': status}

    def _refresh_document_library(self, force: bool = False) -> None:
        try:
            import os
            flt = (self.doc_library_filter.get().strip().lower() if hasattr(self, 'doc_library_filter') else '')
            if hasattr(self, 'doc_library_tree'):
                for it in self.doc_library_tree.get_children():
                    self.doc_library_tree.delete(it)
            directory = getattr(self, 'doc_library_dir', None)
            if not directory or not os.path.exists(directory):
                return
            files = []
            try:
                for fn in os.listdir(directory):
                    if fn.lower().endswith(('.pdf','.jpg','.jpeg','.png','.doc','.docx','.xls','.xlsx','.json','.csv')):
                        files.append(os.path.join(directory, fn))
            except Exception:
                files = []
            # Basic pagination/truncation for performance
            files = sorted(files)[:500]
            for fp in files:
                meta = self._get_doc_metadata(fp)
                if flt and (flt not in meta['name'].lower() and flt not in meta['type'].lower() and flt not in meta['date'].lower() and flt not in meta['status'].lower()):
                    continue
                self.doc_library_tree.insert('', 'end', values=(meta['name'], meta['date'], meta['type'], meta['status']), tags=(fp,))
        except Exception:
            pass

    def _on_doc_library_select(self) -> None:
        try:
            sel = self.doc_library_tree.selection()
            if not sel:
                return
            # Use first selected to preview
            iid = sel[0]
            # Retrieve path from tags
            tags = self.doc_library_tree.item(iid, 'tags')
            if tags:
                path = tags[0]
                self._load_document_preview(path)
        except Exception:
            pass

    def _doc_library_selected_paths(self) -> list:
        try:
            paths = []
            for iid in self.doc_library_tree.selection():
                tags = self.doc_library_tree.item(iid, 'tags')
                if tags:
                    paths.append(tags[0])
            return paths
        except Exception:
            return []

    def _doc_library_open_external(self) -> None:
        try:
            import os
            paths = self._doc_library_selected_paths()
            for p in paths:
                if os.path.exists(p):
                    os.startfile(p)
        except Exception:
            pass

    def _doc_library_bulk_download(self) -> None:
        try:
            from tkinter import filedialog
            import os, shutil
            paths = self._doc_library_selected_paths()
            if not paths:
                messagebox.showwarning('Documents', 'Select document(s) first')
                return
            dest = filedialog.askdirectory()
            if not dest:
                return
            for p in paths:
                try:
                    shutil.copy2(p, dest)
                except Exception:
                    pass
            messagebox.showinfo('Documents', 'Downloaded selected documents')
        except Exception:
            pass

    def _doc_library_bulk_archive(self) -> None:
        try:
            import os, shutil
            paths = self._doc_library_selected_paths()
            if not paths:
                messagebox.showwarning('Documents', 'Select document(s) first')
                return
            arc_dir = os.path.join(str(db.DB_DIR), 'archive')
            os.makedirs(arc_dir, exist_ok=True)
            for p in paths:
                try:
                    shutil.move(p, os.path.join(arc_dir, os.path.basename(p)))
                except Exception:
                    pass
            self._refresh_document_library(force=True)
            messagebox.showinfo('Documents', 'Archived selected documents')
        except Exception:
            pass

    def _document_zoom(self, delta: int) -> None:
        """No-op: document viewer removed."""
        pass

    def _document_search(self) -> None:
        """No-op: document viewer removed."""
        pass

    def _document_add_annotation(self) -> None:
        """No-op: document viewer removed."""
        pass

    def _capture_ui_snapshot(self, *, label: str = "snapshot") -> None:
        try:
            snap = {}
            for name in ('txn_date','txn_desc','debit_acct','credit_acct','debit_amt','credit_amt','txn_doc_ref','txn_external_ref','txn_source_type'):
                w = getattr(self, name, None)
                if w is None:
                    continue
                try:
                    snap[name] = {
                        'w': w.winfo_width(),
                        'h': w.winfo_height(),
                        'x': w.winfo_rootx(),
                        'y': w.winfo_rooty(),
                    }
                except Exception:
                    pass
            self._audit('ui_snapshot', {'label': label, 'widgets': snap})
        except Exception:
            pass

    def _default_accounts_for_source(self, source: str) -> tuple[str | None, str | None]:
        try:
            s = (source or '').strip().lower()
            pairs = self._rules_map.get('source_pairs') or {
                'invoice': ('102 - Accounts Receivable', '401 - Service Revenue'),
                'receipt': ('101 - Cash', '401 - Service Revenue'),
                'sales': ('101 - Cash', '401 - Service Revenue'),
                'bank': ('101 - Cash', '401 - Service Revenue'),
                'payroll': ('402 - Salaries & Wages', '101 - Cash'),
                'adjust': ('405 - Supplies Expense', '124 - Supplies'),
            }
            if s in pairs:
                da_disp, ca_disp = pairs[s]
                try:
                    # Use flexible account matching instead of exact display string matching
                    da = self._match_account_display(da_disp) if da_disp else None
                    ca = self._match_account_display(ca_disp) if ca_disp else None
                    # If matching failed, try the original display strings as fallback
                    if not da and da_disp:
                        # Try to verify the account exists by checking if display format matches
                        accs = db.get_accounts(conn=self.engine.conn)
                        displays = {f"{a['code']} - {a['name']}" for a in accs}
                        if da_disp in displays:
                            da = da_disp
                    if not ca and ca_disp:
                        accs = db.get_accounts(conn=self.engine.conn)
                        displays = {f"{a['code']} - {a['name']}" for a in accs}
                        if ca_disp in displays:
                            ca = ca_disp
                    return (da, ca)
                except Exception:
                    # Fallback to original strings if matching fails
                    return (da_disp, ca_disp)
            return (None, None)
        except Exception:
            return (None, None)

    def _default_fallback_accounts(self) -> tuple[str | None, str | None]:
        try:
            accs = db.get_accounts(conn=self.engine.conn)
            cash = next((a for a in accs if a['code'] == '101' or a['name'].lower() == 'cash'), None)
            srv = next((a for a in accs if a['code'] == '401' or a['name'].lower() == 'service revenue'), None)
            asset = next((a for a in accs if str(a['type']).lower() == 'asset'), None)
            revenue = next((a for a in accs if str(a['type']).lower() == 'revenue'), None)
            da = f"{asset['code']} - {asset['name']}" if asset else (f"{cash['code']} - {cash['name']}" if cash else None)
            ca = f"{revenue['code']} - {revenue['name']}" if revenue else (f"{srv['code']} - {srv['name']}" if srv else None)
            return (da, ca)
        except Exception:
            return (None, None)

    def _load_inference_rules(self) -> None:
        try:
            import json, os
            rules_path = os.path.join(str(db.DB_DIR), 'rules.json')
            if os.path.exists(rules_path):
                with open(rules_path, 'r', encoding='utf-8') as f:
                    self._rules_map = json.load(f) or {}
        except Exception:
            pass

    def _is_duplicate_transaction(
        self,
        *,
        date: str,
        desc: str,
        debit_account_id: int,
        credit_account_id: int,
        debit_amt: float,
        credit_amt: float,
        doc_ref: str | None,
        ext_ref: str | None,
        status: str,
    ) -> bool:
        """
        Best‑effort duplicate protection for the Transactions tab.

        We treat an entry as a duplicate if there is already a journal entry in
        the current period with the same:
          - date
          - description
          - document_ref (if provided)
          - external_ref (if provided)
          - status (draft/posted)
        and which has one debit line and one credit line matching the selected
        accounts and amounts.
        """
        try:
            conn = getattr(self.engine, "conn", None)
            period_id = getattr(self.engine, "current_period_id", None)
            if not conn:
                return False

            params: list[object] = [date, desc]
            where_extra: list[str] = []

            if doc_ref:
                where_extra.append("je.document_ref = ?")
                params.append(doc_ref)
            else:
                where_extra.append("je.document_ref IS NULL")

            if ext_ref:
                where_extra.append("je.external_ref = ?")
                params.append(ext_ref)
            else:
                where_extra.append("je.external_ref IS NULL")

            where_extra.append("je.status = ?")
            params.append(status)

            if period_id:
                where_extra.append("je.period_id = ?")
                params.append(int(period_id))

            # Amounts and account ids (rounded for safety)
            params.extend(
                [
                    float(round(debit_amt, 2)),
                    int(debit_account_id),
                    float(round(credit_amt, 2)),
                    int(credit_account_id),
                ]
            )

            where_clause = " AND ".join(where_extra)
            sql = f"""
                SELECT 1
                FROM journal_entries je
                JOIN journal_lines jl_deb
                    ON jl_deb.entry_id = je.id
                JOIN journal_lines jl_cred
                    ON jl_cred.entry_id = je.id
                WHERE
                    je.date = ?
                    AND je.description = ?
                    AND {where_clause}
                    AND jl_deb.debit = ?
                    AND jl_deb.credit = 0
                    AND jl_deb.account_id = ?
                    AND jl_cred.credit = ?
                    AND jl_cred.debit = 0
                    AND jl_cred.account_id = ?
                LIMIT 1
            """
            cur = conn.execute(sql, params)
            return bool(cur.fetchone())
        except Exception:
            # Never block posting due to a duplicate‑check failure
            return False

    def _record_transaction(self, status: str) -> None:
        # Check permission - viewers cannot create transactions
        if not self._has_permission('create'):
            messagebox.showerror(
                'Access Denied',
                'You do not have permission to create transactions. '
                'Only users with create permission (Accountant, Manager, Admin) can record transactions.'
            )
            return
        
        try:
            date = self.txn_date.get().strip() if hasattr(self, 'txn_date') else ''
            desc = self.txn_desc.get().strip() if hasattr(self, 'txn_desc') else ''
            debit_acct = self.debit_acct.get().strip() if hasattr(self, 'debit_acct') else ''
            credit_acct = self.credit_acct.get().strip() if hasattr(self, 'credit_acct') else ''
            debit_amt_txt = self.debit_amt.get().strip() if hasattr(self, 'debit_amt') else ''
            credit_amt_txt = self.credit_amt.get().strip() if hasattr(self, 'credit_amt') else ''
            doc_ref = self.txn_doc_ref.get().strip() if hasattr(self, 'txn_doc_ref') else None
            ext_ref = self.txn_external_ref.get().strip() if hasattr(self, 'txn_external_ref') else None
            memo = self.txn_memo.get('1.0', tk.END).strip() if hasattr(self, 'txn_memo') else None
            source_type = self.txn_source_type.get().strip() if hasattr(self, 'txn_source_type') else None
            attach = self.txn_attachment_path.get().strip() if hasattr(self, 'txn_attachment_path') else None
            is_adjust = bool(self.txn_is_adjust.get()) if hasattr(self, 'txn_is_adjust') else False
            schedule = bool(self.txn_schedule_reverse.get()) if hasattr(self, 'txn_schedule_reverse') else False
            reverse_on = self.txn_reverse_date.get().strip() if schedule and hasattr(self, 'txn_reverse_date') else None
            def _parse_amt(txt: str) -> float | None:
                try:
                    import re
                    if txt is None:
                        return None
                    s = str(txt).strip()
                    neg = False
                    if s.startswith('(') and s.endswith(')'):
                        neg = True
                        s = s[1:-1]
                    s = re.sub(r"[^0-9.\-]", "", s)
                    if not s:
                        return None
                    val = float(s)
                    if neg:
                        val = -abs(val)
                    return val
                except Exception:
                    return None

            if status == 'draft':
                if not date:
                    date = datetime.now().strftime('%Y-%m-%d')
                if not desc:
                    desc = 'Draft'
                lines: list[JournalLine] = []
                try:
                    if debit_acct and credit_acct and debit_amt_txt and credit_amt_txt:
                        debit_amt = _parse_amt(debit_amt_txt)
                        credit_amt = _parse_amt(credit_amt_txt)
                        if debit_amt is None or credit_amt is None:
                            raise ValueError('Amounts must be numeric')
                        if round(debit_amt - credit_amt, 2) == 0:
                            did = self._resolve_account_id(debit_acct)
                            cid = self._resolve_account_id(credit_acct)
                            if did and cid:
                                lines = [JournalLine(account_id=did, debit=debit_amt), JournalLine(account_id=cid, credit=credit_amt)]
                except Exception:
                    lines = []
                entry_id = self.engine.record_entry(
                    date,
                    desc,
                    lines,
                    is_adjusting=is_adjust,
                    document_ref=doc_ref or None,
                    external_ref=ext_ref or None,
                    memo=memo or None,
                    source_type=source_type or None,
                    status=status,
                    attachments=[('document', attach)] if attach else None,
                    schedule_reverse_on=reverse_on or None,
                )
                messagebox.showinfo('Saved', f'Draft {entry_id} saved')
                try:
                    if self.current_period_id:
                        self.period_form_cache[int(self.current_period_id)] = self._snapshot_txn_form()
                except Exception:
                    pass
                self._clear_transaction_form()
                try:
                    self._load_recent_transactions()
                except Exception:
                    pass
                return
            if not date:
                date = datetime.now().strftime('%Y-%m-%d')
            if not desc:
                desc = 'Transaction'
            if (not debit_acct and not credit_acct) and not getattr(self, '_accounts_prefilled', False):
                try:
                    da, ca = self._infer_accounts_from_context(desc, source_type)
                except Exception:
                    da, ca = (None, None)
                if (not debit_acct) and da:
                    try:
                        self.debit_acct.set(da)
                    except Exception:
                        pass
                    debit_acct = da
                if (not credit_acct) and ca:
                    try:
                        self.credit_acct.set(ca)
                    except Exception:
                        pass
                    credit_acct = ca
            if not debit_acct or not credit_acct:
                messagebox.showerror('Error', 'Select both debit and credit accounts')
                return
            # If only one amount provided, mirror to the other to form a balanced entry
            if debit_amt_txt and not credit_amt_txt:
                credit_amt_txt = debit_amt_txt
                try:
                    self.credit_amt.delete(0, tk.END); self.credit_amt.insert(0, credit_amt_txt)
                except Exception:
                    pass
            elif credit_amt_txt and not debit_amt_txt:
                debit_amt_txt = credit_amt_txt
                try:
                    self.debit_amt.delete(0, tk.END); self.debit_amt.insert(0, debit_amt_txt)
                except Exception:
                    pass
            if not debit_acct or not credit_acct:
                messagebox.showerror('Error', 'Select both debit and credit accounts')
                return
            debit_amt = _parse_amt(debit_amt_txt)
            credit_amt = _parse_amt(credit_amt_txt)
            if debit_amt is None or credit_amt is None:
                messagebox.showerror('Error', 'Amounts must be numeric (e.g., 1500 or 1,500.00)')
                return
            if round(debit_amt - credit_amt, 2) != 0:
                messagebox.showerror('Error', 'Debits must equal credits')
                return
            did = self._resolve_account_id(debit_acct)
            cid = self._resolve_account_id(credit_acct)
            
            # Validate that both account IDs were resolved successfully
            if not did:
                messagebox.showerror('Error', f'Could not resolve debit account: "{debit_acct}". Please select a valid account.')
                return
            if not cid:
                messagebox.showerror('Error', f'Could not resolve credit account: "{credit_acct}". Please select a valid account.')
                return
            
            lines = [JournalLine(account_id=did, debit=debit_amt), JournalLine(account_id=cid, credit=credit_amt)]
            # Duplicate-entry protection: prevent accidentally posting the same
            # transaction twice from the Transactions tab.
            if did and cid:
                try:
                    if self._is_duplicate_transaction(
                        date=date,
                        desc=desc,
                        debit_account_id=int(did),
                        credit_account_id=int(cid),
                        debit_amt=float(debit_amt),
                        credit_amt=float(credit_amt),
                        doc_ref=doc_ref or None,
                        ext_ref=ext_ref or None,
                        status=status,
                    ):
                        messagebox.showerror(
                            "Duplicate Transaction",
                            "This transaction appears to be a duplicate of one that is already "
                            "recorded for this period (same date, description, reference, accounts, "
                            "and amounts).\n\n"
                            "The entry has NOT been recorded again.",
                        )
                        return
                except Exception:
                    # If duplicate check fails, continue with normal posting.
                    pass
            entry_id = self.engine.record_entry(
                date,
                desc,
                lines,
                is_adjusting=is_adjust,
                document_ref=doc_ref or None,
                external_ref=ext_ref or None,
                memo=memo or None,
                source_type=source_type or None,
                status=status,
                attachments=[('document', attach)] if attach else None,
                schedule_reverse_on=reverse_on or None,
            )
            messagebox.showinfo('Recorded', f'Journal entry {entry_id} recorded')
            try:
                if self.current_period_id:
                    self.period_form_cache[int(self.current_period_id)] = self._snapshot_txn_form()
            except Exception:
                pass
            self._refresh_after_post()
            self._clear_transaction_form()
        except Exception as e:
            # Show the existing user‑facing message, but also log for diagnostics.
            try:
                self._handle_exception("record_transaction", e)
            except Exception:
                pass
            messagebox.showerror('Error', f'Failed to record transaction: {e}')

    def _on_ledger_entry_double_click(self, event):
        """Handle double-click on a ledger entry to show detailed activity for that account."""
        try:
            if not hasattr(self, "ledger_tree"):
                return
            sel = self.ledger_tree.selection()
            if not sel:
                return
            item = sel[0]
            values = self.ledger_tree.item(item, "values")
            if not values:
                return
            raw_label = values[0]
            # Ledger displays either "CODE - Name" or just "Name"
            account_display = (raw_label or "").split(" (", 1)[0]
            code = None
            name = account_display
            if " - " in account_display:
                code, name = account_display.split(" - ", 1)
            # Prefer lookup by code if available, otherwise by name.
            acc = None
            try:
                if code:
                    acc = db.get_account_by_code(code, conn=self.engine.conn)
                if not acc:
                    acc = db.get_account_by_name(name, self.engine.conn)
            except Exception:
                acc = None
            if not acc:
                messagebox.showwarning("Account Not Found", f"Could not resolve account: {account_display}")
                return
            self._view_account_details(acc)
        except Exception as e:
            try:
                self._handle_exception("ledger_entry_double_click", e)
            except Exception:
                pass

    def _view_account_details(self, account_row):
        """
        Show detailed transactions for the selected account in a modal dialog.

        account_row is expected to be a sqlite3.Row (or mapping) with at least
        id, code, and name fields.
        """
        try:
            acc_id = int(account_row["id"])
            acc_code = str(account_row.get("code") if hasattr(account_row, "keys") else account_row["code"])
            acc_name = str(account_row.get("name") if hasattr(account_row, "keys") else account_row["name"])
        except Exception:
            # Fallback: try basic dict-style access
            try:
                acc_id = int(account_row["id"])
                acc_code = str(account_row["code"])
                acc_name = str(account_row["name"])
            except Exception as e:
                try:
                    self._handle_exception("view_account_details_resolve", e)
                except Exception:
                    pass
                messagebox.showerror("Error", "Unable to resolve selected account details.")
                return

        # Build dialog
        dlg = tk.Toplevel(self)
        dlg.title(f"Account Details – {acc_code} {acc_name}")
        try:
            dlg.configure(bg=self.palette.get("surface_bg", "#ffffff"))
        except Exception:
            pass
        dlg.transient(self)
        dlg.grab_set()

        # Header
        header = ttk.Frame(dlg, style="Techfix.Surface.TFrame")
        header.pack(fill=tk.X, padx=12, pady=(10, 4))
        ttk.Label(
            header,
            text=f"{acc_code} – {acc_name}",
            style="Techfix.Headline.TLabel",
        ).pack(side=tk.LEFT, fill=tk.X, expand=True)

        # Date range + controls
        controls = ttk.Frame(dlg, style="Techfix.Surface.TFrame")
        controls.pack(fill=tk.X, padx=12, pady=4)

        ttk.Label(controls, text="From (YYYY-MM-DD):", style="Techfix.TLabel").pack(side=tk.LEFT, padx=(0, 4))
        from_entry = ttk.Entry(controls, width=12, style="Techfix.TEntry")
        from_entry.pack(side=tk.LEFT, padx=(0, 8))

        ttk.Label(controls, text="To (YYYY-MM-DD):", style="Techfix.TLabel").pack(side=tk.LEFT, padx=(0, 4))
        to_entry = ttk.Entry(controls, width=12, style="Techfix.TEntry")
        to_entry.pack(side=tk.LEFT, padx=(0, 8))

        # Default to period bounds if available
        try:
            period = getattr(self.engine, "current_period", None)
            if period and "start_date" in period.keys() and period["start_date"]:
                from_entry.insert(0, period["start_date"])
            if period and "end_date" in period.keys() and period["end_date"]:
                to_entry.insert(0, period["end_date"])
        except Exception:
            pass

        btn_frame = ttk.Frame(controls, style="Techfix.Surface.TFrame")
        btn_frame.pack(side=tk.RIGHT)
        refresh_btn = ttk.Button(
            btn_frame,
            text="🔄 Refresh",
            style="Techfix.TButton",
        )
        export_btn = ttk.Button(
            btn_frame,
            text="💾 Export to Excel",
            style="Techfix.TButton",
        )
        close_btn = ttk.Button(
            btn_frame,
            text="Close",
            command=dlg.destroy,
            style="Techfix.TButton",
        )
        refresh_btn.pack(side=tk.LEFT, padx=4)
        export_btn.pack(side=tk.LEFT, padx=4)
        close_btn.pack(side=tk.LEFT, padx=4)

        # Treeview for detailed transactions
        body = ttk.Frame(dlg, style="Techfix.Surface.TFrame")
        body.pack(fill=tk.BOTH, expand=True, padx=12, pady=(0, 12))

        columns = ("date", "entry_id", "description", "debit", "credit", "balance")
        tree = ttk.Treeview(
            body,
            columns=columns,
            show="headings",
            selectmode="browse",
            style="Techfix.Treeview",
        )
        for col, text, anchor in [
            ("date", "Date", tk.W),
            ("entry_id", "Entry #", tk.W),
            ("description", "Description", tk.W),
            ("debit", "Debit", tk.E),
            ("credit", "Credit", tk.E),
            ("balance", "Running Balance", tk.E),
        ]:
            tree.heading(col, text=text, anchor=anchor)
        tree.column("date", width=100, anchor=tk.W)
        tree.column("entry_id", width=70, anchor=tk.W)
        tree.column("description", width=260, anchor=tk.W)
        tree.column("debit", width=100, anchor=tk.E)
        tree.column("credit", width=100, anchor=tk.E)
        tree.column("balance", width=140, anchor=tk.E)

        vsb = ttk.Scrollbar(body, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        # Data loader
        def load_rows() -> None:
            for item in tree.get_children():
                tree.delete(item)
            try:
                date_from = from_entry.get().strip() or None
                date_to = to_entry.get().strip() or None
                params = [acc_id]
                where = ["jl.account_id = ?"]
                if date_from:
                    where.append("je.date >= ?")
                    params.append(date_from)
                if date_to:
                    where.append("je.date <= ?")
                    params.append(date_to)
                where_sql = " AND ".join(where)
                cur = self.engine.conn.execute(
                    f"""
                    SELECT
                        je.id AS entry_id,
                        je.date,
                        je.description,
                        jl.debit,
                        jl.credit
                    FROM journal_lines jl
                    JOIN journal_entries je ON je.id = jl.entry_id
                    WHERE {where_sql}
                    ORDER BY je.date, je.id, jl.id
                    """,
                    params,
                )
                bal = 0.0
                for row in cur.fetchall():
                    d = float(row["debit"] or 0.0)
                    c = float(row["credit"] or 0.0)
                    bal += d - c
                    tree.insert(
                        "",
                        "end",
                        values=(
                            row["date"],
                            row["entry_id"],
                            row["description"],
                            f"{d:,.2f}" if d else "",
                            f"{c:,.2f}" if c else "",
                            f"{bal:,.2f}",
                        ),
                    )
            except Exception as e:
                try:
                    self._handle_exception("view_account_details_load", e)
                except Exception:
                    pass
                messagebox.showerror("Error", f"Failed to load account details: {e}")

        # Wire buttons after definition so they can call load_rows
        refresh_btn.configure(command=load_rows)

        def do_export() -> None:
            try:
                default_name = f"account_{acc_code}_{acc_name}.xlsx".replace(" ", "_")
                self._export_tree_to_excel(tree, default_name=default_name)
            except Exception as e:
                try:
                    self._handle_exception("view_account_details_export", e)
                except Exception:
                    pass
                messagebox.showerror("Error", f"Failed to export account details: {e}")

        export_btn.configure(command=do_export)

        # Initial load and focus
        load_rows()
        try:
            dlg.focus_set()
        except Exception:
            pass

    def _build_fs_tab(self) -> None:
        frame = self.tab_fs
        
        # Date range controls
        controls = ttk.Frame(frame, style="Techfix.Surface.TFrame")
        controls.pack(fill=tk.X, padx=12, pady=12)
        
        # Date range selection
        date_frame = ttk.Frame(controls, style="Techfix.Surface.TFrame")
        date_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # Use surface-style labels so text blends with the bar in all themes
        ttk.Label(date_frame, text="From (YYYY-MM-DD):", style="Techfix.TLabel").pack(side=tk.LEFT, padx=(0, 4))
        self.fs_date_from = ttk.Entry(date_frame, width=12, style="Techfix.TEntry")
        self.fs_date_from.pack(side=tk.LEFT, padx=(0, 12))
        
        ttk.Label(date_frame, text="To (YYYY-MM-DD):", style="Techfix.TLabel").pack(side=tk.LEFT, padx=(0, 4))
        self.fs_date_to = ttk.Entry(date_frame, width=12, style="Techfix.TEntry")
        self.fs_date_to.pack(side=tk.LEFT)
        
        # Default to last day of current month
        import datetime
        import calendar
        today = datetime.date.today()
        # Get last day of current month
        last_day = calendar.monthrange(today.year, today.month)[1]
        last_day_of_month = today.replace(day=last_day).strftime("%Y-%m-%d")
        self.fs_date_to.insert(0, last_day_of_month)
        
        # Action buttons + presets - use grid for responsive layout
        btn_frame = ttk.Frame(controls, style="Techfix.Surface.TFrame")
        btn_frame.pack(side=tk.RIGHT)
        
        # Configure btn_frame for responsive wrapping
        btn_frame.columnconfigure(0, weight=0)
        btn_frame.columnconfigure(1, weight=0)
        btn_frame.columnconfigure(2, weight=0)
        btn_frame.columnconfigure(3, weight=0)
        btn_frame.rowconfigure(0, weight=0)
        btn_frame.rowconfigure(1, weight=0)  # Second row for wrapping

        # Preset selector (e.g. Last Month, YTD)
        preset_var = tk.StringVar(value="Custom")
        self.fs_preset_var = preset_var
        preset_box = ttk.Combobox(
            btn_frame,
            textvariable=preset_var,
            state="readonly",
            width=14,
            values=["Custom", "This Month", "Last Month", "Year to Date"],
            style="Techfix.TCombobox",
        )
        preset_box.grid(row=0, column=0, padx=(0, 4), pady=2, sticky="w")

        def _apply_preset(name: str) -> None:
            try:
                import datetime as _dt
                import calendar
                today = _dt.date.today()
                if name == "This Month":
                    start = today.replace(day=1)
                    # Use last day of current month
                    last_day = calendar.monthrange(today.year, today.month)[1]
                    end = today.replace(day=last_day)
                elif name == "Last Month":
                    first_this = today.replace(day=1)
                    last_month_end = first_this - _dt.timedelta(days=1)
                    start = last_month_end.replace(day=1)
                    end = last_month_end
                elif name == "Year to Date":
                    start = today.replace(month=1, day=1)
                    # Use last day of current month for year-to-date
                    last_day = calendar.monthrange(today.year, today.month)[1]
                    end = today.replace(day=last_day)
                else:
                    return
                self.fs_date_from.delete(0, tk.END)
                self.fs_date_from.insert(0, start.isoformat())
                self.fs_date_to.delete(0, tk.END)
                self.fs_date_to.insert(0, end.isoformat())
            except Exception as e:
                try:
                    self._handle_exception("fs_apply_preset", e)
                except Exception:
                    pass

        def _on_preset_change(event=None) -> None:
            name = preset_var.get()
            _apply_preset(name)

        preset_box.bind("<<ComboboxSelected>>", _on_preset_change)
        
        def load_financials_with_loading():
            loading = self.show_loading("Please wait... Generating financial statements...")
            try:
                self.update()  # Update UI to show loading indicator
                self._load_financials()
            finally:
                if loading:
                    self.hide_loading(loading)
        
        run_btn = ttk.Button(
            btn_frame, 
            text="📊 Run Report", 
            command=load_financials_with_loading, 
            style="Techfix.TButton"
        )
        run_btn.grid(row=0, column=1, padx=4, pady=2, sticky="w")
        
        export_xls_btn = ttk.Button(
            btn_frame,
            text="💾 Export to Excel",
            command=self._export_fs,
            style="Techfix.TButton"
        )
        export_xls_btn.grid(row=0, column=2, padx=4, pady=2, sticky="w")
        
        export_txt_btn = ttk.Button(
            btn_frame,
            text="💾 Export to Text",
            command=self._export_financials,
            style="Techfix.TButton"
        )
        export_txt_btn.grid(row=0, column=3, padx=4, pady=2, sticky="w")
        
        # Store references for responsive layout
        self.fs_btn_frame = btn_frame
        self.fs_preset_box = preset_box
        self.fs_run_btn = run_btn
        self.fs_export_xls_btn = export_xls_btn
        self.fs_export_txt_btn = export_txt_btn
        
        # Create notebook for different financial statements
        self.fs_notebook = ttk.Notebook(frame, style="Techfix.TNotebook")
        self.fs_notebook.pack(fill=tk.BOTH, expand=True, padx=12, pady=(0, 12))
        
        # Income Statement Tab
        self.income_statement_frame = ttk.Frame(self.fs_notebook, style="Techfix.Surface.TFrame")
        self.fs_notebook.add(self.income_statement_frame, text="Income Statement")
        
        # Balance Sheet Tab
        self.balance_sheet_frame = ttk.Frame(self.fs_notebook, style="Techfix.Surface.TFrame")
        self.fs_notebook.add(self.balance_sheet_frame, text="Balance Sheet")
        
        # Cash Flow Tab (placeholder for future implementation)
        self.cash_flow_frame = ttk.Frame(self.fs_notebook, style="Techfix.Surface.TFrame")
        self.fs_notebook.add(self.cash_flow_frame, text="Cash Flow")
        try:
            self.fs_notebook.bind("<<NotebookTabChanged>>", self._on_fs_tab_changed)
        except Exception:
            pass
        
        # Create text widgets for each statement
        self._create_fs_text_widgets()

        # Keyboard shortcuts for Financial Statements
        try:
            # F5 runs financials, Ctrl+E exports current statement to Excel, Ctrl+T exports to text
            frame.bind_all("<F5>", lambda e: self._load_financials())
            frame.bind_all("<Control-e>", lambda e: self._export_fs())
            frame.bind_all("<Control-E>", lambda e: self._export_fs())
            frame.bind_all("<Control-t>", lambda e: self._export_financials())
            frame.bind_all("<Control-T>", lambda e: self._export_financials())
        except Exception:
            pass

    def _on_fs_tab_changed(self, event=None):
        try:
            w = event.widget if event and hasattr(event, 'widget') else getattr(self, 'fs_notebook', None)
            if not w:
                return
            tab_id = w.select()
            if not tab_id:
                return
            try:
                frame = self.nametowidget(tab_id)
            except Exception:
                frame = None
            if frame:
                try:
                    self._animate_tab_pulse(frame)
                except Exception:
                    pass
        except Exception:
            pass

    def _build_trial_tab(self) -> None:
        """Build the Trial Balance tab"""
        frame = self.tab_trial

        # Controls frame with refresh
        controls = ttk.Frame(frame, style="Techfix.Surface.TFrame")
        controls.pack(fill=tk.X, padx=12, pady=8)

        ttk.Label(controls, text="As of (YYYY-MM-DD):", style="Techfix.TLabel").pack(side=tk.LEFT)
        self.tb_date = ttk.Entry(controls, width=14, style="Techfix.TEntry")
        self.tb_date.pack(side=tk.LEFT, padx=(6, 12))
        # Make this date field blend with the toolbar in both themes (no bright OS border)
        try:
            self.tb_date.configure(
                highlightthickness=0,
                bd=0,
                relief=tk.FLAT,
                highlightbackground=self.palette.get("surface_bg", "#ffffff"),
                highlightcolor=self.palette.get("surface_bg", "#ffffff"),
            )
        except Exception:
            pass

        ttk.Button(controls, text="Refresh", command=self._load_trial_balances, style="Techfix.TButton").pack(side=tk.LEFT)
        # Export trial balance to Excel
        ttk.Button(controls, text="Export to Excel", command=lambda: self._export_tree_to_excel(self.trial_tree, default_name=f"trial_balance_{self.tb_date.get() if hasattr(self, 'tb_date') else ''}.xlsx"), style="Techfix.TButton").pack(side=tk.LEFT, padx=(6,0))
        self.tb_status_label = ttk.Label(controls, text="Trial Balance", style="Techfix.TLabel")
        self.tb_status_label.pack(side=tk.RIGHT)

        cols = ("code", "name", "debit", "credit")
        self.trial_tree = ttk.Treeview(frame, columns=cols, show="headings", style="Techfix.Treeview")
        for c in cols:
            anchor = tk.E if c in ("debit", "credit") else tk.W
            width = 120 if c in ("debit", "credit") else 220
            self.trial_tree.heading(c, text=c.title(), anchor=anchor)
            self.trial_tree.column(c, stretch=True, width=width, anchor=anchor)

        vsb = ttk.Scrollbar(frame, orient="vertical", command=self.trial_tree.yview)
        self.trial_tree.configure(yscrollcommand=vsb.set)
        self.trial_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=12, pady=(0, 12))
        vsb.pack(side=tk.RIGHT, fill=tk.Y, pady=(0, 12))

        # Load initial data
        try:
            self._load_trial_balances()
        except Exception:
            pass

    def _load_trial_balances(self) -> None:
        """Compute and display the trial balance"""
        # Clear existing items
        if hasattr(self, 'trial_tree'):
            for item in self.trial_tree.get_children():
                self.trial_tree.delete(item)

        try:
            as_of = None
            if hasattr(self, 'tb_date'):
                d = self.tb_date.get().strip()
                as_of = d or None

            rows = db.compute_trial_balance(up_to_date=as_of, include_temporary=True, period_id=self.engine.current_period_id, conn=self.engine.conn)
            try:
                # Check if there are any adjusting entries in the current period
                # If adjusting entries exist, show "Adjusted Trial Balance", otherwise "Unadjusted Trial Balance"
                cur = self.engine.conn.execute("""
                    SELECT COUNT(*) as count
                    FROM journal_entries
                    WHERE period_id = ? 
                      AND is_adjusting = 1 
                      AND status = 'posted'
                """, (self.engine.current_period_id,))
                result = cur.fetchone()
                has_adjusting_entries = result and result['count'] > 0
                
                if has_adjusting_entries:
                    self.tb_status_label.configure(text="Adjusted Trial Balance")
                else:
                    self.tb_status_label.configure(text="Unadjusted Trial Balance")
            except Exception:
                # Default to showing "Adjusted Trial Balance" since compute_trial_balance includes all entries
                self.tb_status_label.configure(text="Adjusted Trial Balance")
            # Show only accounts with non-zero balance/activity for the active period
            def _has_activity(r: dict) -> bool:
                dcol, ccol = self._balance_to_columns(r)
                return bool((dcol or 0) != 0 or (ccol or 0) != 0)
            rows = [r for r in rows if _has_activity(r)]

            for r in rows:
                dcol, ccol = self._balance_to_columns(r)
                code = r['code'] if 'code' in r.keys() else ''
                name = r['name'] if 'name' in r.keys() else ''
                self.trial_tree.insert('', 'end', values=(code, name, f"{dcol:,.2f}" if dcol else "", f"{ccol:,.2f}" if ccol else ""))
            # Add totals row
            try:
                total_d = 0.0
                total_c = 0.0
                for item in self.trial_tree.get_children():
                    vals = self.trial_tree.item(item, 'values')
                    try:
                        if vals and vals[2]:
                            total_d += float(str(vals[2]).replace(',', ''))
                    except Exception:
                        pass
                    try:
                        if vals and vals[3]:
                            total_c += float(str(vals[3]).replace(',', ''))
                    except Exception:
                        pass
                self.trial_tree.insert('', 'end', values=("", "Totals", f"{total_d:,.2f}", f"{total_c:,.2f}"), tags=('totals',))
                self.trial_tree.tag_configure('totals', background=self.palette.get('tab_selected_bg', '#e0ecff'))
            except Exception:
                pass

        except Exception as e:
            messagebox.showerror("Error", f"Failed to load trial balances: {str(e)}")

    def _export_tree_to_excel(self, tree: ttk.Treeview, *, default_name: str = "export.xlsx") -> None:
        """Export the contents of a Treeview to an Excel file using `db.export_rows_to_excel`.

        Rows are exported in the order they appear in the tree. Column headers are taken
        from `tree['columns']`. The totals row (if present) will be exported as-is.
        """
        try:
            cols = list(tree['columns']) if tree and 'columns' in tree.keys() else []
            headers = [str(c).title() for c in cols]
            rows = []
            for iid in tree.get_children():
                vals = tree.item(iid, 'values') or ()
                # Ensure the row is a simple list matching headers length
                row = [vals[i] if i < len(vals) else '' for i in range(len(headers))]
                rows.append(row)

            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
                initialfile=default_name,
            )
            if not filename:
                return
            from pathlib import Path
            try:
                from openpyxl import Workbook
                from openpyxl.utils import get_column_letter
                from openpyxl.styles import Font
            except Exception:
                # fallback to simple CSV/Excel writer
                db.export_rows_to_excel(rows, headers, Path(filename), sheet_name=Path(filename).stem)
                messagebox.showinfo("Export Successful", f"Exported to {filename}")
                return

            wb = Workbook()
            ws = wb.active
            ws.title = Path(filename).stem[:31]

            # write headers
            for cidx, h in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=cidx, value=h)
                cell.font = Font(bold=True)

            # write rows
            for ridx, row in enumerate(rows, start=2):
                for cidx, val in enumerate(row, start=1):
                    # try to write numeric values as numbers
                    if isinstance(val, (int, float)):
                        ws.cell(row=ridx, column=cidx, value=val)
                    else:
                        # attempt to parse numeric strings
                        s = str(val).replace(',', '') if val is not None else ''
                        try:
                            num = float(s)
                            ws.cell(row=ridx, column=cidx, value=num)
                        except Exception:
                            ws.cell(row=ridx, column=cidx, value=val)

            # Add totals for numeric columns (columns where all except header are numeric)
            max_row = ws.max_row
            for cidx in range(1, len(headers) + 1):
                # check if column appears numeric
                is_numeric = True
                for rr in range(2, max_row + 1):
                    v = ws.cell(row=rr, column=cidx).value
                    if v is None or isinstance(v, (int, float)):
                        continue
                    try:
                        float(str(v).replace(',', ''))
                    except Exception:
                        is_numeric = False
                        break
                if is_numeric and max_row >= 2:
                    col_letter = get_column_letter(cidx)
                    total_row = max_row + 1
                    ws.cell(row=total_row, column=cidx, value=f"=SUM({col_letter}2:{col_letter}{max_row})")

            # auto-width
            for i in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(i)].width = 15

            wb.save(filename)
            messagebox.showinfo("Export Successful", f"Exported to {filename}")
        except Exception as e:
            messagebox.showerror("Export Failed", f"Failed to export to Excel: {e}")

    def _balance_to_columns(self, row: dict) -> tuple[float, float]:
        """Convert a trial balance row into debit/credit columns.

        Returns (debit_amount, credit_amount)
        """
        try:
            # Handle both dict and sqlite3.Row objects
            acc_type = (row['type'] if 'type' in row.keys() and row['type'] else '').lower()
            normal = (row['normal_side'] if 'normal_side' in row.keys() and row['normal_side'] else 'debit').lower()
            net_debit = float(row['net_debit'] if 'net_debit' in row.keys() and row['net_debit'] is not None else 0)
            net_credit = float(row['net_credit'] if 'net_credit' in row.keys() and row['net_credit'] is not None else 0)

            # Contra-assets always have credit balances and should be shown on credit side
            if acc_type == 'contra asset':
                # For contra-assets, credit balance is positive (normal)
                # Show the credit amount on credit side, debit amount (if any) on debit side
                if net_credit > 0:
                    return (0.0, net_credit)
                elif net_debit > 0:
                    # Abnormal debit balance (shouldn't happen, but handle it)
                    return (net_debit, 0.0)
                else:
                    return (0.0, 0.0)

            if normal == 'debit':
                bal = net_debit - net_credit
                if bal >= 0:
                    return (bal, 0.0)
                else:
                    return (0.0, abs(bal))
            else:
                bal = net_credit - net_debit
                if bal >= 0:
                    return (0.0, bal)
                else:
                    return (abs(bal), 0.0)
        except Exception:
            return (0.0, 0.0)
        
    def _create_fs_text_widgets(self):
        """Create and configure text widgets for financial statements"""
        # Create text widgets directly in their respective frames
        self.income_text = self._create_fs_text_widget(self.income_statement_frame)
        self.balance_sheet_text = self._create_fs_text_widget(self.balance_sheet_frame)
        self.cash_flow_text = self._create_fs_text_widget(self.cash_flow_frame)
        
        # Add some content to the cash flow tab
        self.cash_flow_text.insert(tk.END, "Cash Flow Statement\n" + "="*20 + "\n\n")
        self.cash_flow_text.insert(tk.END, "Cash flow statement will be implemented in a future update.\n")
        
    def _create_fs_text_widget(self, parent):
        """Helper to create a consistent text widget for financial statements"""
        # Create a frame to hold everything
        container = ttk.Frame(parent, style="Techfix.Surface.TFrame")
        container.pack(fill=tk.BOTH, expand=True)
        
        # Create text widget with improved styling
        text_widget = tk.Text(
            container,
            wrap=tk.WORD,
            font=FONT_MONO,
            bg=self.palette["surface_bg"],
            fg=self.palette["text_primary"],
            insertbackground=self.palette["text_primary"],
            selectbackground=self.palette["accent_color"],
            selectforeground="white",
            bd=0,
            highlightthickness=0,
            padx=20,
            pady=20
        )
        
        # Configure tags for formatting
        text_widget.tag_configure("header", font=("Segoe UI", 16, "bold"), 
                                foreground=self.palette["accent_color"])
        text_widget.tag_configure("subheader", font=("Segoe UI", 10), 
                                foreground=self.palette["text_secondary"])
        text_widget.tag_configure("section", font=("Segoe UI", 12, "bold"), 
                                foreground=self.palette["text_primary"])
        text_widget.tag_configure("total", font=("Segoe UI", 10, "bold"), 
                                foreground=self.palette["text_primary"])
        text_widget.tag_configure("net", font=("Segoe UI", 12, "bold"), 
                                foreground=self.palette["accent_color"])
        text_widget.tag_configure("warning", foreground="red")
        
        # Add scrollbar
        scrollbar = ttk.Scrollbar(container, orient=tk.VERTICAL, command=text_widget.yview)
        text_widget.configure(yscrollcommand=scrollbar.set)
        
        # Pack them with scrollbar on right, text on left
        text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Make text widget read-only
        text_widget.config(state=tk.DISABLED)
        
        # Add a method to update the text widget
        def update_text(content, tags=None):
            text_widget.config(state=tk.NORMAL)
            text_widget.delete(1.0, tk.END)
            if tags:
                text_widget.insert(tk.END, content, tags)
            else:
                text_widget.insert(tk.END, content)
            text_widget.config(state=tk.DISABLED)
            text_widget.see(tk.END)  # Scroll to the end
            
        # Add the update method to the text widget
        text_widget.update_text = update_text
        
        # Configure grid weights to make the container expand properly
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)
        
        return text_widget
        
    def _export_financials(self):
        """Export the current financial statements to a file"""
        # Get the currently selected tab
        current_tab = self.fs_notebook.select()
        if not current_tab:
            messagebox.showinfo("Export", "Please select a financial statement to export.")
            return
            
        tab_index = self.fs_notebook.index(current_tab)
        tab_name = self.fs_notebook.tab(tab_index, "text")
        
        # Get the text widget for the current tab
        if tab_index == 0:  # Income Statement
            content = self.income_text.get(1.0, tk.END)
        elif tab_index == 1:  # Balance Sheet
            content = self.balance_sheet_text.get(1.0, tk.END)
        else:  # Cash Flow or other tabs
            messagebox.showinfo("Export", "Export not available for this statement yet.")
            return
        
        # Ask for save location
        filename = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Text Files", "*.txt"), ("All Files", "*.*")],
            initialfile=f"{tab_name.replace(' ', '_')}_{self.fs_date_to.get() or 'report'}.txt"
        )
        
        if filename:
            try:
                with open(filename, 'w') as f:
                    f.write(content)
                messagebox.showinfo("Export Successful", f"{tab_name} exported successfully!")
            except Exception as e:
                messagebox.showerror("Export Failed", f"Failed to export {tab_name}: {str(e)}")

    def _format_amount(self, amount: float) -> str:
        """Shared helper to format amounts (use parentheses for negative values)"""
        try:
            amount = float(amount or 0)
        except (ValueError, TypeError):
            return "0.00"
        if amount < 0:
            return f"({abs(amount):,.2f})"
        return f"{amount:,.2f}"
    
    def _update_text_widget(self, widget: tk.Text, content: List[tuple]) -> None:
        """Shared helper to update a text widget with formatted content"""
        widget.config(state=tk.NORMAL)
        widget.delete(1.0, tk.END)
        for text, tag in content:
            if tag:
                widget.insert(tk.END, text, tag)
            else:
                widget.insert(tk.END, text)
        widget.config(state=tk.DISABLED)
    
    def _generate_income_statement(self, trial_balance: list, as_of_date: str = None, start_date: str = None) -> None:
        """Generate and display the income statement using backend method for consistency"""
        try:
            # Check if income_text widget exists
            if not hasattr(self, 'income_text'):
                logger.warning("Income statement text widget not initialized yet")
                return
            
            # Determine date range for income statement
            # Use provided start_date if available, otherwise determine it
            end_date = as_of_date or date.today().isoformat()
            
            if not start_date:
                try:
                    if self.engine.current_period and 'start_date' in self.engine.current_period:
                        start_date = self.engine.current_period['start_date']
                    else:
                        # Get earliest entry date as fallback
                        cur = self.engine.conn.execute("""
                            SELECT MIN(date) as min_date FROM journal_entries 
                            WHERE period_id = ?
                        """, (self.engine.current_period_id,))
                        result = cur.fetchone()
                        start_date = result['min_date'] if result and result['min_date'] else '1900-01-01'
                except Exception:
                    start_date = '1900-01-01'
            
            # Use backend method for consistent calculation (handles contra-revenue, etc.)
            # Pass period_id=None to allow cross-period reporting when dates are set
            income_stmt = self.engine.generate_income_statement(start_date, end_date, period_id=None)
            
            # Extract data from backend result
            revenues = income_stmt.get('revenues', [])
            expenses = income_stmt.get('expenses', [])
            total_revenue = income_stmt.get('total_revenue', 0.0)
            total_expense = income_stmt.get('total_expense', 0.0)
            net_income = income_stmt.get('net_income', 0.0)
            
            # Build the income statement content
            content = []
            content.append(('Income Statement\n', 'header'))
            content.append((f'For the period ended {end_date}\n\n', 'subheader'))
            
            # Add revenues section
            content.append(('Revenues\n', 'section'))
            if revenues:
                for rev in revenues:
                    amount = rev.get('amount', 0.0)
                    if abs(amount) > 0.005:  # Only show non-zero amounts
                        content.append((f"{rev.get('name', 'Unknown')}: {self._format_amount(amount)}\n", None))
            else:
                content.append(('(none)\n', None))
            content.append((f'\nTotal Revenue: {self._format_amount(total_revenue)}\n\n', 'total'))
            
            # Add expenses section
            content.append(('Expenses\n', 'section'))
            if expenses:
                for exp in expenses:
                    amount = exp.get('amount', 0.0)
                    if abs(amount) > 0.005:  # Only show non-zero amounts
                        content.append((f"{exp.get('name', 'Unknown')}: {self._format_amount(amount)}\n", None))
            else:
                content.append(('(none)\n', None))
            content.append((f'\nTotal Expenses: {self._format_amount(total_expense)}\n\n', 'total'))
            
            # Add net income
            content.append((f'Net Income: {self._format_amount(net_income)}\n', 'net'))
            
            # Update the text widget (check if it exists first)
            if hasattr(self, 'income_text') and self.income_text:
                try:
                    self._update_text_widget(self.income_text, content)
                except Exception as widget_error:
                    logger.error(f"Failed to update income statement widget: {widget_error}", exc_info=True)
            else:
                logger.warning("Income statement text widget not available for update")
            
        except Exception as e:
            logger.error(f"Failed to generate income statement: {str(e)}", exc_info=True)
            # Only show error dialog if widget exists and is visible
            try:
                if hasattr(self, 'income_text') and self.income_text:
                    # Show error in widget
                    self._show_error_in_text_widget(self.income_text, str(e), "income statement")
                    # Only show messagebox if Financial Statements tab is visible
                    try:
                        if hasattr(self, 'fs_notebook') and self.fs_notebook.winfo_viewable():
                            messagebox.showerror("Error", f"Failed to generate income statement: {str(e)}")
                    except Exception:
                        pass
            except Exception:
                pass
    
    def _diagnose_balance_sheet_imbalance(self) -> list[str]:
        """Diagnose common issues that cause balance sheet imbalances."""
        issues = []
        try:
            conn = self.engine.conn
            
            # Check 1: Draft transactions that aren't included in balance sheet
            draft_count = conn.execute("""
                SELECT COUNT(*) as cnt FROM journal_entries 
                WHERE period_id = ? AND status = 'draft'
            """, (self.engine.current_period_id,)).fetchone()
            if draft_count and draft_count['cnt'] > 0:
                issues.append(f"You have {draft_count['cnt']} draft transaction(s). Drafts are NOT included in the balance sheet. Post them using 'Record & Post' button.")
            
            # Check 2: Unbalanced journal entries (debits != credits)
            unbalanced = conn.execute("""
                SELECT je.id, je.date, je.description, 
                       SUM(jl.debit) as total_debit, SUM(jl.credit) as total_credit
                FROM journal_entries je
                JOIN journal_lines jl ON jl.entry_id = je.id
                WHERE je.period_id = ? AND je.status = 'posted'
                GROUP BY je.id
                HAVING ABS(SUM(jl.debit) - SUM(jl.credit)) > 0.01
            """, (self.engine.current_period_id,)).fetchall()
            if unbalanced:
                issues.append(f"You have {len(unbalanced)} unbalanced journal entry/entries. Each entry must have equal debits and credits.")
            
            # Check 3: Entries with only one line (missing debit or credit)
            single_line = conn.execute("""
                SELECT je.id, je.date, je.description, COUNT(jl.id) as line_count
                FROM journal_entries je
                JOIN journal_lines jl ON jl.entry_id = je.id
                WHERE je.period_id = ? AND je.status = 'posted'
                GROUP BY je.id
                HAVING COUNT(jl.id) < 2
            """, (self.engine.current_period_id,)).fetchall()
            if single_line:
                issues.append(f"You have {len(single_line)} entry/entries with only one line. Each transaction needs both a debit and credit.")
            
            # Check 4: Entries with zero amounts
            zero_amount = conn.execute("""
                SELECT je.id, je.date, je.description
                FROM journal_entries je
                JOIN journal_lines jl ON jl.entry_id = je.id
                WHERE je.period_id = ? AND je.status = 'posted'
                  AND jl.debit = 0 AND jl.credit = 0
            """, (self.engine.current_period_id,)).fetchall()
            if zero_amount:
                issues.append(f"You have {len(zero_amount)} entry/entries with zero amounts. These won't affect the balance sheet.")
            
            # Check 5: Missing closing entries (if period should be closed)
            # This is informational only - skip for now as it requires complex aggregation
            
            # Check 6: Transactions not assigned to current period
            wrong_period = conn.execute("""
                SELECT COUNT(*) as cnt FROM journal_entries 
                WHERE period_id IS NULL OR period_id != ?
            """, (self.engine.current_period_id,)).fetchone()
            if wrong_period and wrong_period['cnt'] > 0:
                issues.append(f"You have {wrong_period['cnt']} transaction(s) not assigned to the current period. They won't appear in this period's balance sheet.")
            
        except Exception as e:
            issues.append(f"Error running diagnostics: {str(e)}")
        
        return issues
    
    def _generate_balance_sheet(self, trial_balance: list, as_of_date: str = None) -> None:
        """Generate and display the balance sheet using the backend engine method
        
        Note: The trial_balance parameter is kept for API consistency but is not used.
        The backend generate_balance_sheet method recalculates the trial balance to ensure accuracy.
        """
        try:
            # Determine the date to use
            if not as_of_date:
                try:
                    cur = self.engine.conn.execute("""
                        SELECT MAX(date) as max_date FROM journal_entries 
                        WHERE period_id = ?
                    """, (self.engine.current_period_id,))
                    result = cur.fetchone()
                    as_of_date = result['max_date'] if result and result['max_date'] else date.today().isoformat()
                except Exception:
                    as_of_date = date.today().isoformat()
            
            # If closing entries exist, check if the requested date is before the latest entry
            # This prevents imbalance when viewing balance sheet before all transactions are included
            try:
                closing_entries_exist = False
                closing_count = self.engine.conn.execute("""
                    SELECT COUNT(*) as cnt FROM journal_entries 
                    WHERE is_closing = 1 AND period_id = ?
                """, (self.engine.current_period_id,)).fetchone()
                closing_entries_exist = closing_count and closing_count['cnt'] > 0
                
                if closing_entries_exist:
                    # Get the latest entry date
                    latest_date = self.engine.conn.execute("""
                        SELECT MAX(date) as max_date FROM journal_entries 
                        WHERE period_id = ? AND (status = 'posted' OR status IS NULL)
                    """, (self.engine.current_period_id,)).fetchone()
                    
                    if latest_date and latest_date['max_date']:
                        latest_entry_date = latest_date['max_date']
                        # If requested date is before latest entry, use latest entry date to avoid imbalance
                        if as_of_date < latest_entry_date:
                            as_of_date = latest_entry_date
            except Exception:
                pass  # If check fails, just use the original date
            
            # Generate balance sheet using backend method
            # The backend method recalculates the trial balance with include_temporary=False
            # to ensure only permanent accounts are included in the balance sheet
            balance_sheet = self.engine.generate_balance_sheet(as_of_date)
            
            # Extract data from backend result
            assets = balance_sheet.get('assets', [])
            liabilities = balance_sheet.get('liabilities', [])
            equity = balance_sheet.get('equity', [])
            total_assets = balance_sheet.get('total_assets', 0.0)
            total_liabilities = balance_sheet.get('total_liabilities', 0.0)
            total_equity = balance_sheet.get('total_equity', 0.0)
            balance_check = balance_sheet.get('balance_check', 0.0)
            
            # ALWAYS recalculate totals from individual items to ensure accuracy
            # This fixes any discrepancies between backend calculation and display
            calculated_assets = sum(asset.get('amount', 0.0) for asset in assets)
            calculated_liabilities = sum(liab.get('amount', 0.0) for liab in liabilities)
            calculated_equity = sum(eq.get('amount', 0.0) for eq in equity)
            
            # Log if there's a discrepancy (for debugging)
            if abs(total_assets - calculated_assets) > 0.01:
                try:
                    logger.warning(
                        f"Balance sheet asset total mismatch: "
                        f"reported={total_assets:.2f}, calculated={calculated_assets:.2f}, "
                        f"difference={abs(total_assets - calculated_assets):.2f}"
                    )
                except Exception:
                    pass
            
            if abs(total_liabilities - calculated_liabilities) > 0.01:
                try:
                    logger.warning(
                        f"Balance sheet liability total mismatch: "
                        f"reported={total_liabilities:.2f}, calculated={calculated_liabilities:.2f}"
                    )
                except Exception:
                    pass
            
            if abs(total_equity - calculated_equity) > 0.01:
                try:
                    logger.warning(
                        f"Balance sheet equity total mismatch: "
                        f"reported={total_equity:.2f}, calculated={calculated_equity:.2f}"
                    )
                except Exception:
                    pass
            
            # ALWAYS use calculated totals (sum of displayed items) for display
            # This ensures the displayed total matches what the user sees
            # This fixes the issue where backend might return incorrect totals
            total_assets = round(calculated_assets, 2)
            total_liabilities = round(calculated_liabilities, 2)
            total_equity = round(calculated_equity, 2)
            
            # Recalculate balance_check with corrected totals
            # Formula: Assets = Liabilities + Equity (should equal 0)
            balance_check = round(total_assets - (total_liabilities + total_equity), 2)
            
            # Final verification - if still unbalanced, add diagnostic info
            if abs(balance_check) > 0.01:
                # Double-check by recalculating from displayed items
                displayed_assets_sum = sum(
                    asset.get('amount', 0.0) 
                    for asset in assets 
                    if abs(asset.get('amount', 0.0)) > 0.005
                )
                displayed_liab_sum = sum(
                    liab.get('amount', 0.0) 
                    for liab in liabilities 
                    if abs(liab.get('amount', 0.0)) > 0.005
                )
                displayed_equity_sum = sum(
                    eq.get('amount', 0.0) 
                    for eq in equity 
                    if abs(eq.get('amount', 0.0)) > 0.005
                )
                
                # Use displayed sums if they differ
                if abs(total_assets - displayed_assets_sum) > 0.01:
                    total_assets = round(displayed_assets_sum, 2)
                if abs(total_liabilities - displayed_liab_sum) > 0.01:
                    total_liabilities = round(displayed_liab_sum, 2)
                if abs(total_equity - displayed_equity_sum) > 0.01:
                    total_equity = round(displayed_equity_sum, 2)
                
                # Recalculate balance check one more time
                balance_check = round(total_assets - (total_liabilities + total_equity), 2)
            
            # Build the balance sheet content
            content = []
            content.append(("Balance Sheet\n", "header"))
            content.append((f"As of {as_of_date}\n", "subheader"))
            content.append(("\n", None))
            
            # Add assets section
            content.append(("Assets\n", "section"))
            if assets:
                for asset in assets:
                    amount = asset.get('amount', 0.0)
                    if abs(amount) > 0.005:  # Only show accounts with non-zero balances
                        content.append((f"{asset.get('name', 'Unknown')}: {self._format_amount(amount)}\n", None))
            else:
                content.append(("(none)\n", None))
            # Add Total Assets row
            content.append(("─" * 50 + "\n", None))
            content.append((f"Total Assets: {self._format_amount(total_assets)}\n\n", "total"))
            
            # Add liabilities section
            content.append(("Liabilities\n", "section"))
            if liabilities:
                for liab in liabilities:
                    amount = liab.get('amount', 0.0)
                    if abs(amount) > 0.005:  # Only show accounts with non-zero balances
                        content.append((f"{liab.get('name', 'Unknown')}: {self._format_amount(amount)}\n", None))
            else:
                content.append(("(none)\n", None))
            # Add Total Liabilities row
            content.append(("─" * 50 + "\n", None))
            content.append((f"Total Liabilities: {self._format_amount(total_liabilities)}\n\n", "total"))
            
            # Add equity section
            content.append(("Equity\n", "section"))
            if equity:
                for eq in equity:
                    amount = eq.get('amount', 0.0)
                    if abs(amount) > 0.005:  # Only show accounts with non-zero balances
                        content.append((f"{eq.get('name', 'Unknown')}: {self._format_amount(amount)}\n", None))
            else:
                content.append(("(none)\n", None))
            # Add Total Equity row
            content.append(("─" * 50 + "\n", None))
            content.append((f"Total Equity: {self._format_amount(total_equity)}\n\n", "total"))
            
            # Check if closing entries have been completed (step 8)
            # After closing entries, Net Income is already included in Owner's Capital
            # So we need different balance check logic for open vs closed periods
            closing_completed = False
            try:
                statuses = self.engine.get_cycle_status()
                step8 = next((r for r in statuses if int(r['step']) == 8), None)
                # Check both status and if closing entries actually exist
                status_completed = step8 and (step8['status'] == 'completed')
                
                # Also check if closing entries exist in database (more reliable)
                closing_entries_exist = False
                try:
                    closing_count = self.engine.conn.execute("""
                        SELECT COUNT(*) as cnt FROM journal_entries 
                        WHERE is_closing = 1 AND period_id = ?
                    """, (self.engine.current_period_id,)).fetchone()
                    closing_entries_exist = closing_count and closing_count['cnt'] > 0
                except Exception:
                    pass
                
                # Closing is completed if status says so OR if closing entries exist
                closing_completed = status_completed or closing_entries_exist
            except Exception:
                pass
            
            # Calculate net income from income statement for clearer breakdown
            # Get net income from the current period's income statement
            net_income = 0.0
            try:
                if hasattr(self, 'fs_date_from') and hasattr(self, 'fs_date_to'):
                    date_from = self.fs_date_from.get().strip() or None
                    date_to = self.fs_date_to.get().strip() or None
                    if date_from and date_to:
                        income_stmt = self.engine.generate_income_statement(date_from, date_to, period_id=None)
                        net_income = income_stmt.get('net_income', 0.0)
            except Exception:
                pass
            
            # Add summary section at the bottom with all totals (matching FINAL_ACCOUNTING.py format)
            content.append(("=" * 50 + "\n", None))
            content.append(("SUMMARY\n", "section"))
            content.append(("=" * 50 + "\n", None))
            content.append((f"Total Assets: {self._format_amount(total_assets)}\n", "total"))
            content.append((f"Total Liabilities: {self._format_amount(total_liabilities)}\n", "total"))
            
            if closing_completed:
                # After closing entries, Net Income is already in Owner's Capital
                content.append((f"Total Equity (Net Income already included): {self._format_amount(total_equity)}\n", "total"))
                content.append(("─" * 50 + "\n", None))
                total_liab_equity = total_liabilities + total_equity
                content.append((f"Total Liabilities + Equity: {self._format_amount(total_liab_equity)}\n", "total"))
                content.append(("=" * 50 + "\n", None))
                # GRAND TOTAL line (should equal Total Assets when balanced)
                content.append((f"GRAND TOTAL (Assets = L + E): {self._format_amount(total_assets)}\n", "total"))
                content.append(("=" * 50 + "\n\n", None))
                
                # Balance check for closed periods: Assets = Liabilities + Equity
                balance_check = total_assets - (total_liabilities + total_equity)
            else:
                # Before closing entries, Net Income is separate
                content.append((f"Total Equity (excluding Net Income): {self._format_amount(total_equity)}\n", "total"))
                content.append((f"Total Equity (including Net Income): {self._format_amount(total_equity + net_income)}\n", "total"))
                content.append(("─" * 50 + "\n", None))
                total_liab_equity_net = total_liabilities + total_equity + net_income
                content.append((f"Total Liabilities + Equity + Net Income: {self._format_amount(total_liab_equity_net)}\n", "total"))
                content.append(("=" * 50 + "\n", None))
                # GRAND TOTAL line (should equal Total Assets when balanced)
                content.append((f"GRAND TOTAL (Assets = L + E + Net Income): {self._format_amount(total_assets)}\n", "total"))
                content.append(("=" * 50 + "\n\n", None))
                
                # Balance check for open periods: Assets = Liabilities + Equity + Net Income
                balance_check = total_assets - (total_liabilities + total_equity + net_income)
            
            # Check accounting equation (balance_check should be 0.00)
            if abs(balance_check) > 0.05:  # Allow for small floating point differences
                if closing_completed:
                    # Check if the imbalance is due to viewing balance sheet before all transactions
                    try:
                        latest_date = self.engine.conn.execute("""
                            SELECT MAX(date) as max_date FROM journal_entries 
                            WHERE period_id = ? AND (status = 'posted' OR status IS NULL)
                        """, (self.engine.current_period_id,)).fetchone()
                        
                        if latest_date and latest_date['max_date'] and as_of_date < latest_date['max_date']:
                            content.append((
                                f"\n⚠ Warning: Accounting equation does not balance!\n"
                                f"Assets ({self._format_amount(total_assets)}) ≠ Liabilities + Equity ({self._format_amount(total_liabilities + total_equity)})\n"
                                f"Difference: {self._format_amount(abs(balance_check))}\n\n"
                                f"Note: You are viewing the balance sheet 'As of {as_of_date}', but closing entries\n"
                                f"closed revenue/expenses from the entire period (up to {latest_date['max_date']}).\n"
                                f"Assets only include transactions up to {as_of_date}, creating this imbalance.\n\n"
                                f"To see a balanced sheet, view 'As of {latest_date['max_date']}' (latest entry date).\n",
                                "warning"
                            ))
                        else:
                            content.append((
                                f"\n⚠ Warning: Accounting equation does not balance!\n"
                                f"Assets ({self._format_amount(total_assets)}) ≠ Liabilities + Equity ({self._format_amount(total_liabilities + total_equity)})\n"
                                f"Difference: {self._format_amount(abs(balance_check))}\n", 
                                "warning"
                            ))
                            # Add diagnostic information
                            try:
                                diagnostics = self._diagnose_balance_sheet_imbalance()
                                if diagnostics:
                                    content.append(("\nPossible Issues:\n", "section"))
                                    for issue in diagnostics:
                                        content.append((f"  • {issue}\n", "warning"))
                            except Exception:
                                pass
                    except Exception:
                        content.append((
                            f"\n⚠ Warning: Accounting equation does not balance!\n"
                            f"Assets ({self._format_amount(total_assets)}) ≠ Liabilities + Equity ({self._format_amount(total_liabilities + total_equity)})\n"
                            f"Difference: {self._format_amount(abs(balance_check))}\n", 
                            "warning"
                        ))
                else:
                    content.append((
                        f"\n⚠ Warning: Accounting equation does not balance!\n"
                        f"Assets ({self._format_amount(total_assets)}) ≠ Liabilities + Equity + Net Income ({self._format_amount(total_liabilities + total_equity + net_income)})\n"
                        f"Difference: {self._format_amount(abs(balance_check))}\n", 
                        "warning"
                    ))
                    # Add diagnostic information
                    try:
                        diagnostics = self._diagnose_balance_sheet_imbalance()
                        if diagnostics:
                            content.append(("\nPossible Issues:\n", "section"))
                            for issue in diagnostics:
                                content.append((f"  • {issue}\n", "warning"))
                    except Exception:
                        pass
            else:
                content.append((
                    f"\n✓ Balance Sheet balances ✅\n",
                    None
                ))
            
            # Update the text widget
            self._update_text_widget(self.balance_sheet_text, content)
            
        except Exception as e:
            logger.error(f"Failed to generate balance sheet: {str(e)}", exc_info=True)
            messagebox.showerror("Error", f"Failed to generate balance sheet: {str(e)}")
            # Show error in widget
            self._show_error_in_text_widget(self.balance_sheet_text, str(e), "balance sheet")
    
    def _generate_cash_flow_statement(self, start_date: str = None, end_date: str = None) -> None:
        """Generate and display the cash flow statement"""
        try:
            # Determine safe start/end for cash flow
            start = start_date or (
                (self.engine.current_period['start_date'] if self.engine.current_period and 'start_date' in self.engine.current_period else None)
            ) or '1900-01-01'
            end = end_date or date.today().isoformat()
            
            # Generate cash flow using backend engine
            # Cash flow already filters by date, so it should work across periods
            cf = self.engine.generate_cash_flow(start, end)
            
            # Check for errors
            if isinstance(cf, dict) and cf.get('error'):
                content = []
                content.append(("Cash Flow Statement\n", 'header'))
                content.append((f"Period: {start} → {end}\n\n", 'subheader'))
                content.append((f"Error: {cf.get('error')}\n", 'warning'))
                self._update_text_widget(self.cash_flow_text, content)
                return
            
            # Extract data from backend result
            sections = cf.get('sections', {}) if isinstance(cf, dict) else {}
            totals = cf.get('totals', {}) if isinstance(cf, dict) else {}
            net_change = cf.get('net_change_in_cash', 0.0)
            
            # Build the cash flow statement content
            content = []
            content.append(("Cash Flow Statement\n", 'header'))
            content.append((f"Period: {start} → {end}\n\n", 'subheader'))
            
            # Add each section (Operating, Investing, Financing)
            for sec in ('Operating', 'Investing', 'Financing'):
                items = sections.get(sec, [])
                content.append((f"{sec}\n", 'section'))
                
                if not items:
                    content.append(("  (no activity)\n\n", None))
                else:
                    for it in items:
                        try:
                            amt = float(it.get('amount', 0))
                        except (ValueError, TypeError):
                            amt = 0.0
                        entry_date = it.get('date', '')
                        entry_id = it.get('entry_id', '')
                        content.append((f"  {entry_date}: Entry #{entry_id}: {self._format_amount(amt)}\n", None))
                    
                    # Add section total
                    section_total = totals.get(sec, 0.0)
                    content.append((f"\n  Total {sec}: {self._format_amount(section_total)}\n\n", 'total'))
            
            # Add net change in cash
            content.append((f"Net Change in Cash: {self._format_amount(net_change)}\n", 'net'))
            
            # Add simpler cash receipts/payments summary (matching FINAL_ACCOUNTING.py approach)
            try:
                # Get cash account ID
                cash_acc = db.get_account_by_name("Cash", conn=self.engine.conn)
                if cash_acc:
                    cash_id = int(cash_acc["id"])
                    # Calculate simple cash receipts (debits) and payments (credits) for the period
                    period_filter = self._get_period_filter_for_dates(start, end)
                    period_clause = " AND je.period_id = ?" if period_filter else ""
                    period_params = [start, end] + ([period_filter] if period_filter else [])
                    
                    cur = self.engine.conn.execute(f"""
                        SELECT COALESCE(SUM(jl.debit), 0) as cash_in,
                               COALESCE(SUM(jl.credit), 0) as cash_out
                        FROM journal_lines jl
                        JOIN journal_entries je ON je.id = jl.entry_id
                        WHERE jl.account_id = ?
                          AND date(je.date) BETWEEN date(?) AND date(?)
                          AND (je.status = 'posted' OR je.status IS NULL)
                          {period_clause}
                    """, [cash_id] + period_params)
                    
                    result = cur.fetchone()
                    if result:
                        cash_in = float(result['cash_in'] or 0)
                        cash_out = float(result['cash_out'] or 0)
                        net_cash_simple = cash_in - cash_out
                        
                        content.append(("\n" + "=" * 50 + "\n", None))
                        content.append(("SIMPLE CASH SUMMARY\n", 'section'))
                        content.append(("=" * 50 + "\n", None))
                        content.append((f"Cash Receipts (debits to Cash): {self._format_amount(cash_in)}\n", 'total'))
                        content.append((f"Cash Payments (credits from Cash): {self._format_amount(cash_out)}\n", 'total'))
                        content.append(("─" * 50 + "\n", None))
                        content.append((f"Net Cash Change: {self._format_amount(net_cash_simple)}\n", 'net'))
            except Exception as e:
                logger.debug(f"Could not add simple cash summary: {e}")
            
            # Update the text widget
            self._update_text_widget(self.cash_flow_text, content)
            
        except Exception as e:
            logger.error(f"Error generating cash flow statement: {e}", exc_info=True)
            # Show error in widget
            self._show_error_in_text_widget(self.cash_flow_text, str(e), "cash flow statement")

    def _get_period_filter_for_dates(self, date_from: str | None, date_to: str | None) -> int | None:
        """Determine period filter based on date range.
        
        Returns None for cross-period reporting when both dates are set,
        otherwise returns current_period_id.
        """
        if date_from and date_to:
            return None  # Cross-period reporting
        elif date_to:
            return None  # Include all entries up to date_to
        else:
            return self.engine.current_period_id  # Fallback to period filter
    
    def _show_error_in_text_widget(self, text_widget, error_message: str, statement_name: str) -> None:
        """Display an error message in a text widget.
        
        Args:
            text_widget: The tkinter Text widget to update
            error_message: The error message to display
            statement_name: Name of the statement (e.g., "income statement")
        """
        try:
            if text_widget and hasattr(text_widget, 'config'):
                text_widget.config(state=tk.NORMAL)
                text_widget.delete(1.0, tk.END)
                text_widget.insert(tk.END, f"Error generating {statement_name}:\n{error_message}\n", 'warning')
                text_widget.config(state=tk.DISABLED)
        except Exception:
            pass
    
    def _should_include_temporary_accounts_in_bs(self) -> bool:
        """Check if temporary accounts should be included in balance sheet.
        
        Returns False if step 8 (closing entries) is completed, True otherwise.
        """
        try:
            statuses = self.engine.get_cycle_status()
            step8 = next((r for r in statuses if int(r['step']) == 8), None)
            if step8 and (step8['status'] == 'completed'):
                return False
        except Exception:
            pass
        return True
    
    def _regenerate_financial_statements(self, as_of_date: str | None = None) -> None:
        """Re-generate income statement and balance sheet text without modifying cycle status.

        This is used when switching themes to refresh the displayed text and tags without
        calling engine methods that change cycle step statuses.
        """
        try:
            # Use the same date-range logic as _load_financials so that
            # "From" and "To" behave consistently across themes.
            date_from = self.fs_date_from.get().strip() if hasattr(self, 'fs_date_from') else None
            if date_from == "":
                date_from = None
            date_to = as_of_date or (self.fs_date_to.get().strip() if hasattr(self, 'fs_date_to') else None)
            if date_to == "":
                date_to = None

            # Use helper methods to determine period filter and temporary account inclusion
            period_filter = self._get_period_filter_for_dates(date_from, date_to)
            inc_temp_bs = self._should_include_temporary_accounts_in_bs()
            
            rows_is = db.compute_trial_balance(
                from_date=date_from,
                up_to_date=date_to,
                include_temporary=True,
                period_id=period_filter,
                conn=self.engine.conn,
            )
            rows_bs = db.compute_trial_balance(
                up_to_date=date_to,
                include_temporary=inc_temp_bs,
                period_id=period_filter,
                conn=self.engine.conn,
            )

            # Re-generate text widgets only (these functions write to Text widgets)
            try:
                # Pass date_from to ensure correct date range
                self._generate_income_statement(rows_is, date_to, start_date=date_from)
            except Exception as e:
                logger.exception("Error generating income statement in _regenerate_financial_statements")
                pass
            try:
                self._generate_balance_sheet(rows_bs, date_to)
            except Exception:
                pass
        except Exception:
            # Fail silently during theme refresh to avoid interrupting UI
            pass
    
    def _load_financials(self, mark_status: bool = True) -> None:
        """Load and display financial statements based on date range"""
        try:
            # Get date range from the UI (with fallback if fields don't exist yet)
            try:
                date_from = self.fs_date_from.get().strip() or None if hasattr(self, 'fs_date_from') else None
                date_to = self.fs_date_to.get().strip() or None if hasattr(self, 'fs_date_to') else None
            except Exception:
                date_from = None
                date_to = None
            
            # Set default dates if not provided
            # Use last day of current month as default for "To" date
            if not date_to:
                import calendar
                today = datetime.now().date()
                last_day = calendar.monthrange(today.year, today.month)[1]
                date_to = today.replace(day=last_day).isoformat()
                # Update UI field if it exists
                try:
                    if hasattr(self, 'fs_date_to'):
                        self.fs_date_to.delete(0, tk.END)
                        self.fs_date_to.insert(0, date_to)
                except Exception:
                    pass
            if not date_from:
                # Try to get period start date
                try:
                    period = db.get_accounting_period_by_id(self.engine.current_period_id, conn=self.engine.conn)
                    if period and period.get('start_date'):
                        date_from = period['start_date']
                    else:
                        # Default to start of current year
                        date_from = datetime.now().date().replace(month=1, day=1).isoformat()
                except Exception:
                    # Default to start of current year
                    date_from = datetime.now().date().replace(month=1, day=1).isoformat()
                # Update UI field if it exists
                try:
                    if hasattr(self, 'fs_date_from'):
                        self.fs_date_from.delete(0, tk.END)
                        self.fs_date_from.insert(0, date_from)
                except Exception:
                    pass
            
            # Clear previous content using the update_text method
            try:
                self.income_text.update_text("")
                self.balance_sheet_text.update_text("")
                self.cash_flow_text.update_text("")
            except Exception:
                # Text widgets might not exist if tab hasn't been viewed yet
                pass
            
            # Use helper methods to determine period filter and temporary account inclusion
            period_filter = self._get_period_filter_for_dates(date_from, date_to)
            inc_temp_bs = self._should_include_temporary_accounts_in_bs()
            
            # Get trial balance data for the specified date range
            # Exclude closing entries for income statement to show revenue/expenses before closing
            rows_is = db.compute_trial_balance(
                from_date=date_from,
                up_to_date=date_to,
                include_temporary=True,
                period_id=period_filter,
                exclude_closing=True,
                conn=self.engine.conn
            )
            rows_bs = db.compute_trial_balance(
                up_to_date=date_to,
                include_temporary=inc_temp_bs,
                period_id=period_filter,
                conn=self.engine.conn
            )
            
            # Process data for financial statements
            # Pass date_from to income statement so it uses the correct date range
            self._generate_income_statement(rows_is, date_to, start_date=date_from)
            self._generate_balance_sheet(rows_bs, date_to)
            # Generate cash flow using dedicated method
            self._generate_cash_flow_statement(date_from, date_to)
            
            # Check balance sheet balance and show notification (matching FINAL_ACCOUNTING.py approach)
            try:
                balance_sheet = self.engine.generate_balance_sheet(date_to or date.today().isoformat())
                total_assets = balance_sheet.get('total_assets', 0.0)
                total_liabilities = balance_sheet.get('total_liabilities', 0.0)
                total_equity = balance_sheet.get('total_equity', 0.0)
                
                # Check if closing entries have been completed
                closing_completed = False
                try:
                    statuses = self.engine.get_cycle_status()
                    step8 = next((r for r in statuses if int(r['step']) == 8), None)
                    # Check both status and if closing entries actually exist
                    status_completed = step8 and (step8['status'] == 'completed')
                    
                    # Also check if closing entries exist in database (more reliable)
                    closing_entries_exist = False
                    try:
                        closing_count = self.engine.conn.execute("""
                            SELECT COUNT(*) as cnt FROM journal_entries 
                            WHERE is_closing = 1 AND period_id = ?
                        """, (self.engine.current_period_id,)).fetchone()
                        closing_entries_exist = closing_count and closing_count['cnt'] > 0
                    except Exception:
                        pass
                    
                    # Closing is completed if status says so OR if closing entries exist
                    closing_completed = status_completed or closing_entries_exist
                except Exception:
                    pass
                
                # Balance check depends on whether closing entries are completed
                if closing_completed:
                    # After closing: Assets = Liabilities + Equity (Net Income already in Capital)
                    balance_check = abs(total_assets - (total_liabilities + total_equity))
                else:
                    # Before closing: Assets = Liabilities + Equity + Net Income
                    income_stmt = self.engine.generate_income_statement(
                        date_from or '1900-01-01', 
                        date_to or date.today().isoformat(), 
                        period_id=None
                    )
                    net_income = income_stmt.get('net_income', 0.0)
                    balance_check = abs(total_assets - (total_liabilities + total_equity + net_income))
                
                if balance_check < 0.05:
                    # Only show success message if mark_status is True (to avoid spam during theme changes)
                    if mark_status:
                        try:
                            messagebox.showinfo("Balance Check", "Balance Sheet balances ✅")
                        except Exception:
                            pass
                else:
                    if mark_status:
                        try:
                            messagebox.showwarning(
                                "Balance Check", 
                                f"Balance Sheet does NOT balance ❌ — check entries and adjustments.\n"
                                f"Difference: {self._format_amount(balance_check)}"
                            )
                        except Exception:
                            pass
            except Exception:
                pass  # Fail silently if balance check fails
            
            # Only mark as completed if we got this far without errors and caller allows status changes
            if mark_status:
                self.engine.set_cycle_step_status(
                    7,
                    "completed",
                    note=f"Financial statements generated as of {date_to or 'latest'}",
                )
                self.engine.set_cycle_step_status(
                    8,
                    "in_progress",
                    note="Ready to prepare closing entries",
                )
            
        except Exception as e:
            try:
                self._handle_exception("load_financials", e)
            except Exception:
                pass
            messagebox.showerror("Error", f"Failed to generate financial statements: {str(e)}")
            # Don't mark as completed or change cycle status if caller disallowed status changes
            if mark_status:
                try:
                    self.engine.set_cycle_step_status(
                        7,
                        "in_progress",
                        note=f"Error generating statements: {str(e)[:100]}",
                    )
                except Exception:
                    pass
        # Always refresh cycle status display in the UI
        try:
            self._load_cycle_status()
        except Exception:
            pass

    # --------------------- Closing Tab ---------------------
    def _build_closing_tab(self) -> None:
        frame = self.tab_closing
        frame.grid_rowconfigure(1, weight=2)
        frame.grid_rowconfigure(2, weight=1)
        frame.grid_columnconfigure(0, weight=1)

        top = ttk.Frame(frame, style="Techfix.Surface.TFrame")
        top.grid(row=0, column=0, sticky="ew", padx=12, pady=(12, 6))
        ttk.Label(top, text="Closing Date (YYYY-MM-DD)").pack(side=tk.LEFT, padx=(0, 6))
        self.close_date = ttk.Entry(top, width=16, style="Techfix.TEntry")
        self.close_date.pack(side=tk.LEFT, padx=(0, 12))
        ttk.Button(top, text="Make Closing Entries", command=self._do_close, style="Techfix.TButton").pack(side=tk.LEFT)
        ttk.Button(top, text="Refresh Preview", command=self._load_closing_preview, style="Techfix.TButton").pack(
            side=tk.LEFT, padx=(12, 0)
        )

        preview = ttk.Labelframe(frame, text="Closing Entry Preview", style="Techfix.TLabelframe")
        preview.grid(row=1, column=0, sticky="nsew", padx=12, pady=(0, 12))
        cols = ("code", "name", "action", "amount")
        self.closing_preview_tree = ttk.Treeview(preview, columns=cols, show="headings", style="Techfix.Treeview")
        for c in cols:
            width = 100 if c == "amount" else 160
            self.closing_preview_tree.heading(c, text=c.title(), anchor="w")
            self.closing_preview_tree.column(c, width=width, stretch=True)
        self.closing_preview_tree.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)
        self._load_closing_preview()

        log_frame = ttk.Labelframe(frame, text="Closing Log", style="Techfix.TLabelframe")
        log_frame.grid(row=2, column=0, sticky="nsew", padx=12, pady=(0, 12))
        self.close_log = tk.Text(
            log_frame,
            height=8,
            bg=self.palette["surface_bg"],
            fg=self.palette["text_primary"],
            font=FONT_MONO,
            bd=0,
            highlightthickness=0,
            highlightbackground=self.palette.get("surface_bg", "#ffffff"),
            relief=tk.FLAT,
        )
        self.close_log.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)
        self._load_closing_log()

    def _do_close(self) -> None:
        date = self.close_date.get().strip()
        try:
            if not date:
                date = datetime.utcnow().date().isoformat()
            else:
                datetime.strptime(date, '%Y-%m-%d')
        except Exception:
            messagebox.showerror('Error', 'Enter closing date as YYYY-MM-DD')
            return
        ids = self.engine.make_closing_entries(date)
        self.close_log.insert(tk.END, f"Created closing entries: {ids}\n")
        self._refresh_after_post()
        self._load_closing_preview()
        self._load_closing_log()

    def _load_closing_preview(self) -> None:
        """Populate the Closing Entry Preview with amounts that would be posted."""
        # Clear existing rows
        if hasattr(self, 'closing_preview_tree'):
            for it in self.closing_preview_tree.get_children():
                self.closing_preview_tree.delete(it)

        try:
            if not self.engine.current_period_id:
                return
            pid = self.engine.current_period_id
            conn = self.engine.conn
            cur = conn.cursor()
            inserted = False

            # Revenues to close
            cur.execute(
                """
                SELECT a.code, a.name, ROUND(COALESCE(SUM(jl.credit) - SUM(jl.debit),0),2) AS balance
                FROM accounts a
                LEFT JOIN journal_lines jl ON jl.account_id = a.id
                LEFT JOIN journal_entries je ON je.id = jl.entry_id
                WHERE a.type = 'Revenue' AND a.is_active=1 AND je.period_id = ?
                GROUP BY a.id, a.code, a.name
                HAVING ABS(balance) > 0.005
                """,
                (pid,)
            )
            for r in cur.fetchall():
                amt = float(r['balance'])
                action = 'Close revenue → Capital (credit)' if amt >= 0 else 'Close revenue (reverse-sign) → Capital (debit)'
                self.closing_preview_tree.insert('', 'end', values=(r['code'], r['name'], action, f"{abs(amt):,.2f}"))
                inserted = True

            # Expenses to close
            cur.execute(
                """
                SELECT a.code, a.name, ROUND(COALESCE(SUM(jl.debit) - SUM(jl.credit),0),2) AS balance
                FROM accounts a
                LEFT JOIN journal_lines jl ON jl.account_id = a.id
                LEFT JOIN journal_entries je ON je.id = jl.entry_id
                WHERE a.type = 'Expense' AND a.is_active=1 AND je.period_id = ?
                GROUP BY a.id, a.code, a.name
                HAVING ABS(balance) > 0.005
                """,
                (pid,)
            )
            for e in cur.fetchall():
                amt = float(e['balance'])
                action = 'Close expense → Capital (debit)' if amt > 0 else 'Close expense (reverse-sign) → Capital (credit)'
                self.closing_preview_tree.insert('', 'end', values=(e['code'], e['name'], action, f"{abs(amt):,.2f}"))
                inserted = True

            # Owner's Drawings (close to capital)
            drawings = db.get_account_by_name("Owner's Drawings", conn)
            if drawings:
                cur.execute(
                    """
                    SELECT ROUND(COALESCE(SUM(debit) - SUM(credit),0),2) AS balance
                    FROM journal_lines jl
                    JOIN journal_entries je ON je.id = jl.entry_id
                    WHERE jl.account_id = ? AND je.period_id = ?
                    """,
                    (drawings['id'], pid),
                )
                bal = float(cur.fetchone()[0] or 0)
                if bal > 0.005:
                    self.closing_preview_tree.insert('', 'end', values=(drawings['code'], "Owner's Drawings", 'Close drawings → Capital (debit)', f"{bal:,.2f}"))
                    inserted = True

            if not inserted:
                self.closing_preview_tree.insert('', 'end', values=("", "", "No amounts to close", ""))

        except Exception as e:
            messagebox.showerror("Error", f"Failed to load closing preview: {e}")

    def _load_closing_log(self) -> None:
        try:
            if hasattr(self, 'close_log'):
                self.close_log.delete('1.0', tk.END)
            rows = db.list_audit_log(limit=200, conn=self.engine.conn)
            for r in rows:
                action = r['action']
                det = r['details']
                try:
                    d = json.loads(det) if det else {}
                except Exception:
                    d = {}
                if action == 'journal_entry_created' and bool(d.get('is_closing')) and (not self.engine.current_period_id or int(d.get('period_id', 0) or 0) == int(self.engine.current_period_id)):
                    ts = r['timestamp']
                    eid = d.get('entry_id')
                    desc = d.get('description')
                    self.close_log.insert(tk.END, f"[{ts}] entry {eid} {desc}\n")
            if hasattr(self, 'close_log'):
                self.close_log.see(tk.END)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load closing log: {e}")

    # --------------------- Post-Closing Tab ---------------------
    def _build_postclosing_tab(self) -> None:
        frame = self.tab_postclosing
        controls = ttk.Frame(frame, style="Techfix.Surface.TFrame")
        controls.pack(fill=tk.X, padx=12, pady=(12, 6))
        ttk.Label(controls, text="As of (YYYY-MM-DD)").pack(side=tk.LEFT)
        self.pctb_date = ttk.Entry(controls, width=16, style="Techfix.TEntry")
        self.pctb_date.pack(side=tk.LEFT, padx=6)
        ttk.Button(controls, text="Refresh", command=self._load_postclosing_tb, style="Techfix.TButton").pack(side=tk.LEFT, padx=6)
        ttk.Button(controls, text="Complete Post-Closing TB", command=self._complete_postclosing_tb_action, style="Techfix.TButton").pack(side=tk.LEFT, padx=6)

        # Two-pane layout: left = Post-Closing TB, right = Reversing Schedule
        content = ttk.Frame(frame, style="Techfix.App.TFrame")
        content.pack(fill=tk.BOTH, expand=True, padx=12, pady=(0, 12))
        content.grid_columnconfigure(0, weight=1)
        content.grid_columnconfigure(1, weight=1)
        content.grid_rowconfigure(0, weight=1)

        # Left pane: Post-Closing Trial Balance
        tb_wrap = ttk.Labelframe(content, text="Post-Closing Trial Balance", style="Techfix.TLabelframe")
        tb_wrap.grid(row=0, column=0, sticky="nsew", padx=0)
        cols = ("code", "name", "debit", "credit")
        self.pctb_tree = ttk.Treeview(tb_wrap, columns=cols, show="headings", style="Techfix.Treeview")
        for c in cols:
            self.pctb_tree.heading(c, text=c.title(), anchor="w")
            self.pctb_tree.column(c, stretch=True, width=140, anchor="w")
        pctb_scroll = ttk.Scrollbar(tb_wrap, orient=tk.VERTICAL, command=self.pctb_tree.yview)
        self.pctb_tree.configure(yscrollcommand=pctb_scroll.set)
        self.pctb_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=4, pady=4)
        pctb_scroll.pack(side=tk.LEFT, fill=tk.Y, padx=(0,4), pady=4)

        # Right pane: Reversing Schedule
        sched_wrap = ttk.Labelframe(content, text="Reversing Entry Schedule", style="Techfix.TLabelframe")
        sched_wrap.grid(row=0, column=1, sticky="nsew", padx=0)

        # Inner frame to keep the treeview and its scrollbar together at the top
        sched_inner = ttk.Frame(sched_wrap, style="Techfix.Surface.TFrame")
        sched_inner.pack(fill=tk.BOTH, expand=True, padx=0, pady=(0, 4))

        rcols = ("id", "original_entry", "reverse_on", "reversal_entry", "status")
        self.reversing_tree = ttk.Treeview(sched_inner, columns=rcols, show="headings", style="Techfix.Treeview")
        for c in rcols:
            width = 80 if c == "id" else 140
            self.reversing_tree.heading(c, text=c.replace("_", " ").title(), anchor="w")
            self.reversing_tree.column(c, width=width, stretch=(c != "status"))
        rev_scroll = ttk.Scrollbar(sched_inner, orient=tk.VERTICAL, command=self.reversing_tree.yview)
        self.reversing_tree.configure(yscrollcommand=rev_scroll.set)
        self.reversing_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=4, pady=(4, 0))
        rev_scroll.pack(side=tk.LEFT, fill=tk.Y, padx=(0,4), pady=(4, 0))

        # Controls row directly under the treeview (bottom of this labelframe)
        schedule_controls = ttk.Frame(sched_wrap, style="Techfix.Surface.TFrame")
        schedule_controls.pack(fill=tk.X, padx=4, pady=(0, 6))
        ttk.Button(
            schedule_controls,
            text="Refresh Schedule",
            command=self._load_reversing_queue,
            style="Techfix.TButton",
        ).pack(side=tk.LEFT, padx=8)
        ttk.Button(
            schedule_controls,
            text="Complete Reversing Schedule",
            command=self._complete_reversing_schedule_action,
            style="Techfix.TButton",
        ).pack(side=tk.LEFT, padx=8)

    def _load_reversing_queue(self) -> None:
        """Load the reversing entry schedule into the treeview."""
        if hasattr(self, 'reversing_tree'):
            for it in self.reversing_tree.get_children():
                self.reversing_tree.delete(it)

        try:
            rows = self.engine.list_reversing_queue()
            for r in rows:
                self.reversing_tree.insert('', 'end', values=(
                    r['id'],
                    r['original_entry_id'],
                    r['reverse_on'],
                    r['reversed_entry_id'] if 'reversed_entry_id' in r.keys() else '',
                    r['status'],
                ))
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load reversing queue: {e}")

    def _complete_reversing_schedule_action(self) -> None:
        try:
            as_of = (self.pctb_date.get().strip() if hasattr(self, 'pctb_date') else '') or None
            created = self.engine.process_reversing_schedule(as_of)
            self._load_reversing_queue()
            self._load_cycle_status()
            # Refresh financial statements after processing reversing entries
            try:
                self._load_financials(mark_status=False)
            except Exception:
                pass  # Don't fail if refresh fails
            messagebox.showinfo("Completed", f"Posted {len(created)} reversing entr(ies)")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to complete reversing schedule: {e}")

    # --------------------- Export Tab ---------------------
    def _build_export_tab(self) -> None:
        frame = self.tab_export
        
        # Import/Backup/Restore section
        import_wrapper = ttk.Labelframe(frame, text="Import & Backup", style="Techfix.TLabelframe")
        import_wrapper.pack(fill=tk.X, padx=12, pady=(12, 6))
        
        # Button container for horizontal layout
        import_btn_container = ttk.Frame(import_wrapper, style="Techfix.Surface.TFrame")
        import_btn_container.pack(fill=tk.X, padx=12, pady=12)
        
        # Import button
        import_btn = ttk.Button(
            import_btn_container,
            text=f"{ICONS.get('add', '📥')} Import",
            command=self._show_import_dialog,
            style="Techfix.Enhanced.TButton",
        )
        import_btn.pack(side=tk.LEFT, padx=(0, 6))
        
        # Backup button
        backup_btn = ttk.Button(
            import_btn_container,
            text=f"{ICONS.get('save', '💾')} Backup",
            command=self._show_backup_dialog,
            style="Techfix.Enhanced.TButton",
        )
        backup_btn.pack(side=tk.LEFT, padx=(0, 6))
        
        # Restore button
        restore_btn = ttk.Button(
            import_btn_container,
            text=f"{ICONS.get('refresh', '📂')} Restore",
            command=self._show_restore_dialog,
            style="Techfix.Enhanced.TButton",
        )
        restore_btn.pack(side=tk.LEFT, padx=(0, 6))
        
        # Export section
        wrapper = ttk.Labelframe(frame, text="Export Options", style="Techfix.TLabelframe")
        wrapper.pack(fill=tk.X, padx=12, pady=12)

        for col in range(2):
            wrapper.columnconfigure(col, weight=1)

        buttons = [
            ("Export Journal (Excel)", self._export_journal),
            ("Export Ledger (Excel)", self._export_ledger),
            ("Export Trial Balance (Excel)", self._export_tb),
            ("Export Financials (Excel)", self._export_fs),
            ("Export All (Excel)", self._export_all_to_excel),
            ("Export All to Email", self._export_all_to_email),
            ("Export Documentation (PDF)", self._export_program_docs_pdf),
        ]

        for idx, (label, command) in enumerate(buttons):
            r, c = divmod(idx, 2)
            ttk.Button(wrapper, text=label, command=command, style="Techfix.TButton").grid(
                row=r, column=c, padx=8, pady=8, sticky="ew"
            )

    # --------------------- Audit Tab ---------------------
    def _build_audit_tab(self) -> None:
        frame = self.tab_audit
        controls = ttk.Frame(frame, style="Techfix.Surface.TFrame")
        controls.pack(fill=tk.X, padx=12, pady=8)
        ttk.Label(controls, text="Filter:", style="Techfix.AppBar.TLabel").pack(side=tk.LEFT)
        self.audit_filter_var = tk.StringVar(value="")
        ttk.Entry(controls, textvariable=self.audit_filter_var, width=24, style="Techfix.TEntry").pack(side=tk.LEFT, padx=(6,12))
        ttk.Button(controls, text="Refresh", command=lambda: self._load_audit_log(), style="Techfix.TButton").pack(side=tk.LEFT)
        ttk.Button(controls, text="Export CSV", command=lambda: self._export_audit_csv(), style="Techfix.TButton").pack(side=tk.LEFT, padx=(6,0))

        cols = ("id","timestamp","user","action","details")
        self.audit_tree = ttk.Treeview(frame, columns=cols, show="headings", style="Techfix.Treeview")
        for c in cols:
            anchor = tk.W
            width = 100 if c in ("id","user") else (160 if c=="timestamp" else 640)
            self.audit_tree.heading(c, text=c.title(), anchor=anchor)
            self.audit_tree.column(c, stretch=True, width=width, anchor=anchor)
        vsb = ttk.Scrollbar(frame, orient="vertical", command=self.audit_tree.yview)
        self.audit_tree.configure(yscrollcommand=vsb.set)
        self.audit_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=12, pady=(0, 12))
        vsb.pack(side=tk.RIGHT, fill=tk.Y, pady=(0, 12))
        try:
            self._load_audit_log()
        except Exception:
            pass

    def _load_audit_log(self) -> None:
        try:
            if hasattr(self, 'audit_tree'):
                for it in self.audit_tree.get_children():
                    self.audit_tree.delete(it)
            filt = (self.audit_filter_var.get().strip().lower() if hasattr(self, 'audit_filter_var') else '')
            rows = db.list_audit_log(limit=500, conn=self.engine.conn)
            for r in rows:
                act = r['action'] if 'action' in r.keys() else ''
                det = r['details'] if 'details' in r.keys() else ''
                if filt and (filt not in str(act).lower() and filt not in str(det).lower()):
                    continue
                self.audit_tree.insert('', 'end', values=(r['id'], r['timestamp'], r['user'], act, det))
        except Exception:
            pass

    def _export_audit_csv(self) -> None:
        try:
            from pathlib import Path
            rows = []
            for it in self.audit_tree.get_children():
                rows.append(self.audit_tree.item(it, 'values'))
            out = Path(str(db.DB_DIR)) / 'audit_export.csv'
            import csv
            with out.open('w', newline='', encoding='utf-8') as f:
                w = csv.writer(f)
                w.writerow(["id","timestamp","user","action","details"])
                for r in rows:
                    w.writerow(list(r))
            messagebox.showinfo('Export', f'CSV exported: {out}')
        except Exception as e:
            messagebox.showerror('Export', f'Failed to export audit: {e}')

    # --------------------- Help / How To Use Tab ---------------------
    def _build_help_tab(self) -> None:
        """Build a styled in‑app 'How to Use?' guide."""
        frame = self.tab_help

        container = ttk.Frame(frame, style="Techfix.Surface.TFrame")
        container.pack(fill=tk.BOTH, expand=True, padx=24, pady=24)

        # Header / title
        header = ttk.Frame(container, style="Techfix.Surface.TFrame")
        header.pack(fill=tk.X, pady=(0, 12))

        ttk.Label(
            header,
            text="How to Use TechFix",
            style="Techfix.TLabel",
        ).pack(side=tk.LEFT, anchor=tk.W)

        ttk.Label(
            header,
            text="Quick guide to the full accounting cycle inside TechFix",
            style="Techfix.TLabel",
        ).pack(side=tk.LEFT, anchor=tk.W, padx=(16, 0))

        # Card‑style body to match other surfaces, works in light & dark themes
        body = ttk.Frame(container, style="Techfix.Surface.TFrame")
        body.pack(fill=tk.BOTH, expand=True)

        # Add a subtle border using a Canvas background color
        body_inner = ttk.Frame(body, style="Techfix.Surface.TFrame")
        body_inner.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)

        # Scrollable text area for instructions
        text_frame = ttk.Frame(body_inner, style="Techfix.Surface.TFrame")
        text_frame.pack(fill=tk.BOTH, expand=True)

        bg_color = self.palette.get("surface_bg", "#ffffff")
        fg_color = self.palette.get("text_primary", "#1f2937")

        help_text = tk.Text(
            text_frame,
            wrap=tk.WORD,
            bg=bg_color,
            fg=fg_color,
            font=FONT_BASE,
            relief=tk.FLAT,
            bd=0,
            padx=10,
            pady=10,
            highlightthickness=0,
            insertbackground=self.palette.get("accent_color", "#2563eb"),
            state=tk.NORMAL,
        )
        # Store reference so theme updates can modify it
        self.help_text = help_text
        vsb = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=help_text.yview)
        help_text.configure(yscrollcommand=vsb.set)
        help_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        # High‑level step‑by‑step guide with clear sections
        guide = (
            "Welcome to TechFix\n"
            "===================\n\n"
            "TechFix walks you through the full accounting cycle, from daily transactions\n"
            "all the way to post‑closing trial balance. Use the tabs in the left sidebar\n"
            "from top to bottom for the smoothest workflow.\n\n"
            "1. Set up your accounting period\n"
            "--------------------------------\n"
            "• Make sure the current accounting period is correct (for example: January 2025).\n"
            "• All journal entries will be recorded into this active period.\n\n"
            "2. Record daily transactions (Transactions tab)\n"
            "----------------------------------------------\n"
            "• Open the Transactions tab.\n"
            "• Choose or attach the source document (invoice, receipt, bill, etc.).\n"
            "• Optionally scan / prefill from a document or QR/barcode.\n"
            "• Or click the manual entry option and paste structured data (JSON or key=value).\n"
            "• Confirm that Date, Description, Debit account, Credit account, and Amounts\n"
            "  are all filled in.\n"
            "• Click Record / Post to save the journal entry.\n\n"
            "3. Review the Journal (Journal tab)\n"
            "-----------------------------------\n"
            "• View all posted entries in chronological order.\n"
            "• Use filters or exports (where available) to review specific transactions.\n\n"
            "4. Check account activity (Ledger tab)\n"
            "--------------------------------------\n"
            "• Open the Ledger tab and select an account (e.g., Cash, Accounts Receivable).\n"
            "• Review each posting and the running balance for that account.\n\n"
            "5. Prepare the Trial Balance (Trial Balance tab)\n"
            "-----------------------------------------------\n"
            "• Confirm that total debits equal total credits for the period.\n"
            "• If they do not match, drill back into the Journal or Ledger to locate and\n"
            "  fix errors.\n\n"
            "6. Post Adjusting Entries (Adjustments tab)\n"
            "-------------------------------------------\n"
            "• Record adjustments such as supplies used, depreciation, prepaid expenses,\n"
            "  and accruals.\n"
            "• After you post adjustments, regenerate the adjusted trial balance if needed.\n\n"
            "7. Generate Financial Statements (Fin. Statements tab)\n"
            "------------------------------------------------------\n"
            "• Review the Income Statement, Statement of Owner's Equity, and Balance Sheet\n"
            "  for the active period.\n\n"
            "8. Perform Closing Entries (Closing & Post‑Closing tabs)\n"
            "--------------------------------------------------------\n"
            "• Use the Closing tab to close revenue, expense, and drawings accounts.\n"
            "• Then open the Post‑Closing tab to generate the post‑closing trial balance.\n\n"
            "9. Review system history (Audit Log tab)\n"
            "----------------------------------------\n"
            "• Use the Audit Log to see a technical history of key actions (prefill attempts,\n"
            "  document loading, exports, UI changes, etc.).\n"
            "• Use the filter box to search by action or details, or export to CSV.\n\n"
            "Helpful tips\n"
            "------------\n"
            "• If manual entry data does not fill accounts automatically, include\n"
            "  debit_account and credit_account (or valid account codes/names) in your data.\n"
            "• Use the Export options to back up your journal, trial balance, and audit log.\n"
            "• Working through the tabs in order (Transactions → … → Post‑Closing) mirrors\n"
            "  the standard accounting cycle and keeps your workflow simple.\n"
        )

        help_text.insert("1.0", guide)
        help_text.configure(state=tk.DISABLED)

    def _export_journal(self) -> None:
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")])
        if not path:
            return
        rows = db.fetch_journal(period_id=self.engine.current_period_id, conn=self.engine.conn)
        headers = ["date","entry_id","description","code","name","debit","credit"]
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font

            wb = Workbook()
            ws = wb.active
            ws.title = "Journal"

            for cidx, h in enumerate(headers, start=1):
                ws.cell(row=1, column=cidx, value=h).font = Font(bold=True)

            for ridx, r in enumerate(rows, start=2):
                ws.cell(row=ridx, column=1, value=r["date"])
                ws.cell(row=ridx, column=2, value=r["entry_id"])
                ws.cell(row=ridx, column=3, value=r["description"])
                ws.cell(row=ridx, column=4, value=r["code"])
                ws.cell(row=ridx, column=5, value=r["name"])
                try:
                    ws.cell(row=ridx, column=6, value=float(r["debit"] or 0))
                except Exception:
                    ws.cell(row=ridx, column=6, value=r["debit"])
                try:
                    ws.cell(row=ridx, column=7, value=float(r["credit"] or 0))
                except Exception:
                    ws.cell(row=ridx, column=7, value=r["credit"])

            last = ws.max_row
            # Add SUM formulas for debit (col 6) and credit (col 7) if there are data rows
            if last >= 2:
                col_f = get_column_letter(6)
                col_g = get_column_letter(7)
                total_row = last + 1
                ws.cell(row=total_row, column=5, value="Totals:").font = Font(bold=True)
                ws.cell(row=total_row, column=6, value=f"=SUM({col_f}2:{col_f}{last})")
                ws.cell(row=total_row, column=7, value=f"=SUM({col_g}2:{col_g}{last})")

            # auto-width
            for i in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(i)].width = 15

            wb.save(path)
            messagebox.showinfo("Exported", "Journal exported to Excel.")
        except Exception as e:
            messagebox.showerror("Export Error", str(e))

    def _export_ledger(self) -> None:
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")])
        if not path:
            return
        rows = db.fetch_ledger(period_id=self.engine.current_period_id, conn=self.engine.conn)
        headers = ["code","name","date","description","debit","credit"]
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font

            wb = Workbook()
            ws = wb.active
            ws.title = "Ledger"

            for cidx, h in enumerate(headers, start=1):
                ws.cell(row=1, column=cidx, value=h).font = Font(bold=True)

            for ridx, r in enumerate(rows, start=2):
                ws.cell(row=ridx, column=1, value=r["code"]) if "code" in r.keys() else ws.cell(row=ridx, column=1, value="")
                ws.cell(row=ridx, column=2, value=r["name"]) if "name" in r.keys() else ws.cell(row=ridx, column=2, value="")
                ws.cell(row=ridx, column=3, value=r["date"]) if "date" in r.keys() else ws.cell(row=ridx, column=3, value="")
                ws.cell(row=ridx, column=4, value=r["description"]) if "description" in r.keys() else ws.cell(row=ridx, column=4, value="")
                try:
                    ws.cell(row=ridx, column=5, value=float(r["debit"] or 0))
                except Exception:
                    ws.cell(row=ridx, column=5, value=r["debit"] if "debit" in r.keys() else "")
                try:
                    ws.cell(row=ridx, column=6, value=float(r["credit"] or 0))
                except Exception:
                    ws.cell(row=ridx, column=6, value=r["credit"] if "credit" in r.keys() else "")

            last = ws.max_row
            if last >= 2:
                col_e = get_column_letter(5)
                col_f = get_column_letter(6)
                total_row = last + 1
                ws.cell(row=total_row, column=4, value="Totals:").font = Font(bold=True)
                ws.cell(row=total_row, column=5, value=f"=SUM({col_e}2:{col_e}{last})")
                ws.cell(row=total_row, column=6, value=f"=SUM({col_f}2:{col_f}{last})")

            for i in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(i)].width = 15

            wb.save(path)
            messagebox.showinfo("Exported", "Ledger exported to Excel.")
        except Exception as e:
            messagebox.showerror("Export Error", str(e))

    def _export_tb(self) -> None:
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")])
        if not path:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font

            wb = Workbook()
            
            # --- Combined Trial Balance sheet ---
            ws_tb = wb.active
            ws_tb.title = "Trial Balance"
            
            # Title row
            title_cell = ws_tb.cell(row=1, column=1, value="TRIAL BALANCE")
            title_cell.font = Font(bold=True, size=14)
            ws_tb.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
            
            # Headers with clear labels
            headers = [
                "Code", 
                "Account Name", 
                "Unadjusted Debit", 
                "Unadjusted Credit",
                "Adjusted Debit",
                "Adjusted Credit"
            ]
            for cidx, h in enumerate(headers, start=1):
                cell = ws_tb.cell(row=2, column=cidx, value=h)
                cell.font = Font(bold=True)
            
            # Get both unadjusted and adjusted trial balances
            unadj_rows = db.compute_unadjusted_trial_balance(
                period_id=self.engine.current_period_id,
                include_temporary=True,
                conn=self.engine.conn
            )
            adj_rows = db.compute_trial_balance(
                period_id=self.engine.current_period_id,
                conn=self.engine.conn
            )
            
            # Create dictionaries for easy lookup
            unadj_by_code = {r["code"]: r for r in unadj_rows if "code" in r.keys() and r["code"]}
            adj_by_code = {r["code"]: r for r in adj_rows if "code" in r.keys() and r["code"]}
            
            # Get all unique account codes
            all_codes = sorted(set(list(unadj_by_code.keys()) + list(adj_by_code.keys())))
            
            # Write data rows
            for ridx, code in enumerate(all_codes, start=3):
                unadj_r = unadj_by_code.get(code)
                adj_r = adj_by_code.get(code)
                
                # Get account name (prefer from adjusted, fallback to unadjusted)
                if adj_r:
                    name = adj_r["name"] if "name" in adj_r.keys() else ""
                elif unadj_r:
                    name = unadj_r["name"] if "name" in unadj_r.keys() else ""
                else:
                    name = ""
                
                # Get unadjusted balances
                if unadj_r:
                    unadj_d, unadj_c = self._balance_to_columns(unadj_r)
                else:
                    unadj_d, unadj_c = 0.0, 0.0
                
                # Get adjusted balances
                if adj_r:
                    adj_d, adj_c = self._balance_to_columns(adj_r)
                else:
                    adj_d, adj_c = 0.0, 0.0
                
                ws_tb.cell(row=ridx, column=1, value=code)
                ws_tb.cell(row=ridx, column=2, value=name)
                ws_tb.cell(row=ridx, column=3, value=float(unadj_d or 0))
                ws_tb.cell(row=ridx, column=4, value=float(unadj_c or 0))
                ws_tb.cell(row=ridx, column=5, value=float(adj_d or 0))
                ws_tb.cell(row=ridx, column=6, value=float(adj_c or 0))
            
            # Add totals row
            last_row = ws_tb.max_row
            if last_row >= 3:
                total_row = last_row + 1
                ws_tb.cell(row=total_row, column=2, value="Totals:").font = Font(bold=True)
                
                # Unadjusted totals
                col_unadj_d = get_column_letter(3)
                col_unadj_c = get_column_letter(4)
                ws_tb.cell(row=total_row, column=3, value=f"=SUM({col_unadj_d}3:{col_unadj_d}{last_row})").font = Font(bold=True)
                ws_tb.cell(row=total_row, column=4, value=f"=SUM({col_unadj_c}3:{col_unadj_c}{last_row})").font = Font(bold=True)
                
                # Adjusted totals
                col_adj_d = get_column_letter(5)
                col_adj_c = get_column_letter(6)
                ws_tb.cell(row=total_row, column=5, value=f"=SUM({col_adj_d}3:{col_adj_d}{last_row})").font = Font(bold=True)
                ws_tb.cell(row=total_row, column=6, value=f"=SUM({col_adj_c}3:{col_adj_c}{last_row})").font = Font(bold=True)
            
            # Set column widths
            ws_tb.column_dimensions[get_column_letter(1)].width = 10  # Code
            ws_tb.column_dimensions[get_column_letter(2)].width = 25  # Account Name
            ws_tb.column_dimensions[get_column_letter(3)].width = 18  # Unadjusted Debit
            ws_tb.column_dimensions[get_column_letter(4)].width = 18  # Unadjusted Credit
            ws_tb.column_dimensions[get_column_letter(5)].width = 18  # Adjusted Debit
            ws_tb.column_dimensions[get_column_letter(6)].width = 18  # Adjusted Credit

            wb.save(path)
            messagebox.showinfo("Exported", "Trial balance exported to Excel.")
        except Exception as e:
            messagebox.showerror("Export Error", str(e))

    def _export_fs(self) -> None:
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", 
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"financial_statements_{self.fs_date_to.get() or 'report'}.xlsx"
        )
        if not path:
            return
            
        try:
            # Get the content from all financial statement tabs
            statements = []
            
            # Get Income Statement
            self.income_text.config(state=tk.NORMAL)
            income_content = self.income_text.get("1.0", tk.END).splitlines()
            self.income_text.config(state=tk.DISABLED)
            if any(line.strip() for line in income_content):
                statements.append(("Income Statement", income_content))
            
            # Get Balance Sheet
            self.balance_sheet_text.config(state=tk.NORMAL)
            balance_content = self.balance_sheet_text.get("1.0", tk.END).splitlines()
            self.balance_sheet_text.config(state=tk.DISABLED)
            if any(line.strip() for line in balance_content):
                statements.append(("Balance Sheet", balance_content))
            
            # Get Cash Flow Statement if available
            if hasattr(self, 'cash_flow_text'):
                self.cash_flow_text.config(state=tk.NORMAL)
                cash_flow_content = self.cash_flow_text.get("1.0", tk.END).splitlines()
                self.cash_flow_text.config(state=tk.DISABLED)
                if any(line.strip() for line in cash_flow_content):
                    statements.append(("Cash Flow", cash_flow_content))
            
            if not statements:
                messagebox.showinfo("No Data", "No financial statement data available to export.")
                return
            
            # Create Excel workbook with multiple sheets
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            
            # Remove default sheet if not needed
            if statements:
                wb.remove(wb.active)
            
            # Add each statement as a separate sheet
            for sheet_name, lines in statements:
                ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name max 31 chars
                
                # Add header with formatting
                header = ws.cell(row=1, column=1, value=sheet_name)
                header.font = Font(bold=True, size=14)
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
                ws.row_dimensions[1].height = 24
                
                # Add content
                for row_idx, line in enumerate(lines, start=3):  # Start from row 3 to leave space for header
                    if not line.strip():
                        continue
                    ws.cell(row=row_idx, column=1, value=line)
                
                # Auto-adjust column width
                for col in ws.columns:
                    if not col:
                        continue  # Skip empty columns
                        
                    max_length = 0
                    try:
                        # Get the column letter from the first cell in the column
                        column_letter = col[0].column_letter
                        
                        # Find the maximum content length in the column
                        for cell in col:
                            try:
                                if cell.value and len(str(cell.value).strip()) > max_length:
                                    max_length = len(str(cell.value).strip())
                            except:
                                pass
                                
                        # Set column width with some padding and a maximum
                        if max_length > 0:
                            adjusted_width = (max_length + 2) * 1.2
                            ws.column_dimensions[column_letter].width = min(adjusted_width, 50)  # Cap at 50
                    except (IndexError, AttributeError) as e:
                        print(f"Warning: Could not adjust column width: {e}")
            
            # Save the workbook
            wb.save(path)
            messagebox.showinfo("Export Successful", "Financial statements exported to Excel successfully!")
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export financial statements: {str(e)}")
            raise

    def _export_all_to_excel(self, path: Optional[str] = None, notify: bool = True) -> Optional[str]:
        """Export Journal, Ledger, Trial Balance, and Financial Statements into one Excel workbook."""
        if path is None:
            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")],
                initialfile=f"techfix_export_{datetime.now().date().isoformat()}.xlsx",
            )
        if not path:
            return None

        try:
            try:
                from openpyxl import Workbook
                from openpyxl.utils import get_column_letter
                from openpyxl.styles import Font
            except ImportError as e:
                raise RuntimeError("openpyxl is required for Excel export. Install with: pip install openpyxl") from e

            wb = Workbook()

            # --- Journal sheet ---
            ws_j = wb.active
            ws_j.title = "Journal"
            journal_rows = db.fetch_journal(period_id=self.engine.current_period_id, conn=self.engine.conn)
            j_headers = ["entry_id", "date", "description", "code", "name", "debit", "credit"]
            for cidx, h in enumerate(j_headers, start=1):
                cell = ws_j.cell(row=1, column=cidx, value=h)
                cell.font = Font(bold=True)
            for ridx, r in enumerate(journal_rows, start=2):
                ws_j.cell(row=ridx, column=1, value=r["entry_id"])
                ws_j.cell(row=ridx, column=2, value=r["date"])
                ws_j.cell(row=ridx, column=3, value=r["description"])
                ws_j.cell(row=ridx, column=4, value=r["code"])
                ws_j.cell(row=ridx, column=5, value=r["name"])
                # numeric values
                try:
                    ws_j.cell(row=ridx, column=6, value=float(r["debit"] or 0))
                except Exception:
                    ws_j.cell(row=ridx, column=6, value=r["debit"])
                try:
                    ws_j.cell(row=ridx, column=7, value=float(r["credit"] or 0))
                except Exception:
                    ws_j.cell(row=ridx, column=7, value=r["credit"])

            # Add totals row for journal debit/credit
            last = ws_j.max_row
            if last >= 2:
                fcol = get_column_letter(6)
                gcol = get_column_letter(7)
                total_row = last + 1
                ws_j.cell(row=total_row, column=5, value="Totals:").font = Font(bold=True)
                ws_j.cell(row=total_row, column=6, value=f"=SUM({fcol}2:{fcol}{last})")
                ws_j.cell(row=total_row, column=7, value=f"=SUM({gcol}2:{gcol}{last})")

            # Auto-width
            for i, _ in enumerate(j_headers, start=1):
                ws_j.column_dimensions[get_column_letter(i)].width = 15

            # --- Ledger sheet ---
            ws_l = wb.create_sheet(title="Ledger")
            l_headers = ["account_id", "code", "name", "date", "description", "debit", "credit"]
            for cidx, h in enumerate(l_headers, start=1):
                cell = ws_l.cell(row=1, column=cidx, value=h)
                cell.font = Font(bold=True)
            ledger_rows = db.fetch_ledger(period_id=self.engine.current_period_id, conn=self.engine.conn)
            for ridx, r in enumerate(ledger_rows, start=2):
                # r is a sqlite3.Row; access fields by key with safety checks
                if "account_id" in r.keys():
                    ws_l.cell(row=ridx, column=1, value=r["account_id"])
                if "code" in r.keys():
                    ws_l.cell(row=ridx, column=2, value=r["code"])
                else:
                    # fallback blank
                    ws_l.cell(row=ridx, column=2, value="")
                ws_l.cell(row=ridx, column=3, value=r["name"])
                if "date" in r.keys():
                    ws_l.cell(row=ridx, column=4, value=r["date"])
                if "description" in r.keys():
                    ws_l.cell(row=ridx, column=5, value=r["description"])
                # debit/credit as numeric when possible
                try:
                    ws_l.cell(row=ridx, column=6, value=float(r["debit"] or 0))
                except Exception:
                    try:
                        ws_l.cell(row=ridx, column=6, value=r["debit"])
                    except Exception:
                        ws_l.cell(row=ridx, column=6, value="")
                try:
                    ws_l.cell(row=ridx, column=7, value=float(r["credit"] or 0))
                except Exception:
                    try:
                        ws_l.cell(row=ridx, column=7, value=r["credit"])
                    except Exception:
                        ws_l.cell(row=ridx, column=7, value="")

            # Add totals row for ledger debit/credit
            last_l = ws_l.max_row
            if last_l >= 2:
                ecol = get_column_letter(6)
                fcol = get_column_letter(7)
                total_row = last_l + 1
                ws_l.cell(row=total_row, column=5, value="Totals:").font = Font(bold=True)
                ws_l.cell(row=total_row, column=6, value=f"=SUM({ecol}2:{ecol}{last_l})")
                ws_l.cell(row=total_row, column=7, value=f"=SUM({fcol}2:{fcol}{last_l})")

            for i, _ in enumerate(l_headers, start=1):
                ws_l.column_dimensions[get_column_letter(i)].width = 15

            # --- Combined Trial Balance sheet ---
            ws_tb = wb.create_sheet(title="Trial Balance")
            
            # Title row
            title_cell = ws_tb.cell(row=1, column=1, value="TRIAL BALANCE")
            title_cell.font = Font(bold=True, size=14)
            ws_tb.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
            
            # Headers with clear labels
            tb_headers = [
                "Code", 
                "Account Name", 
                "Unadjusted Debit", 
                "Unadjusted Credit",
                "Adjusted Debit",
                "Adjusted Credit"
            ]
            for cidx, h in enumerate(tb_headers, start=1):
                cell = ws_tb.cell(row=2, column=cidx, value=h)
                cell.font = Font(bold=True)
            
            # Get both unadjusted and adjusted trial balances
            unadj_tb_data = db.compute_unadjusted_trial_balance(
                period_id=self.engine.current_period_id,
                include_temporary=True,
                conn=self.engine.conn
            )
            adj_tb_rows = db.compute_trial_balance(
                period_id=self.engine.current_period_id,
                conn=self.engine.conn
            )
            
            # Create dictionaries for easy lookup
            unadj_by_code = {r["code"]: r for r in unadj_tb_data if "code" in r.keys() and r["code"]}
            adj_by_code = {r["code"]: r for r in adj_tb_rows if "code" in r.keys() and r["code"]}
            
            # Get all unique account codes
            all_codes = sorted(set(list(unadj_by_code.keys()) + list(adj_by_code.keys())))
            
            # Write data rows
            for ridx, code in enumerate(all_codes, start=3):
                unadj_r = unadj_by_code.get(code)
                adj_r = adj_by_code.get(code)
                
                # Get account name (prefer from adjusted, fallback to unadjusted)
                if adj_r:
                    name = adj_r["name"] if "name" in adj_r.keys() else ""
                elif unadj_r:
                    name = unadj_r["name"] if "name" in unadj_r.keys() else ""
                else:
                    name = ""
                
                # Get unadjusted balances
                if unadj_r:
                    unadj_d, unadj_c = self._balance_to_columns(unadj_r)
                else:
                    unadj_d, unadj_c = 0.0, 0.0
                
                # Get adjusted balances
                if adj_r:
                    adj_d, adj_c = self._balance_to_columns(adj_r)
                else:
                    adj_d, adj_c = 0.0, 0.0
                
                ws_tb.cell(row=ridx, column=1, value=code)
                ws_tb.cell(row=ridx, column=2, value=name)
                ws_tb.cell(row=ridx, column=3, value=float(unadj_d or 0))
                ws_tb.cell(row=ridx, column=4, value=float(unadj_c or 0))
                ws_tb.cell(row=ridx, column=5, value=float(adj_d or 0))
                ws_tb.cell(row=ridx, column=6, value=float(adj_c or 0))
            
            # Add totals row
            last_tb = ws_tb.max_row
            if last_tb >= 3:
                total_row = last_tb + 1
                ws_tb.cell(row=total_row, column=2, value="Totals:").font = Font(bold=True)
                
                # Unadjusted totals
                col_unadj_d = get_column_letter(3)
                col_unadj_c = get_column_letter(4)
                ws_tb.cell(row=total_row, column=3, value=f"=SUM({col_unadj_d}3:{col_unadj_d}{last_tb})").font = Font(bold=True)
                ws_tb.cell(row=total_row, column=4, value=f"=SUM({col_unadj_c}3:{col_unadj_c}{last_tb})").font = Font(bold=True)
                
                # Adjusted totals
                col_adj_d = get_column_letter(5)
                col_adj_c = get_column_letter(6)
                ws_tb.cell(row=total_row, column=5, value=f"=SUM({col_adj_d}3:{col_adj_d}{last_tb})").font = Font(bold=True)
                ws_tb.cell(row=total_row, column=6, value=f"=SUM({col_adj_c}3:{col_adj_c}{last_tb})").font = Font(bold=True)
            
            # Set column widths
            ws_tb.column_dimensions[get_column_letter(1)].width = 10  # Code
            ws_tb.column_dimensions[get_column_letter(2)].width = 25  # Account Name
            ws_tb.column_dimensions[get_column_letter(3)].width = 18  # Unadjusted Debit
            ws_tb.column_dimensions[get_column_letter(4)].width = 18  # Unadjusted Credit
            ws_tb.column_dimensions[get_column_letter(5)].width = 18  # Adjusted Debit
            ws_tb.column_dimensions[get_column_letter(6)].width = 18  # Adjusted Credit

            # --- Post-Closing Trial Balance sheet ---
            ws_pctb = wb.create_sheet(title="Post-Closing TB")
            
            # Title row
            title_cell = ws_pctb.cell(row=1, column=1, value="POST-CLOSING TRIAL BALANCE")
            title_cell.font = Font(bold=True, size=14)
            ws_pctb.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
            
            # Headers
            pctb_headers = [
                "Code", 
                "Account Name", 
                "Debit", 
                "Credit"
            ]
            for cidx, h in enumerate(pctb_headers, start=1):
                cell = ws_pctb.cell(row=2, column=cidx, value=h)
                cell.font = Font(bold=True)
            
            # Get post-closing trial balance (exclude temporary accounts)
            pctb_rows = db.compute_trial_balance(
                period_id=self.engine.current_period_id,
                include_temporary=False,
                conn=self.engine.conn
            )
            
            # Filter to only show accounts with non-zero balances
            def _has_activity_pc(r: dict) -> bool:
                d, c = self._balance_to_columns(r)
                return bool((d or 0) != 0 or (c or 0) != 0)
            
            pctb_rows = [r for r in pctb_rows if _has_activity_pc(r)]
            
            # Write data rows
            for ridx, r in enumerate(pctb_rows, start=3):
                code = r["code"] if "code" in r.keys() else ""
                name = r["name"] if "name" in r.keys() else ""
                d, c = self._balance_to_columns(r)
                
                ws_pctb.cell(row=ridx, column=1, value=code)
                ws_pctb.cell(row=ridx, column=2, value=name)
                ws_pctb.cell(row=ridx, column=3, value=float(d or 0))
                ws_pctb.cell(row=ridx, column=4, value=float(c or 0))
            
            # Add totals row
            last_pctb = ws_pctb.max_row
            if last_pctb >= 3:
                total_row = last_pctb + 1
                ws_pctb.cell(row=total_row, column=2, value="Totals:").font = Font(bold=True)
                
                col_d = get_column_letter(3)
                col_c = get_column_letter(4)
                ws_pctb.cell(row=total_row, column=3, value=f"=SUM({col_d}3:{col_d}{last_pctb})").font = Font(bold=True)
                ws_pctb.cell(row=total_row, column=4, value=f"=SUM({col_c}3:{col_c}{last_pctb})").font = Font(bold=True)
            
            # Set column widths
            ws_pctb.column_dimensions[get_column_letter(1)].width = 10  # Code
            ws_pctb.column_dimensions[get_column_letter(2)].width = 25  # Account Name
            ws_pctb.column_dimensions[get_column_letter(3)].width = 18  # Debit
            ws_pctb.column_dimensions[get_column_letter(4)].width = 18  # Credit

            # Helper function for adjusting entries (used by Worksheet)
            def _tb_filtered_adjusting():
                period_id = self.engine.current_period_id
                # The balance expression uses 2 placeholders, and it's used 4 times in the SQL
                # (twice in net_debit calculation, twice in net_credit calculation)
                # So we need 8 placeholders total: 2 * 4 = 8
                balance = """
                    COALESCE(SUM(CASE 
                        WHEN je.period_id = ? AND je.is_adjusting = 1 THEN jl.debit 
                        ELSE 0 
                    END), 0) - 
                    COALESCE(SUM(CASE 
                        WHEN je.period_id = ? AND je.is_adjusting = 1 THEN jl.credit 
                        ELSE 0 
                    END), 0)
                """
                sql = f"""
                    SELECT a.code, a.name, a.type,
                           ROUND(CASE WHEN ({balance}) > 0 THEN ({balance}) ELSE 0 END, 2) AS net_debit,
                           ROUND(CASE WHEN ({balance}) < 0 THEN -(({balance})) ELSE 0 END, 2) AS net_credit
                    FROM accounts a
                    LEFT JOIN journal_lines jl ON jl.account_id = a.id
                    LEFT JOIN journal_entries je ON je.id = jl.entry_id
                    WHERE a.is_active = 1
                    GROUP BY a.id, a.code, a.name, a.type
                    ORDER BY a.code
                """
                # Supply 8 parameters: period_id for each of the 8 placeholders
                cur = self.engine.conn.execute(sql, [period_id] * 8)
                return list(cur.fetchall())

            ws_w = wb.create_sheet(title="Worksheet")
            ws_w.append([
                "Account No.",
                "Account Title",
                "Unadjusted Trial Balance Dr",
                "Unadjusted Trial Balance Cr",
                "Adjustments Dr",
                "Adjustments Cr",
                "Adjusted Trial Balance Dr",
                "Adjusted Trial Balance Cr",
                "Statement of Financial Performance Dr",
                "Statement of Financial Performance Cr",
                "Statement of Financial Position Dr",
                "Statement of Financial Position Cr",
            ])
            # Use the data we already fetched for Trial Balance sheet
            # unadj_by_code and adj_by_code are already created above
            # Use helper function for adjusting entries only
            adjs = _tb_filtered_adjusting()
            # adj_tb_rows is already fetched above as adj_tb_rows
            # unadj_by_code and adj_by_code are already created above for Trial Balance sheet
            adj_by_code_adjs = {r["code"]: r for r in adjs}
            adjtb_by_code = adj_by_code  # Already created above
            codes = sorted(set(list(unadj_by_code.keys()) + list(adj_by_code_adjs.keys()) + list(adjtb_by_code.keys())))
            totals = {"un_dr":0.0,"un_cr":0.0,"aj_dr":0.0,"aj_cr":0.0,"ad_dr":0.0,"ad_cr":0.0,"is_dr":0.0,"is_cr":0.0,"sfp_dr":0.0,"sfp_cr":0.0}
            for code in codes:
                ru = unadj_by_code.get(code)
                ra = adj_by_code_adjs.get(code)
                rt = adjtb_by_code.get(code)
                # Get name and type from first available row
                row_for_name = rt or ru or ra
                if row_for_name:
                    name = row_for_name["name"] if "name" in row_for_name.keys() else ""
                    typ = row_for_name["type"] if "type" in row_for_name.keys() else ""
                else:
                    name = ""
                    typ = ""
                un_dr = float(ru["net_debit"]) if ru else 0.0
                un_cr = float(ru["net_credit"]) if ru else 0.0
                aj_dr = float(ra["net_debit"]) if ra else 0.0
                aj_cr = float(ra["net_credit"]) if ra else 0.0
                ad_dr = float(rt["net_debit"]) if rt else 0.0
                ad_cr = float(rt["net_credit"]) if rt else 0.0
                is_dr = ad_dr if typ.lower() in ("revenue","expense") and ad_dr>0 else 0.0
                is_cr = ad_cr if typ.lower() in ("revenue","expense") and ad_cr>0 else 0.0
                sfp_dr = ad_dr if typ.lower() not in ("revenue","expense") and ad_dr>0 else 0.0
                sfp_cr = ad_cr if typ.lower() not in ("revenue","expense") and ad_cr>0 else 0.0
                ws_w.append([code, name, un_dr, un_cr, aj_dr, aj_cr, ad_dr, ad_cr, is_dr, is_cr, sfp_dr, sfp_cr])
                totals["un_dr"] += un_dr; totals["un_cr"] += un_cr
                totals["aj_dr"] += aj_dr; totals["aj_cr"] += aj_cr
                totals["ad_dr"] += ad_dr; totals["ad_cr"] += ad_cr
                totals["is_dr"] += is_dr; totals["is_cr"] += is_cr
                totals["sfp_dr"] += sfp_dr; totals["sfp_cr"] += sfp_cr
            ws_w.append(["TOTAL","", totals["un_dr"], totals["un_cr"], totals["aj_dr"], totals["aj_cr"], totals["ad_dr"], totals["ad_cr"], totals["is_dr"], totals["is_cr"], totals["sfp_dr"], totals["sfp_cr"]])
            net_income = round(totals["is_cr"] - totals["is_dr"], 2)
            if net_income != 0:
                ws_w.append(["INCOME STATEMENT","", "", "", "", "", "", "", 0.0 if net_income>0 else abs(net_income), net_income if net_income>0 else 0.0, "", ""])
                ws_w.append(["TOTAL","", totals["un_dr"], totals["un_cr"], totals["aj_dr"], totals["aj_cr"], totals["ad_dr"], totals["ad_cr"], totals["is_dr"] + (0.0 if net_income>0 else abs(net_income)), totals["is_cr"] + (net_income if net_income>0 else 0.0), totals["sfp_dr"], totals["sfp_cr"]])

            # --- Financial statements sheets ---
            # Income Statement
            self.income_text.config(state=tk.NORMAL)
            income_lines = self.income_text.get("1.0", tk.END).splitlines()
            self.income_text.config(state=tk.DISABLED)
            if any(line.strip() for line in income_lines):
                ws_inc = wb.create_sheet(title="Income Statement")
                for ridx, line in enumerate(income_lines, start=1):
                    ws_inc.cell(row=ridx, column=1, value=line)
                ws_inc.column_dimensions[get_column_letter(1)].width = 120

            # Balance Sheet
            self.balance_sheet_text.config(state=tk.NORMAL)
            bs_lines = self.balance_sheet_text.get("1.0", tk.END).splitlines()
            self.balance_sheet_text.config(state=tk.DISABLED)
            if any(line.strip() for line in bs_lines):
                ws_bs = wb.create_sheet(title="Balance Sheet")
                for ridx, line in enumerate(bs_lines, start=1):
                    ws_bs.cell(row=ridx, column=1, value=line)
                ws_bs.column_dimensions[get_column_letter(1)].width = 120

            # Cash Flow
            if hasattr(self, 'cash_flow_text'):
                self.cash_flow_text.config(state=tk.NORMAL)
                cf_lines = self.cash_flow_text.get("1.0", tk.END).splitlines()
                self.cash_flow_text.config(state=tk.DISABLED)
                if any(line.strip() for line in cf_lines):
                    ws_cf = wb.create_sheet(title="Cash Flow")
                    for ridx, line in enumerate(cf_lines, start=1):
                        ws_cf.cell(row=ridx, column=1, value=line)
                    ws_cf.column_dimensions[get_column_letter(1)].width = 120

            # Save workbook
            wb.save(path)
            if notify:
                messagebox.showinfo("Export Successful", f"All data exported to {path}")

        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export all data: {str(e)}")
            raise
        
        return path

    def _export_all_to_email(self) -> None:
        """Export all reports to Excel and send as an email attachment."""
        # Allow fully automatic mode via environment variables; otherwise show a nicer dialog.
        # Default sender is empty (no placeholder). If TECHFIX_SMTP_USER is set, use it,
        # and optionally apply suffix for plus-addressing.
        default_from = os.environ.get("TECHFIX_SMTP_USER", "").strip()
        suffix = os.environ.get("TECHFIX_EMAIL_SUFFIX", "").strip()
        if suffix and default_from:
            local, at, domain = default_from.partition("@")
            default_from = f"{local}+{suffix}{at}{domain}"

        last_account = None
        if hasattr(self, "email_accounts") and self.email_accounts:
            if isinstance(self.email_accounts[0], dict):
                last_account = self.email_accounts[0]

        env_values = {
            "recipient": os.environ.get("TECHFIX_EMAIL_TO", ""),
            "smtp_server": os.environ.get("TECHFIX_SMTP_SERVER", ""),
            "smtp_port": os.environ.get("TECHFIX_SMTP_PORT", ""),
            "smtp_user": os.environ.get("TECHFIX_SMTP_USER", ""),
            "smtp_password": os.environ.get("TECHFIX_SMTP_PASS", ""),
        }

        # Defaults for dialog: prefer env if set, otherwise last saved, otherwise fallback.
        dialog_defaults = {
            "recipient": env_values["recipient"] or (last_account or {}).get("recipient", ""),
            "smtp_server": env_values["smtp_server"] or (last_account or {}).get("smtp_server", "smtp.gmail.com"),
            "smtp_port": env_values["smtp_port"] or (last_account or {}).get("smtp_port", "587"),
            "smtp_user": env_values["smtp_user"] or (last_account or {}).get("smtp_user", default_from),
            "smtp_password": env_values["smtp_password"] or (last_account or {}).get("smtp_password", ""),
        }

        # If all env vars are provided, skip dialog; otherwise allow picking saved accounts.
        if all(env_values.values()):
            recipient = env_values["recipient"]
            smtp_server = env_values["smtp_server"]
            smtp_port = env_values["smtp_port"]
            smtp_user = env_values["smtp_user"]
            smtp_password = env_values["smtp_password"]
        else:
            config = self._prompt_email_settings(dialog_defaults)
            if not config:
                return
            recipient = config["recipient"]
            smtp_server = config["smtp_server"]
            smtp_port = config["smtp_port"]
            smtp_user = config["smtp_user"]
            smtp_password = config["smtp_password"]

        temp_path = None
        try:
            # Create a temporary Excel export without showing a save dialog
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                temp_path = tmp.name
            
            exported_path = self._export_all_to_excel(path=temp_path, notify=False)
            if not exported_path or not Path(exported_path).exists():
                raise RuntimeError("Failed to generate export file.")

            # Compose the email with attachment
            msg = EmailMessage()
            msg["Subject"] = f"TechFix Export - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            msg["From"] = smtp_user
            msg["To"] = recipient

            plain_body = (
                "Hello,\n\n"
                "Your TechFix full export is attached (journal, ledger, trial balance, financial statements).\n"
                "If you have questions, reply to this email.\n\n"
                "Thanks,\nTechFix"
            )
            html_body = f"""
            <html>
              <body style="font-family: 'Segoe UI', Arial, sans-serif; color: #1f2937; background: #f8fafc; padding: 16px;">
                <div style="max-width: 520px; margin: 0 auto; background: #ffffff; border: 1px solid #e5e7eb; border-radius: 10px; padding: 18px;">
                  <h2 style="margin: 0 0 12px 0; color: #2563eb;">TechFix Export</h2>
                  <p style="margin: 0 0 12px 0;">Your full export is attached:</p>
                  <ul style="margin: 0 0 12px 18px; padding: 0; color: #374151;">
                    <li>Journal</li>
                    <li>Ledger</li>
                    <li>Trial Balance</li>
                    <li>Financial Statements</li>
                  </ul>
                  <p style="margin: 0 0 12px 0; color: #4b5563;">Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}</p>
                  <p style="margin: 0; color: #111827;">Thanks,<br/>TechFix</p>
                </div>
              </body>
            </html>
            """
            msg.set_content(plain_body)
            msg.add_alternative(html_body, subtype="html")

            with open(exported_path, "rb") as f:
                data = f.read()
            filename = Path(exported_path).name
            msg.add_attachment(
                data,
                maintype="application",
                subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=filename,
            )

            with smtplib.SMTP(smtp_server, int(smtp_port)) as server:
                server.starttls()
                server.login(smtp_user, smtp_password)
                server.send_message(msg)

            messagebox.showinfo("Email Sent", f"Export emailed to {recipient}.")
        except Exception as e:
            messagebox.showerror("Email Error", f"Failed to email export: {e}")
        finally:
            if temp_path:
                try:
                    Path(temp_path).unlink(missing_ok=True)
                except Exception:
                    pass

    def _prompt_email_settings(self, defaults: dict) -> Optional[dict]:
        """Display a styled, single-form dialog to collect email settings."""
        dialog = tk.Toplevel(self)
        dialog.title("Send Export via Email")
        dialog.transient(self)
        dialog.grab_set()
        dialog.resizable(False, False)

        # Close on Esc
        dialog.bind("<Escape>", lambda *_: dialog.destroy())

        container = ttk.Frame(dialog, padding=12, style="Techfix.Surface.TFrame")
        container.grid(row=0, column=0, sticky="nsew")

        saved_accounts = [
            acct for acct in getattr(self, "email_accounts", [])
            if isinstance(acct, dict) and acct.get("smtp_user")
        ]
        saved_labels = [acct.get("smtp_user", "") for acct in saved_accounts]
        saved_lookup = {label: acct for label, acct in zip(saved_labels, saved_accounts)}

        saved_email_frame = ttk.Frame(container, style="Techfix.Surface.TFrame")
        saved_email_frame.grid(row=0, column=0, sticky="ew", pady=(0,8))
        saved_email_frame.columnconfigure(0, weight=1)
        
        ttk.Label(saved_email_frame, text="Saved Emails", style="Techfix.TLabel").grid(row=0, column=0, sticky="w", pady=(0,4))
        saved_var = tk.StringVar(value=saved_labels[0] if saved_labels else "")
        saved_combo = ttk.Combobox(
            saved_email_frame,
            textvariable=saved_var,
            values=saved_labels,
            width=30,
            state="readonly" if saved_labels else "disabled",
            style="Techfix.TEntry"
        )
        saved_combo.grid(row=1, column=0, sticky="ew", padx=(0,6))
        
        def delete_selected_email():
            selected_label = saved_var.get().strip()
            if not selected_label or not messagebox.askyesno(
                "Delete Email",
                f"Delete saved email account '{selected_label}'?",
                parent=dialog
            ):
                return
            
            # Remove from accounts list
            if hasattr(self, "email_accounts") and isinstance(self.email_accounts, list):
                self.email_accounts = [
                    acct for acct in self.email_accounts
                    if isinstance(acct, dict) and acct.get("smtp_user", "") != selected_label
                ]
                self._save_window_settings()
            
            # Refresh the combobox
            updated_accounts = [
                acct for acct in getattr(self, "email_accounts", [])
                if isinstance(acct, dict) and acct.get("smtp_user")
            ]
            updated_labels = [acct.get("smtp_user", "") for acct in updated_accounts]
            saved_lookup.clear()
            saved_lookup.update({label: acct for label, acct in zip(updated_labels, updated_accounts)})
            
            saved_combo['values'] = updated_labels
            if updated_labels:
                saved_var.set(updated_labels[0])
                saved_combo.config(state="readonly")
                delete_btn.config(state="normal")
                apply_saved_account(updated_labels[0])
            else:
                saved_var.set("")
                saved_combo.config(state="disabled")
                delete_btn.config(state="disabled")
                # Clear form fields
                recipient_var.set("")
                server_var.set("smtp.gmail.com")
                port_var.set("587")
                user_var.set("")
                pass_var.set("")
                remember_var.set(False)
                user_manually_edited[0] = False
        
        delete_btn = ttk.Button(
            saved_email_frame,
            text="Delete",
            command=delete_selected_email,
            style="Techfix.Tertiary.TButton",
            state="normal" if saved_labels else "disabled"
        )
        delete_btn.grid(row=1, column=1, sticky="e")

        ttk.Label(container, text="Recipient Email", style="Techfix.TLabel").grid(row=2, column=0, sticky="w", pady=(0,4))
        recipient_var = tk.StringVar(value=defaults.get("recipient",""))
        recipient_entry = ttk.Entry(container, textvariable=recipient_var, width=36, style="Techfix.TEntry")
        recipient_entry.grid(row=3, column=0, sticky="ew", pady=(0,8))

        ttk.Label(container, text="SMTP Server", style="Techfix.TLabel").grid(row=4, column=0, sticky="w", pady=(0,4))
        server_var = tk.StringVar(value=defaults.get("smtp_server","smtp.gmail.com"))
        ttk.Entry(container, textvariable=server_var, width=36, style="Techfix.TEntry").grid(row=5, column=0, sticky="ew", pady=(0,8))

        ttk.Label(container, text="Port", style="Techfix.TLabel").grid(row=6, column=0, sticky="w", pady=(0,4))
        port_var = tk.StringVar(value=defaults.get("smtp_port","587"))
        ttk.Entry(container, textvariable=port_var, width=10, style="Techfix.TEntry").grid(row=7, column=0, sticky="w", pady=(0,8))

        ttk.Label(container, text="Username", style="Techfix.TLabel").grid(row=8, column=0, sticky="w", pady=(0,4))
        user_var = tk.StringVar(value=defaults.get("smtp_user","") or defaults.get("recipient",""))
        user_entry = ttk.Entry(container, textvariable=user_var, width=36, style="Techfix.TEntry")
        user_entry.grid(row=9, column=0, sticky="ew", pady=(0,8))
        
        # Track if user manually edited username to avoid overwriting
        user_manually_edited = [False]
        loading_saved_account = [False]
        
        def on_user_focus_in(event):
            user_manually_edited[0] = True
        
        def on_recipient_change(*args):
            # Skip auto-fill when loading saved account or if user manually edited username
            if loading_saved_account[0] or user_manually_edited[0]:
                return
            recipient_val = recipient_var.get().strip()
            if recipient_val:
                user_var.set(recipient_val)
        
        user_entry.bind("<FocusIn>", on_user_focus_in)
        recipient_var.trace_add("write", on_recipient_change)

        ttk.Label(container, text="Password", style="Techfix.TLabel").grid(row=10, column=0, sticky="w", pady=(0,4))
        pass_var = tk.StringVar(value=defaults.get("smtp_password",""))
        ttk.Entry(container, textvariable=pass_var, width=36, show="*", style="Techfix.TEntry").grid(row=11, column=0, sticky="ew", pady=(0,8))

        remember_var = tk.BooleanVar(value=bool(defaults.get("smtp_password")))
        ttk.Checkbutton(
            container,
            text="Remember password",
            variable=remember_var,
            style="Techfix.TCheckbutton"
        ).grid(row=12, column=0, sticky="w", pady=(0,8))

        status_var = tk.StringVar(value="")
        ttk.Label(container, textvariable=status_var, style="Techfix.AppBar.TLabel", foreground=self.palette.get("text_secondary", "#4b5563")).grid(row=13, column=0, sticky="w", pady=(0,8))

        btns = ttk.Frame(container, style="Techfix.Surface.TFrame")
        btns.grid(row=14, column=0, sticky="e")

        result: dict | None = None

        def apply_saved_account(label: str) -> None:
            acct = saved_lookup.get(label)
            if not acct:
                return
            loading_saved_account[0] = True  # Disable auto-fill while loading
            user_manually_edited[0] = False  # Reset flag when loading saved account
            recipient_var.set(acct.get("recipient", ""))
            server_var.set(acct.get("smtp_server", "smtp.gmail.com"))
            port_var.set(acct.get("smtp_port", "587"))
            user_var.set(acct.get("smtp_user", ""))
            pwd = acct.get("smtp_password", "")
            pass_var.set(pwd)
            remember_var.set(bool(pwd))
            loading_saved_account[0] = False  # Re-enable auto-fill after loading

        if saved_labels:
            apply_saved_account(saved_labels[0])
            saved_combo.bind("<<ComboboxSelected>>", lambda *_: apply_saved_account(saved_var.get()))

        def submit() -> None:
            nonlocal result
            rec = recipient_var.get().strip()
            svr = server_var.get().strip()
            prt = port_var.get().strip() or "587"
            usr = user_var.get().strip()
            pwd = pass_var.get()
            if not rec:
                status_var.set("Recipient is required.")
                return
            if not svr:
                status_var.set("SMTP server is required.")
                return
            if not usr:
                status_var.set("Username is required.")
                return
            if not pwd:
                status_var.set("Password is required.")
                return
            result = {
                "recipient": rec,
                "smtp_server": svr,
                "smtp_port": prt,
                "smtp_user": usr,
                "smtp_password": pwd,
            }
            # Save recent account (optional password)
            self._save_email_account(result, remember_var.get())
            dialog.destroy()

        ttk.Button(btns, text="Send", command=submit, style="Techfix.TButton").grid(row=0, column=0, padx=(0,6))
        ttk.Button(btns, text="Cancel", command=dialog.destroy, style="Techfix.Tertiary.TButton").grid(row=0, column=1)

        dialog.columnconfigure(0, weight=1)
        container.columnconfigure(0, weight=1)
        recipient_var.set(recipient_var.get())  # ensure var binding
        dialog.wait_window()
        return result

    def _export_program_docs_pdf(self) -> None:
        """Generate a PDF with program documentation (overview, usage, modules)."""
        path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")],
            initialfile=f"techfix_documentation_{__import__('datetime').date.today().isoformat()}.pdf",
        )
        if not path:
            return
        try:
            try:
                from reportlab.lib.pagesizes import letter
                from reportlab.lib.units import inch
                from reportlab.pdfgen import canvas
            except ImportError as e:
                raise RuntimeError("reportlab is required for PDF export. Install with: pip install reportlab") from e

            c = canvas.Canvas(path, pagesize=letter)
            width, height = letter
            margin = 0.75 * inch

            def draw_title(title: str, y: float) -> float:
                c.setFont("Helvetica-Bold", 16)
                c.drawString(margin, y, title)
                return y - 0.3 * inch

            def draw_text(lines: list[str], y: float, font="Helvetica", size=10) -> float:
                c.setFont(font, size)
                for ln in lines:
                    if y < margin + 0.5 * inch:
                        c.showPage(); y = height - margin
                        c.setFont(font, size)
                    c.drawString(margin, y, ln)
                    y -= 0.18 * inch
                return y

            from datetime import date
            y = height - margin
            y = draw_title("TechFix Documentation", y)
            c.setFont("Helvetica", 10)
            c.drawString(margin, y, f"Generated: {date.today().isoformat()}")
            y -= 0.3 * inch

            # Overview
            y = draw_title("Overview", y)
            overview = [
                "TechFix is a desktop accounting practice app (Tkinter + SQLite).",
                "It guides users through the 10-step accounting cycle and produces statements.",
            ]
            y = draw_text(overview, y)

            # Features
            y = draw_title("Key Features", y)
            features = [
                "• Journalization, Ledger, Trial Balance (adjusted/unadjusted)",
                "• Financial Statements (Income Statement, Balance Sheet, Cash Flow)",
                "• Closing and Post-Closing Trial Balance, Reversing Schedule",
                "• Exports: Journal/Ledger/TB/Financials to Excel; All-in-one workbook",
                "• Configurable data directory via TECHFIX_DATA_DIR",
            ]
            y = draw_text(features, y)

            # Modules
            y = draw_title("Modules", y)
            mods = [
                "• techfix/db.py – schema, queries, trial balance, exports",
                "• techfix/accounting.py – orchestration, cycle status, closing/reversing",
                "• techfix/gui.py – Tkinter UI, tabs, theme, exports",
                "• main.py – entry point",
            ]
            y = draw_text(mods, y)

            # Usage
            y = draw_title("Usage", y)
            usage = [
                "Run: python -m techfix or python TECHFIX/TECHFIX/main.py",
                "Configure DB location: set TECHFIX_DATA_DIR to a writable folder",
                "Use the Transactions tab to record entries; Refresh Cycle to update views",
                "Use Export tab for Excel and Documentation outputs",
            ]
            y = draw_text(usage, y)

            # Current Period & Cycle
            try:
                pid = int(self.engine.current_period_id or 0)
                y = draw_title("Current Period", y)
                cp = self.engine.current_period or {}
                cp_lines = [
                    f"• Name: {cp.get('name','')}",
                    f"• Start: {cp.get('start_date','')}  End: {cp.get('end_date','')}  Closed: {cp.get('is_closed',0)}",
                ]
                y = draw_text(cp_lines, y)
                y = draw_title("Cycle Status", y)
                rows = self.engine.get_cycle_status()
                cyc_lines = [f"• {int(r['step'])}. {r['step_name']}: {r['status']}" for r in rows]
                y = draw_text(cyc_lines or ["• (no data)"], y)
            except Exception:
                pass

            y = draw_title("Workflow", y)
            wf_lines = [
                "1. Record transactions in the Transactions pane (debits/credits).",
                "2. Use Refresh Cycle to update Journal, Ledger and Trial Balance.",
                "3. Generate Financial Statements and review the accounting equation.",
                "4. Make Closing Entries, prepare Post‑Closing TB, schedule reversals.",
                "5. Export reports (Excel) and documentation (PDF) from the Export tab.",
            ]
            y = draw_text(wf_lines, y)

            y = draw_title("Equations & Treatment", y)
            eq_lines = [
                "• Income Statement: Net Income = Revenues − Expenses.",
                "• Owner’s Equity Statement: Ending Capital = Beginning Capital + Net Income − Withdrawals.",
                "• Balance Sheet: Assets = Liabilities + Ending Owner’s Equity.",
                "• Contra‑assets reduce total assets (e.g., accumulated depreciation).",
                "• Trial Balance displays net debit/credit by account based on normal side.",
            ]
            y = draw_text(eq_lines, y)

            # Trial Balance Totals
            try:
                y = draw_title("Trial Balance Totals", y)
                inc_temp_bs = True
                try:
                    statuses = self.engine.get_cycle_status()
                    step8 = next((r for r in statuses if int(r['step']) == 8), None)
                    if step8 and (step8['status'] == 'completed'):
                        inc_temp_bs = False
                except Exception:
                    pass
                tb_rows = db.compute_trial_balance(include_temporary=inc_temp_bs, period_id=self.engine.current_period_id, conn=self.engine.conn)
                total_d = 0.0
                total_c = 0.0
                for r in tb_rows:
                    d_amt, c_amt = self._balance_to_columns(r)
                    total_d += float(d_amt or 0)
                    total_c += float(c_amt or 0)
                y = draw_text([f"• Total Debit: {total_d:,.2f}", f"• Total Credit: {total_c:,.2f}"], y)
            except Exception:
                pass

            # Config & Paths
            try:
                y = draw_title("Config", y)
                from pathlib import Path
                data_dir = str(db.DB_DIR.resolve())
                db_path = str(db.DB_PATH.resolve())
                y = draw_text([f"• TECHFIX_DATA_DIR: {data_dir}", f"• Database: {db_path}"], y)
            except Exception:
                pass

            # Read README.md if available and append
            try:
                from pathlib import Path
                readme = Path(__file__).resolve().parents[2] / "README.md"
                if readme.exists():
                    y = draw_title("README Summary", y)
                    # Take first ~40 non-empty lines
                    lines = [ln.strip() for ln in readme.read_text(encoding="utf-8").splitlines() if ln.strip()][:40]
                    y = draw_text(lines, y)
            except Exception:
                pass

            c.showPage(); c.save()
            messagebox.showinfo("Exported", f"Documentation exported to PDF: {path}")
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export documentation: {e}")

    # --------------------- Shared helpers ---------------------
    def _load_all_views(self) -> None:
        self._load_cycle_status()
        if hasattr(self, 'journal_tree'):
            self._load_journal_entries()
        if hasattr(self, 'ledger_tree'):
            self._load_ledger_entries()
        if hasattr(self, 'trial_tree'):
            self._load_trial_balances()
        try:
            self._load_postclosing_tb()
        except Exception:
            pass
        # Load financials without changing cycle step statuses during startup
        self._load_financials(mark_status=False)
        self._load_adjustments()
        self._load_closing_preview()
        self._load_reversing_queue()
        try:
            self._load_recent_transactions()
        except Exception:
            pass

    def _refresh_after_post(self) -> None:
        self._load_journal_entries()
        self._load_ledger_entries()
        self._load_trial_balances()
        self._load_postclosing_tb()
        # Ensure financial statements reflect newly posted entries without
        # requiring the user to re-open the app or manually hit Generate.
        try:
            self._load_financials(mark_status=False)
        except Exception as e:
            # Log the error but don't block other refresh actions
            logger.exception("Error refreshing financial statements after post")
            # Do not block other refresh actions if FS generation fails
            pass
        self._load_cycle_status()
        self._load_adjustments()
        self._load_closing_preview()
        self._load_reversing_queue()
        try:
            self._load_recent_transactions()
        except Exception:
            pass

    def _load_postclosing_tb(self) -> None:
        if not hasattr(self, 'pctb_tree'):
            return
        for item in self.pctb_tree.get_children():
            self.pctb_tree.delete(item)
        try:
            as_of = (self.pctb_date.get().strip() if hasattr(self, 'pctb_date') else '') or None
            rows = db.compute_trial_balance(up_to_date=as_of, include_temporary=False, period_id=self.engine.current_period_id, conn=self.engine.conn)
            # Only show accounts with non-zero balances in post-closing snapshot
            def _has_activity_pc(r: dict) -> bool:
                d, c = self._balance_to_columns(r)
                return bool((d or 0) != 0 or (c or 0) != 0)
            rows = [r for r in rows if _has_activity_pc(r)]
            for r in rows:
                code = r['code'] if 'code' in r.keys() else ''
                name = r['name'] if 'name' in r.keys() else ''
                d, c = self._balance_to_columns(r)
                self.pctb_tree.insert('', 'end', values=(code, name, f"{d:,.2f}" if d else '', f"{c:,.2f}" if c else ''))
            try:
                total_d = 0.0
                total_c = 0.0
                for iid in self.pctb_tree.get_children():
                    vals = self.pctb_tree.item(iid, 'values')
                    try:
                        if vals and vals[2]:
                            total_d += float(str(vals[2]).replace(',', ''))
                    except Exception:
                        pass
                    try:
                        if vals and vals[3]:
                            total_c += float(str(vals[3]).replace(',', ''))
                    except Exception:
                        pass
                self.pctb_tree.insert('', 'end', values=("", "Totals", f"{total_d:,.2f}", f"{total_c:,.2f}"), tags=('totals',))
                self.pctb_tree.tag_configure('totals', background=self.palette.get('tab_selected_bg', '#e0ecff'))
            except Exception:
                pass
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load post-closing TB: {e}")

    def _complete_postclosing_tb_action(self) -> None:
        try:
            as_of = (self.pctb_date.get().strip() if hasattr(self, 'pctb_date') else '') or None
            rows = db.compute_trial_balance(up_to_date=as_of, include_temporary=False, period_id=self.engine.current_period_id, conn=self.engine.conn)
            self._load_postclosing_tb()
            self.engine.capture_trial_balance_snapshot("post_closing", as_of or "latest", rows)
            self.engine.set_cycle_step_status(9, "completed", f"Post-closing TB prepared as of {as_of or 'latest'}")
            self.engine.set_cycle_step_status(10, "in_progress", "Schedule reversing entries pending")
            self._load_cycle_status()
            messagebox.showinfo("Completed", "Post-closing trial balance marked completed")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to complete post-closing TB: {e}")


if __name__ == "__main__":
    app = TechFixApp()
    app.mainloop()
